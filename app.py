import hashlib
import math
import os
import secrets
import threading
import time
import traceback
import urllib.request
from datetime import date
from functools import wraps

import psycopg
from psycopg.rows import dict_row
from flask import (
    Flask, request, jsonify, render_template,
    session, redirect, url_for, abort
)
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage,
    PostbackEvent, LocationMessage
)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

LINE_CHANNEL_ACCESS_TOKEN = os.environ.get('LINE_CHANNEL_ACCESS_TOKEN', '')
LINE_CHANNEL_SECRET       = os.environ.get('LINE_CHANNEL_SECRET', '')
ADMIN_PASSWORD            = os.environ.get('ADMIN_PASSWORD', 'admin123')
_raw_db_url               = os.environ.get('DATABASE_URL', '')
DATABASE_URL              = _raw_db_url.replace('postgres://', 'postgresql://', 1) if _raw_db_url.startswith('postgres://') else _raw_db_url
RENDER_EXTERNAL_URL       = os.environ.get('RENDER_EXTERNAL_URL', '')

print(f"[startup] DATABASE_URL prefix: {DATABASE_URL[:20] if DATABASE_URL else 'NOT SET'}")

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler      = WebhookHandler(LINE_CHANNEL_SECRET)

# ─── Imports ──────────────────────────────────────────────────────────────────
import json as _json
from datetime import datetime as _dt, timedelta as _td, timezone as _tz

TW_TZ = _tz(_td(hours=8))   # Asia/Taipei (UTC+8)

WEEKDAY_ZH = ['一', '二', '三', '四', '五', '六', '日']

# ─── PostgreSQL ───────────────────────────────────────────────────────────────

def get_db():
    return psycopg.connect(DATABASE_URL, row_factory=dict_row)

def _hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()


def init_db():
    if not DATABASE_URL:
        print("[WARNING] DATABASE_URL not set — skipping init_db()")
        return
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_staff (
                    id              SERIAL PRIMARY KEY,
                    name            TEXT NOT NULL UNIQUE,
                    username        TEXT UNIQUE,
                    password_hash   TEXT DEFAULT '',
                    role            TEXT DEFAULT '',
                    active          BOOLEAN DEFAULT TRUE,
                    employee_code   TEXT DEFAULT '',
                    department      TEXT DEFAULT '',
                    position_title  TEXT DEFAULT '',
                    hire_date       DATE,
                    birth_date      DATE,
                    base_salary     NUMERIC(12,2) DEFAULT 0,
                    insured_salary  NUMERIC(12,2) DEFAULT 0,
                    daily_hours     NUMERIC(4,1) DEFAULT 8,
                    ot_rate1        NUMERIC(4,2) DEFAULT 1.33,
                    ot_rate2        NUMERIC(4,2) DEFAULT 1.67,
                    salary_type     TEXT DEFAULT 'monthly',
                    hourly_rate     NUMERIC(12,2) DEFAULT 0,
                    vacation_quota  INT DEFAULT NULL,
                    salary_notes    TEXT DEFAULT '',
                    line_user_id    TEXT,
                    bind_code       TEXT,
                    created_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_records (
                    id            SERIAL PRIMARY KEY,
                    staff_id      INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    punch_type    TEXT NOT NULL,
                    punched_at    TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    note          TEXT DEFAULT '',
                    is_manual     BOOLEAN DEFAULT FALSE,
                    manual_by     TEXT DEFAULT '',
                    latitude      NUMERIC(10,6),
                    longitude     NUMERIC(10,6),
                    gps_distance  INT,
                    location_name TEXT DEFAULT '',
                    created_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_locations (
                    id            SERIAL PRIMARY KEY,
                    location_name TEXT NOT NULL DEFAULT '打卡地點',
                    lat           NUMERIC(10,6) NOT NULL,
                    lng           NUMERIC(10,6) NOT NULL,
                    radius_m      INT DEFAULT 100,
                    active        BOOLEAN DEFAULT TRUE,
                    created_at    TIMESTAMPTZ DEFAULT NOW(),
                    updated_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_config (
                    id           INT PRIMARY KEY DEFAULT 1,
                    gps_required BOOLEAN DEFAULT FALSE,
                    updated_at   TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                INSERT INTO punch_config (id, gps_required)
                VALUES (1, FALSE)
                ON CONFLICT (id) DO NOTHING
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS line_punch_config (
                    id                   INT PRIMARY KEY DEFAULT 1,
                    channel_access_token TEXT DEFAULT '',
                    channel_secret       TEXT DEFAULT '',
                    enabled              BOOLEAN DEFAULT FALSE,
                    updated_at           TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                INSERT INTO line_punch_config (id)
                VALUES (1)
                ON CONFLICT (id) DO NOTHING
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS schedule_config (
                    month           TEXT PRIMARY KEY,
                    max_off_per_day INT DEFAULT 2,
                    vacation_quota  INT DEFAULT 8,
                    notes           TEXT DEFAULT '',
                    updated_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS schedule_requests (
                    id           SERIAL PRIMARY KEY,
                    staff_id     INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    month        TEXT NOT NULL,
                    dates        JSONB NOT NULL DEFAULT '[]',
                    status       TEXT DEFAULT 'pending',
                    submit_note  TEXT DEFAULT '',
                    reviewed_by  TEXT DEFAULT '',
                    reviewed_at  TIMESTAMPTZ,
                    review_note  TEXT DEFAULT '',
                    created_at   TIMESTAMPTZ DEFAULT NOW(),
                    updated_at   TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(staff_id, month)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS punch_requests (
                    id            SERIAL PRIMARY KEY,
                    staff_id      INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    punch_type    TEXT NOT NULL,
                    requested_at  TIMESTAMPTZ NOT NULL,
                    reason        TEXT DEFAULT '',
                    status        TEXT DEFAULT 'pending',
                    reviewed_by   TEXT DEFAULT '',
                    review_note   TEXT DEFAULT '',
                    reviewed_at   TIMESTAMPTZ,
                    created_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS shift_types (
                    id          SERIAL PRIMARY KEY,
                    name        TEXT NOT NULL,
                    start_time  TIME NOT NULL,
                    end_time    TIME NOT NULL,
                    color       TEXT DEFAULT '#4a7bda',
                    departments TEXT DEFAULT '',
                    active      BOOLEAN DEFAULT TRUE,
                    sort_order  INT DEFAULT 0,
                    created_at  TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS shift_assignments (
                    id            SERIAL PRIMARY KEY,
                    staff_id      INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    shift_type_id INT REFERENCES shift_types(id) ON DELETE CASCADE,
                    shift_date    DATE NOT NULL,
                    note          TEXT DEFAULT '',
                    created_at    TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(staff_id, shift_date)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS overtime_requests (
                    id              SERIAL PRIMARY KEY,
                    staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    request_date    DATE NOT NULL,
                    start_time      TIME NOT NULL,
                    end_time        TIME NOT NULL,
                    ot_hours        NUMERIC(5,2),
                    reason          TEXT DEFAULT '',
                    status          TEXT DEFAULT 'pending',
                    reviewed_by     TEXT DEFAULT '',
                    review_note     TEXT DEFAULT '',
                    ot_pay          NUMERIC(12,2) DEFAULT 0,
                    day_type        TEXT DEFAULT 'weekday',
                    reviewed_at     TIMESTAMPTZ,
                    created_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)

            # Seed default shifts if empty
            existing_shifts = conn.execute("SELECT COUNT(*) as cnt FROM shift_types").fetchone()
            if existing_shifts['cnt'] == 0:
                defaults = [
                    ('吧台班',  '08:00', '16:00', '#8b5cf6', '吧台', 1),
                    ('外場A班', '09:00', '17:00', '#2e9e6b', '外場', 2),
                    ('外場B班', '14:00', '22:00', '#0ea5e9', '外場', 3),
                    ('廚房A班', '08:00', '16:00', '#e07b2a', '廚房', 4),
                    ('廚房B班', '12:00', '20:00', '#d64242', '廚房', 5),
                ]
                for name, st, et, color, dept, sort in defaults:
                    conn.execute(
                        "INSERT INTO shift_types (name,start_time,end_time,color,departments,sort_order) VALUES (%s,%s,%s,%s,%s,%s)",
                        (name, st, et, color, dept, sort)
                    )

        print("[OK] Database tables created")
    except Exception as e:
        print(f"[ERROR] init_db failed: {e}")
        raise

    # Schema migrations (each in its own connection to avoid transaction abort)
    migrations = [
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS username TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS password_hash TEXT DEFAULT ''",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS latitude NUMERIC(10,6)",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS longitude NUMERIC(10,6)",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS gps_distance INT",
        "ALTER TABLE punch_records ADD COLUMN IF NOT EXISTS location_name TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS line_user_id TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bind_code TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS employee_code TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS department TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS position_title TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS hire_date DATE",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS birth_date DATE",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS base_salary NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS insured_salary NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_notes TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS daily_hours NUMERIC(4,1) DEFAULT 8",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS ot_rate1 NUMERIC(4,2) DEFAULT 1.33",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS ot_rate2 NUMERIC(4,2) DEFAULT 1.67",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS ot_rate3 NUMERIC(4,2) DEFAULT 2.0",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS document_id INT REFERENCES finance_documents(id) ON DELETE SET NULL",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS leave_start_time TEXT DEFAULT NULL",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS leave_end_time TEXT DEFAULT NULL",
        "ALTER TABLE finance_documents ADD COLUMN IF NOT EXISTS image_data TEXT",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_type TEXT DEFAULT 'monthly'",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS hourly_rate NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS vacation_quota INT DEFAULT NULL",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_code TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_name TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_branch TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS bank_account TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS account_holder TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS password_plain TEXT DEFAULT ''",
        "ALTER TABLE overtime_requests ADD COLUMN IF NOT EXISTS day_type TEXT DEFAULT 'weekday'",
        "ALTER TABLE overtime_requests ALTER COLUMN start_time DROP NOT NULL",
        "ALTER TABLE overtime_requests ALTER COLUMN end_time DROP NOT NULL",
        # 員工個人/保險欄位
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS national_id TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS gender TEXT DEFAULT ''",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS insurance_type TEXT DEFAULT 'regular'",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS address TEXT DEFAULT ''",
        # 多店
        """CREATE TABLE IF NOT EXISTS stores (
            id         SERIAL PRIMARY KEY,
            name       TEXT NOT NULL,
            code       TEXT UNIQUE,
            address    TEXT DEFAULT '',
            active     BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS store_id INT REFERENCES stores(id) ON DELETE SET NULL",
        "ALTER TABLE punch_locations ADD COLUMN IF NOT EXISTS store_id INT REFERENCES stores(id) ON DELETE SET NULL",
        "ALTER TABLE admin_accounts ADD COLUMN IF NOT EXISTS store_ids JSONB DEFAULT '[]'",
        "ALTER TABLE admin_accounts ADD COLUMN IF NOT EXISTS password_plain TEXT DEFAULT ''",
        "ALTER TABLE schedule_requests ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()",
        """CREATE TABLE IF NOT EXISTS shift_staffing_requirements (
            id            SERIAL PRIMARY KEY,
            shift_type_id INT REFERENCES shift_types(id) ON DELETE CASCADE,
            day_of_week   SMALLINT NOT NULL,
            required_count INT NOT NULL DEFAULT 1,
            created_at    TIMESTAMPTZ DEFAULT NOW(),
            updated_at    TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(shift_type_id, day_of_week)
        )""",
        """CREATE TABLE IF NOT EXISTS admin_accounts (
            id              SERIAL PRIMARY KEY,
            username        TEXT NOT NULL UNIQUE,
            password_hash   TEXT NOT NULL,
            display_name    TEXT DEFAULT '',
            permissions     JSONB DEFAULT '[]',
            is_super        BOOLEAN DEFAULT FALSE,
            active          BOOLEAN DEFAULT TRUE,
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            last_login_at   TIMESTAMPTZ
        )""",
        # 效能索引
        "CREATE INDEX IF NOT EXISTS idx_punch_records_staff_time   ON punch_records(staff_id, punched_at)",
        "CREATE INDEX IF NOT EXISTS idx_shift_assignments_staff_date ON shift_assignments(staff_id, shift_date)",
        "CREATE INDEX IF NOT EXISTS idx_leave_requests_staff_status ON leave_requests(staff_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_leave_requests_staff_date   ON leave_requests(staff_id, start_date)",
        "CREATE INDEX IF NOT EXISTS idx_overtime_requests_staff     ON overtime_requests(staff_id, status)",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS sort_order INT DEFAULT 0",
    ]
    for sql in migrations:
        try:
            with get_db() as mc:
                mc.execute(sql)
        except Exception as me:
            print(f"[MIGRATION SKIP] {sql[:70]}: {me}")

    # Seed default super admin; always sync password from ADMIN_PASSWORD env var
    try:
        all_modules = _json.dumps(['punch','sched','leave','salary','ann','holiday','finance'])
        pw_hash = _hash_pw(ADMIN_PASSWORD)
        with get_db() as conn:
            existing = conn.execute(
                "SELECT id FROM admin_accounts WHERE username='admin'"
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE admin_accounts SET password_hash=%s, is_super=TRUE WHERE username='admin'",
                    (pw_hash,)
                )
                print("[OK] admin password synced from ADMIN_PASSWORD env var")
            else:
                conn.execute("""
                    INSERT INTO admin_accounts (username, password_hash, display_name, permissions, is_super)
                    VALUES (%s,%s,'超級管理員',%s,TRUE)
                """, ('admin', pw_hash, all_modules))
                print("[OK] Default super admin seeded (username: admin)")
    except Exception as e:
        print(f"[WARN] admin seed: {e}")

    # 確保預設店家存在，並補齊舊資料
    try:
        with get_db() as conn:
            conn.execute("INSERT INTO stores (name, code) VALUES ('主店','main') ON CONFLICT (code) DO NOTHING")
            conn.execute("UPDATE punch_staff     SET store_id=(SELECT id FROM stores WHERE code='main') WHERE store_id IS NULL")
            conn.execute("UPDATE punch_locations SET store_id=(SELECT id FROM stores WHERE code='main') WHERE store_id IS NULL")
    except Exception as e:
        print(f"[WARN] store seed: {e}")

    print("[OK] Database initialised")


init_db()

# ─── Keep-Alive ───────────────────────────────────────────────────────────────

def keep_alive():
    time.sleep(10)
    while True:
        try:
            base = RENDER_EXTERNAL_URL.rstrip('/') if RENDER_EXTERNAL_URL else 'http://localhost:5000'
            urllib.request.urlopen(
                urllib.request.Request(f'{base}/health', headers={'User-Agent': 'KeepAlive/1.0'}),
                timeout=10
            )
        except Exception as e:
            print(f"[keep-alive] ping failed: {e}")
        time.sleep(14 * 60)

threading.Thread(target=keep_alive, daemon=True).start()


# ─── 特休自動同步 ─────────────────────────────────────────────────────────────

def _run_annual_leave_sync():
    """依勞基法第38條，依到職日計算特休天數，寫入 leave_balances。每日午夜自動執行。"""
    from datetime import date as _d_sync
    import json as _json_sync
    year = str(_d_sync.today().year)
    try:
        with get_db() as conn:
            staff_list = conn.execute(
                "SELECT id, name, hire_date FROM punch_staff WHERE active=TRUE AND hire_date IS NOT NULL"
            ).fetchall()
            lt = conn.execute("SELECT id FROM leave_types WHERE code='annual'").fetchone()
            if not lt:
                return
            lt_id = lt['id']
            for s in staff_list:
                days = _calc_annual_leave_days(s['hire_date'])
                conn.execute("""
                    INSERT INTO leave_balances (staff_id, leave_type_id, year, total_days, used_days)
                    VALUES (%s,%s,%s,%s,0)
                    ON CONFLICT (staff_id, leave_type_id, year) DO UPDATE
                      SET total_days=EXCLUDED.total_days, updated_at=NOW()
                """, (s['id'], lt_id, int(year), days))
    except Exception as e:
        print(f"[annual_leave_sync] {e}")


def _annual_leave_sync_loop():
    import time as _time_sync
    # 啟動時立即執行一次
    _run_annual_leave_sync()
    while True:
        # 計算距離明天 00:05 台北時間的秒數
        now = _dt.now(TW_TZ)
        tmr = (now + _td(days=1)).date()
        tomorrow_05 = _dt(tmr.year, tmr.month, tmr.day, 0, 5, tzinfo=TW_TZ)
        sleep_secs = (tomorrow_05 - now).total_seconds()
        if sleep_secs < 0:
            sleep_secs = 3600
        _time_sync.sleep(sleep_secs)
        _run_annual_leave_sync()


threading.Thread(target=_annual_leave_sync_loop, daemon=True).start()


# ─── Health ───────────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    try:
        with get_db() as conn:
            conn.execute('SELECT 1')
        return jsonify({'status': 'ok', 'db': 'connected'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'detail': str(e)}), 500

# ─── Admin Auth ───────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': '請先登入'}), 401
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

def require_module(module):
    """確認已登入且擁有指定模組權限（超級管理員跳過模組檢查）。"""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get('logged_in'):
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'error': '請先登入'}), 401
                return redirect(url_for('admin_login'))
            if not session.get('admin_is_super'):
                perms = session.get('admin_permissions') or []
                if module not in perms:
                    return jsonify({'error': f'無「{module}」模組的存取權限'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

def require_super(f):
    """只允許超級管理員存取。"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'error': '請先登入'}), 401
        if not session.get('admin_is_super'):
            return jsonify({'error': '需要超級管理員權限'}), 403
        return f(*args, **kwargs)
    return decorated

@app.route('/')
def index():
    return redirect(url_for('admin_login'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            error = '請輸入帳號與密碼'
        else:
            with get_db() as conn:
                row = conn.execute(
                    "SELECT * FROM admin_accounts WHERE username=%s AND active=TRUE",
                    (username,)
                ).fetchone()
            if row and row['password_hash'] == _hash_pw(password):
                perms = row['permissions']
                if isinstance(perms, str):
                    try: perms = _json.loads(perms)
                    except: perms = []
                session['logged_in']          = True
                session['admin_id']           = row['id']
                session['admin_username']     = row['username']
                session['admin_display_name'] = row['display_name'] or row['username']
                session['admin_permissions']  = perms
                session['admin_is_super']     = bool(row['is_super'])
                with get_db() as conn:
                    conn.execute("UPDATE admin_accounts SET last_login_at=NOW() WHERE id=%s", (row['id'],))
                return redirect(url_for('admin_dashboard'))
            error = '帳號或密碼錯誤'
    return render_template('login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin')
@app.route('/admin/')
@login_required
def admin_dashboard():
    perms    = session.get('admin_permissions') or []
    is_super = bool(session.get('admin_is_super'))
    return render_template('admin.html',
        admin_display_name=session.get('admin_display_name',''),
        admin_permissions=perms,
        admin_is_super=is_super,
    )

# ── Admin Accounts API ────────────────────────────────────────────────────────

def _admin_row(r):
    if not r: return None
    d = dict(r)
    d.pop('password_hash', None)
    if d.get('password_plain') is None: d['password_plain'] = ''
    perms = d.get('permissions')
    if isinstance(perms, str):
        try: d['permissions'] = _json.loads(perms)
        except: d['permissions'] = []
    if d.get('created_at'):   d['created_at']   = d['created_at'].isoformat()
    if d.get('last_login_at'): d['last_login_at'] = d['last_login_at'].isoformat()
    return d

@app.route('/api/admin/me', methods=['GET'])
@login_required
def api_admin_me():
    return jsonify({
        'id':           session.get('admin_id'),
        'username':     session.get('admin_username'),
        'display_name': session.get('admin_display_name'),
        'permissions':  session.get('admin_permissions') or [],
        'is_super':     bool(session.get('admin_is_super')),
    })

@app.route('/api/admin/accounts', methods=['GET'])
@require_super
def api_admin_accounts_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM admin_accounts ORDER BY id").fetchall()
    return jsonify([_admin_row(r) for r in rows])

@app.route('/api/admin/accounts', methods=['POST'])
@require_super
def api_admin_account_create():
    b = request.get_json(force=True)
    username = b.get('username','').strip()
    password = b.get('password','').strip()
    if not username: return jsonify({'error': '帳號為必填'}), 400
    if not password or len(password) < 4: return jsonify({'error': '密碼至少 4 個字元'}), 400
    perms = b.get('permissions', [])
    with get_db() as conn:
        try:
            row = conn.execute("""
                INSERT INTO admin_accounts (username, password_hash, password_plain, display_name, permissions, is_super, active)
                VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
            """, (username, _hash_pw(password), password, b.get('display_name','').strip(),
                  _json.dumps(perms), bool(b.get('is_super', False)), True)).fetchone()
        except Exception as e:
            if 'unique' in str(e).lower(): return jsonify({'error': '帳號已存在'}), 409
            return jsonify({'error': str(e)}), 500
    return jsonify(_admin_row(row)), 201

@app.route('/api/admin/accounts/<int:aid>', methods=['PUT'])
@require_super
def api_admin_account_update(aid):
    b = request.get_json(force=True)
    username = b.get('username','').strip()
    if not username: return jsonify({'error': '帳號為必填'}), 400
    password = b.get('password','').strip()
    perms = b.get('permissions', [])
    with get_db() as conn:
        if password:
            if len(password) < 4: return jsonify({'error': '密碼至少 4 個字元'}), 400
            row = conn.execute("""
                UPDATE admin_accounts SET username=%s, password_hash=%s, password_plain=%s, display_name=%s,
                  permissions=%s, is_super=%s, active=%s WHERE id=%s RETURNING *
            """, (username, _hash_pw(password), password, b.get('display_name','').strip(),
                  _json.dumps(perms), bool(b.get('is_super', False)),
                  bool(b.get('active', True)), aid)).fetchone()
        else:
            row = conn.execute("""
                UPDATE admin_accounts SET username=%s, display_name=%s,
                  permissions=%s, is_super=%s, active=%s WHERE id=%s RETURNING *
            """, (username, b.get('display_name','').strip(),
                  _json.dumps(perms), bool(b.get('is_super', False)),
                  bool(b.get('active', True)), aid)).fetchone()
    return jsonify(_admin_row(row)) if row else ('', 404)

@app.route('/api/admin/accounts/<int:aid>', methods=['DELETE'])
@require_super
def api_admin_account_delete(aid):
    if aid == session.get('admin_id'):
        return jsonify({'error': '不能刪除自己的帳號'}), 400
    with get_db() as conn:
        conn.execute("DELETE FROM admin_accounts WHERE id=%s", (aid,))
    return jsonify({'deleted': aid})

# ─── Shared Helpers ───────────────────────────────────────────────────────────

def _gps_distance(lat1, lng1, lat2, lng2):
    R = 6371000
    p = math.pi / 180
    a = (math.sin((lat2 - lat1) * p / 2) ** 2 +
         math.cos(lat1 * p) * math.cos(lat2 * p) *
         math.sin((lng2 - lng1) * p / 2) ** 2)
    return int(2 * R * math.asin(math.sqrt(a)))

def punch_staff_row(row):
    if not row: return None
    d = dict(row)
    d.pop('password_hash', None)
    # password_plain is kept so admin can view it; keep as empty string if null
    if d.get('password_plain') is None: d['password_plain'] = ''
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('hire_date'):  d['hire_date']  = d['hire_date'].isoformat()
    if d.get('birth_date'): d['birth_date'] = d['birth_date'].isoformat()
    return d

def _parse_tw_datetime(s):
    """Parse a datetime string (with or without tz) treating naive strings as Taiwan time (UTC+8)."""
    if not s:
        return None
    dt = _dt.fromisoformat(str(s).replace('Z', '+00:00'))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TW_TZ)
    return dt


def punch_record_row(row):
    if not row: return None
    d = dict(row)
    for f in ['latitude', 'longitude']:
        if d.get(f) is not None: d[f] = float(d[f])
    for f in ['punched_at', 'created_at']:
        if d.get(f):
            dt = d[f]
            if dt.tzinfo is None:
                from datetime import timezone as _utz
                dt = dt.replace(tzinfo=_utz.utc)
            d[f] = dt.astimezone(TW_TZ).isoformat()
    return d

def loc_row(row):
    if not row: return None
    d = dict(row)
    for f in ['lat', 'lng']:
        if d.get(f) is not None: d[f] = float(d[f])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    return d

def punch_req_row(row):
    if not row: return None
    d = dict(row)
    if d.get('requested_at'): d['requested_at'] = d['requested_at'].isoformat()
    if d.get('reviewed_at'):  d['reviewed_at']  = d['reviewed_at'].isoformat()
    if d.get('created_at'):   d['created_at']   = d['created_at'].isoformat()
    return d

def ot_req_row(row):
    if not row: return None
    d = dict(row)
    if d.get('request_date'): d['request_date'] = d['request_date'].isoformat()
    if d.get('start_time'):   d['start_time']   = str(d['start_time'])[:5]
    if d.get('end_time'):     d['end_time']      = str(d['end_time'])[:5]
    if d.get('ot_pay'):       d['ot_pay']        = float(d['ot_pay'])
    if d.get('ot_hours'):     d['ot_hours']      = float(d['ot_hours'])
    if d.get('reviewed_at'):  d['reviewed_at']   = d['reviewed_at'].isoformat()
    if d.get('created_at'):   d['created_at']    = d['created_at'].isoformat()
    return d

def shift_type_row(row):
    if not row: return None
    d = dict(row)
    if d.get('start_time'): d['start_time'] = str(d['start_time'])[:5]
    if d.get('end_time'):   d['end_time']   = str(d['end_time'])[:5]
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

def shift_assign_row(row):
    if not row: return None
    d = dict(row)
    if d.get('shift_date'): d['shift_date'] = d['shift_date'].isoformat()
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

def sched_req_row(row):
    if not row: return None
    d = dict(row)
    if isinstance(d.get('dates'), str):
        try: d['dates'] = _json.loads(d['dates'])
        except: d['dates'] = []
    if d.get('reviewed_at'): d['reviewed_at'] = d['reviewed_at'].isoformat()
    if d.get('created_at'):  d['created_at']  = d['created_at'].isoformat()
    if d.get('updated_at'):  d['updated_at']  = d['updated_at'].isoformat()
    return d

def get_schedule_config(conn, month):
    row = conn.execute("SELECT * FROM schedule_config WHERE month=%s", (month,)).fetchone()
    if not row:
        return {'month': month, 'max_off_per_day': 2, 'vacation_quota': 8, 'notes': ''}
    return dict(row)

def get_off_counts(conn, month):
    rows = conn.execute("""
        SELECT elem as d, COUNT(*) as cnt
        FROM schedule_requests,
             jsonb_array_elements_text(dates) as elem
        WHERE month=%s AND status IN ('approved','pending')
        GROUP BY elem
    """, (month,)).fetchall()
    return {r['d']: int(r['cnt']) for r in rows}

# ═══════════════════════════════════════════════════════════════════
# Employee Punch Page
# ═══════════════════════════════════════════════════════════════════

@app.route('/punch')
@app.route('/staff')
def punch_page():
    return render_template('staff.html')

# ── Employee Session ──────────────────────────────────────────────

@app.route('/api/punch/login', methods=['POST'])
def api_punch_login():
    b = request.get_json(force=True)
    username = b.get('username', '').strip()
    password = b.get('password', '').strip()
    if not username or not password:
        return jsonify({'error': '請輸入帳號及密碼'}), 400
    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE username=%s AND active=TRUE", (username,)
        ).fetchone()
    if not staff or staff['password_hash'] != _hash_pw(password):
        return jsonify({'error': '帳號或密碼錯誤'}), 401
    session['punch_staff_id']   = staff['id']
    session['punch_staff_name'] = staff['name']
    return jsonify({'id': staff['id'], 'name': staff['name'], 'role': staff['role']})

@app.route('/api/punch/logout', methods=['POST'])
def api_punch_logout():
    session.pop('punch_staff_id', None)
    session.pop('punch_staff_name', None)
    return jsonify({'ok': True})

@app.route('/api/punch/me', methods=['GET'])
def api_punch_me():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        staff = conn.execute(
            "SELECT id,name,role FROM punch_staff WHERE id=%s AND active=TRUE", (sid,)
        ).fetchone()
    if not staff:
        session.pop('punch_staff_id', None)
        return jsonify({'error': 'not logged in'}), 401
    return jsonify(dict(staff))

# ── GPS Settings ──────────────────────────────────────────────────

@app.route('/api/punch/settings', methods=['GET'])
def api_punch_settings_get():
    """Public: GPS config + active locations for the punch page."""
    with get_db() as conn:
        cfg  = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
        locs = conn.execute(
            "SELECT * FROM punch_locations WHERE active=TRUE ORDER BY id"
        ).fetchall()
    return jsonify({
        'gps_required': cfg['gps_required'] if cfg else False,
        'locations': [loc_row(r) for r in locs]
    })

@app.route('/api/punch/config', methods=['PUT'])
@login_required
def api_punch_config_update():
    b = request.get_json(force=True)
    gps_required = bool(b.get('gps_required', False))
    with get_db() as conn:
        conn.execute(
            "UPDATE punch_config SET gps_required=%s, updated_at=NOW() WHERE id=1",
            (gps_required,)
        )
    return jsonify({'gps_required': gps_required})

@app.route('/api/punch/locations', methods=['GET'])
@login_required
def api_punch_locations_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM punch_locations ORDER BY id").fetchall()
    return jsonify([loc_row(r) for r in rows])

@app.route('/api/punch/locations', methods=['POST'])
@login_required
def api_punch_locations_create():
    b = request.get_json(force=True)
    name = b.get('location_name', '').strip() or '打卡地點'
    try:
        lat = float(b['lat']); lng = float(b['lng'])
    except Exception:
        return jsonify({'error': '請填入有效的緯度和經度'}), 400
    radius_m = int(b.get('radius_m') or 100)
    with get_db() as conn:
        row = conn.execute(
            "INSERT INTO punch_locations (location_name, lat, lng, radius_m) VALUES (%s,%s,%s,%s) RETURNING *",
            (name, lat, lng, radius_m)
        ).fetchone()
    return jsonify(loc_row(row)), 201

@app.route('/api/punch/locations/<int:lid>', methods=['PUT'])
@login_required
def api_punch_locations_update(lid):
    b = request.get_json(force=True)
    name = b.get('location_name', '').strip() or '打卡地點'
    try:
        lat = float(b['lat']); lng = float(b['lng'])
    except Exception:
        return jsonify({'error': '請填入有效的緯度和經度'}), 400
    radius_m = int(b.get('radius_m') or 100)
    active   = bool(b.get('active', True))
    with get_db() as conn:
        row = conn.execute(
            "UPDATE punch_locations SET location_name=%s,lat=%s,lng=%s,radius_m=%s,active=%s,updated_at=NOW() WHERE id=%s RETURNING *",
            (name, lat, lng, radius_m, active, lid)
        ).fetchone()
    return jsonify(loc_row(row)) if row else ('', 404)

@app.route('/api/punch/locations/<int:lid>', methods=['DELETE'])
@login_required
def api_punch_locations_delete(lid):
    with get_db() as conn:
        conn.execute("DELETE FROM punch_locations WHERE id=%s", (lid,))
    return jsonify({'deleted': lid})

# ── Clock In/Out ──────────────────────────────────────────────────

@app.route('/api/punch/clock', methods=['POST'])
def api_punch_clock():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': '請先登入'}), 401

    b          = request.get_json(force=True)
    punch_type = b.get('punch_type')
    lat        = b.get('lat')
    lng        = b.get('lng')

    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400

    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE id=%s AND active=TRUE", (sid,)
        ).fetchone()
        if not staff:
            return jsonify({'error': '員工不存在'}), 404
        cfg  = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
        locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()

    gps_required = cfg['gps_required'] if cfg else False
    gps_distance = None
    matched_loc  = None

    if lat is not None and lng is not None and locs:
        for loc in locs:
            d = _gps_distance(lat, lng, float(loc['lat']), float(loc['lng']))
            if gps_distance is None or d < gps_distance:
                gps_distance = d
                matched_loc  = loc

    if gps_required:
        if lat is None or lng is None:
            return jsonify({'error': '無法取得 GPS，請允許定位權限後重試'}), 403
        if not locs:
            return jsonify({'error': '管理員尚未設定任何打卡地點'}), 403
        if gps_distance is None or gps_distance > int(matched_loc['radius_m']):
            return jsonify({
                'error': f'距離最近地點「{matched_loc["location_name"]}」{gps_distance} 公尺，超出允許範圍（{matched_loc["radius_m"]} 公尺）',
                'distance': gps_distance,
                'radius': int(matched_loc['radius_m'])
            }), 403

    with get_db() as conn:
        recent = conn.execute("""
            SELECT id FROM punch_records
            WHERE staff_id=%s AND punch_type=%s
              AND punched_at > NOW() - INTERVAL '1 minute'
        """, (sid, punch_type)).fetchone()
        if recent:
            return jsonify({'error': '1 分鐘內已打過卡'}), 429

        matched_name = matched_loc['location_name'] if matched_loc else ''
        row = conn.execute("""
            INSERT INTO punch_records
              (staff_id, punch_type, latitude, longitude, gps_distance, location_name)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (sid, punch_type, lat, lng, gps_distance, matched_name)).fetchone()

    d = punch_record_row(row)
    d['staff_name']   = staff['name']
    d['gps_distance'] = gps_distance
    return jsonify(d), 201

@app.route('/api/punch/today', methods=['GET'])
def api_punch_today():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify([])
    with get_db() as conn:
        rows = conn.execute("""
            SELECT pr.*, ps.name as staff_name
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE pr.staff_id=%s
              AND (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
                = (NOW() AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY pr.punched_at ASC
        """, (sid,)).fetchall()
    return jsonify([punch_record_row(r) for r in rows])

@app.route('/api/punch/my-records', methods=['GET'])
def api_punch_my_records():
    """Employee self-service: own punch records for a month."""
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': 'not logged in'}), 401
    month = request.args.get('month', '')
    if not month:
        from datetime import timezone as _tz, timedelta as _tda
        month = _dt.now(_tz(_tda(hours=8))).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT punch_type, punched_at, gps_distance, location_name, is_manual
            FROM punch_records
            WHERE staff_id=%s
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei', 'YYYY-MM') = %s
            ORDER BY punched_at ASC
        """, (sid, month)).fetchall()
    from datetime import timezone as _tz2, timedelta as _tdb
    TW = _tz2(_tdb(hours=8))
    LABEL = {'in': '上班', 'out': '下班', 'break_out': '休息開始', 'break_in': '休息結束'}
    result = {}
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            from datetime import timezone as _utz
            pa = pa.replace(tzinfo=_utz.utc)
        pa_tw    = pa.astimezone(TW)
        date_str = pa_tw.strftime('%Y-%m-%d')
        time_str = pa_tw.strftime('%H:%M')
        if date_str not in result:
            result[date_str] = []
        result[date_str].append({
            'type':          r['punch_type'],
            'label':         LABEL.get(r['punch_type'], r['punch_type']),
            'time':          time_str,
            'gps_distance':  r['gps_distance'],
            'location_name': r['location_name'] or '',
            'is_manual':     bool(r['is_manual']),
        })
    return jsonify({'month': month, 'records': result})

# ── Admin: Staff CRUD ─────────────────────────────────────────────

@app.route('/api/punch/staff', methods=['GET'])
@login_required
def api_punch_staff_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM punch_staff ORDER BY sort_order, name").fetchall()
    return jsonify([punch_staff_row(r) for r in rows])

@app.route('/api/punch/staff/reorder', methods=['POST'])
@login_required
def api_punch_staff_reorder():
    """批次更新員工排列順序，接受 [{id, sort_order}, ...]"""
    items = request.get_json(force=True) or []
    if not isinstance(items, list):
        return jsonify({'error': '格式錯誤'}), 400
    with get_db() as conn:
        for item in items:
            conn.execute(
                "UPDATE punch_staff SET sort_order=%s WHERE id=%s",
                (int(item.get('sort_order', 0)), int(item['id']))
            )
    return jsonify({'ok': True})

@app.route('/api/punch/staff', methods=['POST'])
@login_required
def api_punch_staff_create():
    b        = request.get_json(force=True)
    name     = b.get('name', '').strip()
    username = b.get('username', '').strip()
    password = b.get('password', '').strip()
    if not name:     return jsonify({'error': '姓名為必填'}), 400
    if not username: return jsonify({'error': '帳號為必填'}), 400
    if not password:
        return jsonify({'error': '請設定密碼'}), 400
    employee_code = b.get('employee_code', '') or None
    if employee_code: employee_code = employee_code.strip() or None
    department     = (b.get('department') or '').strip()
    role           = b.get('role', '').strip()
    hire_date      = b.get('hire_date') or None
    birth_date     = b.get('birth_date') or None
    bank_code      = (b.get('bank_code') or '').strip()
    bank_name      = (b.get('bank_name') or '').strip()
    bank_branch    = (b.get('bank_branch') or '').strip()
    bank_account   = (b.get('bank_account') or '').strip()
    account_holder = (b.get('account_holder') or '').strip()
    try:
        with get_db() as conn:
            row = conn.execute("""
                INSERT INTO punch_staff
                  (name, username, password_hash, password_plain, role, position_title, employee_code,
                   department, hire_date, birth_date,
                   bank_code, bank_name, bank_branch, bank_account, account_holder)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
            """, (name, username, _hash_pw(password), password, role, role, employee_code,
                  department, hire_date, birth_date,
                  bank_code, bank_name, bank_branch, bank_account, account_holder)).fetchone()
        return jsonify(punch_staff_row(row)), 201
    except psycopg.errors.UniqueViolation:
        return jsonify({'error': '姓名或帳號已存在，請換一個'}), 409
    except Exception as e:
        print(f"[punch_staff_create] error: {e}")
        # Check if it's a unique constraint in the error message
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': '姓名或帳號已存在，請換一個'}), 409
        return jsonify({'error': f'新增失敗：{str(e)}'}), 500

@app.route('/api/punch/staff/<int:sid>', methods=['PUT'])
@login_required
def api_punch_staff_update(sid):
    b             = request.get_json(force=True)
    name          = b.get('name', '').strip()
    username      = b.get('username', '').strip()
    password      = b.get('password', '').strip()
    role          = b.get('role', '').strip()
    active        = bool(b.get('active', True))
    employee_code = b.get('employee_code', '') or None
    if employee_code: employee_code = employee_code.strip() or None
    bank_code      = (b.get('bank_code') or '').strip()
    bank_name      = (b.get('bank_name') or '').strip()
    bank_branch    = (b.get('bank_branch') or '').strip()
    bank_account   = (b.get('bank_account') or '').strip()
    account_holder = (b.get('account_holder') or '').strip()
    department     = (b.get('department') or '').strip()
    hire_date      = b.get('hire_date') or None
    birth_date     = b.get('birth_date') or None
    if not name or not username:
        return jsonify({'error': '姓名和帳號為必填'}), 400
    with get_db() as conn:
        if password:
            row = conn.execute("""
                UPDATE punch_staff
                SET name=%s,username=%s,password_hash=%s,password_plain=%s,role=%s,position_title=%s,active=%s,employee_code=%s,
                    department=%s,hire_date=%s,birth_date=%s,
                    bank_code=%s,bank_name=%s,bank_branch=%s,bank_account=%s,account_holder=%s
                WHERE id=%s RETURNING *
            """, (name, username, _hash_pw(password), password, role, role, active, employee_code,
                  department, hire_date, birth_date,
                  bank_code, bank_name, bank_branch, bank_account, account_holder, sid)).fetchone()
        else:
            row = conn.execute("""
                UPDATE punch_staff
                SET name=%s,username=%s,role=%s,position_title=%s,active=%s,employee_code=%s,
                    department=%s,hire_date=%s,birth_date=%s,
                    bank_code=%s,bank_name=%s,bank_branch=%s,bank_account=%s,account_holder=%s
                WHERE id=%s RETURNING *
            """, (name, username, role, role, active, employee_code,
                  department, hire_date, birth_date,
                  bank_code, bank_name, bank_branch, bank_account, account_holder, sid)).fetchone()
    return jsonify(punch_staff_row(row)) if row else ('', 404)

@app.route('/api/punch/staff/<int:sid>', methods=['DELETE'])
@login_required
def api_punch_staff_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM punch_staff WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

# ── Admin: Punch Records ──────────────────────────────────────────

@app.route('/api/punch/records', methods=['GET'])
@login_required
def api_punch_records():
    staff_id  = request.args.get('staff_id')
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    month     = request.args.get('month')

    conds, params = ["TRUE"], []
    if staff_id: conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    if month:
        conds.append("TO_CHAR(pr.punched_at,'YYYY-MM')=%s"); params.append(month)
    elif date_from:
        conds.append("(pr.punched_at AT TIME ZONE 'Asia/Taipei')::date>=%s"); params.append(date_from)
        if date_to:
            conds.append("(pr.punched_at AT TIME ZONE 'Asia/Taipei')::date<=%s"); params.append(date_to)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*, ps.name as staff_name, ps.role as staff_role
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.punched_at DESC LIMIT 500
        """, params).fetchall()
    return jsonify([punch_record_row(r) for r in rows])

@app.route('/api/punch/records', methods=['POST'])
@login_required
def api_punch_record_manual():
    b          = request.get_json(force=True)
    staff_id   = b.get('staff_id')
    punch_type = b.get('punch_type')
    punched_at = b.get('punched_at')
    note       = b.get('note', '').strip()
    manual_by  = b.get('manual_by', '').strip()
    if not all([staff_id, punch_type, punched_at]):
        return jsonify({'error': '缺少必要欄位'}), 400
    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400
    punched_at_parsed = _parse_tw_datetime(punched_at)
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO punch_records
              (staff_id, punch_type, punched_at, note, is_manual, manual_by)
            VALUES (%s,%s,%s,%s,TRUE,%s) RETURNING *
        """, (staff_id, punch_type, punched_at_parsed, note, manual_by)).fetchone()
        staff = conn.execute("SELECT name FROM punch_staff WHERE id=%s", (staff_id,)).fetchone()
    d = punch_record_row(row)
    if staff: d['staff_name'] = staff['name']
    return jsonify(d), 201

@app.route('/api/punch/records/<int:rid>', methods=['PUT'])
@login_required
def api_punch_record_update(rid):
    b = request.get_json(force=True)
    if b.get('punch_type') not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400
    punched_at_parsed = _parse_tw_datetime(b.get('punched_at'))
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_records
            SET punch_type=%s, punched_at=%s, note=%s, is_manual=TRUE, manual_by=%s
            WHERE id=%s RETURNING *
        """, (b.get('punch_type'), punched_at_parsed,
              b.get('note', ''), b.get('manual_by', ''), rid)).fetchone()
    return jsonify(punch_record_row(row)) if row else ('', 404)

@app.route('/api/punch/records/<int:rid>', methods=['DELETE'])
@login_required
def api_punch_record_delete(rid):
    with get_db() as conn:
        pr = conn.execute("SELECT staff_id, punched_at FROM punch_records WHERE id=%s", (rid,)).fetchone()
        conn.execute("DELETE FROM punch_records WHERE id=%s", (rid,))
        if pr and pr['staff_id']:
            punch_month = (pr['punched_at'].strftime('%Y-%m')
                           if hasattr(pr['punched_at'], 'strftime')
                           else str(pr['punched_at'])[:7])
            conn.execute("""
                DELETE FROM salary_records
                WHERE staff_id=%s AND month=%s AND status='draft'
            """, (pr['staff_id'], punch_month))
    return jsonify({'deleted': rid})

@app.route('/api/punch/summary', methods=['GET'])
@login_required
def api_punch_summary():
    month = request.args.get('month') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_out,
                   COUNT(*) as punch_count,
                   BOOL_OR(pr.is_manual) as has_manual
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.id, ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date DESC, ps.name
        """, (month,)).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['work_date']  = d['work_date'].isoformat()  if d['work_date']  else None
        d['clock_in']   = d['clock_in'].isoformat()   if d['clock_in']   else None
        d['clock_out']  = d['clock_out'].isoformat()  if d['clock_out']  else None
        if d['clock_in'] and d['clock_out']:
            from datetime import datetime as _dt2
            ci = _dt2.fromisoformat(d['clock_in'].replace('Z', ''))
            co = _dt2.fromisoformat(d['clock_out'].replace('Z', ''))
            d['duration_min'] = max(0, int((co - ci).total_seconds() / 60))
        else:
            d['duration_min'] = None
        result.append(d)

    # Merge cross-midnight pairs: day N has clock_in only → day N+1 has clock_out only
    from datetime import date as _date2, timedelta as _td2, datetime as _dt2m
    result.sort(key=lambda x: (x.get('staff_id', 0), x.get('work_date', '')))
    merged = []
    skip_idx = set()
    for i, d in enumerate(result):
        if i in skip_idx:
            continue
        if d['clock_in'] and not d['clock_out'] and i + 1 < len(result):
            nd = result[i + 1]
            if (nd['staff_id'] == d['staff_id']
                    and d['work_date'] and nd['work_date']
                    and nd['work_date'] == (
                        _date2.fromisoformat(d['work_date']) + _td2(days=1)
                    ).isoformat()
                    and nd['clock_out'] and not nd['clock_in']):
                d = dict(d)
                d['clock_out']   = nd['clock_out']
                ci = _dt2m.fromisoformat(d['clock_in'].replace('Z', ''))
                co = _dt2m.fromisoformat(d['clock_out'].replace('Z', ''))
                d['duration_min'] = max(0, int((co - ci).total_seconds() / 60))
                d['punch_count']  = d.get('punch_count', 0) + nd.get('punch_count', 0)
                d['has_manual']   = bool(d.get('has_manual')) or bool(nd.get('has_manual'))
                skip_idx.add(i + 1)
        merged.append(d)
    result = merged

    return jsonify(result)

@app.route('/api/attendance/monthly-stats', methods=['GET'])
@login_required
def api_attendance_monthly_stats():
    """
    月出勤統計報表（每位員工匯總）
    回傳：出勤天數、總工時、遲到次數、缺打卡次數、平均工時
    """
    month = request.args.get('month') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        # 每人每日打卡彙整
        rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   ps.department, ps.role,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id AND ps.active = TRUE
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM') = %s
            GROUP BY ps.id, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, work_date
        """, (month,)).fetchall()

        # 班別指派（用於遲到判斷）
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.date, st.start_time, st.end_time
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE TO_CHAR(sa.date,'YYYY-MM') = %s
        """, (month,)).fetchall()
        shift_map = {(r['staff_id'], str(r['date'])): r for r in shift_rows}

    # Convert to mutable dicts; merge cross-midnight pairs before stats computation
    rows = [dict(r) for r in rows]
    from datetime import timedelta as _td_cm
    rows.sort(key=lambda x: (x['staff_id'], x['work_date']))
    merged_rows = []
    skip_cm = set()
    for _i, _r in enumerate(rows):
        if _i in skip_cm:
            continue
        if _r['has_in'] and not _r['has_out'] and _i + 1 < len(rows):
            _nr = rows[_i + 1]
            _nxt_date = _r['work_date'] + _td_cm(days=1) if _r['work_date'] else None
            if (_nr['staff_id'] == _r['staff_id']
                    and _nxt_date and _nr['work_date'] == _nxt_date
                    and _nr['has_out'] and not _nr['has_in']):
                _r = dict(_r)
                _r['has_out']   = True
                _r['clock_out'] = _nr['clock_out']
                skip_cm.add(_i + 1)
        merged_rows.append(_r)
    rows = merged_rows

    from collections import defaultdict
    stats = defaultdict(lambda: {
        'staff_id': None, 'staff_name': '', 'department': '', 'role': '',
        'days_worked': 0, 'total_minutes': 0,
        'late_count': 0, 'early_count': 0, 'missing_in_count': 0, 'missing_out_count': 0,
        'anomaly_dates': [],
    })

    for r in rows:
        sid  = r['staff_id']
        ds   = str(r['work_date'])
        s    = stats[sid]
        s['staff_id']   = sid
        s['staff_name'] = r['staff_name']
        s['department'] = r['department'] or ''
        s['role']       = r['role']       or ''

        has_in  = bool(r['has_in'])
        has_out = bool(r['has_out'])

        if has_in or has_out:
            s['days_worked'] += 1

        if r['clock_in'] and r['clock_out']:
            diff = (r['clock_out'] - r['clock_in']).total_seconds() / 60
            if diff > 0:
                s['total_minutes'] += int(diff)

        # 缺打卡
        if has_in and not has_out:
            s['missing_out_count'] += 1
            s['anomaly_dates'].append({'date': ds, 'type': 'missing_out', 'label': '缺下班卡'})
        if not has_in and has_out:
            s['missing_in_count'] += 1
            s['anomaly_dates'].append({'date': ds, 'type': 'missing_in', 'label': '缺上班卡'})

        # 遲到（比對班別）
        if has_in and r['clock_in']:
            shift = shift_map.get((sid, ds))
            if shift and shift['start_time']:
                try:
                    sh, sm = map(int, str(shift['start_time'])[:5].split(':'))
                    ci_local = r['clock_in']
                    ih, im   = ci_local.hour, ci_local.minute
                    late_mins = (ih * 60 + im) - (sh * 60 + sm)
                    if late_mins > 10:
                        s['late_count'] += 1
                        s['anomaly_dates'].append({'date': ds, 'type': 'late',
                                                   'label': f'遲到 {late_mins} 分鐘'})
                except Exception:
                    pass

        # 早退（比對班別）
        if has_out and r['clock_out']:
            shift = shift_map.get((sid, ds))
            if shift and shift['end_time']:
                try:
                    eh, em = map(int, str(shift['end_time'])[:5].split(':'))
                    co_local = r['clock_out']
                    oh, om   = co_local.hour, co_local.minute
                    early_mins = (eh * 60 + em) - (oh * 60 + om)
                    if early_mins > 15:
                        s['early_count'] += 1
                        s['anomaly_dates'].append({'date': ds, 'type': 'early',
                                                   'label': f'早退 {early_mins} 分鐘'})
                except Exception:
                    pass

    result = []
    for s in sorted(stats.values(), key=lambda x: (x['department'], x['staff_name'])):
        h   = s['total_minutes'] // 60
        m   = s['total_minutes'] % 60
        avg = round(s['total_minutes'] / s['days_worked'] / 60, 1) if s['days_worked'] else 0
        s['total_hours']   = round(s['total_minutes'] / 60, 1)
        s['avg_hours_day'] = avg
        s['total_hm']      = f"{h}h {m:02d}m"
        result.append(s)
    return jsonify({'month': month, 'stats': result})

# ── Punch Requests (補打卡申請) ───────────────────────────────────

@app.route('/api/punch/request', methods=['POST'])
def api_punch_req_submit():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    b            = request.get_json(force=True)
    punch_type   = b.get('punch_type')
    requested_at = b.get('requested_at')
    reason       = b.get('reason', '').strip()
    if punch_type not in ('in', 'out', 'break_out', 'break_in'):
        return jsonify({'error': '無效的打卡類型'}), 400
    if not requested_at:
        return jsonify({'error': '請選擇補打時間'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO punch_requests (staff_id, punch_type, requested_at, reason)
            VALUES (%s,%s,%s,%s) RETURNING *
        """, (sid, punch_type, requested_at, reason)).fetchone()
    return jsonify(punch_req_row(row)), 201

@app.route('/api/punch/request/my', methods=['GET'])
def api_punch_req_my():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM punch_requests WHERE staff_id=%s ORDER BY requested_at DESC LIMIT 20",
            (sid,)
        ).fetchall()
    return jsonify([punch_req_row(r) for r in rows])

@app.route('/api/punch/requests', methods=['GET'])
@login_required
def api_punch_reqs_list():
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if status: conds.append('pr.status=%s'); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*, ps.name as staff_name, ps.role as staff_role
            FROM punch_requests pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.created_at DESC LIMIT 200
        """, params).fetchall()
    return jsonify([punch_req_row(r) for r in rows])

@app.route('/api/punch/requests/<int:rid>', methods=['DELETE'])
@login_required
def api_punch_req_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM punch_requests WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

# ═══════════════════════════════════════════════════════════════════
# LINE Punch Clock
# ═══════════════════════════════════════════════════════════════════

CUSTOM_RICHMENU_IMAGE_PATH = '/tmp/custom_richmenu.png'
_pending_line_punches = {}   # {line_user_id: punch_type}
import threading as _threading
_line_reply_ctx = _threading.local()   # holds reply_token per request thread


def _use_reply_token():
    """Return and consume the per-request LINE reply token (single-use)."""
    token = getattr(_line_reply_ctx, 'token', None)
    if token:
        _line_reply_ctx.token = None
    return token


def get_line_punch_config():
    if not DATABASE_URL: return None
    try:
        with get_db() as conn:
            row = conn.execute("SELECT * FROM line_punch_config WHERE id=1").fetchone()
        return dict(row) if row else None
    except Exception:
        return None


def _send_line_punch(user_id, text):
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('enabled') or not cfg.get('channel_access_token'):
        return
    api = LineBotApi(cfg['channel_access_token'])
    msg = TextSendMessage(text=text)
    try:
        token = _use_reply_token()
        if token:
            api.reply_message(token, msg)
        else:
            api.push_message(user_id, msg)
    except Exception as e:
        print(f"[LINE PUNCH] send error: {e}")


def _send_line_with_quick_reply(user_id, text, items):
    """Send a message with Quick Reply buttons.
    items: [{'label': str (≤20 chars), 'text': str (message to send on tap)}, ...]
    """
    from linebot.models import QuickReply, QuickReplyButton, MessageAction
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('enabled') or not cfg.get('channel_access_token'):
        return
    qr_items = [
        QuickReplyButton(action=MessageAction(label=it['label'][:20], text=it['text']))
        for it in items[:13]
    ]
    msg = TextSendMessage(text=text, quick_reply=QuickReply(items=qr_items))
    api = LineBotApi(cfg['channel_access_token'])
    try:
        token = _use_reply_token()
        if token:
            api.reply_message(token, msg)
        else:
            api.push_message(user_id, msg)
    except Exception as e:
        print(f"[LINE PUNCH] send (qr) error: {e}")


@app.route('/line-punch/webhook', methods=['POST'])
def line_punch_webhook():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('enabled') or not cfg.get('channel_secret'):
        return 'disabled', 200

    signature = request.headers.get('X-Line-Signature', '')
    body      = request.get_data(as_text=True)

    import hmac, hashlib as _hl, base64 as _b64
    secret   = cfg['channel_secret'].encode('utf-8')
    computed = _b64.b64encode(
        hmac.new(secret, body.encode('utf-8'), _hl.sha256).digest()
    ).decode('utf-8')
    if not hmac.compare_digest(computed, signature):
        return 'Invalid signature', 400

    events = _json.loads(body).get('events', [])
    for event in events:
        try:
            _handle_line_punch_event(event, cfg)
        except Exception as e:
            print(f"[LINE PUNCH] event handler error: {e}\n{traceback.format_exc()}")
    return 'OK', 200


def _handle_line_punch_event(event, cfg):
    source   = event.get('source', {})
    user_id  = source.get('userId')
    evt_type = event.get('type')
    if not user_id: return
    _line_reply_ctx.token = event.get('replyToken')

    msg      = event.get('message', {})
    msg_type = msg.get('type', '')

    if evt_type == 'follow':
        _send_line_punch(user_id,
            '歡迎使用員工打卡系統！👋\n\n'
            '請輸入您的登入帳號完成綁定。\n\n'
            '✏️ 輸入範例：\n  綁定 mary123\n'
            '（請將 mary123 換成您自己的帳號）\n\n'
            '不知道帳號？請詢問管理員。')
        return

    if evt_type != 'message': return

    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE line_user_id=%s AND active=TRUE", (user_id,)
        ).fetchone()

    # ── Not bound yet ─────────────────────────────────────────
    if not staff:
        if msg_type == 'text':
            text = msg.get('text', '').strip()
            if text.startswith('綁定 ') or text.startswith('绑定 '):
                username = text.split(' ', 1)[1].strip()
                if username in ('帳號', '您的帳號', '[您的帳號]', 'username', '帳號名稱'):
                    _send_line_punch(user_id,
                        '請輸入您「實際的」登入帳號，而非說明文字。\n\n'
                        '範例：綁定 mary123')
                    return
                with get_db() as conn:
                    candidate = conn.execute(
                        "SELECT * FROM punch_staff WHERE username=%s AND active=TRUE",
                        (username,)
                    ).fetchone()
                if not candidate:
                    _send_line_punch(user_id,
                        f'找不到帳號「{username}」\n\n'
                        '請確認帳號是否正確，或詢問管理員您的登入帳號。')
                    return
                if candidate['line_user_id']:
                    _send_line_punch(user_id, '此帳號已綁定其他 LINE 帳號，請聯絡管理員。')
                    return
                with get_db() as conn:
                    conn.execute(
                        "UPDATE punch_staff SET line_user_id=%s WHERE id=%s",
                        (user_id, candidate['id'])
                    )
                _send_line_punch(user_id,
                    f'✅ 綁定成功！\n歡迎 {candidate["name"]}！\n\n'
                    '打卡方式：\n📍 傳送位置訊息 → 自動打卡\n'
                    '💬 或輸入：上班 / 下班 / 休息 / 回來\n\n'
                    '輸入「狀態」可查看今日打卡記錄。')
            else:
                _send_line_punch(user_id,
                    '您尚未綁定打卡帳號。\n\n'
                    '請輸入您的登入帳號：\n  綁定 [您的帳號]\n\n'
                    '範例：綁定 mary123')
        return

    # ── Bound staff ───────────────────────────────────────────
    PUNCH_CMDS = {
        '上班': 'in', '上班打卡': 'in',
        '下班': 'out', '下班打卡': 'out',
        '休息': 'break_out', '休息開始': 'break_out',
        '回來': 'break_in', '休息結束': 'break_in',
    }
    PUNCH_LABEL = {
        'in': '上班打卡', 'out': '下班打卡',
        'break_out': '休息開始', 'break_in': '休息結束',
    }

    if msg_type == 'location':
        lat = msg.get('latitude'); lng = msg.get('longitude')
        _do_line_punch(staff, user_id, lat, lng, None, PUNCH_LABEL)

    elif msg_type == 'text':
        text = msg.get('text', '').strip()

        if text in ('狀態', '打卡記錄'):
            _send_status(staff, user_id); return

        if text == '解除綁定':
            with get_db() as conn:
                conn.execute("UPDATE punch_staff SET line_user_id=NULL WHERE id=%s", (staff['id'],))
            _send_line_punch(user_id, '已解除 LINE 帳號綁定。'); return

        punch_type = PUNCH_CMDS.get(text)
        if punch_type:
            with get_db() as conn:
                pcfg = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
                locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()
            gps_required = pcfg['gps_required'] if pcfg else False
            if gps_required and locs:
                from linebot.models import QuickReply, QuickReplyButton, LocationAction
                cfg_lp = get_line_punch_config()
                if cfg_lp and cfg_lp.get('enabled') and cfg_lp.get('channel_access_token'):
                    qr = QuickReply(items=[QuickReplyButton(action=LocationAction(label='📍 傳送位置'))])
                    msg = TextSendMessage(
                        text=f'請傳送您的位置來完成{PUNCH_LABEL[punch_type]}\n點下方「傳送位置」按鈕即可打卡',
                        quick_reply=qr)
                    _api = LineBotApi(cfg_lp['channel_access_token'])
                    try:
                        token = _use_reply_token()
                        if token:
                            _api.reply_message(token, msg)
                        else:
                            _api.push_message(user_id, msg)
                    except Exception as _e:
                        print(f"[LINE PUNCH] location qr error: {_e}")
                _pending_line_punches[user_id] = punch_type
            else:
                _do_line_punch(staff, user_id, None, None, punch_type, PUNCH_LABEL)
        elif text in ('查餘假', '餘假', '假期', '查假', '特休'):
            _line_query_leave_balance(staff, user_id)
        elif text in ('查薪資', '薪資', '薪水', '薪資單', '查薪水'):
            _line_query_salary(staff, user_id)
        elif text.startswith('請假'):
            _line_submit_leave(staff, user_id, text)
        elif text in ('績效', '考核', '我的考核', '查績效'):
            _line_query_performance(staff, user_id)
        elif text in ('假別', '假別清單', '假別列表'):
            _line_show_leave_types(staff, user_id)
        elif (text in ('出勤紀錄', '出勤記錄', '月出勤', '打卡紀錄', '打卡記錄', '出勤查詢')
              or text.startswith('出勤紀錄 ') or text.startswith('出勤記錄 ')
              or text.startswith('打卡紀錄 ') or text.startswith('打卡記錄 ')):
            _line_query_monthly_records(staff, user_id, text)
        elif text == '加班':
            _line_overtime_start(staff, user_id)
        elif text.startswith('申請加班'):
            _line_submit_overtime(staff, user_id, text)
        elif text in ('選單', '功能', '菜單', '?', '？', 'help', 'Help', 'HELP'):
            _line_show_help(staff, user_id)
        else:
            _line_show_help(staff, user_id)


def _do_line_punch(staff, user_id, lat, lng, forced_type, PUNCH_LABEL):
    from datetime import datetime as _dt3, timezone as _tz3, timedelta as _td3
    TW = _tz3(_td3(hours=8))

    # Determine punch type
    if forced_type:
        punch_type = forced_type
    elif user_id in _pending_line_punches:
        punch_type = _pending_line_punches.pop(user_id)
    else:
        with get_db() as conn:
            last = conn.execute("""
                SELECT punch_type, punched_at FROM punch_records
                WHERE staff_id=%s
                  AND punched_at >= NOW() - INTERVAL '24 hours'
                ORDER BY punched_at DESC LIMIT 1
            """, (staff['id'],)).fetchone()
        if not last:
            punch_type = 'in'
        elif last['punch_type'] == 'in':
            punch_type = 'out'
        elif last['punch_type'] == 'break_out':
            punch_type = 'break_in'
        else:
            # last was 'out' or 'break_in' → next would be 'in',
            # but reject if punched within 5 minutes to prevent accidental double-punch
            now_utc = _dt3.now(_tz3.utc)
            last_at = last['punched_at']
            if last_at.tzinfo is None:
                last_at = last_at.replace(tzinfo=_tz3.utc)
            elapsed = int((now_utc - last_at).total_seconds())
            if elapsed < 300:
                _send_line_punch(user_id,
                    f'⚠️ 您剛於 {elapsed} 秒前完成打卡\n'
                    '若確認要重新上班，請等候 5 分鐘後再打卡，\n'
                    '或使用「上班」指令強制打卡。')
                return
            punch_type = 'in'

    label = PUNCH_LABEL.get(punch_type, punch_type)

    # GPS check
    gps_distance = None; matched_name = ''
    if lat is not None and lng is not None:
        with get_db() as conn:
            pcfg = conn.execute("SELECT * FROM punch_config WHERE id=1").fetchone()
            locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()
        gps_required = pcfg['gps_required'] if pcfg else False
        if locs:
            min_dist = None; min_loc = None
            for loc in locs:
                d = _gps_distance(lat, lng, float(loc['lat']), float(loc['lng']))
                if min_dist is None or d < min_dist:
                    min_dist = d; min_loc = loc
            gps_distance = min_dist
            matched_name = min_loc['location_name'] if min_loc else ''
            if gps_required and min_dist > int(min_loc['radius_m']):
                _send_line_punch(user_id,
                    f'❌ {label}失敗\n'
                    f'您距離「{min_loc["location_name"]}」{min_dist} 公尺\n'
                    f'超出允許範圍 {min_loc["radius_m"]} 公尺\n\n'
                    '請確認您在正確地點後重試。')
                return

    # Duplicate guard
    with get_db() as conn:
        recent = conn.execute("""
            SELECT id FROM punch_records
            WHERE staff_id=%s AND punch_type=%s
              AND punched_at > NOW() - INTERVAL '1 minute'
        """, (staff['id'], punch_type)).fetchone()
        if recent:
            _send_line_punch(user_id, f'⚠️ 1 分鐘內已打過{label}，請勿重複打卡。'); return

        conn.execute("""
            INSERT INTO punch_records
              (staff_id, punch_type, latitude, longitude, gps_distance, location_name)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (staff['id'], punch_type, lat, lng, gps_distance, matched_name))

    now      = _dt3.now(TW)
    gps_info = f'\n📍 {matched_name} ({gps_distance}m)' if gps_distance is not None else ''
    _send_line_punch(user_id,
        f'✅ {label}成功\n'
        f'👤 {staff["name"]}\n'
        f'🕐 {now.strftime("%Y/%m/%d %H:%M")}'
        f'{gps_info}')


def _send_status(staff, user_id):
    from datetime import timezone as _tz4, timedelta as _td4
    TW = _tz4(_td4(hours=8))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT punch_type, punched_at, gps_distance, location_name, is_manual
            FROM punch_records
            WHERE staff_id=%s
              AND (punched_at AT TIME ZONE 'Asia/Taipei')::date
                = (NOW() AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY punched_at ASC
        """, (staff['id'],)).fetchall()
    LABEL = {'in': '上班', 'out': '下班', 'break_out': '休息開始', 'break_in': '休息結束'}
    if not rows:
        _send_line_punch(user_id, f'📋 {staff["name"]} 今日尚無打卡記錄。'); return
    lines = [f'📋 {staff["name"]} 今日打卡記錄']
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            from datetime import timezone as _utz2
            pa = pa.replace(tzinfo=_utz2.utc)
        t    = pa.astimezone(TW).strftime('%H:%M')
        dist = f' ({r["gps_distance"]}m)' if r['gps_distance'] is not None else ''
        man  = ' [補打]' if r['is_manual'] else ''
        lines.append(f'• {LABEL.get(r["punch_type"], r["punch_type"])} {t}{dist}{man}')
    _send_line_punch(user_id, '\n'.join(lines))

# ── Admin LINE Punch Config API ────────────────────────────────────

@app.route('/api/line-punch/config', methods=['GET'])
@login_required
def api_line_punch_config_get():
    with get_db() as conn:
        row = conn.execute("SELECT * FROM line_punch_config WHERE id=1").fetchone()
    if not row:
        return jsonify({'enabled': False, 'channel_access_token': '', 'channel_secret': ''})
    d = dict(row)
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    return jsonify(d)

@app.route('/api/line-punch/config', methods=['PUT'])
@login_required
def api_line_punch_config_put():
    b       = request.get_json(force=True)
    token   = b.get('channel_access_token', '').strip()
    secret  = b.get('channel_secret', '').strip()
    enabled = bool(b.get('enabled', False))
    with get_db() as conn:
        conn.execute("""
            UPDATE line_punch_config
            SET channel_access_token=%s, channel_secret=%s, enabled=%s, updated_at=NOW()
            WHERE id=1
        """, (token, secret, enabled))
    return jsonify({'ok': True, 'enabled': enabled})

@app.route('/api/line-punch/staff', methods=['GET'])
@login_required
def api_line_punch_staff():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,name,username,role,active,line_user_id FROM punch_staff ORDER BY name"
        ).fetchall()
    return jsonify([{
        'id': r['id'], 'name': r['name'], 'username': r['username'],
        'role': r['role'], 'active': r['active'],
        'line_bound': bool(r['line_user_id']),
        'line_user_id': r['line_user_id'] or ''
    } for r in rows])

@app.route('/api/line-punch/staff/<int:sid>/unbind', methods=['POST'])
@login_required
def api_line_punch_unbind(sid):
    with get_db() as conn:
        conn.execute("UPDATE punch_staff SET line_user_id=NULL WHERE id=%s", (sid,))
    return jsonify({'ok': True})

# ── Rich Menu ──────────────────────────────────────────────────────

def _call_line_api(cfg, method, path, body=None):
    token = cfg.get('channel_access_token', '')
    url   = 'https://api.line.me/v2/bot' + path
    data  = _json.dumps(body).encode('utf-8') if body else None
    req   = urllib.request.Request(
        url, data=data, method=method,
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {token}'}
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.status, _json.loads(resp.read())
    except urllib.error.HTTPError as e:
        return e.code, {'error': e.read().decode('utf-8', errors='replace')}
    except Exception as e:
        return 0, {'error': str(e)}


def _make_richmenu_png():
    """Generate a simple 2500×1686 PNG with 4 colored quadrants."""
    import struct, zlib
    W, H = 2500, 1686
    colors = [(0x2e,0x9e,0x6b), (0xd6,0x42,0x42), (0xe0,0x7b,0x2a), (0x4a,0x7b,0xda)]
    rows = []
    for y in range(H):
        row = bytearray()
        for x in range(W):
            p = (0 if y < 843 else 1) * 2 + (0 if x < 1250 else 1)
            r, g, b = colors[p]
            if x in (1249, 1250) or y in (842, 843):
                r, g, b = 0x0f, 0x1c, 0x3a
            row += bytes([r, g, b])
        rows.append(bytes([0]) + bytes(row))
    compressed = zlib.compress(b''.join(rows), 1)

    def chunk(name, data):
        c = struct.pack('>I', len(data)) + name + data
        return c + struct.pack('>I', zlib.crc32(c[4:]) & 0xffffffff)

    return (b'\x89PNG\r\n\x1a\n'
            + chunk(b'IHDR', struct.pack('>IIBBBBB', W, H, 8, 2, 0, 0, 0))
            + chunk(b'IDAT', compressed)
            + chunk(b'IEND', b''))


@app.route('/api/line-punch/richmenu/create', methods=['POST'])
@login_required
def api_richmenu_create():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'error': '請先設定 Channel Access Token'}), 400

    b = request.get_json(force=True) or {}
    gdrive_url = b.get('gdrive_url', '').strip()
    btn_texts  = b.get('button_texts') or []
    defaults   = ['上班', '下班', '請假', '加班']
    btn_texts  = [(btn_texts[i].strip() if i < len(btn_texts) and btn_texts[i].strip() else defaults[i]) for i in range(4)]

    body = {
        "size": {"width": 2500, "height": 1686},
        "selected": True,
        "name": "打卡選單",
        "chatBarText": "打卡",
        "areas": [
            {"bounds": {"x": 0,    "y": 0,   "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[0]}},
            {"bounds": {"x": 1250, "y": 0,   "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[1]}},
            {"bounds": {"x": 0,    "y": 843, "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[2]}},
            {"bounds": {"x": 1250, "y": 843, "width": 1250, "height": 843}, "action": {"type": "message", "text": btn_texts[3]}},
        ]
    }

    status, data = _call_line_api(cfg, 'POST', '/richmenu', body)
    if status != 200:
        return jsonify({'error': f'建立失敗 ({status}): {data.get("error","")}'}), 500

    rich_menu_id = data.get('richMenuId', '')

    # Upload image — 1) Google Drive  2) custom local file  3) auto-generate
    png_bytes = None

    if gdrive_url:
        try:
            import re as _re
            file_id = None
            m = _re.search(r'/file/d/([^/?]+)', gdrive_url)
            if m:
                file_id = m.group(1)
            else:
                m = _re.search(r'[?&]id=([^&]+)', gdrive_url)
                if m:
                    file_id = m.group(1)
            if file_id:
                dl_url = f'https://drive.google.com/uc?export=download&id={file_id}'
                req = urllib.request.Request(dl_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=30) as resp:
                    png_bytes = resp.read()
                # Google Drive may return an HTML warning page for large files
                if png_bytes and png_bytes[:1] not in (b'\x89', b'\xff', b'\x47', b'BM'):
                    print(f"[LINE PUNCH] gdrive returned non-image content, ignoring")
                    png_bytes = None
        except Exception as e:
            print(f"[LINE PUNCH] gdrive download error: {e}")

    if not png_bytes:
        try:
            import os
            for _cp in [CUSTOM_RICHMENU_IMAGE_PATH,
                        CUSTOM_RICHMENU_IMAGE_PATH.replace('.png', '.jpg')]:
                if os.path.exists(_cp):
                    with open(_cp, 'rb') as f:
                        png_bytes = f.read()
                    break
        except Exception:
            pass

    if not png_bytes:
        try:
            png_bytes = _make_richmenu_png()
        except Exception:
            pass

    img_ok = False
    if png_bytes:
        content_type = 'image/jpeg' if png_bytes[:2] == b'\xff\xd8' else 'image/png'
        upload_url = f'https://api-data.line.me/v2/bot/richmenu/{rich_menu_id}/content'
        req = urllib.request.Request(
            upload_url, data=png_bytes, method='POST',
            headers={'Content-Type': content_type, 'Authorization': f'Bearer {cfg["channel_access_token"]}'}
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                img_ok = resp.status in (200, 204)
        except Exception:
            pass

    # Set as default for all users
    _call_line_api(cfg, 'POST', f'/user/all/richmenu/{rich_menu_id}')

    return jsonify({'ok': True, 'rich_menu_id': rich_menu_id, 'image_uploaded': img_ok})


@app.route('/api/line-punch/richmenu/list', methods=['GET'])
@login_required
def api_richmenu_list():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'menus': []})
    status, data = _call_line_api(cfg, 'GET', '/richmenu/list')
    if status != 200:
        return jsonify({'menus': [], 'error': data.get('error', '')})
    return jsonify({'menus': data.get('richmenus', [])})


@app.route('/api/line-punch/richmenu/<rich_menu_id>', methods=['DELETE'])
@login_required
def api_richmenu_delete(rich_menu_id):
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'error': '未設定 Token'}), 400
    _call_line_api(cfg, 'DELETE', '/user/all/richmenu')
    status, _ = _call_line_api(cfg, 'DELETE', f'/richmenu/{rich_menu_id}')
    return jsonify({'ok': status in (200, 204), 'status': status})


@app.route('/api/line-punch/richmenu/default', methods=['DELETE'])
@login_required
def api_richmenu_unset_default():
    cfg = get_line_punch_config()
    if not cfg or not cfg.get('channel_access_token'):
        return jsonify({'error': '未設定 Token'}), 400
    status, _ = _call_line_api(cfg, 'DELETE', '/user/all/richmenu')
    return jsonify({'ok': status in (200, 204)})

# ═══════════════════════════════════════════════════════════════════
# Schedule / Shift API
# ═══════════════════════════════════════════════════════════════════

# ── Employee: schedule config + my request ────────────────────────

@app.route('/api/schedule/config/<month>', methods=['GET'])
def api_sched_config_get(month):
    sid = session.get('punch_staff_id')
    with get_db() as conn:
        cfg    = dict(get_schedule_config(conn, month))
        counts = get_off_counts(conn, month)
        if sid:
            row = conn.execute(
                "SELECT vacation_quota FROM punch_staff WHERE id=%s", (sid,)
            ).fetchone()
            if row and row['vacation_quota'] is not None:
                cfg['vacation_quota']  = int(row['vacation_quota'])
                cfg['quota_personal']  = True
    return jsonify({**cfg, 'off_counts': counts})


@app.route('/api/schedule/my-request/<month>', methods=['GET'])
def api_sched_my_request(month):
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        row = conn.execute("""
            SELECT sr.*, ps.name as staff_name
            FROM schedule_requests sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.staff_id=%s AND sr.month=%s
        """, (sid, month)).fetchone()
    return jsonify(sched_req_row(row)) if row else jsonify(None)


@app.route('/api/schedule/my-request', methods=['POST'])
def api_sched_submit():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    b     = request.get_json(force=True)
    month = b.get('month', '').strip()
    dates = b.get('dates', [])
    note  = b.get('submit_note', '').strip()

    if not month: return jsonify({'error': '請選擇月份'}), 400
    if not isinstance(dates, list): return jsonify({'error': '日期格式錯誤'}), 400
    for d in dates:
        if not d.startswith(month):
            return jsonify({'error': f'日期 {d} 不屬於 {month}'}), 400

    try:
        with get_db() as conn:
            cfg = get_schedule_config(conn, month)

            staff_row = conn.execute(
                "SELECT vacation_quota FROM punch_staff WHERE id=%s", (sid,)
            ).fetchone()
            personal_quota  = staff_row['vacation_quota'] if staff_row and staff_row['vacation_quota'] is not None else None
            effective_quota = personal_quota if personal_quota is not None else cfg['vacation_quota']

            if len(dates) > effective_quota:
                quota_source = '個人配額' if personal_quota is not None else '月份預設配額'
                return jsonify({'error': f'申請天數（{len(dates)}天）超過{quota_source}（{effective_quota}天）'}), 422

            overcrowded = []
            for d in dates:
                try:
                    others = conn.execute("""
                        SELECT COUNT(*) as cnt
                        FROM schedule_requests,
                             jsonb_array_elements_text(dates) as elem
                        WHERE month=%s AND status IN ('approved','pending')
                          AND staff_id != %s AND elem=%s
                    """, (month, sid, d)).fetchone()
                    others_count = int(others['cnt']) if others else 0
                except Exception:
                    others_count = 0
                if others_count >= cfg['max_off_per_day']:
                    dt_obj = _dt.strptime(d, '%Y-%m-%d')
                    overcrowded.append({
                        'date': d,
                        'weekday': WEEKDAY_ZH[dt_obj.weekday()],
                        'count': others_count,
                        'max': cfg['max_off_per_day']
                    })
            if overcrowded:
                msgs = [f"{x['date']}（{x['weekday']}）已有 {x['count']} 人排休" for x in overcrowded]
                return jsonify({'error': '以下日期休假人數已達上限：' + '、'.join(msgs), 'overcrowded': overcrowded}), 422

            prev = conn.execute(
                "SELECT status FROM schedule_requests WHERE staff_id=%s AND month=%s",
                (sid, month)
            ).fetchone()
            new_status = 'modified_pending' if prev and prev['status'] == 'approved' else 'pending'
            dates_json = _json.dumps(dates, ensure_ascii=False)

            row = conn.execute("""
                INSERT INTO schedule_requests
                  (staff_id, month, dates, status, submit_note, updated_at)
                VALUES (%s, %s, %s::jsonb, %s, %s, NOW())
                ON CONFLICT (staff_id, month) DO UPDATE
                  SET dates=EXCLUDED.dates, status=EXCLUDED.status,
                      submit_note=EXCLUDED.submit_note, updated_at=NOW()
                RETURNING *
            """, (sid, month, dates_json, new_status, note)).fetchone()

        return jsonify(sched_req_row(row)), 201
    except Exception as e:
        import traceback as _tb
        print(f"[SCHED SUBMIT ERROR] {e}\n{_tb.format_exc()}")
        return jsonify({'error': f'系統錯誤：{str(e)}'}), 500

# ── Admin: schedule config ────────────────────────────────────────

@app.route('/api/schedule/admin/config/<month>', methods=['GET'])
@require_module('sched')
def api_sched_admin_config_get(month):
    with get_db() as conn:
        cfg    = get_schedule_config(conn, month)
        counts = get_off_counts(conn, month)
    return jsonify({**cfg, 'off_counts': counts})


@app.route('/api/schedule/admin/config/<month>', methods=['PUT'])
@require_module('sched')
def api_sched_admin_config_put(month):
    b       = request.get_json(force=True)
    max_off = int(b.get('max_off_per_day') or 2)
    quota   = int(b.get('vacation_quota')   or 8)
    notes   = b.get('notes', '').strip()
    with get_db() as conn:
        conn.execute("""
            INSERT INTO schedule_config (month, max_off_per_day, vacation_quota, notes)
            VALUES (%s,%s,%s,%s)
            ON CONFLICT (month) DO UPDATE
              SET max_off_per_day=%s, vacation_quota=%s, notes=%s, updated_at=NOW()
        """, (month, max_off, quota, notes, max_off, quota, notes))
    return jsonify({'month': month, 'max_off_per_day': max_off,
                    'vacation_quota': quota, 'notes': notes})

# ── Admin: schedule requests ──────────────────────────────────────

@app.route('/api/schedule/admin/requests', methods=['GET'])
@require_module('sched')
def api_sched_admin_requests():
    month  = request.args.get('month', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:  conds.append('sr.month=%s');  params.append(month)
    if status: conds.append('sr.status=%s'); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT sr.*, ps.name as staff_name, ps.role as staff_role
            FROM schedule_requests sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY sr.month DESC, ps.name
        """, params).fetchall()
    return jsonify([sched_req_row(r) for r in rows])


@app.route('/api/schedule/admin/requests/<int:rid>', methods=['PUT'])
@require_module('sched')
def api_sched_admin_review(rid):
    b           = request.get_json(force=True)
    action      = b.get('action')
    reviewed_by = b.get('reviewed_by', '').strip()
    review_note = b.get('review_note', '').strip()
    if action not in ('approve', 'reject', 'revoke'):
        return jsonify({'error': 'action must be approve / reject / revoke'}), 400

    if action == 'revoke':
        with get_db() as conn:
            row = conn.execute("""
                UPDATE schedule_requests
                SET status='pending', reviewed_by='', review_note=%s,
                    reviewed_at=NULL, updated_at=NOW()
                WHERE id=%s RETURNING *
            """, (review_note or '主管已撤銷核准', rid)).fetchone()
        return jsonify(sched_req_row(row)) if row else ('', 404)

    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as conn:
        row = conn.execute("""
            UPDATE schedule_requests
            SET status=%s, reviewed_by=%s, review_note=%s,
                reviewed_at=NOW(), updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (new_status, reviewed_by, review_note, rid)).fetchone()
    if row:
        dates = row['dates'] if isinstance(row['dates'], list) else _json.loads(row['dates'] or '[]')
        extra = f"{row['month']} 排休 {len(dates)} 天"
        if review_note: extra += f"\n審核意見：{review_note}"
        _notify_review_result(row['staff_id'], '排休申請', action, extra)
    return jsonify(sched_req_row(row)) if row else ('', 404)


@app.route('/api/schedule/admin/requests/<int:rid>', methods=['DELETE'])
@require_module('sched')
def api_sched_admin_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM schedule_requests WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})


@app.route('/api/schedule/admin/calendar/<month>', methods=['GET'])
@require_module('sched')
def api_sched_admin_calendar(month):
    with get_db() as conn:
        cfg   = get_schedule_config(conn, month)
        staff = conn.execute(
            "SELECT id,name,role FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        reqs  = conn.execute("""
            SELECT sr.staff_id, sr.dates, sr.status, ps.name
            FROM schedule_requests sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s AND sr.status IN ('approved','pending','modified_pending')
        """, (month,)).fetchall()

    year_int, month_int = int(month[:4]), int(month[5:])
    import calendar as _cal
    days_in_month = _cal.monthrange(year_int, month_int)[1]

    staff_off = {}
    for r in reqs:
        dates_val = r['dates']
        if isinstance(dates_val, str):
            try: dates_val = _json.loads(dates_val)
            except: dates_val = []
        for d in (dates_val or []):
            if r['staff_id'] not in staff_off:
                staff_off[r['staff_id']] = {}
            staff_off[r['staff_id']][d] = r['status']

    days = []
    for day in range(1, days_in_month + 1):
        date_str = f"{month}-{day:02d}"
        dt       = _dt(year_int, month_int, day)
        off_list = []
        for s in staff:
            st = staff_off.get(s['id'], {}).get(date_str)
            if st:
                off_list.append({'staff_id': s['id'], 'name': s['name'],
                                  'role': s['role'], 'status': st})
        days.append({
            'date':          date_str,
            'day':           day,
            'weekday':       WEEKDAY_ZH[dt.weekday()],
            'is_weekend':    dt.weekday() >= 5,
            'off_count':     len(off_list),
            'off_list':      off_list,
            'working_count': len(staff) - len(off_list),
            'over_limit':    len(off_list) > cfg['max_off_per_day'],
        })
    return jsonify({'month': month, 'config': cfg, 'staff_count': len(staff), 'days': days})


@app.route('/api/schedule/admin/summary/<month>', methods=['GET'])
@require_module('sched')
def api_sched_admin_summary(month):
    with get_db() as conn:
        cfg   = get_schedule_config(conn, month)
        staff = conn.execute(
            "SELECT id,name,role FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        reqs  = conn.execute(
            "SELECT sr.* FROM schedule_requests sr WHERE sr.month=%s", (month,)
        ).fetchall()
    req_map = {r['staff_id']: sched_req_row(r) for r in reqs}
    result  = []
    for s in staff:
        req = req_map.get(s['id'])
        result.append({
            'staff_id':   s['id'],
            'name':       s['name'],
            'role':       s['role'],
            'status':     req['status']  if req else 'not_submitted',
            'days_off':   len(req['dates']) if req else 0,
            'quota':      cfg['vacation_quota'],
            'dates':      req['dates']   if req else [],
            'request_id': req['id']      if req else None,
        })
    return jsonify({'config': cfg, 'staff': result})

# ── Shift Types CRUD ──────────────────────────────────────────────

@app.route('/api/shifts/types', methods=['GET'])
@require_module('sched')
def api_shift_types_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM shift_types ORDER BY sort_order, id").fetchall()
    return jsonify([shift_type_row(r) for r in rows])

@app.route('/api/shifts/types/public', methods=['GET'])
def api_shift_types_public():
    """Public endpoint for employee page."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM shift_types WHERE active=TRUE ORDER BY sort_order, id"
        ).fetchall()
    return jsonify([shift_type_row(r) for r in rows])

@app.route('/api/shifts/types', methods=['POST'])
@require_module('sched')
def api_shift_type_create():
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO shift_types (name, start_time, end_time, color, departments, sort_order)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'], b['start_time'], b['end_time'],
              b.get('color', '#4a7bda'), b.get('departments', ''),
              int(b.get('sort_order', 0)))).fetchone()
    return jsonify(shift_type_row(row)), 201

@app.route('/api/shifts/types/<int:sid>', methods=['PUT'])
@require_module('sched')
def api_shift_type_update(sid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE shift_types
            SET name=%s, start_time=%s, end_time=%s, color=%s,
                departments=%s, sort_order=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b['name'], b['start_time'], b['end_time'],
              b.get('color', '#4a7bda'), b.get('departments', ''),
              int(b.get('sort_order', 0)), bool(b.get('active', True)),
              sid)).fetchone()
    return jsonify(shift_type_row(row)) if row else ('', 404)

@app.route('/api/shifts/types/<int:sid>', methods=['DELETE'])
@require_module('sched')
def api_shift_type_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM shift_types WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

# ── Shift Assignments ─────────────────────────────────────────────

@app.route('/api/shifts/assignments', methods=['GET'])
@require_module('sched')
def api_shift_assignments_list():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("to_char(sa.shift_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT sa.*,
                   ps.name as staff_name, ps.role as staff_role,
                   st.name as shift_name, st.start_time, st.end_time,
                   st.color, st.departments
            FROM shift_assignments sa
            JOIN punch_staff ps ON ps.id=sa.staff_id
            JOIN shift_types  st ON st.id=sa.shift_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY sa.shift_date, ps.name
        """, params).fetchall()
    result = []
    for r in rows:
        d = shift_assign_row(r)
        d['staff_name'] = r['staff_name']
        d['staff_role'] = r['staff_role']
        d['shift_name'] = r['shift_name']
        d['start_time'] = str(r['start_time'])[:5]
        d['end_time']   = str(r['end_time'])[:5]
        d['color']      = r['color']
        result.append(d)
    return jsonify(result)


@app.route('/api/shifts/assignments', methods=['POST'])
@require_module('sched')
def api_shift_assignment_create():
    b             = request.get_json(force=True)
    staff_ids     = b.get('staff_ids', [])
    shift_type_id = b.get('shift_type_id')
    dates         = b.get('dates', [])
    note          = b.get('note', '').strip()
    force         = bool(b.get('force', False))

    if not staff_ids or not shift_type_id or not dates:
        return jsonify({'error': '請選擇員工、班別及日期'}), 400

    created = 0
    blocked = []

    with get_db() as conn:
        # Build leave lookup for all involved staff
        leave_lookup = {}
        if not force:
            for sid in staff_ids:
                months = list({d[:7] for d in dates})
                for month in months:
                    row = conn.execute("""
                        SELECT dates FROM schedule_requests
                        WHERE staff_id=%s AND month=%s AND status IN ('approved', 'pending')
                    """, (sid, month)).fetchone()
                    if row:
                        approved_dates = row['dates'] or []
                        if isinstance(approved_dates, str):
                            try: approved_dates = _json.loads(approved_dates)
                            except: approved_dates = []
                        if sid not in leave_lookup:
                            leave_lookup[sid] = set()
                        leave_lookup[sid].update(approved_dates)

        staff_names = {}
        for r in conn.execute(
            "SELECT id,name FROM punch_staff WHERE id = ANY(%s::int[])", (staff_ids,)
        ).fetchall():
            staff_names[r['id']] = r['name']

        for sid in staff_ids:
            leave_dates = leave_lookup.get(sid, set())
            for date_str in dates:
                if date_str in leave_dates and not force:
                    blocked.append({'staff_name': staff_names.get(sid, str(sid)), 'date': date_str})
                    continue
                conn.execute("""
                    INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date, note)
                    VALUES (%s,%s,%s,%s)
                    ON CONFLICT (staff_id, shift_date) DO UPDATE
                      SET shift_type_id=%s, note=%s, created_at=NOW()
                """, (sid, shift_type_id, date_str, note, shift_type_id, note))
                created += 1

    if blocked and created == 0:
        return jsonify({
            'error': '以下日期員工已有核准的排休，無法指派班別：' +
                     '、'.join([f'{x["staff_name"]} {x["date"]}' for x in blocked]),
            'blocked': blocked
        }), 422

    # Notify each assigned staff via LINE
    if created > 0:
        with get_db() as conn:
            shift_info = conn.execute(
                "SELECT name, start_time, end_time FROM shift_types WHERE id=%s", (shift_type_id,)
            ).fetchone()
        if shift_info:
            date_range = f"{min(dates)} ~ {max(dates)}" if len(dates) > 1 else dates[0]
            msg = (f"[排班通知] 已為您安排班別\n"
                   f"班別：{shift_info['name']}（{str(shift_info['start_time'])[:5]}～{str(shift_info['end_time'])[:5]}）\n"
                   f"日期：{date_range}\n"
                   f"共 {len(dates)} 天，請至員工系統查看完整排班。")
            for sid in staff_ids:
                _notify_staff_line(sid, msg)

    result = {'created': created}
    if blocked:
        result['warning'] = f'已指派 {created} 筆，跳過 {len(blocked)} 筆（員工當日有核准排休）'
        result['blocked'] = blocked
    return jsonify(result), 201


@app.route('/api/shifts/assignments/batch-delete', methods=['POST'])
@require_module('sched')
def api_shift_assignment_batch_delete():
    b         = request.get_json(force=True)
    staff_ids = b.get('staff_ids', [])
    dates     = b.get('dates', [])
    if not staff_ids or not dates:
        return jsonify({'error': '請選擇員工及日期'}), 400
    deleted = 0
    with get_db() as conn:
        affected = set()
        for sid in staff_ids:
            for date_str in dates:
                r = conn.execute(
                    "DELETE FROM shift_assignments WHERE staff_id=%s AND shift_date=%s RETURNING id",
                    (sid, date_str)
                ).fetchone()
                if r:
                    deleted += 1
                    affected.add((sid, str(date_str)[:7]))
        for _sid, _month in affected:
            conn.execute("""
                DELETE FROM salary_records
                WHERE staff_id=%s AND month=%s AND status='draft'
            """, (_sid, _month))
    return jsonify({'deleted': deleted})


@app.route('/api/shifts/import', methods=['POST'])
@require_module('sched')
def api_shift_import():
    """
    匯入班表 CSV 或 Excel (.xlsx)。
    表頭（第一列）：姓名,日期,班別,備註  或  代碼,日期,班別,備註
    日期格式：YYYY-MM-DD
    force=1 query param 可強制覆蓋排休衝突。
    """
    import csv, io as _io
    force = request.args.get('force', '0') == '1'
    rows = []

    if 'file' in request.files:
        f = request.files['file']
        fname = (f.filename or '').lower()
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            # ── Excel 解析 ────────────────────────────────────
            import openpyxl as _opx
            wb = _opx.load_workbook(_io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.values)
            if not all_rows:
                return jsonify({'error': '檔案內容為空'}), 400
            headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
            for row in all_rows[1:]:
                if all(v is None or str(v).strip() == '' for v in row):
                    continue  # skip blank rows
                d = {}
                for i, h in enumerate(headers):
                    d[h] = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
                rows.append(d)
        else:
            raw = f.read().decode('utf-8-sig')
            if not raw.strip():
                return jsonify({'error': '檔案內容為空'}), 400
            reader = csv.DictReader(_io.StringIO(raw))
            if reader.fieldnames is None:
                return jsonify({'error': '無法解析 CSV 欄位'}), 400
            reader.fieldnames = [h.strip() for h in reader.fieldnames]
            rows = list(reader)
    else:
        raw = request.get_data(as_text=True)
        if not raw.strip():
            return jsonify({'error': '檔案內容為空'}), 400
        reader = csv.DictReader(_io.StringIO(raw))
        if reader.fieldnames is None:
            return jsonify({'error': '無法解析 CSV 欄位'}), 400
        reader.fieldnames = [h.strip() for h in reader.fieldnames]
        rows = list(reader)

    if not rows:
        return jsonify({'error': '無資料列'}), 400

    # 確認必要欄位
    all_keys = rows[0].keys() if rows else []
    has_name = '姓名' in all_keys
    has_code = '代碼' in all_keys
    if not (has_name or has_code):
        return jsonify({'error': '檔案缺少「姓名」或「代碼」欄位'}), 400
    if '日期' not in all_keys:
        return jsonify({'error': '檔案缺少「日期」欄位'}), 400
    if '班別' not in all_keys:
        return jsonify({'error': '檔案缺少「班別」欄位'}), 400
    with get_db() as conn:
        # 預先建立索引，避免逐列查詢
        staff_by_name = {r['name']: r['id'] for r in conn.execute(
            "SELECT id, name FROM punch_staff WHERE active=TRUE"
        ).fetchall()}
        staff_by_code = {r['employee_code']: r['id'] for r in conn.execute(
            "SELECT id, employee_code FROM punch_staff WHERE active=TRUE AND employee_code IS NOT NULL AND employee_code!=''",
        ).fetchall()}
        shift_by_name = {r['name']: r['id'] for r in conn.execute(
            "SELECT id, name FROM shift_types WHERE active=TRUE"
        ).fetchall()}

        # 預先讀取所有涉及員工的核准排休（僅在非強制時）
        leave_lookup = {}   # { staff_id: set(date_str) }
        if not force:
            leave_rows = conn.execute("""
                SELECT staff_id, dates FROM schedule_requests
                WHERE status='approved'
            """).fetchall()
            for lr in leave_rows:
                sid = lr['staff_id']
                dates_val = lr['dates']
                if isinstance(dates_val, str):
                    try: dates_val = _json.loads(dates_val)
                    except: dates_val = []
                if sid not in leave_lookup:
                    leave_lookup[sid] = set()
                leave_lookup[sid].update(dates_val or [])

        created = 0
        skipped = []   # 衝突（排休）
        errors  = []   # 找不到員工/班別、日期格式錯誤

        for i, row in enumerate(rows, start=2):   # 從第2列計算（第1列是表頭）
            name_val = row.get('姓名', '').strip()
            code_val = row.get('代碼', '').strip()
            date_str = row.get('日期', '').strip()
            shift_name = row.get('班別', '').strip()
            note = row.get('備註', '').strip()

            # 找員工 ID
            staff_id = None
            if code_val:
                staff_id = staff_by_code.get(code_val)
            if staff_id is None and name_val:
                staff_id = staff_by_name.get(name_val)
            if staff_id is None:
                errors.append({'row': i, 'reason': f'找不到員工：{code_val or name_val}'})
                continue

            # 找班別 ID
            shift_id = shift_by_name.get(shift_name)
            if shift_id is None:
                errors.append({'row': i, 'reason': f'找不到班別：{shift_name}'})
                continue

            # 驗證日期
            try:
                from datetime import date as _date
                _date.fromisoformat(date_str)
            except ValueError:
                errors.append({'row': i, 'reason': f'日期格式錯誤：{date_str}（應為 YYYY-MM-DD）'})
                continue

            # 排休衝突檢查
            if not force and date_str in leave_lookup.get(staff_id, set()):
                display = name_val or code_val
                skipped.append({'row': i, 'reason': f'{display} 於 {date_str} 有核准排休'})
                continue

            # 寫入（衝突則覆蓋）
            conn.execute("""
                INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date, note)
                VALUES (%s,%s,%s,%s)
                ON CONFLICT (staff_id, shift_date) DO UPDATE
                  SET shift_type_id=%s, note=%s, created_at=NOW()
            """, (staff_id, shift_id, date_str, note, shift_id, note))
            created += 1

    result = {'created': created, 'skipped': skipped, 'errors': errors}
    if errors or skipped:
        result['message'] = f'匯入完成：{created} 筆成功，{len(skipped)} 筆排休衝突跳過，{len(errors)} 筆錯誤'
    else:
        result['message'] = f'匯入完成：共 {created} 筆排班'
    return jsonify(result), 201


@app.route('/api/shifts/conflicts', methods=['GET'])
@require_module('sched')
def api_shift_conflicts():
    """
    偵測班表衝突與警示：
    - overtime_hours : 單班時數 > 10 小時
    - midnight_cross : 跨日班別（結束時間 < 開始時間）
    - consecutive_days : 連續排班 >= 6 天（6天警告，7天以上錯誤）
    """
    month = request.args.get('month', '')
    if not month:
        return jsonify({'error': '請指定月份'}), 400

    from datetime import date as _dc, timedelta as _tdc

    conflicts = []

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date,
                   ps.name  AS staff_name,
                   st.name  AS shift_name,
                   st.start_time, st.end_time
            FROM shift_assignments sa
            JOIN punch_staff  ps ON ps.id = sa.staff_id
            JOIN shift_types  st ON st.id = sa.shift_type_id
            WHERE TO_CHAR(sa.shift_date, 'YYYY-MM') = %s
            ORDER BY sa.staff_id, sa.shift_date
        """, (month,)).fetchall()

    # ── 每班時數 & 跨日 ────────────────────────────────────────────
    for r in rows:
        s = r['start_time'];  e = r['end_time']
        sm = s.hour * 60 + s.minute
        em = e.hour * 60 + e.minute
        cross = em < sm
        dur   = ((24 * 60 - sm) + em) if cross else (em - sm)
        hrs   = dur / 60

        if cross:
            conflicts.append({
                'type':       'midnight_cross',
                'severity':   'info',
                'date':       str(r['shift_date']),
                'staff_name': r['staff_name'],
                'shift_name': r['shift_name'],
                'message':    f"跨日班別 {str(s)[:5]}～{str(e)[:5]}（共 {hrs:.1f} 小時）",
            })

        if hrs > 10:
            conflicts.append({
                'type':       'overtime_hours',
                'severity':   'warning' if hrs <= 12 else 'error',
                'date':       str(r['shift_date']),
                'staff_name': r['staff_name'],
                'shift_name': r['shift_name'],
                'message':    f"單班 {hrs:.1f} 小時，超過 10 小時上限",
            })

    # ── 連續排班天數 ───────────────────────────────────────────────
    staff_dates = {}
    for r in rows:
        sid = r['staff_id']
        if sid not in staff_dates:
            staff_dates[sid] = {'name': r['staff_name'], 'dates': []}
        staff_dates[sid]['dates'].append(_dc.fromisoformat(str(r['shift_date'])))

    for sid, info in staff_dates.items():
        dates = sorted(set(info['dates']))
        streak = [dates[0]]
        for i in range(1, len(dates)):
            if (dates[i] - dates[i-1]).days == 1:
                streak.append(dates[i])
            else:
                # evaluate finished streak
                if len(streak) >= 6:
                    sev = 'error' if len(streak) >= 7 else 'warning'
                    conflicts.append({
                        'type':       'consecutive_days',
                        'severity':   sev,
                        'date':       streak[0].isoformat(),
                        'staff_name': info['name'],
                        'shift_name': '',
                        'message':    (
                            f"連續排班 {len(streak)} 天"
                            f"（{streak[0].isoformat()} ～ {streak[-1].isoformat()}）"
                            + ('，違反勞基法每 7 日至少休 1 日' if len(streak) >= 7 else '，接近法定上限')
                        ),
                    })
                streak = [dates[i]]
        # last streak
        if len(streak) >= 6:
            sev = 'error' if len(streak) >= 7 else 'warning'
            conflicts.append({
                'type':       'consecutive_days',
                'severity':   sev,
                'date':       streak[0].isoformat(),
                'staff_name': info['name'],
                'shift_name': '',
                'message':    (
                    f"連續排班 {len(streak)} 天"
                    f"（{streak[0].isoformat()} ～ {streak[-1].isoformat()}）"
                    + ('，違反勞基法每 7 日至少休 1 日' if len(streak) >= 7 else '，接近法定上限')
                ),
            })

    # sort: error first, then by date
    sev_order = {'error': 0, 'warning': 1, 'info': 2}
    conflicts.sort(key=lambda c: (sev_order.get(c['severity'], 9), c['date']))
    return jsonify({'month': month, 'count': len(conflicts), 'conflicts': conflicts})


@app.route('/api/shifts/export', methods=['GET'])
@require_module('sched')
def api_shift_export():
    """匯出指定月份班表為 Excel (.xlsx)"""
    month = request.args.get('month', '')
    if not month:
        return jsonify({'error': '請指定月份'}), 400

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import calendar as _cal2
    from datetime import date as _de

    y, mo = int(month[:4]), int(month[5:7])
    days_in_month = _cal2.monthrange(y, mo)[1]
    DAYS_CN = ['一','二','三','四','五','六','日']

    with get_db() as conn:
        staff_list = conn.execute(
            "SELECT id, name, employee_code, role FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        assigns = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, sa.note,
                   st.name AS shift_name, st.start_time, st.end_time
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE TO_CHAR(sa.shift_date, 'YYYY-MM') = %s
        """, (month,)).fetchall()
        holidays = {str(r['date']) for r in conn.execute(
            "SELECT date FROM public_holidays WHERE TO_CHAR(date,'YYYY-MM')=%s", (month,)
        ).fetchall()}

    lookup = {}
    for a in assigns:
        key = (a['staff_id'], str(a['shift_date']))
        lookup[key] = f"{a['shift_name']}\n{str(a['start_time'])[:5]}~{str(a['end_time'])[:5]}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month} 班表"

    # ── 樣式定義 ─────────────────────────────────────────────
    navy_fill   = PatternFill('solid', fgColor='0F1C3A')
    grey_fill   = PatternFill('solid', fgColor='F4F6FA')
    wkend_fill  = PatternFill('solid', fgColor='FFF5F5')
    hol_fill    = PatternFill('solid', fgColor='FFF0F0')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    info_font = Font(bold=True, size=10)
    cell_font = Font(size=9)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── 標題列 ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    for col, label in enumerate(['姓名', '代碼', '職稱'], start=1):
        c = ws.cell(1, col, label)
        c.font = hdr_font; c.fill = navy_fill; c.alignment = center; c.border = thin_border
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9

    for d in range(1, days_in_month + 1):
        col  = d + 3
        dt   = _de(y, mo, d)
        wd   = dt.weekday()    # 0=Mon … 6=Sun
        ds   = f"{month}-{d:02d}"
        is_wkend = wd >= 5      # Sat or Sun
        is_hol   = ds in holidays
        c = ws.cell(1, col, f"{d}\n{DAYS_CN[wd]}")
        c.font      = Font(bold=True, color='FF4444' if is_wkend or is_hol else 'FFFFFF', size=9)
        c.fill      = PatternFill('solid', fgColor='1A3060') if not (is_wkend or is_hol) else PatternFill('solid', fgColor='8B2020')
        c.alignment = center
        c.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 11

    # ── 員工列 ───────────────────────────────────────────────
    for row_idx, staff in enumerate(staff_list, start=2):
        ws.row_dimensions[row_idx].height = 30
        for col, val in enumerate([staff['name'], staff['employee_code'] or '', staff['role'] or ''], start=1):
            c = ws.cell(row_idx, col, val)
            c.font = info_font if col == 1 else cell_font
            c.fill = grey_fill; c.alignment = center; c.border = thin_border

        for d in range(1, days_in_month + 1):
            col = d + 3
            ds  = f"{month}-{d:02d}"
            dt  = _de(y, mo, d); wd = dt.weekday()
            val = lookup.get((staff['id'], ds), '')
            c   = ws.cell(row_idx, col, val)
            c.font      = Font(size=8, color='1A1A2E' if val else 'CCCCCC')
            c.alignment = center
            c.border    = thin_border
            if not val:
                c.fill = wkend_fill if wd >= 5 else (hol_fill if ds in holidays else PatternFill('solid', fgColor='FFFFFF'))

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(
        buf, as_attachment=True,
        download_name=f"班表_{month}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/shifts/my-schedule', methods=['GET'])
def api_my_shift_schedule():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    month = request.args.get('month', '')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT sa.shift_date, sa.note,
                   st.name as shift_name, st.start_time, st.end_time, st.color
            FROM shift_assignments sa
            JOIN shift_types st ON st.id=sa.shift_type_id
            WHERE sa.staff_id=%s
              AND to_char(sa.shift_date,'YYYY-MM')=%s
            ORDER BY sa.shift_date
        """, (sid, month)).fetchall()
    result = {}
    for r in rows:
        ds = r['shift_date'].isoformat()
        result[ds] = {
            'shift_name': r['shift_name'],
            'start_time': str(r['start_time'])[:5],
            'end_time':   str(r['end_time'])[:5],
            'color':      r['color'],
            'note':       r['note'],
        }
    return jsonify({'month': month, 'shifts': result})

# ── Overtime Requests ─────────────────────────────────────────────

@app.route('/api/overtime/my-requests', methods=['GET'])
def api_ot_my_list():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM overtime_requests WHERE staff_id=%s ORDER BY request_date DESC LIMIT 30",
            (sid,)
        ).fetchall()
    return jsonify([ot_req_row(r) for r in rows])


@app.route('/api/overtime/my-requests', methods=['POST'])
def api_ot_submit():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    b            = request.get_json(force=True)
    request_date = b.get('request_date', '').strip()
    start_time   = b.get('start_time', '').strip()
    end_time     = b.get('end_time', '').strip()
    reason       = b.get('reason', '').strip()
    day_type     = b.get('day_type', 'weekday').strip()
    if day_type not in ('weekday', 'rest_day', 'holiday', 'special'):
        day_type = 'weekday'
    if not request_date or not start_time or not end_time:
        return jsonify({'error': '請填寫加班日期及時間'}), 400
    if not reason:
        return jsonify({'error': '請填寫加班原因'}), 400
    from datetime import datetime as _dtot, timedelta as _tdot
    try:
        s = _dtot.strptime(start_time, '%H:%M')
        e = _dtot.strptime(end_time,   '%H:%M')
        if e <= s: e += _tdot(days=1)
        ot_hours = round((e - s).total_seconds() / 3600, 2)
    except ValueError:
        return jsonify({'error': '時間格式錯誤'}), 400
    if ot_hours <= 0 or ot_hours > 12:
        return jsonify({'error': '加班時數不合理（0~12小時）'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO overtime_requests
              (staff_id, request_date, start_time, end_time, ot_hours, reason, day_type)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (sid, request_date, start_time, end_time, ot_hours, reason, day_type)).fetchone()
    return jsonify(ot_req_row(row)), 201


@app.route('/api/overtime/requests', methods=['GET'])
@login_required
def api_ot_admin_list():
    status = request.args.get('status', '')
    month  = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if status: conds.append('r.status=%s');                          params.append(status)
    if month:  conds.append("to_char(r.request_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT r.*, ps.name as staff_name, ps.role as staff_role
            FROM overtime_requests r
            JOIN punch_staff ps ON ps.id=r.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY r.request_date DESC, r.created_at DESC
        """, params).fetchall()
    return jsonify([
        ot_req_row(r) | {'staff_name': r['staff_name'], 'staff_role': r['staff_role']}
        for r in rows
    ])


def _calc_ot_pay(staff_row, ot_hours, day_type='weekday'):
    salary_type = staff_row.get('salary_type', 'monthly') or 'monthly'
    base_salary = float(staff_row.get('base_salary')  or 0)
    hourly_rate = float(staff_row.get('hourly_rate')  or 0)
    daily_hours = float(staff_row.get('daily_hours')  or 8)
    ot_rate1    = float(staff_row.get('ot_rate1')     or 1.33)
    ot_rate2    = float(staff_row.get('ot_rate2')     or 1.67)
    ot_rate3    = float(staff_row.get('ot_rate3')     or 2.0)

    if salary_type == 'hourly':
        base_hourly = hourly_rate
    else:
        base_hourly = base_salary / 30 / daily_hours if (base_salary and daily_hours) else 0

    if base_hourly <= 0:
        return 0.0, base_hourly

    h = float(ot_hours)
    if day_type in ('holiday', 'special'):
        pay = round(base_hourly * h * 2.0, 0)
    elif day_type == 'rest_day':
        billed = max(h, 4.0)
        h1 = min(billed, 2.0); h2 = min(max(0.0, billed - 2.0), 2.0); h3 = max(0.0, billed - 4.0)
        pay = round(base_hourly * (h1 * ot_rate1 + h2 * ot_rate2 + h3 * ot_rate3), 0)
    else:
        h1 = min(h, 2.0); h2 = min(max(0.0, h - 2.0), 2.0); h3 = max(0.0, h - 4.0)
        pay = round(base_hourly * (h1 * ot_rate1 + h2 * ot_rate2 + h3 * ot_rate3), 0)

    return pay, base_hourly


@app.route('/api/overtime/requests/<int:rid>', methods=['PUT'])
@login_required
def api_ot_review(rid):
    b           = request.get_json(force=True)
    action      = b.get('action')
    reviewed_by = b.get('reviewed_by', '').strip()
    review_note = b.get('review_note', '').strip()
    if action not in ('approve', 'reject'):
        return jsonify({'error': 'invalid action'}), 400
    new_status   = 'approved' if action == 'approve' else 'rejected'
    ot_pay_final = 0.0

    with get_db() as conn:
        req = conn.execute(
            "SELECT * FROM overtime_requests WHERE id=%s", (rid,)
        ).fetchone()
        if not req: return ('', 404)

        if action == 'approve':
            staff = conn.execute("""
                SELECT base_salary, hourly_rate, daily_hours,
                       ot_rate1, ot_rate2, ot_rate3, salary_type
                FROM punch_staff WHERE id=%s
            """, (req['staff_id'],)).fetchone()
            if staff:
                dtype        = req.get('day_type', 'weekday') or 'weekday'
                ot_pay_final, _ = _calc_ot_pay(staff, req['ot_hours'] or 0, dtype)

        row = conn.execute("""
            UPDATE overtime_requests
            SET status=%s, reviewed_by=%s, review_note=%s,
                ot_pay=%s, reviewed_at=NOW()
            WHERE id=%s RETURNING *
        """, (new_status, reviewed_by, review_note, ot_pay_final, rid)).fetchone()

        # 加班狀態改變 → 刪除該月 draft 薪資，確保下次產生時重算加班費
        if row:
            ot_month = str(req['request_date'])[:7]
            conn.execute("""
                DELETE FROM salary_records
                WHERE staff_id=%s AND month=%s AND status='draft'
            """, (req['staff_id'], ot_month))

        sn = conn.execute(
            "SELECT name FROM punch_staff WHERE id=%s", (req['staff_id'],)
        ).fetchone()

    result = ot_req_row(row)
    result['staff_name'] = sn['name'] if sn else ''
    # LINE notification
    time_str = (f"{row['start_time']}～{row['end_time']}" if row.get('start_time') and row.get('end_time')
                else f"{float(row['ot_hours'])} 小時")
    extra = f"{row['request_date']} {time_str}"
    if action == 'approve' and float(row.get('ot_pay') or 0) > 0:
        extra += f"\n加班費：${float(row['ot_pay']):,.0f}"
    if review_note: extra += f"\n審核意見：{review_note}"
    _notify_review_result(req['staff_id'], '加班申請', action, extra)
    return jsonify(result)


@app.route('/api/overtime/requests/<int:rid>', methods=['DELETE'])
@login_required
def api_ot_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM overtime_requests WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})


@app.route('/api/overtime/monthly-summary', methods=['GET'])
@login_required
def api_ot_monthly_summary():
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                ps.id   AS staff_id,
                ps.name AS staff_name,
                ps.role AS staff_role,
                COUNT(*)                                                      AS request_count,
                SUM(r.ot_hours)                                               AS total_hours,
                SUM(CASE WHEN r.status='approved' THEN r.ot_hours ELSE 0 END) AS approved_hours,
                SUM(CASE WHEN r.status='pending'  THEN r.ot_hours ELSE 0 END) AS pending_hours,
                SUM(CASE WHEN r.status='rejected' THEN r.ot_hours ELSE 0 END) AS rejected_hours,
                COUNT(CASE WHEN r.status='approved' THEN 1 END)               AS approved_count,
                COUNT(CASE WHEN r.status='pending'  THEN 1 END)               AS pending_count,
                COUNT(CASE WHEN r.status='rejected' THEN 1 END)               AS rejected_count
            FROM overtime_requests r
            JOIN punch_staff ps ON ps.id = r.staff_id
            WHERE to_char(r.request_date, 'YYYY-MM') = %s
            GROUP BY ps.id, ps.name, ps.role
            ORDER BY total_hours DESC
        """, (month,)).fetchall()
    return jsonify([{
        'staff_id':       r['staff_id'],
        'staff_name':     r['staff_name'],
        'staff_role':     r['staff_role'] or '',
        'request_count':  r['request_count'],
        'total_hours':    float(r['total_hours']    or 0),
        'approved_hours': float(r['approved_hours'] or 0),
        'pending_hours':  float(r['pending_hours']  or 0),
        'rejected_hours': float(r['rejected_hours'] or 0),
        'approved_count': r['approved_count'],
        'pending_count':  r['pending_count'],
        'rejected_count': r['rejected_count'],
    } for r in rows])


@app.route('/api/overtime/calc-preview', methods=['POST'])
@login_required
def api_ot_calc_preview():
    b        = request.get_json(force=True)
    staff_id = b.get('staff_id')
    ot_hours = float(b.get('ot_hours') or 0)
    if not staff_id: return jsonify({'error': 'staff_id required'}), 400
    with get_db() as conn:
        staff = conn.execute("""
            SELECT name, base_salary, hourly_rate, daily_hours,
                   ot_rate1, ot_rate2, ot_rate3, salary_type
            FROM punch_staff WHERE id=%s
        """, (staff_id,)).fetchone()
    if not staff: return ('', 404)
    day_type     = b.get('day_type', 'weekday') or 'weekday'
    ot_pay, base_hourly = _calc_ot_pay(staff, ot_hours, day_type)

    if day_type == 'rest_day':
        billed = max(ot_hours, 4.0)
        h1 = min(billed, 2.0); h2 = min(max(0.0, billed - 2.0), 2.0); h3 = max(0.0, billed - 4.0)
    elif day_type in ('holiday', 'special'):
        h1 = ot_hours; h2 = 0.0; h3 = 0.0
    else:
        h1 = min(ot_hours, 2.0); h2 = min(max(0.0, ot_hours - 2.0), 2.0); h3 = max(0.0, ot_hours - 4.0)

    return jsonify({
        'staff_name':  staff['name'],
        'salary_type': staff.get('salary_type', 'monthly'),
        'base_salary': float(staff.get('base_salary') or 0),
        'hourly_rate': float(staff.get('hourly_rate') or 0),
        'base_hourly': round(base_hourly, 2),
        'ot_hours':    ot_hours,
        'day_type':    day_type,
        'h1':          h1,
        'h2':          h2,
        'h3':          h3,
        'ot_rate1':    float(staff.get('ot_rate1') or 1.33),
        'ot_rate2':    float(staff.get('ot_rate2') or 1.67),
        'ot_rate3':    float(staff.get('ot_rate3') or 2.0),
        'ot_pay':      ot_pay,
    })


# ═══════════════════════════════════════════════════════════════════
# Leave Management (請假管理)
# 2026 勞基法：
#   特休：到職1年10天、2年15天、3~5年每年+1、滿5年20天(上限)
#   病假：每年30天(半薪)，超過住院病假 365 天內 30 天(全薪)
#   事假：每年14天(無薪)
#   生理假：每月1天(含病假計算，前3天半薪)
#   婚假：8天全薪
#   喪假：父母/配偶/子女8天；祖父母/孫子女/兄弟姐妹6天；曾祖父母3天
#   產假：8週全薪；陪產假：7天全薪
#   公假：全薪
# ═══════════════════════════════════════════════════════════════════

# ── Leave Tables ─────────────────────────────────────────────────────────────

def init_leave_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS leave_types (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL UNIQUE,
            code        TEXT NOT NULL UNIQUE,
            pay_rate    NUMERIC(4,2) DEFAULT 1.0,
            max_days    NUMERIC(5,1),
            description TEXT DEFAULT '',
            color       TEXT DEFAULT '#4a7bda',
            active      BOOLEAN DEFAULT TRUE,
            sort_order  INT DEFAULT 0,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS leave_requests (
            id              SERIAL PRIMARY KEY,
            staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            leave_type_id   INT REFERENCES leave_types(id),
            start_date      DATE NOT NULL,
            end_date        DATE NOT NULL,
            start_half      BOOLEAN DEFAULT FALSE,
            end_half        BOOLEAN DEFAULT FALSE,
            total_days      NUMERIC(5,1) NOT NULL DEFAULT 0,
            reason          TEXT DEFAULT '',
            status          TEXT DEFAULT 'pending',
            reviewed_by     TEXT DEFAULT '',
            review_note     TEXT DEFAULT '',
            reviewed_at     TIMESTAMPTZ,
            substitute_name TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS leave_balances (
            id          SERIAL PRIMARY KEY,
            staff_id    INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            leave_type_id INT REFERENCES leave_types(id),
            year        INT NOT NULL,
            total_days  NUMERIC(5,1) DEFAULT 0,
            used_days   NUMERIC(5,1) DEFAULT 0,
            note        TEXT DEFAULT '',
            updated_at  TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(staff_id, leave_type_id, year)
        )""",
        "ALTER TABLE leave_types ADD COLUMN IF NOT EXISTS allow_hourly BOOLEAN DEFAULT FALSE",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS total_hours NUMERIC(5,1)",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[leave_init] {str(e)[:80]}")

    # Seed default leave types
    # (name, code, pay_rate, max_days, color, sort_order, allow_hourly)
    defaults = [
        ('特休假',   'annual',       1.0,  30,  '#2e9e6b', 1,  False),
        ('病假',     'sick',         0.5,  30,  '#e07b2a', 2,  True),
        ('住院病假', 'hospitalize',  1.0,  30,  '#d64242', 3,  False),
        ('事假',     'personal',     0.0,  14,  '#8892a4', 4,  True),
        ('生理假',   'menstrual',    0.5,  12,  '#c45cb8', 5,  False),
        ('婚假',     'marriage',     1.0,   8,  '#c8a96e', 6,  False),
        ('喪假',     'funeral',      1.0,   8,  '#4a7bda', 7,  False),
        ('產假',     'maternity',    1.0,  56,  '#e05c8a', 8,  False),
        ('陪產假',   'paternity',    1.0,   7,  '#5cb8c4', 9,  False),
        ('公假',     'official',     1.0, None, '#243d6e', 10, False),
        ('補休',     'compensatory', 1.0, None, '#8b5cf6', 11, False),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM leave_types").fetchone()['c']
            if cnt == 0:
                for name, code, pay, maxd, color, sort, allow_hourly in defaults:
                    conn.execute(
                        "INSERT INTO leave_types (name,code,pay_rate,max_days,color,sort_order,allow_hourly) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                        (name, code, pay, maxd, color, sort, allow_hourly)
                    )
    except Exception as e:
        print(f"[leave_seed] {e}")

init_leave_db()

def leave_type_row(row):
    if not row: return None
    d = dict(row)
    if d.get('max_days') is not None: d['max_days'] = float(d['max_days'])
    if d.get('pay_rate') is not None: d['pay_rate'] = float(d['pay_rate'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

def leave_req_row(row):
    if not row: return None
    d = dict(row)
    if d.get('start_date'): d['start_date'] = d['start_date'].isoformat()
    if d.get('end_date'):   d['end_date']   = d['end_date'].isoformat()
    if d.get('total_days'): d['total_days'] = float(d['total_days'])
    if d.get('total_hours') is not None: d['total_hours'] = float(d['total_hours'])
    if d.get('reviewed_at'): d['reviewed_at'] = d['reviewed_at'].isoformat()
    if d.get('created_at'):  d['created_at']  = d['created_at'].isoformat()
    if d.get('updated_at'):  d['updated_at']  = d['updated_at'].isoformat()
    return d

def leave_balance_row(row):
    if not row: return None
    d = dict(row)
    if d.get('total_days') is not None: d['total_days'] = float(d['total_days'])
    if d.get('used_days')  is not None: d['used_days']  = float(d['used_days'])
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    return d

def _calc_annual_leave_days(hire_date_str, ref_date_str=None):
    """
    勞基法第38條特休天數計算（2017年修正版，現行有效）

    到職滿6個月：3天
    到職滿1年：7天
    到職滿2年：10天
    到職滿3年：14天
    到職滿4年：14天（同第3年）
    到職滿5年：15天
    到職滿6～9年：15天（同第5年）
    到職滿10年起：每年+1天，上限30天

    回傳當期應給特休天數（整數）
    """
    if not hire_date_str:
        return 0
    from datetime import date as _date
    try:
        hire = _date.fromisoformat(str(hire_date_str))
    except Exception:
        return 0

    ref = _date.today()
    if ref_date_str:
        try:
            ref = _date.fromisoformat(str(ref_date_str))
        except Exception:
            pass

    # 計算到職滿幾個月（以完整月份計）
    months = (ref.year - hire.year) * 12 + (ref.month - hire.month)
    # 若當月日期未到到職日，扣一個月
    if ref.day < hire.day:
        months -= 1
    if months < 0:
        months = 0

    # 正確換算年數（以整月為準）
    years_complete = months // 12
    months_extra   = months % 12

    # 勞基法第38條逐段對應
    if months < 6:
        return 0
    elif months < 12:
        # 滿6個月未滿1年：3天
        return 3
    elif years_complete < 2:
        # 滿1年未滿2年：7天
        return 7
    elif years_complete < 3:
        # 滿2年未滿3年：10天
        return 10
    elif years_complete < 5:
        # 滿3年未滿5年：14天
        return 14
    elif years_complete < 10:
        # 滿5年未滿10年：15天
        return 15
    else:
        # 滿10年：16天，之後每年+1，上限30天
        # years_complete=10 → extra=1 → 15+1=16 ✓
        extra = years_complete - 9
        return min(15 + extra, 30)


def _calc_annual_leave_schedule(hire_date_str):
    """
    回傳員工特休天數完整排程表，供前端顯示用。
    每一列：{ label, days, date_reached, is_past, is_current }
    """
    if not hire_date_str:
        return []
    from datetime import date as _date
    import calendar as _cal

    try:
        hire = _date.fromisoformat(str(hire_date_str))
    except Exception:
        return []

    today = _date.today()

    milestones = [
        (6,   3,  '滿6個月'),
        (12,  7,  '滿1年'),
        (24, 10,  '滿2年'),
        (36, 14,  '滿3年'),
        (60, 15,  '滿5年'),
        (120,16,  '滿10年'),
        (132,17,  '滿11年'),
        (144,18,  '滿12年'),
        (156,19,  '滿13年'),
        (168,20,  '滿14年'),
        (180,21,  '滿15年'),
        (192,22,  '滿16年'),
        (204,23,  '滿17年'),
        (216,24,  '滿18年'),
        (228,25,  '滿19年'),
        (240,30,  '滿20年（上限30天）'),
    ]

    result      = []
    current_days = _calc_annual_leave_days(hire_date_str)

    for months_needed, days, label in milestones:
        total_m = hire.month + months_needed
        y = hire.year + (total_m - 1) // 12
        m = (total_m - 1) % 12 + 1
        max_day = _cal.monthrange(y, m)[1]
        try:
            reached = _date(y, m, min(hire.day, max_day))
        except Exception:
            continue

        result.append({
            'label':        label,
            'days':         days,
            'date_reached': reached.isoformat(),
            'is_past':      reached <= today,
            'is_current':   (days == current_days and reached <= today),
        })

    return result

def _calc_leave_days(start_date_str, end_date_str, start_half=False, end_half=False):
    """計算請假天數（含半天選項），排除週日"""
    from datetime import date as _date, timedelta as _tdd
    try:
        s = _date.fromisoformat(start_date_str)
        e = _date.fromisoformat(end_date_str)
    except Exception:
        return 0.0
    if e < s: return 0.0
    days = 0.0
    cur  = s
    while cur <= e:
        if cur.weekday() != 6:  # exclude Sunday (勞基法最低標準)
            if cur == s and start_half and cur == e and end_half:
                days += 1.0  # same day: AM half + PM half = full day
            elif cur == s and start_half:
                days += 0.5
            elif cur == e and end_half:
                days += 0.5
            else:
                days += 1.0
        cur += _tdd(days=1)
    return days

# ── Leave Type CRUD ──────────────────────────────────────────────

@app.route('/api/leave/types', methods=['GET'])
@require_module('leave')
def api_leave_types_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM leave_types ORDER BY sort_order, id").fetchall()
    return jsonify([leave_type_row(r) for r in rows])

@app.route('/api/leave/types/public', methods=['GET'])
def api_leave_types_public():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM leave_types WHERE active=TRUE ORDER BY sort_order, id").fetchall()
    return jsonify([leave_type_row(r) for r in rows])

@app.route('/api/leave/types', methods=['POST'])
@require_module('leave')
def api_leave_type_create():
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO leave_types (name,code,pay_rate,max_days,description,color,sort_order)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'], b['code'], float(b.get('pay_rate',1.0)),
              b.get('max_days') or None, b.get('description',''),
              b.get('color','#4a7bda'), int(b.get('sort_order',0)))).fetchone()
    return jsonify(leave_type_row(row)), 201

@app.route('/api/leave/types/<int:tid>', methods=['PUT'])
@require_module('leave')
def api_leave_type_update(tid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE leave_types SET name=%s,code=%s,pay_rate=%s,max_days=%s,
              description=%s,color=%s,sort_order=%s,active=%s
            WHERE id=%s RETURNING *
        """, (b['name'], b['code'], float(b.get('pay_rate',1.0)),
              b.get('max_days') or None, b.get('description',''),
              b.get('color','#4a7bda'), int(b.get('sort_order',0)),
              bool(b.get('active',True)), tid)).fetchone()
    return jsonify(leave_type_row(row)) if row else ('', 404)

@app.route('/api/leave/types/<int:tid>', methods=['DELETE'])
@require_module('leave')
def api_leave_type_delete(tid):
    with get_db() as conn:
        conn.execute("DELETE FROM leave_types WHERE id=%s", (tid,))
    return jsonify({'deleted': tid})

# ── Leave Requests ────────────────────────────────────────────────

@app.route('/api/leave/requests', methods=['GET'])
@require_module('leave')
def api_leave_requests_list():
    status   = request.args.get('status', '')
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    conds, params = ['TRUE'], []
    if status:   conds.append('lr.status=%s');                            params.append(status)
    if staff_id: conds.append('lr.staff_id=%s');                          params.append(int(staff_id))
    if month:    conds.append("to_char(lr.start_date,'YYYY-MM')=%s");     params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.role as staff_role,
                   lt.name as leave_type_name, lt.code as leave_code,
                   lt.pay_rate, lt.color as leave_color
            FROM leave_requests lr
            JOIN punch_staff ps ON ps.id=lr.staff_id
            JOIN leave_types  lt ON lt.id=lr.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY lr.start_date DESC, lr.created_at DESC LIMIT 300
        """, params).fetchall()
    result = []
    for r in rows:
        d = leave_req_row(r)
        d['staff_name']      = r['staff_name']
        d['staff_role']      = r['staff_role']
        d['leave_type_name'] = r['leave_type_name']
        d['leave_code']      = r['leave_code']
        d['pay_rate']        = float(r['pay_rate'])
        d['leave_color']     = r['leave_color']
        result.append(d)
    return jsonify(result)

@app.route('/api/leave/requests', methods=['POST'])
@require_module('leave')
def api_leave_request_admin_create():
    """管理員直接建立請假記錄"""
    b = request.get_json(force=True)
    sid           = b.get('staff_id')
    leave_type_id = b.get('leave_type_id')
    start_date    = b.get('start_date', '').strip()
    end_date      = b.get('end_date', '').strip()
    start_half    = bool(b.get('start_half', False))
    end_half      = bool(b.get('end_half', False))
    reason        = b.get('reason', '').strip()
    status        = b.get('status', 'approved')

    if not all([sid, leave_type_id, start_date, end_date]):
        return jsonify({'error': '缺少必要欄位'}), 400

    total_hours_req = b.get('total_hours')
    if total_hours_req is not None:
        try:
            total_hours_req = float(total_hours_req)
        except (ValueError, TypeError):
            total_hours_req = None

    if total_hours_req:
        if total_hours_req < 0.5 or total_hours_req > 8:
            return jsonify({'error': '時數需介於 0.5～8 小時'}), 400
        total_days = round(total_hours_req / 8, 4)
        end_date   = start_date
        start_half = False
        end_half   = False
    else:
        total_hours_req = None
        total_days = _calc_leave_days(start_date, end_date, start_half, end_half)
        if total_days <= 0:
            return jsonify({'error': '請假天數不合理，請檢查日期'}), 400

    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO leave_requests
              (staff_id, leave_type_id, start_date, end_date, start_half, end_half,
               total_days, total_hours, reason, status, reviewed_by, reviewed_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
              CASE WHEN %s='approved' THEN NOW() ELSE NULL END)
            RETURNING *
        """, (sid, leave_type_id, start_date, end_date, start_half, end_half,
              total_days, total_hours_req, reason, status,
              b.get('reviewed_by','管理員'), status)).fetchone()
        if status == 'approved':
            _update_leave_balance(conn, sid, leave_type_id, start_date[:4], total_days)
    return jsonify(leave_req_row(row)), 201

@app.route('/api/leave/requests/<int:rid>', methods=['PUT'])
@require_module('leave')
def api_leave_request_review(rid):
    b           = request.get_json(force=True)
    action      = b.get('action')
    reviewed_by = b.get('reviewed_by', '').strip()
    review_note = b.get('review_note', '').strip()
    if action not in ('approve', 'reject'):
        return jsonify({'error': 'invalid action'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as conn:
        old = conn.execute("SELECT * FROM leave_requests WHERE id=%s", (rid,)).fetchone()
        if not old: return ('', 404)
        old_status = old['status']
        row = conn.execute("""
            UPDATE leave_requests
            SET status=%s, reviewed_by=%s, review_note=%s,
                reviewed_at=NOW(), updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (new_status, reviewed_by, review_note, rid)).fetchone()
        delta = float(old['total_days'])
        lt = conn.execute("SELECT * FROM leave_types WHERE id=%s", (old['leave_type_id'],)).fetchone()
        if action == 'approve' and old_status != 'approved':
            if lt and lt['max_days'] is not None:
                year = str(old['start_date'])[:4]
                # 確保餘額列存在，再鎖定防止並行核准超額
                conn.execute("""
                    INSERT INTO leave_balances (staff_id, leave_type_id, year, total_days, used_days)
                    VALUES (%s, %s, %s, 0, 0)
                    ON CONFLICT (staff_id, leave_type_id, year) DO NOTHING
                """, (old['staff_id'], old['leave_type_id'], int(year)))
                bal = conn.execute("""
                    SELECT COALESCE(used_days, 0) as used
                    FROM leave_balances
                    WHERE staff_id=%s AND leave_type_id=%s AND year=%s
                    FOR UPDATE
                """, (old['staff_id'], old['leave_type_id'], int(year))).fetchone()
                used = float(bal['used']) if bal else 0.0
                if used + delta > float(lt['max_days']):
                    remaining = float(lt['max_days']) - used
                    return jsonify({'error': f'{lt["name"]}餘額不足（剩 {remaining} 天），無法核准'}), 422
            _update_leave_balance(conn, old['staff_id'], old['leave_type_id'],
                                  str(old['start_date'])[:4], delta)
        elif action == 'reject' and old_status == 'approved':
            # 已核准的假單被改拒絕：補回餘額
            _update_leave_balance(conn, old['staff_id'], old['leave_type_id'],
                                  str(old['start_date'])[:4], -delta)

        # 請假狀態改變 → 刪除受影響月份的 draft 薪資，確保重算扣款
        if old_status != new_status:
            affected_months = {str(old['start_date'])[:7], str(old['end_date'])[:7]}
            for _m in affected_months:
                conn.execute("""
                    DELETE FROM salary_records
                    WHERE staff_id=%s AND month=%s AND status='draft'
                """, (old['staff_id'], _m))
    if row:
        total_hours = old.get('total_hours')
        duration_str = f"{float(total_hours)} 小時" if total_hours else f"{float(old['total_days'])} 天"
        extra = f"{str(old['start_date'])} ~ {str(old['end_date'])} 共 {duration_str}"
        if review_note: extra += f"\n審核意見：{review_note}"
        _notify_review_result(old['staff_id'], '請假申請', action, extra)
    return jsonify(leave_req_row(row)) if row else ('', 404)

@app.route('/api/leave/requests/<int:rid>', methods=['DELETE'])
@require_module('leave')
def api_leave_request_delete(rid):
    with get_db() as conn:
        old = conn.execute("SELECT * FROM leave_requests WHERE id=%s", (rid,)).fetchone()
        if not old:
            return jsonify({'error': '找不到假單'}), 404
        # 若已核准，刪除前先補回餘額並清除 draft 薪資
        if old['status'] == 'approved':
            _update_leave_balance(conn, old['staff_id'], old['leave_type_id'],
                                  str(old['start_date'])[:4], -float(old['total_days']))
            affected_months = {str(old['start_date'])[:7], str(old['end_date'])[:7]}
            for _m in affected_months:
                conn.execute("""
                    DELETE FROM salary_records
                    WHERE staff_id=%s AND month=%s AND status='draft'
                """, (old['staff_id'], _m))
        conn.execute("DELETE FROM leave_requests WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

def _update_leave_balance(conn, staff_id, leave_type_id, year_str, delta_days):
    year = int(year_str)
    conn.execute("""
        INSERT INTO leave_balances (staff_id, leave_type_id, year, total_days, used_days)
        VALUES (%s, %s, %s, 0, %s)
        ON CONFLICT (staff_id, leave_type_id, year) DO UPDATE
          SET used_days = leave_balances.used_days + EXCLUDED.used_days,
              updated_at = NOW()
    """, (staff_id, leave_type_id, year, delta_days))

# ── Employee: submit leave request ────────────────────────────────

@app.route('/api/leave/my-requests', methods=['GET'])
def api_leave_my_list():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lr.*, lt.name as leave_type_name, lt.code as leave_code,
                   lt.color as leave_color, lt.pay_rate
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE lr.staff_id=%s
            ORDER BY lr.start_date DESC LIMIT 30
        """, (sid,)).fetchall()
    result = []
    for r in rows:
        d = leave_req_row(r)
        d['leave_type_name'] = r['leave_type_name']
        d['leave_code']      = r['leave_code']
        d['leave_color']     = r['leave_color']
        d['pay_rate']        = float(r['pay_rate'])
        result.append(d)
    return jsonify(result)

@app.route('/api/leave/my-requests', methods=['POST'])
def api_leave_submit():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    b             = request.get_json(force=True)
    leave_type_id = b.get('leave_type_id')
    start_date    = b.get('start_date', '').strip()
    end_date      = b.get('end_date',   '').strip()
    start_half    = bool(b.get('start_half', False))
    end_half      = bool(b.get('end_half',   False))
    reason        = b.get('reason', '').strip()
    substitute    = b.get('substitute_name', '').strip()
    document_id   = b.get('document_id') or None

    if not all([leave_type_id, start_date, end_date]):
        return jsonify({'error': '缺少必要欄位'}), 400

    total_hours_req = b.get('total_hours')
    if total_hours_req is not None:
        try:
            total_hours_req = float(total_hours_req)
        except (ValueError, TypeError):
            total_hours_req = None

    if total_hours_req:
        if total_hours_req <= 0 or total_hours_req > 24:
            return jsonify({'error': '請假時數不合理（需介於 0～24 小時）'}), 400
        total_days = round(total_hours_req / 8, 4)
        end_date   = start_date
        start_half = False
        end_half   = False
    else:
        total_hours_req = None
        total_days = _calc_leave_days(start_date, end_date, start_half, end_half)
        if total_days <= 0:
            return jsonify({'error': '請假天數不合理，請檢查日期'}), 400

    with get_db() as conn:
        lt = conn.execute("SELECT * FROM leave_types WHERE id=%s", (leave_type_id,)).fetchone()
        if lt and lt['max_days'] is not None:
            year = start_date[:4]
            # 確保餘額列存在，避免 FOR UPDATE 鎖不到列
            conn.execute("""
                INSERT INTO leave_balances (staff_id, leave_type_id, year, total_days, used_days)
                VALUES (%s, %s, %s, 0, 0)
                ON CONFLICT (staff_id, leave_type_id, year) DO NOTHING
            """, (sid, leave_type_id, int(year)))
            # 鎖定餘額列，防止並行請假超額
            bal = conn.execute("""
                SELECT COALESCE(used_days, 0) as used
                FROM leave_balances
                WHERE staff_id=%s AND leave_type_id=%s AND year=%s
                FOR UPDATE
            """, (sid, leave_type_id, int(year))).fetchone()
            used = float(bal['used']) if bal else 0.0
            if used + total_days > float(lt['max_days']):
                remaining = float(lt['max_days']) - used
                if total_hours_req:
                    rem_hours = round(remaining * 8, 1)
                    return jsonify({'error': f'{lt["name"]}剩餘 {rem_hours} 小時，無法申請 {total_hours_req} 小時'}), 422
                return jsonify({'error': f'{lt["name"]}剩餘 {remaining} 天，無法申請 {total_days} 天'}), 422

        row = conn.execute("""
            INSERT INTO leave_requests
              (staff_id, leave_type_id, start_date, end_date, start_half, end_half,
               total_days, total_hours, reason, substitute_name, document_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (sid, leave_type_id, start_date, end_date, start_half, end_half,
              total_days, total_hours_req, reason, substitute, document_id)).fetchone()
    return jsonify(leave_req_row(row)), 201

# ── Leave Balance ─────────────────────────────────────────────────

@app.route('/api/leave/balances', methods=['GET'])
def api_leave_balances():
    """管理員和員工都可以查詢，員工只能查自己的"""
    year     = request.args.get('year', '')
    staff_id = request.args.get('staff_id', '')

    # 員工端：只能查自己
    if not session.get('logged_in'):
        sid = session.get('punch_staff_id')
        if not sid:
            return jsonify({'error': 'not logged in'}), 401
        staff_id = str(sid)   # 強制只查自己
    if not year:
        from datetime import date as _d2
        year = str(_d2.today().year)
    conds, params = ["lb.year=%s"], [int(year)]
    if staff_id: conds.append("lb.staff_id=%s"); params.append(int(staff_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lb.*, ps.name as staff_name, lt.name as leave_type_name,
                   lt.code as leave_code, lt.max_days, lt.color as leave_color
            FROM leave_balances lb
            JOIN punch_staff  ps ON ps.id=lb.staff_id
            JOIN leave_types  lt ON lt.id=lb.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY ps.name, lt.sort_order
        """, params).fetchall()
    result = []
    for r in rows:
        d = leave_balance_row(r)
        d['staff_name']      = r['staff_name']
        d['leave_type_name'] = r['leave_type_name']
        d['leave_code']      = r['leave_code']
        d['leave_color']     = r['leave_color']
        d['max_days']        = float(r['max_days']) if r['max_days'] is not None else None
        result.append(d)
    return jsonify(result)

@app.route('/api/leave/balances/init', methods=['POST'])
@require_module('leave')
def api_leave_balance_init():
    """初始化/更新員工特休天數（依勞基法第38條，以到職日精確計算）"""
    b    = request.get_json(force=True)
    year = b.get('year', '')
    if not year:
        from datetime import date as _d3
        year = str(_d3.today().year)

    with get_db() as conn:
        staff_list = conn.execute(
            "SELECT id, name, hire_date FROM punch_staff WHERE active=TRUE"
        ).fetchall()
        lt = conn.execute("SELECT id FROM leave_types WHERE code='annual'").fetchone()
        if not lt: return jsonify({'error': '找不到特休假類型'}), 404
        lt_id   = lt['id']
        updated = 0
        details = []

        for s in staff_list:
            days = _calc_annual_leave_days(s['hire_date'])
            # 未滿6個月的員工也記錄（0天），方便後續追蹤
            conn.execute("""
                INSERT INTO leave_balances (staff_id, leave_type_id, year, total_days, used_days)
                VALUES (%s,%s,%s,%s,0)
                ON CONFLICT (staff_id, leave_type_id, year) DO UPDATE
                  SET total_days=EXCLUDED.total_days, updated_at=NOW()
            """, (s['id'], lt_id, int(year), days))
            updated += 1
            details.append({
                'name':      s['name'],
                'hire_date': str(s['hire_date']) if s['hire_date'] else None,
                'days':      days,
            })

    return jsonify({'ok': True, 'updated': updated, 'year': year, 'details': details})


@app.route('/api/leave/annual-schedule/<int:staff_id>', methods=['GET'])
@require_module('leave')
def api_annual_leave_schedule(staff_id):
    """回傳員工特休天數完整排程（各里程碑日期與天數）"""
    with get_db() as conn:
        staff = conn.execute(
            "SELECT name, hire_date FROM punch_staff WHERE id=%s", (staff_id,)
        ).fetchone()
    if not staff:
        return ('', 404)
    schedule = _calc_annual_leave_schedule(staff['hire_date'])
    current  = _calc_annual_leave_days(staff['hire_date'])
    return jsonify({
        'staff_id':      staff_id,
        'name':          staff['name'],
        'hire_date':     str(staff['hire_date']) if staff['hire_date'] else None,
        'current_days':  current,
        'schedule':      schedule,
    })


@app.route('/api/leave/annual-schedule/public', methods=['GET'])
def api_annual_leave_schedule_public():
    """員工查看自己的特休排程"""
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        staff = conn.execute(
            "SELECT name, hire_date FROM punch_staff WHERE id=%s", (sid,)
        ).fetchone()
    if not staff:
        return ('', 404)
    schedule = _calc_annual_leave_schedule(staff['hire_date'])
    current  = _calc_annual_leave_days(staff['hire_date'])
    return jsonify({
        'name':         staff['name'],
        'hire_date':    str(staff['hire_date']) if staff['hire_date'] else None,
        'current_days': current,
        'schedule':     schedule,
    })


@app.route('/api/leave/balances/<int:bid>', methods=['PUT'])
@require_module('leave')
def api_leave_balance_update(bid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE leave_balances SET total_days=%s, used_days=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (float(b.get('total_days',0)), float(b.get('used_days',0)),
              b.get('note',''), bid)).fetchone()
    return jsonify(leave_balance_row(row)) if row else ('', 404)

# ── Leave Summary (for salary integration) ───────────────────────

@app.route('/api/leave/summary/<int:staff_id>/<month>', methods=['GET'])
@require_module('leave')
def api_leave_summary(staff_id, month):
    """取得員工某月請假摘要（供薪資計算用）"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lr.*, lt.name as leave_type_name, lt.code, lt.pay_rate
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE lr.staff_id=%s
              AND lr.status='approved'
              AND to_char(lr.start_date,'YYYY-MM')=%s
            ORDER BY lr.start_date
        """, (staff_id, month)).fetchall()
    total_leave_days = 0.0
    unpaid_days      = 0.0
    half_pay_days    = 0.0
    items = []
    for r in rows:
        d = float(r['total_days'])
        pay_r = float(r['pay_rate'])
        total_leave_days += d
        if pay_r == 0:   unpaid_days   += d
        elif pay_r < 1:  half_pay_days += d
        items.append({
            'leave_type': r['leave_type_name'],
            'code':       r['code'],
            'days':       d,
            'pay_rate':   pay_r,
            'start_date': r['start_date'].isoformat(),
            'end_date':   r['end_date'].isoformat(),
        })
    return jsonify({
        'staff_id':         staff_id,
        'month':            month,
        'total_leave_days': total_leave_days,
        'unpaid_days':      unpaid_days,
        'half_pay_days':    half_pay_days,
        'items':            items,
    })

# ═══════════════════════════════════════════════════════════════════
# Salary Management (薪資管理)
# 2026 費率基準：
#   勞保費率 12.5%（員工負擔 20% = 2.5%，含就業保險1%由雇主全額負擔）
#   健保費率 5.17%（員工負擔 30% = 1.551%）
#   勞退提撥 6%（雇主強制提撥，員工自願另計）
#   最低工資 2026年 NT$28,590（月薪）/ NT$190（時薪）
# ═══════════════════════════════════════════════════════════════════

def _get_salary_calc_settings():
    """讀取薪資計算設定（帶預設值）"""
    try:
        with get_db() as conn:
            rows = conn.execute(
                "SELECT setting_key, setting_value FROM salary_calc_settings"
            ).fetchall()
            cfg = {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        cfg = {}
    return {
        'auto_leave_deduction':  cfg.get('auto_leave_deduction',  'true') == 'true',
        'auto_absent_deduction': cfg.get('auto_absent_deduction', 'true') == 'true',
        'auto_income_tax':       cfg.get('auto_income_tax',       'true') == 'true',
    }

def init_salary_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS salary_calc_settings (
            setting_key   TEXT PRIMARY KEY,
            setting_value TEXT NOT NULL DEFAULT 'true'
        )""",
        """CREATE TABLE IF NOT EXISTS salary_items (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            item_type   TEXT NOT NULL DEFAULT 'allowance',
            formula     TEXT DEFAULT '',
            amount      NUMERIC(12,2) DEFAULT 0,
            description TEXT DEFAULT '',
            color       TEXT DEFAULT '#4a7bda',
            active      BOOLEAN DEFAULT TRUE,
            sort_order  INT DEFAULT 0,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_item_ids JSONB DEFAULT NULL",
        "ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS salary_item_overrides JSONB DEFAULT NULL",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS income_tax_withheld    NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS absent_days           NUMERIC(5,1)  DEFAULT 0",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS whole_day_leave_days  NUMERIC(5,1)  DEFAULT 0",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS hourly_base_pay       NUMERIC(12,2) DEFAULT 0",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS actual_work_hours     NUMERIC(8,2)  DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS salary_records (
            id              SERIAL PRIMARY KEY,
            staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            month           TEXT NOT NULL,
            base_salary     NUMERIC(12,2) DEFAULT 0,
            insured_salary  NUMERIC(12,2) DEFAULT 0,
            work_days       NUMERIC(5,1)  DEFAULT 0,
            actual_days     NUMERIC(5,1)  DEFAULT 0,
            leave_days      NUMERIC(5,1)  DEFAULT 0,
            unpaid_days     NUMERIC(5,1)  DEFAULT 0,
            ot_pay          NUMERIC(12,2) DEFAULT 0,
            allowance_total NUMERIC(12,2) DEFAULT 0,
            deduction_total NUMERIC(12,2) DEFAULT 0,
            net_pay         NUMERIC(12,2) DEFAULT 0,
            items           JSONB         DEFAULT '[]',
            status          TEXT          DEFAULT 'draft',
            note            TEXT          DEFAULT '',
            confirmed_by    TEXT          DEFAULT '',
            confirmed_at    TIMESTAMPTZ,
            created_at      TIMESTAMPTZ   DEFAULT NOW(),
            updated_at      TIMESTAMPTZ   DEFAULT NOW(),
            UNIQUE(staff_id, month)
        )""",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[salary_init] {str(e)[:80]}")

    # Seed default salary items
    defaults = [
        ('本薪',        'allowance', 'base_salary+service_years*1000', 0,    '#2e9e6b', 1),
        ('職務加給',    'allowance', '',                                0,    '#0ea5e9', 2),
        ('全勤獎金',    'allowance', '',                                0,    '#c8a96e', 3),
        ('獎金',        'allowance', '',                                0,    '#8b5cf6', 4),
        ('生日禮金',    'allowance', '',                                1000, '#e05c8a', 5),
        ('勞退6%',      'allowance', 'base_salary*0.06+service_years*1000*0.06', 0, '#4a7bda', 6),
        ('病/事/假',    'deduction', '',                                0,    '#8892a4', 7),
        ('勞保費',      'deduction', 'insured_salary*0.125*0.2',       0,    '#d64242', 8),
        ('健保費',      'deduction', 'insured_salary*0.0517*0.3',      0,    '#e07b2a', 9),
        ('勞退提撥6%',  'deduction', 'base_salary*0.06+service_years*1000*0.06', 0, '#4a7bda', 10),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM salary_items").fetchone()['c']
            if cnt == 0:
                for name, itype, formula, amount, color, sort in defaults:
                    conn.execute("""
                        INSERT INTO salary_items (name,item_type,formula,amount,color,sort_order)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (name, itype, formula, amount, color, sort))
    except Exception as e:
        print(f"[salary_seed] {e}")

init_salary_db()

def salary_item_row(row):
    if not row: return None
    d = dict(row)
    if d.get('amount') is not None: d['amount'] = float(d['amount'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

def salary_record_row(row):
    if not row: return None
    d = dict(row)
    for f in ['base_salary','insured_salary','work_days','actual_days','leave_days',
              'unpaid_days','ot_pay','allowance_total','deduction_total','net_pay',
              'income_tax_withheld','absent_days','whole_day_leave_days',
              'hourly_base_pay','actual_work_hours']:
        if d.get(f) is not None: d[f] = float(d[f])
    if isinstance(d.get('items'), str):
        try: d['items'] = _json.loads(d['items'])
        except: d['items'] = []
    if d.get('confirmed_at'): d['confirmed_at'] = d['confirmed_at'].isoformat()
    if d.get('created_at'):   d['created_at']   = d['created_at'].isoformat()
    if d.get('updated_at'):   d['updated_at']   = d['updated_at'].isoformat()
    return d

def _eval_formula(formula, base_salary, insured_salary, service_years, extra=None):
    """安全計算薪資公式
    可用變數：base_salary, insured_salary, service_years,
              actual_days, work_days, leave_days, unpaid_days,
              whole_day_leave_days（整天假天數，小時請假不計入，全勤判斷用此變數）,
              personal_days, sick_days, daily_wage
    支援條件式：例如 3000 if whole_day_leave_days==0 else 0
    """
    if not formula: return 0.0
    try:
        ctx = {
            'base_salary':    float(base_salary or 0),
            'insured_salary': float(insured_salary or 0),
            'service_years':  float(service_years or 0),
        }
        if extra:
            ctx.update({k: float(v or 0) for k, v in extra.items()})
        from simpleeval import simple_eval
        result = float(simple_eval(formula, names=ctx))
        if result != result or abs(result) == float('inf'):  # NaN or Inf（除以零）
            import logging
            logging.warning(f"[FORMULA] 無效結果(NaN/Inf): formula={formula!r} ctx={ctx}")
            return 0.0
        return round(result, 2)
    except ZeroDivisionError:
        import logging
        logging.error(f"[FORMULA] 除以零: formula={formula!r}")
        return 0.0
    except Exception as e:
        import logging
        logging.error(f"[FORMULA] 計算錯誤: formula={formula!r} error={e}")
        return 0.0

def _calc_service_years(hire_date_str):
    if not hire_date_str: return 0.0
    from datetime import date as _d4
    try:
        hire = _d4.fromisoformat(str(hire_date_str))
        return round((_d4.today() - hire).days / 365.25, 2)
    except Exception:
        return 0.0

def _calc_punch_hours(conn, staff_id, month):
    """
    從打卡記錄計算實際工時（時薪制用）
    邏輯：每天找最早 in + 最晚 out，扣除休息時間
    回傳 (total_hours, work_days, details)
    """
    from datetime import datetime as _dth, timezone as _tzh, timedelta as _tdh
    TW = _tzh(_tdh(hours=8))

    rows = conn.execute("""
        SELECT punch_type, punched_at
        FROM punch_records
        WHERE staff_id=%s
          AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
        ORDER BY punched_at ASC
    """, (staff_id, month)).fetchall()

    # Group by date
    day_map = {}
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            pa = pa.replace(tzinfo=_tzh.utc)
        pa_tw  = pa.astimezone(TW)
        ds     = pa_tw.strftime('%Y-%m-%d')
        if ds not in day_map:
            day_map[ds] = []
        day_map[ds].append({'type': r['punch_type'], 'dt': pa_tw})

    # Merge cross-midnight pairs: day N has 'in' but no 'out'; day N+1 has 'out' but no 'in'
    from datetime import date as _dateh2
    _sorted_ds  = sorted(day_map.keys())
    _merged_keys = set()
    for _i in range(len(_sorted_ds) - 1):
        _ds_cur = _sorted_ds[_i]
        _ds_nxt = _sorted_ds[_i + 1]
        if _ds_cur in _merged_keys or _ds_nxt in _merged_keys:
            continue
        try:
            if (_dateh2.fromisoformat(_ds_nxt) - _dateh2.fromisoformat(_ds_cur)).days != 1:
                continue
        except Exception:
            continue
        _cur = day_map[_ds_cur]
        _nxt = day_map[_ds_nxt]
        if (any(p['type'] == 'in'  for p in _cur)
                and not any(p['type'] == 'out' for p in _cur)
                and not any(p['type'] == 'in'  for p in _nxt)
                and any(p['type'] == 'out' for p in _nxt)):
            day_map[_ds_cur] = _cur + _nxt
            _merged_keys.add(_ds_nxt)
    for _k in _merged_keys:
        del day_map[_k]

    total_hours = 0.0
    details     = []
    for ds, punches in sorted(day_map.items()):
        ins   = [p['dt'] for p in punches if p['type'] == 'in']
        outs  = [p['dt'] for p in punches if p['type'] == 'out']
        b_out = [p['dt'] for p in punches if p['type'] == 'break_out']
        b_in  = [p['dt'] for p in punches if p['type'] == 'break_in']

        if not ins or not outs:
            continue

        work_start = min(ins)
        work_end   = max(outs)
        gross_mins = (work_end - work_start).total_seconds() / 60

        # 扣除休息時間（每個 break_in 只能被一個 break_out 使用）
        break_mins = 0.0
        _available_b_in = sorted(b_in)
        for bo in sorted(b_out):
            matched = [bi for bi in _available_b_in if bi > bo]
            if matched:
                bi_used = min(matched)
                break_mins += (bi_used - bo).total_seconds() / 60
                _available_b_in.remove(bi_used)

        net_mins = max(0.0, gross_mins - break_mins)
        net_hrs  = round(net_mins / 60, 2)
        total_hours += net_hrs
        details.append({
            'date':        ds,
            'clock_in':    work_start.strftime('%H:%M'),
            'clock_out':   work_end.strftime('%H:%M'),
            'break_mins':  round(break_mins),
            'net_hours':   net_hrs,
        })

    return round(total_hours, 2), len(day_map), details


def _auto_generate_salary(conn, staff, month, work_days=None):
    """
    自動產生員工月薪資料
    ─ 月薪制：底薪 + 薪資項目公式 + 加班費 - 請假扣款
    ─ 時薪制：打卡實際工時 × 時薪 + 加班費 - 請假扣款
    """
    import calendar as _cal2
    from datetime import date as _d5, timedelta as _td5, datetime as _dts5, timezone as _tz5
    _TW5 = _tz5(_td5(hours=8))
    _today5 = _dts5.now(_TW5).date()
    y, m = int(month[:4]), int(month[5:])
    total_work_days = work_days
    scheduled_dates = set()

    _sal_cfg = _get_salary_calc_settings()

    if total_work_days is None:
        # 1. 優先從排班取工作日
        shift_date_rows = conn.execute("""
            SELECT DISTINCT shift_date FROM shift_assignments
            WHERE staff_id=%s AND TO_CHAR(shift_date,'YYYY-MM')=%s
            ORDER BY shift_date
        """, (staff['id'], month)).fetchall()
        if shift_date_rows:
            scheduled_dates = {r['shift_date'].isoformat() if hasattr(r['shift_date'], 'isoformat') else str(r['shift_date']) for r in shift_date_rows}
            total_work_days = len(scheduled_dates)
        else:
            # 2. 備援：日曆扣除週日 + 國定假日
            holiday_rows = conn.execute("""
                SELECT date FROM public_holidays
                WHERE TO_CHAR(date,'YYYY-MM')=%s
            """, (month,)).fetchall()
            holiday_dates = {r['date'].isoformat() if hasattr(r['date'], 'isoformat') else str(r['date']) for r in holiday_rows}
            days_in_month = _cal2.monthrange(y, m)[1]
            for _d in range(1, days_in_month + 1):
                _dt = _d5(y, m, _d)
                _ds = _dt.isoformat()
                if _dt.weekday() < 5 and _ds not in holiday_dates:
                    scheduled_dates.add(_ds)
            total_work_days = len(scheduled_dates)

    salary_type    = staff.get('salary_type', 'monthly') or 'monthly'
    base_salary    = float(staff.get('base_salary')    or 0)
    hourly_rate    = float(staff.get('hourly_rate')    or 0)
    insured_salary = float(staff.get('insured_salary') or base_salary)
    daily_hours    = float(staff.get('daily_hours')    or 8)
    service_years  = _calc_service_years(staff.get('hire_date'))

    # ── 時薪制：從打卡記錄計算工時 ──────────────────────────
    actual_work_hours = 0.0
    punch_work_days   = 0
    punch_details     = []
    if salary_type == 'hourly':
        actual_work_hours, punch_work_days, punch_details = _calc_punch_hours(
            conn, staff['id'], month
        )
        # 時薪制的 base_salary 等於 實際工時 × 時薪
        hourly_base_pay = round(actual_work_hours * hourly_rate, 2)
    else:
        # 月薪制：daily_wage 用於請假扣款
        hourly_base_pay = 0.0

    # ── 已核准加班費 ────────────────────────────────────────
    ot_rows = conn.execute("""
        SELECT COALESCE(SUM(ot_pay), 0) as total
        FROM overtime_requests
        WHERE staff_id=%s AND status='approved'
          AND to_char(request_date,'YYYY-MM')=%s
    """, (staff['id'], month)).fetchone()
    ot_pay = float(ot_rows['total']) if ot_rows else 0.0

    # ── 請假資訊（只計算本月實際落在該月份的工作天數，排除週日） ──────────────
    _month_first = f'{y}-{m:02d}-01'
    _month_last  = f'{y}-{m:02d}-{_cal2.monthrange(y, m)[1]:02d}'
    _mf_date = _d5.fromisoformat(_month_first)
    _ml_date = _d5.fromisoformat(_month_last)
    _leave_raw = conn.execute("""
        SELECT lt.pay_rate, lt.code, lt.name as leave_name,
               GREATEST(lr.start_date, %s::date) AS eff_start,
               LEAST(lr.end_date, %s::date)      AS eff_end,
               COALESCE(lr.total_hours, 0)        AS leave_hours
        FROM leave_requests lr
        JOIN leave_types lt ON lt.id = lr.leave_type_id
        WHERE lr.staff_id=%s AND lr.status='approved'
          AND lr.start_date <= %s AND lr.end_date >= %s
    """, (_month_first, _month_last, staff['id'], _month_last, _month_first)).fetchall()

    def _count_working_days(start, end):
        """計算 start..end 之間的工作天。
        有排班資料時只計排班日（已含假日/假日補假排除）；
        無排班資料時計週一至週五。"""
        if not start or not end: return 0.0
        if isinstance(start, str): start = _d5.fromisoformat(start)
        if isinstance(end,   str): end   = _d5.fromisoformat(end)
        count = 0.0
        cur = start
        while cur <= end:
            if scheduled_dates:
                if cur.isoformat() in scheduled_dates:
                    count += 1.0
            else:
                if cur.weekday() < 5:
                    count += 1.0
            cur += _td5(days=1)
        return count

    _leave_wd = []
    _hourly_unpaid_hours = 0.0   # 小時無薪假（扣款用）
    _hourly_halfpay_hours = 0.0  # 小時半薪假
    for r in _leave_raw:
        lh = float(r['leave_hours'] or 0)
        is_hourly_leave = lh > 0
        if is_hourly_leave:
            # 小時請假：不計入天數，獨立累計工時
            _leave_wd.append({
                'pay_rate':      float(r['pay_rate']),
                'code':          r['code'],
                'leave_name':    r['leave_name'],
                'days_in_month': 0.0,
                'is_hourly':     True,
                'hours':         lh,
            })
            if float(r['pay_rate']) < 0.001:
                _hourly_unpaid_hours += lh
            elif 0.001 <= float(r['pay_rate']) <= 0.999:
                _hourly_halfpay_hours += lh
        else:
            wd = _count_working_days(r['eff_start'], r['eff_end'])
            _leave_wd.append({
                'pay_rate':      float(r['pay_rate']),
                'code':          r['code'],
                'leave_name':    r['leave_name'],
                'days_in_month': wd,
                'is_hourly':     False,
                'hours':         0.0,
            })

    # leave_days / unpaid_days 只計整天假，不含小時請假 → 全勤判斷正確
    leave_days    = sum(x['days_in_month'] for x in _leave_wd)
    unpaid_days   = sum(x['days_in_month'] for x in _leave_wd if float(x['pay_rate']) < 0.001)
    half_pay_days = sum(x['days_in_month'] for x in _leave_wd if 0.001 <= float(x['pay_rate']) <= 0.999)
    personal_days = sum(x['days_in_month'] for x in _leave_wd if x['code'] == 'personal')
    sick_days     = sum(x['days_in_month'] for x in _leave_wd if x['code'] == 'sick')
    whole_day_leave_days = leave_days   # 只含整天假，公式中判斷全勤用此變數
    leave_rows    = _leave_wd   # 後面 leave_names 使用 r['leave_name'] 字典存取
    if salary_type == 'hourly':
        actual_days = max(0.0, float(punch_work_days) - leave_days)
    else:
        actual_days = max(0.0, total_work_days - leave_days)

    # ── 日薪 / 時薪（用於請假扣款） ───────────────────────
    if salary_type == 'hourly':
        daily_wage  = hourly_rate * daily_hours   # 時薪制日薪 = 時薪 × 每日工時
        hourly_wage = hourly_rate
    else:
        daily_wage  = base_salary / 30 if base_salary > 0 else 0
        hourly_wage = daily_wage / daily_hours if daily_hours > 0 else 0

    # ── 缺勤天數（提前計算，讓公式中的 actual_days 能正確扣除缺勤） ──
    absent_days = 0
    _absent_date_list = []
    if _sal_cfg['auto_absent_deduction'] and salary_type == 'monthly' and scheduled_dates and daily_wage > 0:
        _punch_rows2 = conn.execute("""
            SELECT DISTINCT (punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date
            FROM punch_records WHERE staff_id=%s
              AND TO_CHAR(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
        """, (staff['id'], month)).fetchall()
        _punched_dates2 = {
            r['work_date'].isoformat() if hasattr(r['work_date'], 'isoformat') else str(r['work_date'])
            for r in _punch_rows2
        }
        _mf2 = _d5.fromisoformat(_month_first)
        _ml2 = _d5.fromisoformat(_month_last)
        _leave_date_rows2 = conn.execute("""
            SELECT start_date, end_date FROM leave_requests
            WHERE staff_id=%s AND status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (staff['id'], _month_last, _month_first)).fetchall()
        _leave_date_set2 = set()
        for _lr2 in _leave_date_rows2:
            _ld2 = _lr2['start_date']
            _le2 = _lr2['end_date']
            if isinstance(_ld2, str): _ld2 = _d5.fromisoformat(_ld2)
            if isinstance(_le2, str): _le2 = _d5.fromisoformat(_le2)
            _ld2 = max(_ld2, _mf2)
            _le2 = min(_le2, _ml2)
            while _ld2 <= _le2:
                _leave_date_set2.add(_ld2.isoformat())
                _ld2 += _td5(days=1)
        _absent_date_list = sorted(
            ds for ds in scheduled_dates
            if ds not in _punched_dates2 and ds not in _leave_date_set2
               and _d5.fromisoformat(ds) < _today5
        )
        absent_days = len(_absent_date_list)

    # 公式額外變數（出勤/假別相關）
    _formula_extra = {
        'actual_days':          max(0.0, actual_days - absent_days),
        'work_days':            float(total_work_days),
        'leave_days':           leave_days,
        'whole_day_leave_days': whole_day_leave_days,  # 全勤判斷用，小時請假不影響此值
        'unpaid_days':          unpaid_days,
        'personal_days':        personal_days,
        'sick_days':            sick_days,
        'daily_wage':           daily_wage,
    }

    # ── 組裝薪資項目 ────────────────────────────────────────
    items           = []
    allowance_total = 0.0
    deduction_total = 0.0
    # 員工個人金額覆寫 {str(item_id): amount}
    _overrides = staff.get('salary_item_overrides') or {}
    if isinstance(_overrides, str):
        try: _overrides = _json.loads(_overrides)
        except Exception: _overrides = {}

    def _apply_override(item_id, calculated_amt):
        """若員工設有個人金額，使用個人金額；否則使用計算值"""
        key = str(item_id)
        if key in _overrides and _overrides[key] is not None and _overrides[key] != '':
            return float(_overrides[key]), True   # (amount, is_overridden)
        return calculated_amt, False

    if salary_type == 'hourly':
        # 時薪制：第一筆項目是「本薪（工時計算）」
        items.append({
            'id': 'hourly_base', 'name': '本薪（工時）', 'type': 'allowance',
            'amount': hourly_base_pay, 'formula': '',
            'calc_note': (
                f'{actual_work_hours}h × 時薪${hourly_rate}'
                + (f'（{len(punch_details)}天出勤）' if punch_details else '')
            ),
        })
        allowance_total += hourly_base_pay

        # 時薪制加班費（從打卡計算，若無申請記錄則估算）
        # 先用「加班申請」核准金額；若為 0 則嘗試從工時估算
        if ot_pay == 0 and actual_work_hours > 0:
            # 每天超過 daily_hours 的部分算加班（三段費率：前2h/次2h/超4h）
            rate1 = float(staff.get('ot_rate1') or 1.33)
            rate2 = float(staff.get('ot_rate2') or 1.67)
            rate3 = float(staff.get('ot_rate3') or 2.0)
            for pd in punch_details:
                overtime_h = max(0.0, pd['net_hours'] - daily_hours)
                if overtime_h > 0:
                    h1 = min(overtime_h, 2.0)
                    h2 = min(max(0.0, overtime_h - 2.0), 2.0)
                    h3 = max(0.0, overtime_h - 4.0)
                    ot_pay += round(hourly_rate * (h1 * rate1 + h2 * rate2 + h3 * rate3), 2)

        # 時薪制的保險費以 insured_salary 為準（若未設定則用月薪換算）
        if insured_salary == 0:
            insured_salary = round(hourly_rate * daily_hours * 30, 0)

        # 時薪制只加入保險類扣除項（若員工有指定則只取指定中的保險項）
        staff_item_ids = staff.get('salary_item_ids')
        if staff_item_ids:
            placeholders = ','.join(['%s'] * len(staff_item_ids))
            salary_items_rows = conn.execute(f"""
                SELECT * FROM salary_items
                WHERE active=TRUE AND id IN ({placeholders})
                  AND item_type='deduction'
                  AND (formula LIKE '%insured_salary%' OR formula LIKE '%base_salary%')
                ORDER BY sort_order, id
            """, staff_item_ids).fetchall()
        else:
            salary_items_rows = conn.execute("""
                SELECT * FROM salary_items
                WHERE active=TRUE
                  AND item_type='deduction'
                  AND (formula LIKE '%insured_salary%' OR formula LIKE '%base_salary%')
                ORDER BY sort_order, id
            """).fetchall()
        for it in salary_items_rows:
            calc_amt = _eval_formula(it['formula'] or '', base_salary or insured_salary,
                                     insured_salary, service_years, _formula_extra)
            amt, overridden = _apply_override(it['id'], calc_amt)
            note = f'手動設定 ${amt}' if overridden else (it['formula'] or '')
            items.append({
                'id': it['id'], 'name': it['name'], 'type': 'deduction',
                'amount': round(amt, 2), 'formula': it['formula'] or '',
                'calc_note': note,
            })
            deduction_total += amt

    else:
        # 月薪制：跑啟用的薪資項目（若員工有指定則只跑指定項目）
        staff_item_ids = staff.get('salary_item_ids')
        if staff_item_ids:
            placeholders = ','.join(['%s'] * len(staff_item_ids))
            items_rows = conn.execute(
                f"SELECT * FROM salary_items WHERE active=TRUE AND id IN ({placeholders}) ORDER BY sort_order, id",
                staff_item_ids
            ).fetchall()
        else:
            items_rows = conn.execute(
                "SELECT * FROM salary_items WHERE active=TRUE ORDER BY sort_order, id"
            ).fetchall()
        for it in items_rows:
            formula  = it['formula'] or ''
            calc_amt = float(it['amount'] or 0)
            if formula:
                calc_amt = _eval_formula(formula, base_salary, insured_salary, service_years, _formula_extra)
            amt, overridden = _apply_override(it['id'], calc_amt)
            note = f'手動設定 ${amt}' if overridden else formula
            items.append({
                'id':        it['id'],
                'name':      it['name'],
                'type':      it['item_type'],
                'amount':    round(amt, 2),
                'formula':   formula,
                'calc_note': note,
            })
            if it['item_type'] == 'allowance':
                allowance_total += amt
            else:
                deduction_total += amt

    # ── 加班費（申請核准） ──────────────────────────────────
    if ot_pay > 0:
        items.append({
            'id': 'ot', 'name': '加班費（申請）', 'type': 'allowance',
            'amount': round(ot_pay, 2), 'formula': '',
            'calc_note': '核准加班費合計',
        })
        allowance_total += ot_pay

    # ── 請假扣款（可透過薪資計算設定關閉，改由薪資項目公式處理） ──
    if _sal_cfg['auto_leave_deduction']:
        # 整天無薪假
        if unpaid_days > 0 and daily_wage > 0:
            leave_names = '、'.join(set(
                r['leave_name'] for r in leave_rows
                if float(r['pay_rate']) < 0.001 and not r['is_hourly']
            ))
            deduct = round(daily_wage * unpaid_days, 2)
            items.append({
                'id': 'unpaid', 'name': f'無薪假扣款（{leave_names}）', 'type': 'deduction',
                'amount': deduct, 'formula': '',
                'calc_note': f'{unpaid_days}天 × 日薪${round(daily_wage, 0)}',
            })
            deduction_total += deduct

        # 小時無薪假（按時薪扣，不影響全勤）
        if _hourly_unpaid_hours > 0 and hourly_wage > 0:
            leave_names = '、'.join(set(
                r['leave_name'] for r in leave_rows
                if float(r['pay_rate']) < 0.001 and r['is_hourly']
            ))
            deduct = round(hourly_wage * _hourly_unpaid_hours, 2)
            items.append({
                'id': 'unpaid_hourly', 'name': f'無薪假扣款-小時（{leave_names}）', 'type': 'deduction',
                'amount': deduct, 'formula': '',
                'calc_note': f'{_hourly_unpaid_hours}h × 時薪${round(hourly_wage, 1)}',
            })
            deduction_total += deduct

        # 整天半薪假
        if half_pay_days > 0 and daily_wage > 0:
            leave_names = '、'.join(set(
                r['leave_name'] for r in leave_rows
                if 0.001 <= float(r['pay_rate']) <= 0.999 and not r['is_hourly']
            ))
            deduct = round(daily_wage * half_pay_days * 0.5, 2)
            items.append({
                'id': 'halfpay', 'name': f'半薪假扣款（{leave_names}）', 'type': 'deduction',
                'amount': deduct, 'formula': '',
                'calc_note': f'{half_pay_days}天 × 日薪${round(daily_wage, 0)} × 0.5',
            })
            deduction_total += deduct

        # 小時半薪假（按時薪 × 0.5 扣）
        if _hourly_halfpay_hours > 0 and hourly_wage > 0:
            leave_names = '、'.join(set(
                r['leave_name'] for r in leave_rows
                if 0.001 <= float(r['pay_rate']) <= 0.999 and r['is_hourly']
            ))
            deduct = round(hourly_wage * _hourly_halfpay_hours * 0.5, 2)
            items.append({
                'id': 'halfpay_hourly', 'name': f'半薪假扣款-小時（{leave_names}）', 'type': 'deduction',
                'amount': deduct, 'formula': '',
                'calc_note': f'{_hourly_halfpay_hours}h × 時薪${round(hourly_wage, 1)} × 0.5',
            })
            deduction_total += deduct

    # ── 月薪制：缺勤扣款（可透過薪資計算設定關閉） ───────────────
    if absent_days > 0:
        deduct = round(daily_wage * absent_days, 2)
        sample = '、'.join(_absent_date_list[:3]) + ('等' if absent_days > 3 else '')
        items.append({
            'id': 'absent', 'name': f'缺勤扣款（{absent_days} 天）', 'type': 'deduction',
            'amount': deduct, 'formula': '',
            'calc_note': f'{absent_days} 天 × 日薪 ${round(daily_wage, 0)}（{sample}）',
        })
        deduction_total += deduct

    # ── 薪資所得扣繳稅款（可透過薪資計算設定關閉） ───────────────
    _TAX_THRESHOLD = 88501
    _existing_tax = sum(
        float(it['amount']) for it in items
        if isinstance(it.get('id'), str) and it['id'] == 'income_tax'
    )
    income_tax_withheld = 0.0
    if _sal_cfg['auto_income_tax'] and _existing_tax == 0 and allowance_total > _TAX_THRESHOLD:
        income_tax_withheld = round(allowance_total * 0.05, 0)
        items.append({
            'id': 'income_tax', 'name': '薪資所得扣繳稅款', 'type': 'deduction',
            'amount': income_tax_withheld, 'formula': '',
            'calc_note': f'總支給 ${round(allowance_total,0):,.0f} × 5%（超過起徵點 ${_TAX_THRESHOLD:,}）',
        })
        deduction_total += income_tax_withheld
    else:
        income_tax_withheld = _existing_tax

    net_pay = round(allowance_total - deduction_total, 2)

    return {
        'staff_id':              staff['id'],
        'month':                 month,
        'salary_type':           salary_type,
        'base_salary':           base_salary if salary_type == 'monthly' else 0,
        'hourly_rate':           hourly_rate if salary_type == 'hourly' else 0,
        'hourly_base_pay':       hourly_base_pay if salary_type == 'hourly' else 0,
        'actual_work_hours':     actual_work_hours if salary_type == 'hourly' else 0,
        'insured_salary':        insured_salary,
        'work_days':             total_work_days,
        'actual_days':           max(0, actual_days - absent_days),
        'leave_days':            leave_days,
        'unpaid_days':           unpaid_days,
        'absent_days':           absent_days,
        'whole_day_leave_days':  whole_day_leave_days,
        'ot_pay':                ot_pay,
        'allowance_total':       round(allowance_total, 2),
        'deduction_total':       round(deduction_total, 2),
        'net_pay':               net_pay,
        'income_tax_withheld':   income_tax_withheld,
        'items':                 items,
        'punch_details':         punch_details,
        'status':                'draft',
    }

# ── Employee: view own payslip ────────────────────────────────────

@app.route('/api/salary/my-payslip', methods=['GET'])
def api_my_payslip():
    sid = session.get('punch_staff_id')
    if not sid:
        return jsonify({'error': '請先登入'}), 401
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _dp
        month = _dp.today().strftime('%Y-%m')
    with get_db() as conn:
        row = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.role as staff_role,
                   ps.employee_code, ps.department, ps.salary_type, ps.hourly_rate
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.staff_id = %s AND sr.month = %s
        """, (sid, month)).fetchone()
        if not row:
            # 無薪資記錄時即時計算草稿，讓員工可以提前預覽
            staff = conn.execute("SELECT * FROM punch_staff WHERE id=%s", (sid,)).fetchone()
            if not staff:
                return jsonify({'error': '找不到員工資料'}), 404
            data = _auto_generate_salary(conn, dict(staff), month)
            data['staff_name']    = staff['name']
            data['staff_role']    = staff['role'] or ''
            data['employee_code'] = staff['employee_code'] or ''
            data['department']    = staff['department'] or ''
            data['is_preview']    = True   # 標記為預覽，尚未由管理員確認
            return jsonify(data)
    d = salary_record_row(row)
    d['staff_name']    = row['staff_name']
    d['staff_role']    = row['staff_role']
    d['employee_code'] = row['employee_code'] or ''
    d['department']    = row['department']    or ''
    d['salary_type']   = row['salary_type']   or 'monthly'
    d['hourly_rate']   = float(row['hourly_rate'] or 0)
    d['is_preview']    = False
    return jsonify(d)

# ── Salary Calc Settings ──────────────────────────────────────────

@app.route('/api/salary/calc-settings', methods=['GET'])
@require_module('salary')
def api_salary_calc_settings_get():
    return jsonify(_get_salary_calc_settings())

@app.route('/api/salary/calc-settings', methods=['POST'])
@require_module('salary')
def api_salary_calc_settings_post():
    b = request.get_json(force=True) or {}
    allowed = {'auto_leave_deduction', 'auto_absent_deduction', 'auto_income_tax'}
    with get_db() as conn:
        for key in allowed:
            if key in b:
                val = 'true' if b[key] else 'false'
                conn.execute(
                    "INSERT INTO salary_calc_settings(setting_key,setting_value) VALUES(%s,%s)"
                    " ON CONFLICT(setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                    (key, val)
                )
    return jsonify({'ok': True, 'settings': _get_salary_calc_settings()})

# ── Salary Items CRUD ─────────────────────────────────────────────

@app.route('/api/salary/items', methods=['GET'])
@require_module('salary')
def api_salary_items_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM salary_items ORDER BY sort_order, id").fetchall()
    return jsonify([salary_item_row(r) for r in rows])

@app.route('/api/salary/items', methods=['POST'])
@require_module('salary')
def api_salary_item_create():
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO salary_items (name, item_type, formula, amount, description, color, sort_order)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'], b.get('item_type','allowance'), b.get('formula',''),
              float(b.get('amount',0)), b.get('description',''),
              b.get('color','#4a7bda'), int(b.get('sort_order',0)))).fetchone()
    return jsonify(salary_item_row(row)), 201

@app.route('/api/salary/items/<int:iid>', methods=['PUT'])
@require_module('salary')
def api_salary_item_update(iid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE salary_items SET name=%s, item_type=%s, formula=%s, amount=%s,
              description=%s, color=%s, sort_order=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b['name'], b.get('item_type','allowance'), b.get('formula',''),
              float(b.get('amount',0)), b.get('description',''),
              b.get('color','#4a7bda'), int(b.get('sort_order',0)),
              bool(b.get('active',True)), iid)).fetchone()
    return jsonify(salary_item_row(row)) if row else ('', 404)

@app.route('/api/salary/items/<int:iid>', methods=['DELETE'])
@require_module('salary')
def api_salary_item_delete(iid):
    with get_db() as conn:
        conn.execute("DELETE FROM salary_items WHERE id=%s", (iid,))
    return jsonify({'deleted': iid})

# ── Salary Records ─────────────────────────────────────────────────

@app.route('/api/salary/records', methods=['GET'])
@require_module('salary')
def api_salary_records_list():
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _d6
        month = _d6.today().strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.role as staff_role,
                   ps.employee_code, ps.department
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s
            ORDER BY ps.name
        """, (month,)).fetchall()
    result = []
    for r in rows:
        d = salary_record_row(r)
        d['staff_name']    = r['staff_name']
        d['staff_role']    = r['staff_role']
        d['employee_code'] = r['employee_code'] or ''
        d['department']    = r['department'] or ''
        result.append(d)
    return jsonify(result)

@app.route('/api/salary/records/generate', methods=['POST'])
@require_module('salary')
def api_salary_generate():
    """自動產生或更新該月薪資
    force=True：強制重算已確認薪資（狀態改回 draft）
    force=False（預設）：跳過已確認薪資，不覆蓋
    """
    b     = request.get_json(force=True)
    month = b.get('month', '').strip()
    force = bool(b.get('force', False))
    if not month: return jsonify({'error': '請指定月份'}), 400
    with get_db() as conn:
        try:
            month_int = int(month.replace('-', ''))
        except ValueError:
            return jsonify({'error': '月份格式錯誤，請使用 YYYY-MM'}), 400
        lock_key = 4_200_000_000 + month_int
        locked = conn.execute(
            "SELECT pg_try_advisory_xact_lock(%s) AS locked", (lock_key,)
        ).fetchone()['locked']
        if not locked:
            return jsonify({'error': f'{month} 薪資正在產生中，請稍後再試'}), 409

        staff_list = conn.execute(
            "SELECT * FROM punch_staff WHERE active=TRUE"
        ).fetchall()
        generated = 0
        skipped   = 0
        for staff in staff_list:
            # 非強制模式：跳過已確認的薪資，不覆蓋已確認金額
            if not force:
                existing = conn.execute(
                    "SELECT status FROM salary_records WHERE staff_id=%s AND month=%s",
                    (staff['id'], month)
                ).fetchone()
                if existing and existing['status'] == 'confirmed':
                    skipped += 1
                    continue

            data = _auto_generate_salary(conn, dict(staff), month)
            items_json = _json.dumps(data['items'], ensure_ascii=False)
            conn.execute("""
                INSERT INTO salary_records
                  (staff_id, month, base_salary, insured_salary, work_days, actual_days,
                   leave_days, unpaid_days, ot_pay, allowance_total, deduction_total,
                   net_pay, income_tax_withheld, items, status, updated_at,
                   absent_days, whole_day_leave_days, hourly_base_pay, actual_work_hours)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,'draft',NOW(),%s,%s,%s,%s)
                ON CONFLICT (staff_id, month) DO UPDATE
                  SET base_salary=%s, insured_salary=%s, work_days=%s, actual_days=%s,
                      leave_days=%s, unpaid_days=%s, ot_pay=%s, allowance_total=%s,
                      deduction_total=%s, net_pay=%s, income_tax_withheld=%s, items=%s::jsonb,
                      absent_days=%s, whole_day_leave_days=%s, hourly_base_pay=%s, actual_work_hours=%s,
                      status='draft', updated_at=NOW()
            """, (
                data['staff_id'], month, data['base_salary'], data['insured_salary'],
                data['work_days'], data['actual_days'], data['leave_days'], data['unpaid_days'],
                data['ot_pay'], data['allowance_total'], data['deduction_total'],
                data['net_pay'], data['income_tax_withheld'], items_json,
                data.get('absent_days', 0), data.get('whole_day_leave_days', 0),
                data.get('hourly_base_pay', 0), data.get('actual_work_hours', 0),
                # ON CONFLICT SET
                data['base_salary'], data['insured_salary'], data['work_days'], data['actual_days'],
                data['leave_days'], data['unpaid_days'], data['ot_pay'], data['allowance_total'],
                data['deduction_total'], data['net_pay'], data['income_tax_withheld'], items_json,
                data.get('absent_days', 0), data.get('whole_day_leave_days', 0),
                data.get('hourly_base_pay', 0), data.get('actual_work_hours', 0),
            ))
            generated += 1
    return jsonify({'ok': True, 'generated': generated, 'skipped': skipped, 'month': month})

@app.route('/api/salary/records/<int:rid>', methods=['GET'])
@require_module('salary')
def api_salary_record_get(rid):
    with get_db() as conn:
        row = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.role as staff_role,
                   ps.employee_code, ps.department, ps.hire_date
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.id=%s
        """, (rid,)).fetchone()
    if not row: return ('', 404)
    d = salary_record_row(row)
    d['staff_name']    = row['staff_name']
    d['staff_role']    = row['staff_role']
    d['employee_code'] = row['employee_code'] or ''
    d['department']    = row['department'] or ''
    d['hire_date']     = row['hire_date'].isoformat() if row['hire_date'] else ''
    return jsonify(d)

@app.route('/api/salary/records/<int:rid>', methods=['PUT'])
@require_module('salary')
def api_salary_record_update(rid):
    b = request.get_json(force=True)
    items = b.get('items', [])
    items_json = _json.dumps(items, ensure_ascii=False)
    # 從 items 重新計算扣繳稅款（id='income_tax' 或名稱含「扣繳」）
    tax_withheld = sum(
        float(it.get('amount', 0)) for it in items
        if it.get('type') == 'deduction' and (
            it.get('id') == 'income_tax' or '扣繳' in (it.get('name') or '')
        )
    )
    with get_db() as conn:
        row = conn.execute("""
            UPDATE salary_records SET
              allowance_total=%s, deduction_total=%s, net_pay=%s,
              income_tax_withheld=%s, items=%s::jsonb, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (float(b.get('allowance_total',0)), float(b.get('deduction_total',0)),
              float(b.get('net_pay',0)), tax_withheld, items_json,
              b.get('note',''), rid)).fetchone()
    return jsonify(salary_record_row(row)) if row else ('', 404)

@app.route('/api/salary/records/confirm-all', methods=['POST'])
@require_module('salary')
def api_salary_confirm_all():
    b    = request.get_json(force=True)
    month = b.get('month','').strip()
    by   = b.get('confirmed_by','管理員')
    if not month: return jsonify({'error': '請指定月份'}), 400
    with get_db() as conn:
        rows = conn.execute("""
            UPDATE salary_records SET status='confirmed', confirmed_by=%s,
              confirmed_at=NOW(), updated_at=NOW()
            WHERE month=%s AND status='draft'
            RETURNING id, staff_id, month, net_pay
        """, (by, month)).fetchall()
    confirmed = len(rows)
    for row in rows:
        extra = f"{row['month']} 薪資已確認\n實領金額：${float(row['net_pay'] or 0):,.0f}"
        _notify_review_result(row['staff_id'], '薪資', 'confirmed', extra)
    return jsonify({'ok': True, 'confirmed': confirmed})

@app.route('/api/salary/records/<int:rid>/confirm', methods=['POST'])
@require_module('salary')
def api_salary_confirm(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE salary_records SET status='confirmed', confirmed_by=%s,
              confirmed_at=NOW(), updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b.get('confirmed_by','管理員'), rid)).fetchone()
    if row:
        extra = f"{row['month']} 薪資已確認\n實領金額：${float(row['net_pay'] or 0):,.0f}"
        _notify_review_result(row['staff_id'], '薪資', 'confirmed', extra)
    return jsonify(salary_record_row(row)) if row else ('', 404)

@app.route('/api/salary/records/<int:rid>', methods=['DELETE'])
@require_module('salary')
def api_salary_record_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM salary_records WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

# ── Salary Staff Settings ─────────────────────────────────────────

@app.route('/api/salary/staff', methods=['GET'])
@require_module('salary')
def api_salary_staff_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT id, name, username, role, active, employee_code, department,
                   position_title, hire_date, birth_date, base_salary, insured_salary,
                   daily_hours, ot_rate1, ot_rate2, salary_type, hourly_rate,
                   vacation_quota, salary_notes, salary_item_ids, salary_item_overrides,
                   national_id, gender, insurance_type, address
            FROM punch_staff ORDER BY name
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for f in ['base_salary','insured_salary','daily_hours','ot_rate1','ot_rate2','hourly_rate']:
            if d.get(f) is not None: d[f] = float(d[f])
        if d.get('hire_date'):  d['hire_date']  = d['hire_date'].isoformat()
        if d.get('birth_date'): d['birth_date'] = d['birth_date'].isoformat()
        d['annual_leave_days'] = _calc_annual_leave_days(d.get('hire_date'))
        d['service_years']     = _calc_service_years(d.get('hire_date'))
        result.append(d)
    return jsonify(result)

@app.route('/api/salary/staff/<int:sid>', methods=['PUT'])
@require_module('salary')
def api_salary_staff_update(sid):
    b = request.get_json(force=True)
    def _f(k, default=0): return float(b.get(k, default) or default)
    def _s(k): return b.get(k, '').strip() if b.get(k) else None
    with get_db() as conn:
        salary_item_ids = b.get('salary_item_ids')
        salary_item_ids_json = _json.dumps(salary_item_ids) if salary_item_ids is not None else None
        overrides = b.get('salary_item_overrides')  # dict {str(item_id): amount}
        overrides_json = _json.dumps(overrides) if overrides else None
        # 先讀舊值，判斷薪資相關欄位是否有異動
        _old = conn.execute(
            "SELECT base_salary, insured_salary, daily_hours, hourly_rate, salary_type, ot_rate1, ot_rate2 FROM punch_staff WHERE id=%s", (sid,)
        ).fetchone()
        conn.execute("""
            UPDATE punch_staff SET
              employee_code=%s, department=%s, position_title=%s,
              hire_date=%s, birth_date=%s,
              base_salary=%s, insured_salary=%s, daily_hours=%s,
              ot_rate1=%s, ot_rate2=%s, salary_type=%s,
              hourly_rate=%s, vacation_quota=%s, salary_notes=%s,
              salary_item_ids=%s, salary_item_overrides=%s,
              national_id=%s, gender=%s, insurance_type=%s, address=%s
            WHERE id=%s
        """, (_s('employee_code'), _s('department'), _s('position_title'),
              _s('hire_date'), _s('birth_date'),
              _f('base_salary'), _f('insured_salary'), _f('daily_hours') or 8,
              _f('ot_rate1') or 1.33, _f('ot_rate2') or 1.67,
              b.get('salary_type','monthly'),
              _f('hourly_rate'), b.get('vacation_quota') or None,
              b.get('salary_notes',''), salary_item_ids_json, overrides_json,
              (b.get('national_id') or '').strip(),
              (b.get('gender') or '').strip(),
              (b.get('insurance_type') or 'regular').strip(),
              (b.get('address') or '').strip(),
              sid))
        # 薪資計算參數有異動 → 清除該員工所有 draft 薪資，確保重算
        _salary_keys = ('base_salary','insured_salary','daily_hours','hourly_rate','salary_type','ot_rate1','ot_rate2')
        if _old and any(
            str(b.get(k, '')) != str(float(_old[k] or 0) if k != 'salary_type' else (_old[k] or 'monthly'))
            for k in _salary_keys if k in b
        ):
            conn.execute("DELETE FROM salary_records WHERE staff_id=%s AND status='draft'", (sid,))
        row = conn.execute("SELECT * FROM punch_staff WHERE id=%s", (sid,)).fetchone()
    return jsonify(punch_staff_row(row)) if row else ('', 404)


# ═══════════════════════════════════════════════════════════════════
# Announcement Module (公告管理)
# ═══════════════════════════════════════════════════════════════════

def init_announcement_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS announcements (
                    id          SERIAL PRIMARY KEY,
                    title       TEXT NOT NULL,
                    content     TEXT NOT NULL,
                    category    TEXT DEFAULT 'general',
                    priority    TEXT DEFAULT 'normal',
                    is_pinned   BOOLEAN DEFAULT FALSE,
                    visible_to  TEXT DEFAULT 'all',
                    published_at TIMESTAMPTZ DEFAULT NOW(),
                    expires_at  TIMESTAMPTZ,
                    author      TEXT DEFAULT '管理員',
                    active      BOOLEAN DEFAULT TRUE,
                    view_count  INT DEFAULT 0,
                    created_at  TIMESTAMPTZ DEFAULT NOW(),
                    updated_at  TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f"[announcement_init] {e}")

init_announcement_db()

def ann_row(row):
    if not row: return None
    d = dict(row)
    if d.get('published_at'): d['published_at'] = d['published_at'].isoformat()
    if d.get('expires_at'):   d['expires_at']   = d['expires_at'].isoformat()
    if d.get('created_at'):   d['created_at']   = d['created_at'].isoformat()
    if d.get('updated_at'):   d['updated_at']   = d['updated_at'].isoformat()
    return d

# ── Admin: CRUD ───────────────────────────────────────────────────

@app.route('/api/announcements', methods=['GET'])
@require_module('ann')
def api_ann_list_admin():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM announcements
            ORDER BY is_pinned DESC, published_at DESC
            LIMIT 200
        """).fetchall()
    return jsonify([ann_row(r) for r in rows])

@app.route('/api/announcements', methods=['POST'])
@require_module('ann')
def api_ann_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip():
        return jsonify({'error': '請填寫公告標題'}), 400
    if not b.get('content','').strip():
        return jsonify({'error': '請填寫公告內容'}), 400
    expires = b.get('expires_at') or None
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO announcements
              (title, content, category, priority, is_pinned,
               visible_to, expires_at, author, active)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['title'].strip(), b['content'].strip(),
              b.get('category','general'), b.get('priority','normal'),
              bool(b.get('is_pinned', False)), b.get('visible_to','all'),
              expires, b.get('author','管理員').strip(),
              bool(b.get('active', True)))).fetchone()
    if row and row['active']:
        _broadcast_announcement_line(row['title'], row['content'])
    return jsonify(ann_row(row)), 201

@app.route('/api/announcements/<int:aid>', methods=['PUT'])
@require_module('ann')
def api_ann_update(aid):
    b = request.get_json(force=True)
    if not b.get('title','').strip():
        return jsonify({'error': '請填寫公告標題'}), 400
    expires = b.get('expires_at') or None
    with get_db() as conn:
        row = conn.execute("""
            UPDATE announcements SET
              title=%s, content=%s, category=%s, priority=%s,
              is_pinned=%s, visible_to=%s, expires_at=%s,
              author=%s, active=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b['title'].strip(), b.get('content','').strip(),
              b.get('category','general'), b.get('priority','normal'),
              bool(b.get('is_pinned', False)), b.get('visible_to','all'),
              expires, b.get('author','管理員').strip(),
              bool(b.get('active', True)), aid)).fetchone()
    return jsonify(ann_row(row)) if row else ('', 404)

@app.route('/api/announcements/<int:aid>', methods=['DELETE'])
@require_module('ann')
def api_ann_delete(aid):
    with get_db() as conn:
        conn.execute("DELETE FROM announcements WHERE id=%s", (aid,))
    return jsonify({'deleted': aid})

@app.route('/api/announcements/<int:aid>/pin', methods=['POST'])
@require_module('ann')
def api_ann_toggle_pin(aid):
    with get_db() as conn:
        row = conn.execute(
            "UPDATE announcements SET is_pinned=NOT is_pinned, updated_at=NOW() WHERE id=%s RETURNING *",
            (aid,)
        ).fetchone()
    return jsonify(ann_row(row)) if row else ('', 404)

# ── Public: employee reads ────────────────────────────────────────

@app.route('/api/announcements/public', methods=['GET'])
def api_ann_public():
    """員工端讀取有效公告"""
    from datetime import datetime as _dta
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM announcements
            WHERE active = TRUE
              AND (expires_at IS NULL OR expires_at > NOW())
            ORDER BY is_pinned DESC, published_at DESC
            LIMIT 50
        """).fetchall()
        # 增加閱讀計數（批次）
    return jsonify([ann_row(r) for r in rows])

@app.route('/api/announcements/<int:aid>/view', methods=['POST'])
def api_ann_view(aid):
    with get_db() as conn:
        conn.execute(
            "UPDATE announcements SET view_count = view_count + 1 WHERE id=%s", (aid,)
        )
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════════════
# Feature 1: Taiwan Public Holidays (國定假日)
# ═══════════════════════════════════════════════════════════════════

def init_holiday_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS public_holidays (
                    id          SERIAL PRIMARY KEY,
                    date        DATE NOT NULL UNIQUE,
                    name        TEXT NOT NULL,
                    holiday_type TEXT DEFAULT 'national',
                    note        TEXT DEFAULT '',
                    created_at  TIMESTAMPTZ DEFAULT NOW()
                )
            """)
        # Seed 2025 & 2026 Taiwan holidays
        _seed_holidays()
    except Exception as e:
        print(f"[holiday_init] {e}")

def _seed_holidays():
    """台灣2025-2026國定假日"""
    holidays_2025 = [
        ('2025-01-01', '元旦'),
        ('2025-01-27', '農曆除夕'),
        ('2025-01-28', '春節'),
        ('2025-01-29', '春節'),
        ('2025-01-30', '春節'),
        ('2025-01-31', '春節補假'),
        ('2025-02-28', '和平紀念日'),
        ('2025-04-03', '兒童節補假'),
        ('2025-04-04', '兒童節/清明節'),
        ('2025-05-01', '勞動節'),
        ('2025-05-30', '端午節補假'),
        ('2025-06-02', '端午節'),
        ('2025-10-06', '中秋節補假'),
        ('2025-10-07', '中秋節'),
        ('2025-10-10', '國慶日'),
    ]
    holidays_2026 = [
        ('2026-01-01', '元旦'),
        ('2026-01-28', '農曆除夕'),
        ('2026-01-29', '春節'),
        ('2026-01-30', '春節'),
        ('2026-01-31', '春節'),
        ('2026-02-02', '春節補假'),
        ('2026-02-28', '和平紀念日'),
        ('2026-03-02', '和平紀念日補假'),
        ('2026-04-03', '兒童節'),
        ('2026-04-04', '清明節'),
        ('2026-04-05', '清明節補假'),
        ('2026-05-01', '勞動節'),
        ('2026-06-19', '端午節'),
        ('2026-09-25', '中秋節'),
        ('2026-10-09', '國慶日補假'),
        ('2026-10-10', '國慶日'),
    ]
    all_holidays = holidays_2025 + holidays_2026
    try:
        with get_db() as conn:
            existing = conn.execute("SELECT COUNT(*) as c FROM public_holidays").fetchone()['c']
            if existing == 0:
                for date_str, name in all_holidays:
                    try:
                        conn.execute(
                            "INSERT INTO public_holidays (date, name) VALUES (%s,%s) ON CONFLICT (date) DO NOTHING",
                            (date_str, name)
                        )
                    except Exception:
                        pass
    except Exception as e:
        print(f"[holiday_seed] {e}")

init_holiday_db()

def holiday_row(row):
    if not row: return None
    d = dict(row)
    if d.get('date'):       d['date']       = d['date'].isoformat()
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

def _is_holiday(conn, date_str):
    """Check if a date is a public holiday"""
    row = conn.execute(
        "SELECT id FROM public_holidays WHERE date=%s", (date_str,)
    ).fetchone()
    return row is not None

# ── Holiday CRUD API ─────────────────────────────────────────────

@app.route('/api/holidays', methods=['GET'])
@require_module('holiday')
def api_holidays_list():
    year = request.args.get('year', '')
    conds, params = ['TRUE'], []
    if year:
        conds.append("EXTRACT(YEAR FROM date)=%s")
        params.append(int(year))
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT * FROM public_holidays WHERE {' AND '.join(conds)} ORDER BY date",
            params
        ).fetchall()
    return jsonify([holiday_row(r) for r in rows])

@app.route('/api/holidays/public', methods=['GET'])
def api_holidays_public():
    """Public endpoint for staff page"""
    year = request.args.get('year', '')
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if year:
        conds.append("EXTRACT(YEAR FROM date)=%s"); params.append(int(year))
    if month:
        conds.append("to_char(date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT date, name FROM public_holidays WHERE {' AND '.join(conds)} ORDER BY date",
            params
        ).fetchall()
    return jsonify({r['date'].isoformat(): r['name'] for r in rows})

@app.route('/api/holidays', methods=['POST'])
@require_module('holiday')
def api_holiday_create():
    b = request.get_json(force=True)
    if not b.get('date') or not b.get('name','').strip():
        return jsonify({'error': '請填寫日期和名稱'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO public_holidays (date, name, holiday_type, note)
            VALUES (%s,%s,%s,%s)
            ON CONFLICT (date) DO UPDATE
              SET name=EXCLUDED.name, holiday_type=EXCLUDED.holiday_type, note=EXCLUDED.note
            RETURNING *
        """, (b['date'], b['name'].strip(),
              b.get('holiday_type','national'), b.get('note',''))).fetchone()
    return jsonify(holiday_row(row)), 201

@app.route('/api/holidays/<int:hid>', methods=['DELETE'])
@require_module('holiday')
def api_holiday_delete(hid):
    with get_db() as conn:
        conn.execute("DELETE FROM public_holidays WHERE id=%s", (hid,))
    return jsonify({'deleted': hid})

@app.route('/api/holidays/batch', methods=['POST'])
@require_module('holiday')
def api_holiday_batch():
    """Batch import holidays from JSON list"""
    b    = request.get_json(force=True)
    rows = b.get('holidays', [])
    count = 0
    with get_db() as conn:
        for item in rows:
            try:
                conn.execute("""
                    INSERT INTO public_holidays (date, name, holiday_type, note)
                    VALUES (%s,%s,%s,%s)
                    ON CONFLICT (date) DO UPDATE SET name=EXCLUDED.name
                """, (item['date'], item['name'],
                      item.get('holiday_type','national'), item.get('note','')))
                count += 1
            except Exception:
                pass
    return jsonify({'imported': count})


# ═══════════════════════════════════════════════════════════════════
# Feature 2: LINE Notification Helper
# ═══════════════════════════════════════════════════════════════════

def _notify_staff_line(staff_id, message):
    """
    Send LINE notification to a staff member if they have LINE bound.
    Uses the line_punch_config token (same LINE OA).
    """
    if not DATABASE_URL:
        return
    try:
        with get_db() as conn:
            staff = conn.execute(
                "SELECT line_user_id FROM punch_staff WHERE id=%s", (staff_id,)
            ).fetchone()
            if not staff or not staff['line_user_id']:
                return
            cfg = conn.execute(
                "SELECT * FROM line_punch_config WHERE id=1"
            ).fetchone()
        if not cfg or not cfg.get('enabled') or not cfg.get('channel_access_token'):
            return
        LineBotApi(cfg['channel_access_token']).push_message(
            staff['line_user_id'],
            TextSendMessage(text=message)
        )
    except Exception as e:
        print(f"[LINE notify] staff_id={staff_id}: {e}")


def _notify_review_result(staff_id, category, action, extra_info=''):
    """
    Send a formatted LINE notification for review results.
    category: '補打卡申請', '排休申請', '加班申請', '請假申請', '薪資確認'
    action:   'approved', 'rejected', 'confirmed'
    """
    ACTION_LABEL = {'approved': '核准', 'rejected': '退回', 'confirmed': '確認'}
    ACTION_ICON  = {'approved': '[核准]', 'rejected': '[退回]', 'confirmed': '[確認]'}
    label = ACTION_LABEL.get(action, action)
    icon  = ACTION_ICON.get(action, '')
    msg   = f"{icon} {category}{label}\n{extra_info}\n\n請至員工系統查看詳情。"
    _notify_staff_line(staff_id, msg.strip())


# ═══════════════════════════════════════════════════════════════════
# Feature 3: Export Reports (出勤/薪資報表匯出)
# ═══════════════════════════════════════════════════════════════════

import csv
import io

# ─── Excel 匯出共用工具 ────────────────────────────────────────────
def _xl_workbook(sheet_name='Sheet1'):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws

def _xl_write_header(ws, headers, col_widths):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    hdr_fill  = PatternFill('solid', fgColor='0F1C3A')
    hdr_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def _xl_write_rows(ws, data_rows, num_cols, number_cols=None):
    from openpyxl.styles import Alignment, PatternFill, Border, Side, numbers
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    even_fill = PatternFill('solid', fgColor='F4F6FA')
    center    = Alignment(horizontal='center', vertical='center')
    left      = Alignment(horizontal='left',   vertical='center')
    number_cols = set(number_cols or [])
    for ri, row_vals in enumerate(data_rows, 2):
        fill = even_fill if ri % 2 == 0 else None
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if fill: cell.fill = fill
            cell.border = thin
            if ci in number_cols and isinstance(v, (int, float)):
                cell.alignment = center
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = center if isinstance(v, (int, float, type(None))) else left

def _xl_response(wb, filename):
    from io import BytesIO
    from flask import Response
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )
# ──────────────────────────────────────────────────────────────────


@app.route('/api/export/attendance', methods=['GET'])
@login_required
def api_export_attendance():
    """匯出月度出勤明細 Excel"""
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    if not month:
        month = _dt.now(TW_TZ).strftime('%Y-%m')

    conds, params = ["TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s"], [month]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                ps.employee_code,
                ps.name as staff_name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                pr.punch_type,
                to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei', 'HH24:MI') as punch_time,
                pr.is_manual,
                pr.manual_by,
                pr.gps_distance,
                pr.location_name,
                pr.note
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ps.name, pr.punched_at
        """, params).fetchall()

    PUNCH_LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}

    wb, ws = _xl_workbook(f'{month} 出勤明細')
    headers = ['員工代碼','姓名','部門','職稱','日期','打卡類型','時間','補打卡','操作人','GPS距離(m)','地點','備註']
    widths  = [10, 10, 12, 12, 12, 10, 8, 7, 10, 11, 14, 20]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
         str(r['work_date']), PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
         r['punch_time'], '是' if r['is_manual'] else '',
         r['manual_by'] or '',
         float(r['gps_distance']) if r['gps_distance'] is not None else '',
         r['location_name'] or '', r['note'] or '']
        for r in rows
    ], len(headers))
    return _xl_response(wb, f'attendance_{month}.xlsx')


@app.route('/api/export/attendance-summary', methods=['GET'])
@login_required
def api_export_attendance_summary():
    """匯出月度出勤摘要 Excel（每人每天工時）"""
    month = request.args.get('month', '')
    if not month:
        month = _dt.now(TW_TZ).strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                ps.employee_code,
                ps.name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                BOOL_OR(pr.is_manual) as has_manual,
                COUNT(*) as punch_count
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.employee_code, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (month,)).fetchall()

    wb, ws = _xl_workbook(f'{month} 出勤摘要')
    headers = ['員工代碼','姓名','部門','職稱','日期','上班','下班','工時(h)','打卡次數','含補打']
    widths  = [10, 10, 12, 12, 12, 8, 8, 9, 9, 7]
    _xl_write_header(ws, headers, widths)

    data = []
    for r in rows:
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dt.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dt.fromisoformat(str(r['co_ts']))
                dur_h = round((co - ci).total_seconds() / 3600, 2)
            except Exception:
                pass
        data.append([
            r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
            str(r['work_date']), r['clock_in'] or '', r['clock_out'] or '',
            dur_h, r['punch_count'], '是' if r['has_manual'] else '',
        ])
    _xl_write_rows(ws, data, len(headers), number_cols={8, 9})
    return _xl_response(wb, f'attendance_summary_{month}.xlsx')


@app.route('/api/attendance/anomaly-report', methods=['GET'])
@login_required
def api_anomaly_report_excel():
    """匯出出勤異常報告 Excel（缺打卡、遲到、早退）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    import calendar as _cal
    from datetime import datetime as _dtx, timedelta as _tdx

    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    TW_OFF = _tdx(hours=8)

    with get_db() as conn:
        punch_rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id=pr.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.id, ps.name, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date, ps.name
        """, (month,)).fetchall()

        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date,
                   st.start_time::text as start_time,
                   st.end_time::text   as end_time,
                   ps.name as staff_name, ps.department
            FROM shift_assignments sa
            JOIN shift_types st ON st.id=sa.shift_type_id
            JOIN punch_staff ps ON ps.id=sa.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(sa.shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()

        y_int = int(month[:4]); mo_int = int(month[5:7])
        first_day = f"{y_int}-{mo_int:02d}-01"
        days_in   = _cal.monthrange(y_int, mo_int)[1]
        last_day  = f"{y_int}-{mo_int:02d}-{days_in:02d}"
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (last_day, first_day)).fetchall()

    # Build lookup maps
    shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}
    leave_set = set()
    from datetime import date as _dax, timedelta as _tdax
    for lr in leave_rows:
        s = lr['start_date']; e = lr['end_date']
        cur = s
        while cur <= e:
            leave_set.add((lr['staff_id'], str(cur)))
            cur = _dax.fromisoformat(str(cur)) + _tdax(days=1)
            cur = cur if isinstance(cur, _dax) else cur.date()

    today = _dax.today()

    # Build anomaly rows
    anomalies = []
    for r in punch_rows:
        ds = str(r['work_date'])
        sid = r['staff_id']
        shift = shift_map.get((sid, ds))

        anomaly_type = ''; detail = ''
        late_min = 0; early_min = 0

        if not r['has_in'] and r['has_out']:
            anomaly_type = '缺上班打卡'; detail = f"僅有下班 {str(r['clock_out'])[11:16]}"
        elif r['has_in'] and not r['has_out']:
            if _dax.fromisoformat(ds) < today:
                anomaly_type = '缺下班打卡'; detail = f"上班 {str(r['clock_in'])[11:16]} 無下班"
        elif r['has_in'] and r['has_out'] and shift:
            ci_t = str(r['clock_in'])[11:16]
            co_t = str(r['clock_out'])[11:16]
            sh_s = str(shift['start_time'])[:5]
            sh_e = str(shift['end_time'])[:5]
            try:
                ci_m = int(ci_t[:2])*60 + int(ci_t[3:5])
                sh_s_m = int(sh_s[:2])*60 + int(sh_s[3:5])
                if ci_m - sh_s_m > 10:
                    late_min = ci_m - sh_s_m
                    anomaly_type = '遲到'; detail = f"應 {sh_s}，實際 {ci_t}（+{late_min}分）"
            except Exception:
                pass
            if not anomaly_type:
                try:
                    co_m = int(co_t[:2])*60 + int(co_t[3:5])
                    sh_e_m = int(sh_e[:2])*60 + int(sh_e[3:5])
                    if sh_e_m - co_m > 15:
                        early_min = sh_e_m - co_m
                        anomaly_type = '早退'; detail = f"應 {sh_e}，實際 {co_t}（-{early_min}分）"
                except Exception:
                    pass

        if anomaly_type:
            anomalies.append({
                'staff_name':  r['staff_name'],
                'department':  r['department'] or '',
                'date':        ds,
                'shift_start': str(shift['start_time'])[:5] if shift else '—',
                'shift_end':   str(shift['end_time'])[:5]   if shift else '—',
                'clock_in':    str(r['clock_in'])[11:16]  if r['clock_in']  else '—',
                'clock_out':   str(r['clock_out'])[11:16] if r['clock_out'] else '—',
                'anomaly_type': anomaly_type,
                'detail':       detail,
            })

    # 補抓完全未打卡（有排班但整天沒有任何打卡記錄）
    punched_set = {(r['staff_id'], str(r['work_date'])) for r in punch_rows}
    for sr in shift_rows:
        ds  = str(sr['shift_date'])
        sid = sr['staff_id']
        if _dax.fromisoformat(ds) >= today:
            continue
        if (sid, ds) in punched_set:
            continue
        if (sid, ds) in leave_set:
            continue
        anomalies.append({
            'staff_name':   sr['staff_name'],
            'department':   sr['department'] or '',
            'date':         ds,
            'shift_start':  str(sr['start_time'])[:5],
            'shift_end':    str(sr['end_time'])[:5],
            'clock_in':     '—',
            'clock_out':    '—',
            'anomaly_type': '未打卡',
            'detail':       f"排班 {str(sr['start_time'])[:5]}～{str(sr['end_time'])[:5]}，整日無打卡記錄",
        })
    anomalies.sort(key=lambda x: (x['date'], x['staff_name']))

    # Build Excel
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = f'{month} 異常明細'

    thin = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin',  color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'),
    )
    header_fill   = PatternFill('solid', fgColor='0F1C3A')
    warn_fill     = PatternFill('solid', fgColor='FFF3CD')
    err_fill      = PatternFill('solid', fgColor='FDECEA')
    center_align  = Alignment(horizontal='center', vertical='center')

    headers = ['員工姓名','部門','日期','應上班','應下班','實際上班','實際下班','異常類型','說明']
    col_w   = [12, 10, 12, 8, 8, 8, 8, 12, 30]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    for ri, a in enumerate(anomalies, 2):
        row_fill = err_fill if a['anomaly_type'] in ('缺上班打卡','缺下班打卡') else warn_fill
        vals = [a['staff_name'], a['department'], a['date'],
                a['shift_start'], a['shift_end'],
                a['clock_in'], a['clock_out'],
                a['anomaly_type'], a['detail']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill      = row_fill
            cell.alignment = center_align if ci != 9 else Alignment(vertical='center')
            cell.border    = thin

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Summary sheet
    ws2 = wb.create_sheet('摘要')
    ws2.append(['統計', '數量'])
    ws2.append(['異常總筆數', len(anomalies)])
    by_type = {}
    for a in anomalies:
        by_type[a['anomaly_type']] = by_type.get(a['anomaly_type'], 0) + 1
    for t, c in sorted(by_type.items(), key=lambda x: -x[1]):
        ws2.append([t, c])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response as _FR
    return _FR(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=anomaly_{month}.xlsx'}
    )


@app.route('/api/export/salary', methods=['GET'])
@login_required
def api_export_salary():
    """匯出月度薪資明細 Excel"""
    month = request.args.get('month', '')
    if not month:
        month = _dt.now(TW_TZ).strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, ps.role, ps.salary_type
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month = %s
            ORDER BY ps.name
        """, (month,)).fetchall()

        leave_detail = conn.execute("""
            SELECT lr.staff_id,
                   COALESCE(SUM(CASE WHEN lt.code='personal' THEN lr.total_days ELSE 0 END), 0) AS personal_days,
                   COALESCE(SUM(CASE WHEN lt.code='sick'     THEN lr.total_days ELSE 0 END), 0) AS sick_days
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id = lr.leave_type_id
            WHERE lr.status='approved' AND to_char(lr.start_date,'YYYY-MM')=%s
            GROUP BY lr.staff_id
        """, (month,)).fetchall()
        leave_map = {r['staff_id']: r for r in leave_detail}

    wb, ws = _xl_workbook(f'{month} 薪資明細')
    headers = ['員工代碼','姓名','部門','職稱','薪資制度',
               '工作日','出勤天數','請假天數','無薪假天數','事假天數','病假天數',
               '津貼合計','扣除合計','加班費','實領金額','狀態','備註']
    widths  = [10, 10, 12, 12, 8, 8, 8, 8, 9, 8, 8, 11, 11, 10, 12, 8, 20]
    _xl_write_header(ws, headers, widths)

    data = []
    for r in rows:
        sal_type = r['salary_type'] or 'monthly'
        ld = leave_map.get(r['staff_id'])
        data.append([
            r['employee_code'] or '', r['staff_name'],
            r['department'] or '', r['role'] or '',
            '時薪制' if sal_type == 'hourly' else '月薪制',
            float(r['work_days'] or 0), float(r['actual_days'] or 0),
            float(r['leave_days'] or 0), float(r['unpaid_days'] or 0),
            float(ld['personal_days'] if ld else 0), float(ld['sick_days'] if ld else 0),
            float(r['allowance_total'] or 0), float(r['deduction_total'] or 0),
            float(r['ot_pay'] or 0), float(r['net_pay'] or 0),
            '已確認' if r['status'] == 'confirmed' else '草稿',
            r['note'] or '',
        ])
    _xl_write_rows(ws, data, len(headers), number_cols={6,7,8,9,10,11,12,13,14,15})
    return _xl_response(wb, f'salary_{month}.xlsx')


@app.route('/api/export/leave', methods=['GET'])
@login_required
def api_export_leave():
    """匯出請假記錄 Excel"""
    month    = request.args.get('month', '')
    year     = request.args.get('year',  '')
    staff_id = request.args.get('staff_id', '')
    status   = request.args.get('status', '')  # 空白 = 全部

    conds, params = ['TRUE'], []
    if status:
        conds.append("lr.status=%s"); params.append(status)
    if month:    conds.append("to_char(lr.start_date,'YYYY-MM')=%s"); params.append(month)
    if year:     conds.append("EXTRACT(YEAR FROM lr.start_date)=%s"); params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr
            JOIN punch_staff ps ON ps.id = lr.staff_id
            JOIN leave_types  lt ON lt.id = lr.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY lr.start_date, ps.name
        """, params).fetchall()

    wb, ws = _xl_workbook(f'請假記錄')
    headers = ['員工代碼','姓名','部門','假別','薪資倍率','開始日期','結束日期','天數','原因','代理人','狀態']
    widths  = [10, 10, 12, 10, 8, 12, 12, 7, 24, 10, 8]
    _xl_write_header(ws, headers, widths)

    PAY_LABEL = {1.0:'全薪', 0.5:'半薪', 0.0:'無薪'}
    STATUS_LABEL = {'approved':'已核准','rejected':'已退回','pending':'待審核'}
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '',
         r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
         str(r['start_date']), str(r['end_date']), float(r['total_days']),
         r['reason'] or '', r['substitute_name'] or '',
         STATUS_LABEL.get(r['status'], r['status'])]
        for r in rows
    ], len(headers), number_cols={8})
    return _xl_response(wb, f'leave_{month or year or "all"}.xlsx')


@app.route('/api/export/overtime', methods=['GET'])
@login_required
def api_export_overtime():
    """匯出加班記錄 Excel"""
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    status   = request.args.get('status', '')

    conds, params = ['TRUE'], []
    if month:    conds.append("to_char(r.request_date,'YYYY-MM')=%s"); params.append(month)
    if staff_id: conds.append("r.staff_id=%s"); params.append(int(staff_id))
    if status:   conds.append("r.status=%s"); params.append(status)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT r.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role
            FROM overtime_requests r
            JOIN punch_staff ps ON ps.id = r.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY r.request_date DESC, ps.name
        """, params).fetchall()

    DAY_TYPE = {'weekday':'平日','rest_day':'休息日','holiday':'國定假日','special':'特殊'}
    STATUS_LABEL = {'approved':'已核准','rejected':'已退回','pending':'待審核'}

    wb, ws = _xl_workbook('加班記錄')
    headers = ['員工代碼','姓名','部門','職稱','加班日期','日別','開始時間','結束時間','時數','原因','加班費','狀態','審核人','審核備註']
    widths  = [10, 10, 12, 12, 12, 8, 8, 8, 7, 24, 10, 8, 10, 20]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
         str(r['request_date']), DAY_TYPE.get(r.get('day_type',''), r.get('day_type','')),
         str(r['start_time'] or ''), str(r['end_time'] or ''),
         float(r['ot_hours'] or 0),
         r['reason'] or '', float(r['ot_pay'] or 0) if r.get('ot_pay') else '',
         STATUS_LABEL.get(r['status'], r['status']),
         r.get('reviewed_by') or '', r.get('review_note') or '']
        for r in rows
    ], len(headers), number_cols={9, 11})
    return _xl_response(wb, f'overtime_{month or "all"}.xlsx')


@app.route('/api/export/staff', methods=['GET'])
@login_required
def api_export_staff():
    """匯出員工資料 Excel"""
    dept     = request.args.get('department', '')
    active   = request.args.get('active', '1')

    conds, params = ['TRUE'], []
    if active == '1': conds.append("active=TRUE")
    elif active == '0': conds.append("active=FALSE")
    if dept: conds.append("department=%s"); params.append(dept)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT id, employee_code, name, department, role, position_title,
                   salary_type, base_salary, insured_salary, hourly_rate,
                   daily_hours, hire_date, birth_date, active,
                   username, line_user_id
            FROM punch_staff
            WHERE {' AND '.join(conds)}
            ORDER BY department, name
        """, params).fetchall()

    wb, ws = _xl_workbook('員工資料')
    headers = ['員工代碼','姓名','部門','職稱','職務','薪資制度',
               '底薪','投保薪資','時薪','每日工時','到職日','生日','狀態','帳號','LINE綁定']
    widths  = [10, 10, 12, 12, 12, 8, 11, 11, 10, 8, 12, 12, 6, 12, 8]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
         r['position_title'] or '',
         '時薪制' if r['salary_type'] == 'hourly' else '月薪制',
         float(r['base_salary'] or 0), float(r['insured_salary'] or 0),
         float(r['hourly_rate'] or 0), float(r['daily_hours'] or 8),
         str(r['hire_date']) if r['hire_date'] else '',
         str(r['birth_date']) if r['birth_date'] else '',
         '在職' if r['active'] else '離職',
         r['username'] or '', '已綁定' if r['line_user_id'] else '']
        for r in rows
    ], len(headers), number_cols={7, 8, 9, 10})
    return _xl_response(wb, 'staff_list.xlsx')


@app.route('/api/export/training', methods=['GET'])
@login_required
def api_export_training():
    """匯出訓練記錄 Excel"""
    staff_id = request.args.get('staff_id', '')
    category = request.args.get('category', '')

    conds, params = ['TRUE'], []
    if staff_id: conds.append("tr.staff_id=%s"); params.append(int(staff_id))
    if category: conds.append("tr.category=%s"); params.append(category)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT tr.*, ps.name AS staff_name, ps.department
            FROM training_records tr
            JOIN punch_staff ps ON tr.staff_id = ps.id
            WHERE {' AND '.join(conds)}
            ORDER BY tr.expiry_date ASC NULLS LAST, ps.name
        """, params).fetchall()

    from datetime import date as _today_d
    today = _today_d.today()

    CATEGORY_ZH = {'safety':'安全衛生','fire':'消防','food':'食品衛生',
                   'professional':'專業技能','general':'一般訓練'}

    wb, ws = _xl_workbook('訓練記錄')
    headers = ['員工姓名','部門','課程名稱','類別','完訓日期','到期日','證書號碼','剩餘天數','狀態','備註']
    widths  = [10, 12, 24, 10, 12, 12, 16, 9, 10, 20]
    _xl_write_header(ws, headers, widths)

    from openpyxl.styles import PatternFill
    warn_fill = PatternFill('solid', fgColor='FFF3CD')
    err_fill  = PatternFill('solid', fgColor='FDECEA')

    data = []
    row_colors = []
    for r in rows:
        expiry = str(r['expiry_date']) if r['expiry_date'] else ''
        days_left = ''
        status = '無到期日'
        color = None
        if r['expiry_date']:
            ed = r['expiry_date'] if hasattr(r['expiry_date'], 'year') else _today_d.fromisoformat(str(r['expiry_date']))
            days_left = (ed - today).days
            if days_left < 0:
                status = '已過期'; color = 'err'
            elif days_left <= 60:
                status = '即將到期'; color = 'warn'
            else:
                status = '有效'
        data.append([r['staff_name'], r['department'] or '',
                     r['course_name'], CATEGORY_ZH.get(r['category'], r['category']),
                     str(r['completed_date']) if r['completed_date'] else '',
                     expiry, r['certificate_no'] or '',
                     days_left, status, r['note'] or ''])
        row_colors.append(color)

    from openpyxl.styles import Alignment, Border, Side
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    even_fill = PatternFill('solid', fgColor='F4F6FA')
    center    = Alignment(horizontal='center', vertical='center')
    left_al   = Alignment(horizontal='left', vertical='center')
    for ri, (row_vals, color) in enumerate(zip(data, row_colors), 2):
        fill = err_fill if color == 'err' else warn_fill if color == 'warn' else (even_fill if ri % 2 == 0 else None)
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if fill: cell.fill = fill
            cell.border = thin
            cell.alignment = center if isinstance(v, (int, float, type(None))) else left_al
    return _xl_response(wb, 'training_records.xlsx')


@app.route('/api/export/expense', methods=['GET'])
@login_required
def api_export_expense():
    """匯出費用報帳 Excel"""
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    status   = request.args.get('status', '')

    conds, params = ['TRUE'], []
    if month:    conds.append("to_char(ec.expense_date,'YYYY-MM')=%s"); params.append(month)
    if staff_id: conds.append("ec.staff_id=%s"); params.append(int(staff_id))
    if status:   conds.append("ec.status=%s"); params.append(status)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ec.*, ps.name as staff_name, ps.employee_code, ps.department
            FROM expense_claims ec
            JOIN punch_staff ps ON ps.id = ec.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ec.expense_date DESC
        """, params).fetchall()

    STATUS_LABEL = {'approved':'已核准','rejected':'已退回','pending':'待審核'}

    wb, ws = _xl_workbook('費用報帳')
    headers = ['員工代碼','姓名','部門','費用日期','標題','金額','說明','狀態','審核人','審核意見','申請時間']
    widths  = [10, 10, 12, 12, 24, 11, 30, 8, 10, 24, 16]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '',
         str(r['expense_date']) if r.get('expense_date') else '',
         r['title'] or '', float(r['amount'] or 0),
         r['note'] or '',
         STATUS_LABEL.get(r['status'], r['status']),
         r.get('reviewed_by') or '', r.get('review_note') or '',
         str(r['created_at'])[:16] if r.get('created_at') else '']
        for r in rows
    ], len(headers), number_cols={6})
    return _xl_response(wb, f'expense_{month or "all"}.xlsx')


@app.route('/api/export/leave-balance', methods=['GET'])
@login_required
def api_export_leave_balance():
    """匯出請假餘額 Excel"""
    year = request.args.get('year', '') or str(_dt.now(TW_TZ).year)

    with get_db() as conn:
        rows = conn.execute("""
            SELECT lb.*, ps.name as staff_name, ps.employee_code, ps.department,
                   lt.name as leave_type_name, lt.code as leave_code, lt.max_days
            FROM leave_balances lb
            JOIN punch_staff ps ON ps.id = lb.staff_id
            JOIN leave_types  lt ON lt.id = lb.leave_type_id
            WHERE lb.year = %s
            ORDER BY ps.department, ps.name, lt.sort_order
        """, (int(year),)).fetchall()

    wb, ws = _xl_workbook(f'{year} 請假餘額')
    headers = ['員工代碼','姓名','部門','假別','假別代碼','年度上限(天)','已核准(天)','剩餘(天)']
    widths  = [10, 10, 12, 12, 8, 11, 11, 10]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '',
         r['leave_type_name'], r['leave_code'] or '',
         float(r['max_days']) if r['max_days'] is not None else '無限制',
         float(r['used_days'] or 0),
         round(float(r['max_days'] or 0) - float(r['used_days'] or 0), 2)
         if r['max_days'] is not None else '']
        for r in rows
    ], len(headers), number_cols={6, 7, 8})
    return _xl_response(wb, f'leave_balance_{year}.xlsx')


@app.route('/api/export/performance', methods=['GET'])
@login_required
def api_export_performance():
    """匯出績效考核 Excel"""
    period   = request.args.get('period', '')
    staff_id = request.args.get('staff_id', '')

    conds, params = ['TRUE'], []
    if period:   conds.append("pr.period_label ILIKE %s"); params.append(f'%{period}%')
    if staff_id: conds.append("pr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*,
                   ps.name AS staff_name, ps.employee_code, ps.department, ps.role,
                   pt.name AS template_name
            FROM performance_reviews pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            LEFT JOIN performance_templates pt ON pt.id = pr.template_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.reviewed_at DESC
        """, params).fetchall()

    wb, ws = _xl_workbook('績效考核')
    headers = ['員工代碼','姓名','部門','職稱','考核期間','考核表','分數','滿分','百分比','等級','考核人','備註','考核日期']
    widths  = [10, 10, 12, 12, 16, 16, 8, 8, 9, 6, 10, 30, 16]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
         r['period_label'] or '', r['template_name'] or '',
         float(r['total_score'] or 0), float(r['max_score'] or 100),
         round(float(r['total_score'] or 0) / float(r['max_score'] or 100) * 100, 1),
         r['grade'] or '',
         r['reviewer'] or '', r['comments'] or '',
         str(r['reviewed_at'])[:16] if r.get('reviewed_at') else '']
        for r in rows
    ], len(headers), number_cols={7, 8, 9})
    return _xl_response(wb, f'performance_{period or "all"}.xlsx')


# ── Patch existing review functions with LINE notifications ──────

def _patch_reviews_with_notifications():
    """
    This is called after all route functions are defined.
    We monkey-patch the review endpoints to send LINE notifications.
    The actual patching is done inline in the route handlers below
    via the _notify_review_result helper.
    """
    pass

@app.route('/api/punch/requests/<int:rid>', methods=['PUT'])
@login_required
def api_punch_req_review_v2(rid):
    b           = request.get_json(force=True)
    action      = b.get('action')
    reviewed_by = b.get('reviewed_by', '').strip()
    review_note = b.get('review_note', '').strip()
    if action not in ('approve', 'reject'):
        return jsonify({'error': 'invalid action'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_requests
            SET status=%s, reviewed_by=%s, review_note=%s, reviewed_at=NOW()
            WHERE id=%s
            RETURNING *, (SELECT name FROM punch_staff WHERE id=staff_id) as staff_name
        """, (new_status, reviewed_by, review_note, rid)).fetchone()
        if not row: return ('', 404)
        if action == 'approve':
            conn.execute("""
                INSERT INTO punch_records
                  (staff_id, punch_type, punched_at, note, is_manual, manual_by)
                VALUES (%s,%s,%s,%s,TRUE,%s)
            """, (row['staff_id'], row['punch_type'], row['requested_at'],
                  f'補打卡申請 #{rid}：{row["reason"]}', reviewed_by))
    # LINE notification
    LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}
    dt_str = row['requested_at'].isoformat()[:16].replace('T',' ')
    extra  = f"{LABEL.get(row['punch_type'],'')} {dt_str}"
    if review_note: extra += f"\n審核意見：{review_note}"
    _notify_review_result(row['staff_id'], '補打卡申請', action, extra)
    return jsonify(punch_req_row(row))


# ═══════════════════════════════════════════════════════════════════
# Dashboard API
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/dashboard', methods=['GET'])
@login_required
def api_dashboard():
    from datetime import date as _dd, datetime as _ddt, timezone as _tz, timedelta as _tdd
    TW    = _tz(_tdd(hours=8))
    today = _ddt.now(TW).date()

    # 支援傳入月份參數；預設為當月
    req_month = request.args.get('month', '').strip()
    if req_month and len(req_month) == 7:
        month = req_month
        try:
            y, m = int(month[:4]), int(month[5:])
            import calendar as _cal_d
            last_day = _cal_d.monthrange(y, m)[1]
            from datetime import date as _dcheck
            # 如果查詢的是未來月份，today 仍用實際今天
        except Exception:
            month = today.strftime('%Y-%m')
    else:
        month = today.strftime('%Y-%m')

    with get_db() as conn:

        # ── 今日出勤狀況 ─────────────────────────────────────────
        total_staff = conn.execute(
            "SELECT COUNT(*) as c FROM punch_staff WHERE active=TRUE"
        ).fetchone()['c']

        # 今日已打上班卡的人數
        clocked_in = conn.execute("""
            SELECT COUNT(DISTINCT staff_id) as c
            FROM punch_records
            WHERE punch_type='in'
              AND (punched_at AT TIME ZONE 'Asia/Taipei')::date = %s
        """, (today,)).fetchone()['c']

        # 今日已打下班卡的人數
        clocked_out = conn.execute("""
            SELECT COUNT(DISTINCT staff_id) as c
            FROM punch_records
            WHERE punch_type='out'
              AND (punched_at AT TIME ZONE 'Asia/Taipei')::date = %s
        """, (today,)).fetchone()['c']

        # 今日請假人數（已核准）
        on_leave_today = conn.execute("""
            SELECT COUNT(DISTINCT staff_id) as c
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (today, today)).fetchone()['c']

        # 今日出勤明細（每人狀態）
        today_detail_rows = conn.execute("""
            SELECT ps.id, ps.name, ps.role,
                   MAX(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                   COUNT(pr.id) as punch_count
            FROM punch_staff ps
            LEFT JOIN punch_records pr
              ON pr.staff_id = ps.id
              AND (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date = %s
            WHERE ps.active = TRUE
            GROUP BY ps.id, ps.name, ps.role
            ORDER BY ps.name
        """, (today,)).fetchall()

        today_detail = []
        for r in today_detail_rows:
            # Check if on leave
            leave_row = conn.execute("""
                SELECT lt.name as leave_name
                FROM leave_requests lr
                JOIN leave_types lt ON lt.id = lr.leave_type_id
                WHERE lr.staff_id=%s AND lr.status='approved'
                  AND lr.start_date <= %s AND lr.end_date >= %s
                LIMIT 1
            """, (r['id'], today, today)).fetchone()

            if r['clock_in']:
                if r['clock_out']:
                    status = 'done'
                    status_label = '已下班'
                else:
                    status = 'working'
                    status_label = '上班中'
            elif leave_row:
                status = 'leave'
                status_label = leave_row['leave_name']
            else:
                status = 'absent'
                status_label = '未出勤'

            today_detail.append({
                'id':           r['id'],
                'name':         r['name'],
                'role':         r['role'] or '',
                'clock_in':     r['clock_in']  or '',
                'clock_out':    r['clock_out'] or '',
                'punch_count':  r['punch_count'],
                'status':       status,
                'status_label': status_label,
            })

        # ── 待審申請數 ───────────────────────────────────────────
        pending_punch   = conn.execute("SELECT COUNT(*) as c FROM punch_requests WHERE status='pending'").fetchone()['c']
        pending_ot      = conn.execute("SELECT COUNT(*) as c FROM overtime_requests WHERE status='pending'").fetchone()['c']
        pending_sched   = conn.execute("SELECT COUNT(*) as c FROM schedule_requests WHERE status IN ('pending','modified_pending')").fetchone()['c']
        pending_leave   = conn.execute("SELECT COUNT(*) as c FROM leave_requests WHERE status='pending'").fetchone()['c']

        # ── 本月薪資總覽 ─────────────────────────────────────────
        sal_rows = conn.execute("""
            SELECT COUNT(*) as total_count,
                   COUNT(*) FILTER (WHERE status='confirmed') as confirmed_count,
                   COALESCE(SUM(net_pay),0) as total_net,
                   COALESCE(SUM(allowance_total),0) as total_allow,
                   COALESCE(SUM(deduction_total),0) as total_deduct
            FROM salary_records WHERE month=%s
        """, (month,)).fetchone()

        # ── 本月出勤統計（每天出勤人數，用於折線圖）─────────────
        import calendar as _cal
        days_in_month = _cal.monthrange(today.year, today.month)[1]
        daily_rows = conn.execute("""
            SELECT (punched_at AT TIME ZONE 'Asia/Taipei')::date as d,
                   COUNT(DISTINCT staff_id) as cnt
            FROM punch_records
            WHERE punch_type='in'
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY (punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY d
        """, (month,)).fetchall()
        daily_map = {str(r['d']): r['cnt'] for r in daily_rows}
        daily_attendance = []
        for day in range(1, days_in_month + 1):
            ds = f"{month}-{day:02d}"
            dt = _dd(today.year, today.month, day)
            daily_attendance.append({
                'date':    ds,
                'day':     day,
                'count':   daily_map.get(ds, 0),
                'is_past': dt <= today,
                'weekday': dt.weekday(),
            })

        # ── 本月請假類型分佈（圓餅圖）───────────────────────────
        leave_dist_rows = conn.execute("""
            SELECT lt.name, lt.color, COUNT(*) as cnt,
                   COALESCE(SUM(lr.total_days),0) as days
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id = lr.leave_type_id
            WHERE lr.status='approved'
              AND to_char(lr.start_date,'YYYY-MM')=%s
            GROUP BY lt.name, lt.color
            ORDER BY days DESC
        """, (month,)).fetchall()
        leave_distribution = [
            {'name': r['name'], 'color': r['color'], 'count': r['cnt'], 'days': float(r['days'])}
            for r in leave_dist_rows
        ]

        # ── 本月加班費排行（橫條圖）─────────────────────────────
        ot_rank_rows = conn.execute("""
            SELECT ps.name, ps.role,
                   COALESCE(SUM(r.ot_pay),0) as total_pay,
                   COALESCE(SUM(r.ot_hours),0) as total_hours
            FROM overtime_requests r
            JOIN punch_staff ps ON ps.id = r.staff_id
            WHERE r.status='approved'
              AND to_char(r.request_date,'YYYY-MM')=%s
            GROUP BY ps.name, ps.role
            ORDER BY total_pay DESC
            LIMIT 8
        """, (month,)).fetchall()
        ot_ranking = [
            {'name': r['name'], 'role': r['role'] or '', 'pay': float(r['total_pay']), 'hours': float(r['total_hours'])}
            for r in ot_rank_rows
        ]

    from datetime import date as _ddc
    cur_month = _ddc.today().strftime('%Y-%m')
    return jsonify({
        'month':            month,
        'today':            str(today),
        'is_current_month': month == cur_month,
        # 今日出勤
        'today_summary': {
            'total':       total_staff,
            'working':     clocked_in - clocked_out,
            'clocked_in':  clocked_in,
            'clocked_out': clocked_out,
            'on_leave':    on_leave_today,
            'absent':      total_staff - clocked_in - on_leave_today,
        },
        'today_detail': today_detail,
        # 待審申請
        'pending': {
            'punch':  pending_punch,
            'ot':     pending_ot,
            'sched':  pending_sched,
            'leave':  pending_leave,
            'total':  pending_punch + pending_ot + pending_sched + pending_leave,
        },
        # 本月薪資
        'salary_summary': {
            'total_count':     sal_rows['total_count'],
            'confirmed_count': sal_rows['confirmed_count'],
            'total_net':       float(sal_rows['total_net']),
            'total_allow':     float(sal_rows['total_allow']),
            'total_deduct':    float(sal_rows['total_deduct']),
        },
        # 圖表資料
        'daily_attendance':    daily_attendance,
        'leave_distribution':  leave_distribution,
        'ot_ranking':          ot_ranking,
    })


# ── Dashboard 擴充 API ────────────────────────────────────────────────────────

@app.route('/api/dashboard/labor-cost', methods=['GET'])
@login_required
def api_dashboard_labor_cost():
    """近 12 個月人事費用趨勢"""
    from datetime import date as _dlc
    today = _dlc.today()
    months = []
    for i in range(11, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0: m += 12; y -= 1
        months.append(f'{y}-{m:02d}')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month, COALESCE(SUM(net_pay),0) as total
            FROM salary_records
            WHERE month = ANY(%s)
            GROUP BY month
        """, (months,)).fetchall()
    cost_map = {r['month']: float(r['total']) for r in rows}
    return jsonify({
        'months':     months,
        'labor_cost': [cost_map.get(m, 0) for m in months],
    })


@app.route('/api/dashboard/attendance-heatmap', methods=['GET'])
@login_required
def api_dashboard_attendance_heatmap():
    """本月每日出勤率（熱力圖資料）"""
    from datetime import date as _dah
    import calendar as _calh
    month = request.args.get('month', '') or _dah.today().strftime('%Y-%m')
    y, mo = int(month[:4]), int(month[5:7])
    days_in = _calh.monthrange(y, mo)[1]

    with get_db() as conn:
        total_staff = conn.execute(
            "SELECT COUNT(*) as c FROM punch_staff WHERE active=TRUE"
        ).fetchone()['c']

        punch_rows = conn.execute("""
            SELECT (punched_at AT TIME ZONE 'Asia/Taipei')::date as d,
                   COUNT(DISTINCT staff_id) as cnt
            FROM punch_records
            WHERE punch_type='in'
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY d
        """, (month,)).fetchall()

        leave_rows = conn.execute("""
            SELECT lr.start_date, lr.end_date, COUNT(*) as cnt
            FROM leave_requests lr
            WHERE lr.status='approved'
              AND TO_CHAR(lr.start_date,'YYYY-MM')=%s OR TO_CHAR(lr.end_date,'YYYY-MM')=%s
            GROUP BY lr.start_date, lr.end_date
        """, (month, month)).fetchall()

    punch_map = {str(r['d']): int(r['cnt']) for r in punch_rows}

    from datetime import date as _dah2, timedelta as _tdah
    leave_map = {}
    for lr in leave_rows:
        s = _dah2.fromisoformat(str(lr['start_date']))
        e = _dah2.fromisoformat(str(lr['end_date']))
        cur = s
        while cur <= e:
            ds = str(cur)
            if ds.startswith(month):
                leave_map[ds] = leave_map.get(ds, 0) + 1
            cur += _tdah(days=1)

    days = []
    for d in range(1, days_in + 1):
        ds = f'{y}-{mo:02d}-{d:02d}'
        cnt = punch_map.get(ds, 0)
        rate = round(cnt / total_staff, 3) if total_staff > 0 else 0
        days.append({
            'date': ds,
            'day_of_week': _dah2(y, mo, d).weekday(),
            'count': cnt,
            'attendance_rate': rate,
            'on_leave': leave_map.get(ds, 0),
        })

    return jsonify({'month': month, 'total_staff': total_staff, 'days': days})


@app.route('/api/dashboard/leave-distribution', methods=['GET'])
@login_required
def api_dashboard_leave_distribution():
    """本年度請假類型分佈"""
    from datetime import date as _dld
    year = request.args.get('year', str(_dld.today().year))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lt.name, lt.color,
                   COUNT(*) as cnt,
                   COALESCE(SUM(lr.days), 0) as days
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE lr.status='approved'
              AND EXTRACT(YEAR FROM lr.start_date)=%s
            GROUP BY lt.name, lt.color
            ORDER BY days DESC
        """, (int(year),)).fetchall()
    total = sum(float(r['days']) for r in rows)
    return jsonify({
        'year': year,
        'total_leave_days': total,
        'breakdown': [{
            'name':  r['name'],
            'color': r['color'] or '#4a7bda',
            'days':  float(r['days']),
            'count': int(r['cnt']),
            'pct':   round(float(r['days']) / total * 100, 1) if total > 0 else 0,
        } for r in rows],
    })


# ── 年度扣繳憑單 ────────────────────────────────────────────────────────────

@app.route('/api/export/withholding', methods=['GET'])
@require_module('salary')
def api_export_withholding():
    """年度薪資所得扣繳憑單（所得類別50）"""
    from datetime import date as _dwh
    year   = request.args.get('year', str(_dwh.today().year))
    fmt    = request.args.get('format', 'html')

    fs = _get_finance_settings()
    company_name   = fs.get('company_name', '')
    company_tax_id = fs.get('company_tax_id', '')
    company_address= fs.get('company_address', '')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id, ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total), 0)       AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld), 0)   AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary), 0)        AS avg_insured
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.id, ps.name, ps.national_id, ps.address
            ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()

    # 計算二代健保補充費
    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0

    data = []
    for i, r in enumerate(rows, 1):
        gross = float(r['gross_salary'])
        insured = float(r['avg_insured'])
        data.append({
            'no':          i,
            'name':        r['name'],
            'national_id': r['national_id'] or '—',
            'address':     r['address'] or '—',
            'gross':       gross,
            'supp_nhi':    supp_nhi(gross, insured),
            'tax':         float(r['tax_withheld']),
        })

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        wb = wb2 = openpyxl.Workbook()
        ws = wb.active; ws.title = f'{year}年扣繳憑單'
        hfill = PatternFill('solid', fgColor='0F1C3A')
        thin  = Border(*[Side(style='thin', color='DDDDDD')]*4)
        hdrs  = ['序號','姓名','身分證字號','地址','年度薪資合計','二代健保補充費','扣繳稅額']
        ws.append(hdrs)
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 16; ws.column_dimensions['G'].width = 12
        for d in data:
            ws.append([d['no'], d['name'], d['national_id'], d['address'],
                       d['gross'], d['supp_nhi'], d['tax']])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        from flask import Response as _FR2
        return _FR2(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=withholding_{year}.xlsx'})

    # HTML printable
    rows_html = ''.join(f"""
      <tr>
        <td style="text-align:center">{d['no']}</td>
        <td>{d['name']}</td>
        <td style="font-family:monospace">{d['national_id']}</td>
        <td style="font-size:11px">{d['address']}</td>
        <td style="text-align:right;font-family:monospace">{d['gross']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['supp_nhi']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['tax']:,.0f}</td>
      </tr>""" for d in data)
    html = f"""<!DOCTYPE html><html lang="zh-TW"><head>
<meta charset="UTF-8"><title>{year}年度薪資扣繳憑單</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans TC',sans-serif;font-size:12px;padding:20px;color:#1e2a45}}
h2{{font-size:16px;font-weight:700;margin-bottom:4px}}
.meta{{font-size:11px;color:#666;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse;margin-bottom:20px}}
th{{background:#0f1c3a;color:#fff;padding:7px 10px;font-size:11px;font-weight:600;text-align:left}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:12px}}
tr:nth-child(even){{background:#f8f9fb}}
.note{{font-size:10px;color:#888;border-top:1px solid #ddd;padding-top:8px}}
@media print{{button{{display:none}}}}
</style></head><body>
<button onclick="window.print()" style="margin-bottom:16px;padding:6px 16px;background:#0f1c3a;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">列印</button>
<h2>{year} 年度薪資所得扣繳憑單（所得類別 50）</h2>
<div class="meta">扣繳義務人：{company_name}　統一編號：{company_tax_id}　地址：{company_address}　製表日期：{_dwh.today().isoformat()}</div>
<table>
<thead><tr><th>#</th><th>員工姓名</th><th>身分證字號</th><th>地址</th><th>年度薪資合計(元)</th><th>二代健保補充費(元)</th><th>扣繳稅額(元)</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="note">※ 本報表依薪資紀錄計算，二代健保補充費 = 超出投保薪資部分 × 2.11%。扣繳稅額請依各月薪資記錄人工確認。</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ── 勞健保 EDI 申報 ─────────────────────────────────────────────────────────

def _get_insurance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM insurance_settings").fetchall()
        return {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        return {}

def _roc_date(date_str):
    """Convert YYYY-MM-DD to YYYMMDD (ROC year)"""
    if not date_str: return '0000000'
    try:
        from datetime import date as _dedi
        d = _dedi.fromisoformat(str(date_str)[:10])
        return f'{d.year - 1911:03d}{d.month:02d}{d.day:02d}'
    except Exception:
        return '0000000'

def _edi_bytes(val, width, numeric=False):
    """Encode value to fixed-width bytes (Big5 for text, ASCII-padded for numeric)"""
    s = str(val or '')
    if numeric:
        return s.rjust(width, '0').encode('ascii', errors='replace')[:width]
    try:
        b = s.encode('big5', errors='replace')
    except Exception:
        b = s.encode('ascii', errors='replace')
    if len(b) < width:
        b = b + b' ' * (width - len(b))
    return b[:width]


@app.route('/api/insurance/settings', methods=['GET'])
@require_module('salary')
def api_insurance_settings_get():
    return jsonify(_get_insurance_settings())

@app.route('/api/insurance/settings', methods=['PUT'])
@require_module('salary')
def api_insurance_settings_put():
    b = request.get_json(force=True)
    with get_db() as conn:
        for k in ('labor_insurance_no', 'health_insurance_no', 'employer_name', 'employer_id'):
            conn.execute(
                "INSERT INTO insurance_settings VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(b.get(k, '')).strip()))
    return jsonify({'ok': True})


def _get_edi_staff(staff_ids_str):
    """Fetch staff rows for EDI, optionally filtered by comma-separated IDs."""
    with get_db() as conn:
        if staff_ids_str:
            ids = [int(x) for x in staff_ids_str.split(',') if x.strip().isdigit()]
            rows = conn.execute(
                f"SELECT * FROM punch_staff WHERE id = ANY(%s) AND active=TRUE ORDER BY name",
                (ids,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM punch_staff WHERE active=TRUE ORDER BY name").fetchall()
    return rows


@app.route('/api/export/edi/labor-enroll', methods=['GET'])
@require_module('salary')
def api_edi_labor_enroll():
    """勞工保險加退保申報 EDI（Big5 固定寬度格式）"""
    event_type  = request.args.get('event_type', 'in')   # in=加保 out=退保
    staff_ids   = request.args.get('staff_ids', '')
    event_date  = request.args.get('event_date', '')
    cfg         = _get_insurance_settings()
    labor_no    = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    event_code  = b'1' if event_type == 'in' else b'2'
    event_roc   = _roc_date(event_date).encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(labor_no, 8) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            _roc_date(s.get('birth_date')).encode('ascii') +
            event_roc +
            event_code +
            insured +
            gender_code +
            b'00'   # 職業類別（一般）
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    fname   = f'labor_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    from flask import Response as _FRe
    return _FRe(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/api/export/edi/labor-salary', methods=['GET'])
@require_module('salary')
def api_edi_labor_salary():
    """勞工保險投保薪資調整申報 EDI"""
    month     = request.args.get('month', '')
    staff_ids = request.args.get('staff_ids', '')
    cfg       = _get_insurance_settings()
    labor_no  = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    if not month:
        from datetime import date as _dm2
        month = _dm2.today().strftime('%Y-%m')
    month_roc = f"{int(month[:4]) - 1911:03d}{month[5:7]}".encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(labor_no, 8) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            insured +
            month_roc
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    from flask import Response as _FRs
    return _FRs(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename=labor_salary_{month}.edi'})


@app.route('/api/export/edi/health-enroll', methods=['GET'])
@require_module('salary')
def api_edi_health_enroll():
    """全民健康保險加退保申報 EDI"""
    event_type = request.args.get('event_type', 'in')
    staff_ids  = request.args.get('staff_ids', '')
    event_date = request.args.get('event_date', '')
    cfg        = _get_insurance_settings()
    health_no  = cfg.get('health_insurance_no', '').ljust(10)[:10]
    event_code = b'1' if event_type == 'in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(health_no, 10) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            _roc_date(s.get('birth_date')).encode('ascii') +
            event_roc +
            event_code +
            insured +
            gender_code
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    fname   = f'health_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    from flask import Response as _FRh
    return _FRh(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── 多店管理 ─────────────────────────────────────────────────────────────────

@app.route('/api/stores', methods=['GET'])
@login_required
def api_stores_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stores ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/stores', methods=['POST'])
@login_required
def api_stores_create():
    b = request.get_json(force=True)
    name = (b.get('name') or '').strip()
    code = (b.get('code') or '').strip() or None
    if not name: return jsonify({'error': '店名為必填'}), 400
    with get_db() as conn:
        row = conn.execute(
            "INSERT INTO stores (name, code, address) VALUES (%s,%s,%s) RETURNING *",
            (name, code, (b.get('address') or '').strip())
        ).fetchone()
    return jsonify(dict(row)), 201

@app.route('/api/stores/<int:sid>', methods=['PUT'])
@login_required
def api_stores_update(sid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE stores SET name=%s, code=%s, address=%s, active=%s WHERE id=%s RETURNING *
        """, ((b.get('name') or '').strip(), (b.get('code') or None),
              (b.get('address') or '').strip(), bool(b.get('active', True)), sid)).fetchone()
    return jsonify(dict(row)) if row else ('', 404)

@app.route('/api/stores/<int:sid>', methods=['DELETE'])
@login_required
def api_stores_delete(sid):
    with get_db() as conn:
        conn.execute("UPDATE punch_staff     SET store_id=NULL WHERE store_id=%s", (sid,))
        conn.execute("UPDATE punch_locations SET store_id=NULL WHERE store_id=%s", (sid,))
        conn.execute("DELETE FROM stores WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

@app.route('/api/stores/<int:sid>/staff', methods=['GET'])
@login_required
def api_store_staff(sid):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, role, active FROM punch_staff WHERE store_id=%s ORDER BY name", (sid,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/staff/<int:sid>/store', methods=['PUT'])
@login_required
def api_staff_assign_store(sid):
    b = request.get_json(force=True)
    store_id = b.get('store_id')
    with get_db() as conn:
        conn.execute("UPDATE punch_staff SET store_id=%s WHERE id=%s", (store_id, sid))
    return jsonify({'ok': True})


# ── 排班需求 & 自動排班 ──────────────────────────────────────────────────────

@app.route('/api/shifts/staffing-requirements', methods=['GET'])
@login_required
def api_staffing_req_get():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.id, r.shift_type_id, r.day_of_week, r.required_count,
                   st.name as shift_name, st.color as shift_color
            FROM shift_staffing_requirements r
            JOIN shift_types st ON st.id=r.shift_type_id
            ORDER BY st.sort_order, r.day_of_week
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/shifts/staffing-requirements', methods=['PUT'])
@login_required
def api_staffing_req_put():
    items = request.get_json(force=True)
    if not isinstance(items, list):
        return jsonify({'error': '格式錯誤'}), 400
    count = 0
    with get_db() as conn:
        for it in items:
            stid = int(it.get('shift_type_id', 0))
            dow  = int(it.get('day_of_week', 0))
            req  = max(0, int(it.get('required_count', 1)))
            if req == 0:
                conn.execute(
                    "DELETE FROM shift_staffing_requirements WHERE shift_type_id=%s AND day_of_week=%s",
                    (stid, dow))
            else:
                conn.execute("""
                    INSERT INTO shift_staffing_requirements (shift_type_id, day_of_week, required_count, updated_at)
                    VALUES (%s,%s,%s,NOW())
                    ON CONFLICT (shift_type_id, day_of_week)
                    DO UPDATE SET required_count=EXCLUDED.required_count, updated_at=NOW()
                """, (stid, dow, req))
            count += 1
    return jsonify({'ok': True, 'upserted': count})


@app.route('/api/schedule/auto-generate', methods=['POST'])
@login_required
def api_auto_generate_schedule():
    """自動排班引擎：依人力需求與員工可用性生成班表建議"""
    from datetime import date as _dag, timedelta as _tdag
    import calendar as _calag

    b        = request.get_json(force=True)
    month    = (b.get('month') or '').strip()
    overwrite = bool(b.get('overwrite', False))
    if not month:
        month = _dag.today().strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    days_in   = _calag.monthrange(y, mo)[1]
    all_dates = [_dag(y, mo, d) for d in range(1, days_in + 1)]

    with get_db() as conn:
        shift_types = conn.execute(
            "SELECT * FROM shift_types WHERE active=TRUE ORDER BY sort_order"
        ).fetchall()
        requirements = conn.execute("""
            SELECT shift_type_id, day_of_week, required_count
            FROM shift_staffing_requirements
        """).fetchall()
        staff_list = conn.execute(
            "SELECT id, name FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()

        # 本月已核准休假日期（per staff）
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (f'{y}-{mo:02d}-{days_in:02d}', f'{y}-{mo:02d}-01')).fetchall()

        # 已核准排休
        sched_rows = conn.execute("""
            SELECT staff_id, requested_dates
            FROM schedule_requests
            WHERE status='approved'
              AND to_char(created_at,'YYYY-MM')=%s
        """, (month,)).fetchall()

        # 現有班表
        existing = conn.execute("""
            SELECT staff_id, shift_date FROM shift_assignments
            WHERE TO_CHAR(shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()

    # 建立不可上班日 set: {(staff_id, date_str)}
    off_days = set()
    for lr in leave_rows:
        s = _dag.fromisoformat(str(lr['start_date']))
        e = _dag.fromisoformat(str(lr['end_date']))
        cur = s
        while cur <= e:
            off_days.add((lr['staff_id'], str(cur)))
            cur += _tdag(days=1)
    for sr in sched_rows:
        rdates = sr['requested_dates']
        if isinstance(rdates, str):
            try: rdates = _json.loads(rdates)
            except: rdates = []
        for ds in (rdates or []):
            off_days.add((sr['staff_id'], ds))

    # 已有班表 set（不 overwrite 時跳過）
    existing_set = {(r['staff_id'], str(r['shift_date'])) for r in existing}

    # 需求 map: {(shift_type_id, day_of_week): required_count}
    req_map = {(r['shift_type_id'], r['day_of_week']): r['required_count'] for r in requirements}

    # 排班計數器（避免連續超時）
    assigned_days  = {s['id']: [] for s in staff_list}  # staff_id -> [date]
    assignments    = []
    conflicts      = []
    staff_ids      = [s['id'] for s in staff_list]
    staff_name_map = {s['id']: s['name'] for s in staff_list}

    for date in all_dates:
        dow = date.weekday()  # 0=Mon, 6=Sun
        ds  = str(date)

        for st in shift_types:
            stid     = st['id']
            needed   = req_map.get((stid, dow), 0)
            if needed <= 0:
                continue

            # 可用員工：未請假、未排休
            available = [
                sid for sid in staff_ids
                if (sid, ds) not in off_days
            ]

            # 排除已被指派在其他班（同日）
            already_today = {a['staff_id'] for a in assignments if a['shift_date'] == ds}
            available = [sid for sid in available if sid not in already_today]

            # 排除連續 7 天（含本日）的員工
            def consecutive_days(sid, d):
                days = sorted(assigned_days[sid])
                streak = 0
                check = d
                while check in days:
                    streak += 1
                    check = str(_dag.fromisoformat(check) - _tdag(days=1))
                return streak

            available_ok = [sid for sid in available if consecutive_days(sid, ds) < 6]

            # 按本月已排天數升序（均衡分配）
            available_ok.sort(key=lambda sid: len(assigned_days[sid]))

            assigned_count = 0
            for sid in available_ok:
                if assigned_count >= needed:
                    break
                if not overwrite and (sid, ds) in existing_set:
                    assigned_count += 1
                    continue
                assignments.append({
                    'staff_id':     sid,
                    'staff_name':   staff_name_map[sid],
                    'shift_type_id': stid,
                    'shift_name':   st['name'],
                    'shift_date':   ds,
                })
                assigned_days[sid].append(ds)
                assigned_count += 1

            if assigned_count < needed:
                conflicts.append({
                    'type':   'understaffed',
                    'date':   ds,
                    'shift':  st['name'],
                    'detail': f'{ds} {st["name"]} 需要 {needed} 人，僅能排 {assigned_count} 人',
                })

    # 寫入資料庫
    inserted = 0
    if assignments:
        with get_db() as conn:
            for a in assignments:
                try:
                    if overwrite:
                        conn.execute("""
                            INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date)
                            VALUES (%s,%s,%s)
                            ON CONFLICT (staff_id, shift_date) DO UPDATE
                            SET shift_type_id=EXCLUDED.shift_type_id
                        """, (a['staff_id'], a['shift_type_id'], a['shift_date']))
                    else:
                        conn.execute("""
                            INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date)
                            VALUES (%s,%s,%s)
                            ON CONFLICT DO NOTHING
                        """, (a['staff_id'], a['shift_type_id'], a['shift_date']))
                    inserted += 1
                except Exception:
                    pass

    return jsonify({
        'ok':          True,
        'month':       month,
        'assignments': assignments,
        'conflicts':   conflicts,
        'summary': {
            'assigned':       inserted,
            'conflict_count': len(conflicts),
        },
    })


# ═══════════════════════════════════════════════════════════════════
# 勞基法自動更新通知 (Labor Law Auto-Monitor)
# ═══════════════════════════════════════════════════════════════════

def _init_labor_law_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS labor_law_updates (
                    id          SERIAL PRIMARY KEY,
                    law_name    TEXT NOT NULL DEFAULT '勞動基準法',
                    amend_date  DATE NOT NULL,
                    version_note TEXT DEFAULT '',
                    summary     TEXT DEFAULT '',
                    source_url  TEXT DEFAULT 'https://law.moj.gov.tw/LawClass/LawHistory.aspx?pcode=N0030001',
                    announced   BOOLEAN DEFAULT FALSE,
                    fetched_at  TIMESTAMPTZ DEFAULT NOW(),
                    created_at  TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE(law_name, amend_date)
                )
            """)
    except Exception as e:
        print(f"[labor_law_init] {e}")

_init_labor_law_db()


def _scrape_labor_law_updates():
    """
    Fetch 勞動基準法 amendment history from 全國法規資料庫.
    Returns list of dicts: {amend_date, version_note, summary, source_url}
    """
    import urllib.request as _ur
    import html.parser as _hp
    import re as _re

    LAW_HISTORY_URL = 'https://law.moj.gov.tw/LawClass/LawHistory.aspx?pcode=N0030001'
    LAW_CONTENT_URL = 'https://law.moj.gov.tw/LawClass/LawAll.aspx?pcode=N0030001'

    class _HistoryParser(_hp.HTMLParser):
        def __init__(self):
            super().__init__()
            self.in_td = False
            self.cells = []
            self.current_cell = ''
            self.rows = []
            self.current_row = []

        def handle_starttag(self, tag, attrs):
            if tag == 'tr':
                self.current_row = []
            elif tag in ('td', 'th'):
                self.in_td = True
                self.current_cell = ''

        def handle_endtag(self, tag):
            if tag in ('td', 'th'):
                self.in_td = False
                self.current_row.append(self.current_cell.strip())
            elif tag == 'tr':
                if self.current_row:
                    self.rows.append(self.current_row)

        def handle_data(self, data):
            if self.in_td:
                self.current_cell += data

    results = []
    try:
        req = _ur.Request(
            LAW_HISTORY_URL,
            headers={'User-Agent': 'Mozilla/5.0 (compatible; LaborLawMonitor/1.0)'}
        )
        with _ur.urlopen(req, timeout=20) as resp:
            raw = resp.read()
            try:
                content = raw.decode('utf-8')
            except Exception:
                content = raw.decode('big5', errors='replace')

        parser = _HistoryParser()
        parser.feed(content)

        date_pat = _re.compile(r'(\d{3,4})[./年](\d{1,2})[./月](\d{1,2})')
        seen = set()
        for row in parser.rows:
            text = ' '.join(row)
            m = date_pat.search(text)
            if not m:
                continue
            yr, mo, dy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            # 民國 → 西元
            if yr < 1900:
                yr += 1911
            if not (1984 <= yr <= 2099 and 1 <= mo <= 12 and 1 <= dy <= 31):
                continue
            date_str = f"{yr:04d}-{mo:02d}-{dy:02d}"
            if date_str in seen:
                continue
            seen.add(date_str)
            note = ' '.join(c for c in row if c and not date_pat.search(c))[:200]
            results.append({
                'amend_date': date_str,
                'version_note': note.strip(),
                'summary': f'勞動基準法於 {yr} 年 {mo} 月 {dy} 日修正',
                'source_url': LAW_HISTORY_URL,
            })

        results.sort(key=lambda x: x['amend_date'], reverse=True)
    except Exception as e:
        print(f"[labor_law_scrape] {e}")

    return results


def _run_labor_law_check():
    """
    Scrape and store 勞基法 amendments.
    - First run (empty table): import all history as announced=TRUE (no false alarms).
    - Subsequent runs: new records inserted as announced=FALSE; auto-announcement created.
    """
    updates = _scrape_labor_law_updates()
    if not updates:
        print("[labor_law_check] no data returned from scraper")
        return

    # Detect first-run: if the table is empty, this is an initial historical import
    try:
        with get_db() as conn:
            existing_count = conn.execute(
                "SELECT COUNT(*) AS cnt FROM labor_law_updates WHERE law_name='勞動基準法'"
            ).fetchone()['cnt']
    except Exception as e:
        print(f"[labor_law_check] count error: {e}")
        existing_count = 0

    is_initial_import = (existing_count == 0)
    if is_initial_import:
        print("[labor_law_check] initial import — marking all historical records as announced")

    new_ids = []
    for u in updates:
        try:
            with get_db() as conn:
                # announced=TRUE for initial import (historical data, no notification needed)
                # announced=FALSE for new records found on subsequent runs
                row = conn.execute("""
                    INSERT INTO labor_law_updates
                      (law_name, amend_date, version_note, summary, source_url, announced)
                    VALUES ('勞動基準法', %s, %s, %s, %s, %s)
                    ON CONFLICT (law_name, amend_date) DO NOTHING
                    RETURNING id
                """, (u['amend_date'], u['version_note'], u['summary'],
                      u['source_url'], is_initial_import)).fetchone()
                if row and not is_initial_import:
                    new_ids.append(row['id'])
        except Exception as e:
            print(f"[labor_law_check] db error: {e}")

    if not new_ids:
        print(f"[labor_law_check] no new amendments (initial={is_initial_import})")
        return

    print(f"[labor_law_check] {len(new_ids)} new amendment(s) — creating announcement")
    try:
        with get_db() as conn:
            new_rows = conn.execute("""
                SELECT * FROM labor_law_updates
                WHERE id = ANY(%s)
                ORDER BY amend_date DESC
            """, (new_ids,)).fetchall()
            if not new_rows:
                return
            dates = '、'.join(str(dict(r)['amend_date']) for r in new_rows)
            title = f"勞動基準法修正公告（{dates}）"
            lines = ["系統偵測到勞動基準法新修正版本，請人資部門注意相關條文變動：\n"]
            for r in new_rows:
                rd = dict(r)
                lines.append(f"・修正日期：{rd['amend_date']}")
                if rd.get('version_note'):
                    lines.append(f"  {rd['version_note']}")
                lines.append(f"  來源：{rd['source_url']}")
            lines.append("\n請至全國法規資料庫確認詳細條文內容。")
            content = '\n'.join(lines)
            conn.execute("""
                INSERT INTO announcements
                  (title, content, category, priority, is_pinned, visible_to, author, active)
                VALUES (%s, %s, 'labor_law', 'high', TRUE, 'admin', '勞基法監控系統', TRUE)
            """, (title, content))
            # announced stays FALSE until admin opens the labor-law page (drives the badge)
    except Exception as e:
        print(f"[labor_law_check] announce error: {e}")


def _labor_law_check_loop():
    """Weekly background check for 勞基法 updates."""
    import time as _t
    # 首次啟動延遲 30 秒，避免與其他 init 衝突
    _t.sleep(30)
    _run_labor_law_check()
    while True:
        # 每 7 天檢查一次
        _t.sleep(7 * 24 * 3600)
        _run_labor_law_check()


threading.Thread(target=_labor_law_check_loop, daemon=True).start()


# ── API: 勞基法更新記錄 ──────────────────────────────────────────

@app.route('/api/labor-law/updates', methods=['GET'])
@login_required
def api_labor_law_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM labor_law_updates
            ORDER BY amend_date DESC
            LIMIT 100
        """).fetchall()
        # Admin opened the page — mark all unread as seen (clears the badge)
        conn.execute("""
            UPDATE labor_law_updates SET announced=TRUE
            WHERE announced=FALSE
        """)
    result = []
    for r in rows:
        d = dict(r)
        if d.get('amend_date'):
            d['amend_date'] = str(d['amend_date'])
        if d.get('fetched_at'):
            d['fetched_at'] = d['fetched_at'].isoformat()
        if d.get('created_at'):
            d['created_at'] = d['created_at'].isoformat()
        result.append(d)
    return jsonify(result)


@app.route('/api/labor-law/check', methods=['POST'])
@login_required
def api_labor_law_trigger_check():
    """手動觸發立即檢查勞基法更新"""
    threading.Thread(target=_run_labor_law_check, daemon=True).start()
    return jsonify({'ok': True, 'message': '已開始背景檢查，請稍後重新整理'})


@app.route('/api/labor-law/badge', methods=['GET'])
@login_required
def api_labor_law_badge():
    """Badge counts new amendments not yet seen by admin (announced=FALSE)."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS cnt FROM labor_law_updates WHERE announced=FALSE"
        ).fetchone()
    return jsonify({'unread': row['cnt'] if row else 0})


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

# ═══════════════════════════════════════════════════════════════════
# Feature: Salary PDF (HTML print endpoint)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/salary/records/<int:rid>/pdf', methods=['GET'])
@require_module('salary')
def api_salary_pdf(rid):
    """回傳薪資單 HTML（供瀏覽器列印/另存 PDF）"""
    # 允許員工查看自己的薪資單
    if not session.get('logged_in'):
        sid = session.get('punch_staff_id')
        if not sid:
            return '未登入', 401
    with get_db() as conn:
        row = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, ps.role, ps.salary_type,
                   ps.hourly_rate, ps.hire_date
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.id = %s
        """, (rid,)).fetchone()
    if not row:
        return '找不到薪資記錄', 404
    # 員工只能看自己的
    if not session.get('logged_in'):
        if row['staff_id'] != session.get('punch_staff_id'):
            return '無權限', 403

    d         = salary_record_row(row)
    items     = d.get('items') or []
    allow_items  = [i for i in items if i.get('type') == 'allowance']
    deduct_items = [i for i in items if i.get('type') == 'deduction']
    is_hourly = (row['salary_type'] == 'hourly')

    def money(v):
        try: return f"${float(v):,.0f}"
        except: return '$0'

    def esc_h(s):
        return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

    allow_rows = ''.join(f"""
        <tr>
          <td>{esc_h(i['name'])}</td>
          <td class="num green">{money(i['amount'])}</td>
          <td class="note">{esc_h(i.get('calc_note',''))}</td>
        </tr>""" for i in allow_items)

    deduct_rows = ''.join(f"""
        <tr>
          <td>{esc_h(i['name'])}</td>
          <td class="num red">-{money(i['amount'])}</td>
          <td class="note">{esc_h(i.get('calc_note',''))}</td>
        </tr>""" for i in deduct_items)

    punch_table = ''
    if is_hourly and d.get('punch_details'):
        punch_rows = ''.join(f"""
            <tr>
              <td>{p['date']}</td>
              <td>{p['clock_in']}</td>
              <td>{p['clock_out']}</td>
              <td>{p.get('break_mins',0)} min</td>
              <td class="num">{p['net_hours']} h</td>
            </tr>""" for p in d['punch_details'])
        punch_table = f"""
        <h3>每日工時明細</h3>
        <table>
          <thead><tr><th>日期</th><th>上班</th><th>下班</th><th>休息</th><th>工時</th></tr></thead>
          <tbody>{punch_rows}</tbody>
          <tfoot><tr><td colspan="4"><strong>合計</strong></td><td class="num"><strong>{d.get('actual_work_hours',0)} h</strong></td></tr></tfoot>
        </table>"""

    status_str = '已確認' if row['status'] == 'confirmed' else '草稿（未確認）'
    sal_type   = '時薪制' if is_hourly else '月薪制'
    attend_str = (f"實際工時 {d.get('actual_work_hours',0)}h × 時薪 ${float(row['hourly_rate'] or 0):,.0f}"
                  if is_hourly else
                  f"出勤 {d.get('actual_days',0)} 天 / 工作日 {d.get('work_days',0)} 天")
    if float(d.get('leave_days',0)) > 0:
        attend_str += f"，請假 {d.get('leave_days',0)} 天"
    if float(d.get('unpaid_days',0)) > 0:
        attend_str += f"（無薪 {d.get('unpaid_days',0)} 天）"

    html = f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<title>薪資單 {esc_h(row['staff_name'])} {esc_h(row['month'])}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Noto Sans TC', 'PingFang TC', 'Microsoft JhengHei', sans-serif;
          font-size: 13px; color: #1a2340; background: #fff; padding: 32px; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start;
             border-bottom: 3px solid #1a2340; padding-bottom: 16px; margin-bottom: 24px; }}
  .company {{ font-size: 20px; font-weight: 800; color: #1a2340; }}
  .slip-title {{ font-size: 14px; color: #666; margin-top: 4px; }}
  .staff-info {{ font-size: 12px; color: #444; text-align: right; line-height: 1.8; }}
  .summary {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 12px; margin-bottom: 24px; }}
  .sum-card {{ border: 1.5px solid #e2e8f0; border-radius: 8px; padding: 12px 16px; text-align: center; }}
  .sum-label {{ font-size: 10px; color: #888; margin-bottom: 4px; text-transform: uppercase; letter-spacing: .06em; }}
  .sum-val {{ font-size: 22px; font-weight: 800; font-family: 'DM Mono', monospace; }}
  .sum-val.green {{ color: #2e9e6b; }}
  .sum-val.red   {{ color: #d64242; }}
  .sum-val.navy  {{ color: #1a2340; }}
  .attend {{ background: #f8fafc; border-radius: 6px; padding: 8px 14px;
             font-size: 12px; color: #666; margin-bottom: 20px; }}
  h3 {{ font-size: 12px; font-weight: 700; color: #888; letter-spacing: .08em;
        text-transform: uppercase; margin: 20px 0 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #f1f5f9; padding: 8px 12px; text-align: left;
        font-size: 11px; font-weight: 700; color: #666;
        border-bottom: 2px solid #e2e8f0; }}
  td {{ padding: 7px 12px; border-bottom: 1px solid #f0f2f8; }}
  td.num {{ text-align: right; font-family: 'DM Mono', monospace; font-weight: 600; }}
  td.note {{ font-size: 11px; color: #999; }}
  td.green {{ color: #2e9e6b; }}
  td.red   {{ color: #d64242; }}
  tfoot td {{ font-weight: 700; background: #f8fafc; border-top: 2px solid #e2e8f0; }}
  .net-row td {{ font-size: 16px; font-weight: 800; background: #1a2340; color: #fff; }}
  .net-row td.num {{ color: #f0c040; font-size: 20px; }}
  .footer {{ margin-top: 32px; padding-top: 16px; border-top: 1px solid #e2e8f0;
             display: flex; justify-content: space-between; font-size: 11px; color: #999; }}
  .sign-area {{ display: flex; gap: 48px; margin-top: 40px; }}
  .sign-box {{ flex: 1; border-top: 1px solid #ccc; padding-top: 6px; font-size: 11px; color: #666; }}
  @media print {{
    body {{ padding: 16px; }}
    @page {{ margin: 12mm; size: A4; }}
    .no-print {{ display: none !important; }}
  }}
</style>
</head>
<body>

<div class="no-print" style="text-align:right;margin-bottom:20px">
  <button onclick="window.print()"
    style="padding:10px 24px;background:#1a2340;color:#fff;border:none;border-radius:6px;
           font-size:13px;font-weight:700;cursor:pointer">列印 / 儲存 PDF</button>
</div>

<div class="header">
  <div>
    <div class="company">薪資明細單</div>
    <div class="slip-title">{esc_h(row['month'])} · {sal_type}</div>
  </div>
  <div class="staff-info">
    <div><strong>{esc_h(row['staff_name'])}</strong></div>
    <div>{esc_h(row['employee_code'] or '')}　{esc_h(row['department'] or '')}　{esc_h(row['role'] or '')}</div>
    <div>到職日：{esc_h(str(row['hire_date']) if row['hire_date'] else '—')}</div>
    <div>狀態：<strong>{status_str}</strong></div>
  </div>
</div>

<div class="summary">
  <div class="sum-card">
    <div class="sum-label">津貼合計</div>
    <div class="sum-val green">{money(d.get('allowance_total',0))}</div>
  </div>
  <div class="sum-card">
    <div class="sum-label">扣除合計</div>
    <div class="sum-val red">-{money(d.get('deduction_total',0))}</div>
  </div>
  <div class="sum-card" style="border-color:#1a2340">
    <div class="sum-label">實領金額</div>
    <div class="sum-val navy">{money(d.get('net_pay',0))}</div>
  </div>
</div>

<div class="attend">{attend_str}</div>

<h3>津貼項目</h3>
<table>
  <thead><tr><th>項目</th><th style="text-align:right">金額</th><th>計算說明</th></tr></thead>
  <tbody>{allow_rows}</tbody>
  <tfoot>
    <tr><td><strong>津貼合計</strong></td><td class="num green"><strong>{money(d.get('allowance_total',0))}</strong></td><td></td></tr>
  </tfoot>
</table>

<h3>扣除項目</h3>
<table>
  <thead><tr><th>項目</th><th style="text-align:right">金額</th><th>計算說明</th></tr></thead>
  <tbody>{deduct_rows if deduct_rows else '<tr><td colspan="3" style="color:#ccc;text-align:center;padding:12px">無扣除項目</td></tr>'}</tbody>
  <tfoot>
    <tr><td><strong>扣除合計</strong></td><td class="num red"><strong>-{money(d.get('deduction_total',0))}</strong></td><td></td></tr>
  </tfoot>
</table>

<table style="margin-top:12px">
  <tbody>
    <tr class="net-row">
      <td><strong>實領金額</strong></td>
      <td class="num">{money(d.get('net_pay',0))}</td>
      <td style="color:#ccc;font-size:11px">= 津貼 {money(d.get('allowance_total',0))} - 扣除 {money(d.get('deduction_total',0))}</td>
    </tr>
  </tbody>
</table>

{punch_table}

<div class="sign-area">
  <div class="sign-box">員工簽名</div>
  <div class="sign-box">主管確認</div>
  <div class="sign-box">人資確認</div>
</div>

<div class="footer">
  <span>本薪資單由系統自動產生</span>
  <span>列印日期：<script>document.write(new Date().toLocaleDateString('zh-TW'))</script></span>
</div>

</body>
</html>"""

    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}

# ═══════════════════════════════════════════════════════════════════
# Feature: Batch Review (批次審核)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/punch/requests/batch', methods=['POST'])
@login_required
def api_punch_req_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action')
    by     = b.get('reviewed_by', '管理員')
    note   = b.get('review_note', '')
    if not ids or action not in ('approve', 'reject'):
        return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE punch_requests SET status=%s, reviewed_by=%s,
                  review_note=%s, reviewed_at=NOW()
                WHERE id=%s AND status='pending' RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                if action == 'approve':
                    conn.execute("""
                        INSERT INTO punch_records
                          (staff_id, punch_type, punched_at, note, is_manual, manual_by)
                        VALUES (%s,%s,%s,%s,TRUE,%s)
                    """, (row['staff_id'], row['punch_type'], row['requested_at'],
                          f'補打卡申請#{rid}', by))
                _notify_review_result(row['staff_id'], '補打卡申請', action,
                                      note and f'批次審核意見：{note}' or '')
                done += 1
    return jsonify({'ok': True, 'done': done})


@app.route('/api/overtime/requests/batch', methods=['POST'])
@login_required
def api_ot_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action')
    by     = b.get('reviewed_by', '管理員')
    note   = b.get('review_note', '')
    if not ids or action not in ('approve', 'reject'):
        return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE overtime_requests SET status=%s, reviewed_by=%s,
                  review_note=%s, reviewed_at=NOW()
                WHERE id=%s AND status='pending' RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                if action == 'approve':
                    pay, _ = _calc_ot_pay(dict(row), float(row['ot_hours']),
                                          row.get('day_type','weekday'))
                    conn.execute("""
                        UPDATE overtime_requests SET ot_pay=%s WHERE id=%s
                    """, (pay, rid))
                _notify_review_result(row['staff_id'], '加班申請', action, '')
                done += 1
    return jsonify({'ok': True, 'done': done})


@app.route('/api/schedule/requests/batch', methods=['POST'])
@login_required
def api_sched_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action')
    by     = b.get('reviewed_by', '管理員')
    note   = b.get('review_note', '')
    if not ids or action not in ('approve', 'reject'):
        return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE schedule_requests SET status=%s, reviewed_by=%s,
                  review_note=%s, reviewed_at=NOW(), updated_at=NOW()
                WHERE id=%s AND status IN ('pending','modified_pending') RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                _notify_review_result(row['staff_id'], '排休申請', action, '')
                done += 1
    return jsonify({'ok': True, 'done': done})


@app.route('/api/leave/requests/batch', methods=['POST'])
@login_required
def api_leave_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action')
    by     = b.get('reviewed_by', '管理員')
    note   = b.get('review_note', '')
    if not ids or action not in ('approve', 'reject'):
        return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            old = conn.execute("SELECT * FROM leave_requests WHERE id=%s", (rid,)).fetchone()
            if not old or old['status'] != 'pending':
                continue
            row = conn.execute("""
                UPDATE leave_requests SET status=%s, reviewed_by=%s,
                  review_note=%s, reviewed_at=NOW(), updated_at=NOW()
                WHERE id=%s RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                if action == 'approve':
                    _update_leave_balance(conn, old['staff_id'], old['leave_type_id'],
                                          str(old['start_date'])[:4], float(old['total_days']))
                _notify_review_result(old['staff_id'], '請假申請', action, '')
                done += 1
    return jsonify({'ok': True, 'done': done})


# ═══════════════════════════════════════════════════════════════════
# Feature: Attendance Anomaly Detection (出勤異常)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/attendance/anomalies', methods=['GET'])
@login_required
def api_attendance_anomalies():
    """
    偵測出勤異常：
    - 忘記打下班卡（有上班無下班）
    - 只有下班無上班
    - 遲到（上班時間晚於班別開始時間）
    """
    from datetime import date as _da, datetime as _dta, timezone as _tz, timedelta as _td
    TW    = _tz(_td(hours=8))
    today = _dta.now(TW).date()
    # Check last 7 days
    date_from = today - _td(days=7)

    with get_db() as conn:
        # 取得最近7天打卡記錄（按人、按天）
        rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name, ps.role, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   array_agg(pr.punch_type ORDER BY pr.punched_at) as types,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as first_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as last_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date BETWEEN %s AND %s
              AND ps.active = TRUE
            GROUP BY ps.id, ps.name, ps.role, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date DESC, ps.name
        """, (date_from, today)).fetchall()

        # 取得班別指派（用於遲到／早退判斷）
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.date, st.start_time, st.end_time, st.name as shift_name
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE sa.date BETWEEN %s AND %s
        """, (date_from, today)).fetchall()
        shift_map = {(r['staff_id'], str(r['date'])): r for r in shift_rows}

        # 今日應出勤但未出勤（排除請假）
        all_staff = conn.execute(
            "SELECT id, name, role, department FROM punch_staff WHERE active=TRUE"
        ).fetchall()
        today_punched_ids = {r['staff_id'] for r in rows if str(r['work_date']) == str(today)}
        on_leave_today_ids = set()
        leave_today = conn.execute("""
            SELECT DISTINCT staff_id FROM leave_requests
            WHERE status='approved' AND start_date <= %s AND end_date >= %s
        """, (today, today)).fetchall()
        for r in leave_today:
            on_leave_today_ids.add(r['staff_id'])

    anomalies = []

    # 1. 近7天：有上班但無下班卡
    for r in rows:
        types = list(r['types']) if r['types'] else []
        has_in  = 'in'  in types
        has_out = 'out' in types
        ds = str(r['work_date'])

        if has_in and not has_out and ds != str(today):
            # 昨天或更早沒打下班卡（今天的可能還沒下班）
            anomalies.append({
                'type':       'missing_out',
                'label':      '忘記下班打卡',
                'severity':   'warning',
                'staff_id':   r['staff_id'],
                'name':       r['name'],
                'role':       r['role'] or '',
                'department': r['department'] or '',
                'date':       ds,
                'detail':     f"上班 {r['first_in']}，無下班記錄",
            })

        if not has_in and has_out:
            anomalies.append({
                'type':       'missing_in',
                'label':      '忘記上班打卡',
                'severity':   'warning',
                'staff_id':   r['staff_id'],
                'name':       r['name'],
                'role':       r['role'] or '',
                'department': r['department'] or '',
                'date':       ds,
                'detail':     f"下班 {r['last_out']}，無上班記錄",
            })

        # 遲到判斷（有班別指派）
        if has_in and r['first_in']:
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['start_time']:
                try:
                    sh, sm = map(int, str(shift['start_time'])[:5].split(':'))
                    ih, im = map(int, r['first_in'].split(':'))
                    late_mins = (ih * 60 + im) - (sh * 60 + sm)
                    if late_mins > 10:  # 超過10分鐘算遲到
                        anomalies.append({
                            'type':       'late',
                            'label':      '遲到',
                            'severity':   'warning',
                            'staff_id':   r['staff_id'],
                            'name':       r['name'],
                            'role':       r['role'] or '',
                            'department': r['department'] or '',
                            'date':       ds,
                            'detail':     f"應 {shift['start_time'][:5]} 上班，實際 {r['first_in']}（晚 {late_mins} 分鐘）",
                        })
                except Exception:
                    pass

        # 早退判斷（有班別指派）
        if has_out and r['last_out'] and ds != str(today):
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['end_time']:
                try:
                    eh, em = map(int, str(shift['end_time'])[:5].split(':'))
                    oh, om = map(int, r['last_out'].split(':'))
                    early_mins = (eh * 60 + em) - (oh * 60 + om)
                    if early_mins > 15:  # 超過15分鐘算早退
                        anomalies.append({
                            'type':       'early',
                            'label':      '早退',
                            'severity':   'warning',
                            'staff_id':   r['staff_id'],
                            'name':       r['name'],
                            'role':       r['role'] or '',
                            'department': r['department'] or '',
                            'date':       ds,
                            'detail':     f"應 {shift['end_time'][:5]} 下班，實際 {r['last_out']}（早 {early_mins} 分鐘）",
                        })
                except Exception:
                    pass

    # 2. 今日未出勤（不含請假）
    for s in all_staff:
        if s['id'] not in today_punched_ids and s['id'] not in on_leave_today_ids:
            anomalies.append({
                'type':       'absent',
                'label':      '今日未出勤',
                'severity':   'error',
                'staff_id':   s['id'],
                'name':       s['name'],
                'role':       s['role'] or '',
                'department': s['department'] or '',
                'date':       str(today),
                'detail':     '今日尚無打卡記錄且未請假',
            })

    # Sort: error > warning > info, then by date desc
    sev_order = {'error': 0, 'warning': 1, 'info': 2}
    anomalies.sort(key=lambda x: (sev_order.get(x['severity'], 9), x['date']))
    return jsonify({'anomalies': anomalies, 'count': len(anomalies), 'checked_from': str(date_from)})


# ═══════════════════════════════════════════════════════════════════
# Feature: Staff Termination (離職流程)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/punch/staff/<int:sid>/terminate', methods=['POST'])
@login_required
def api_staff_terminate(sid):
    """辦理離職：設定離職日、停用帳號、記錄備註"""
    b = request.get_json(force=True)
    termination_date = b.get('termination_date', '')
    reason           = b.get('reason', '').strip()
    last_month       = b.get('last_salary_month', '')
    note             = b.get('note', '').strip()

    if not termination_date:
        return jsonify({'error': '請填寫離職日期'}), 400

    with get_db() as conn:
        # Ensure column exists
        try:
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_date DATE")
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_reason TEXT DEFAULT ''")
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_note TEXT DEFAULT ''")
        except Exception:
            pass

        row = conn.execute("""
            UPDATE punch_staff SET
              active = FALSE,
              termination_date   = %s,
              termination_reason = %s,
              termination_note   = %s,
              salary_notes = COALESCE(salary_notes,'') || %s
            WHERE id = %s RETURNING *
        """, (termination_date, reason, note,
              f'\n【離職】{termination_date} {reason}',
              sid)).fetchone()
        if not row:
            return ('', 404)

    return jsonify({
        'ok': True,
        'staff_id': sid,
        'name': row['name'],
        'termination_date': termination_date,
        'last_salary_month': last_month,
    })


@app.route('/api/punch/staff/<int:sid>/reinstate', methods=['POST'])
@login_required
def api_staff_reinstate(sid):
    """復職（重新啟用帳號）"""
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_staff SET active=TRUE,
              termination_date=NULL, termination_reason='', termination_note=''
            WHERE id=%s RETURNING *
        """, (sid,)).fetchone()
    return jsonify(punch_staff_row(row)) if row else ('', 404)


@app.route('/api/punch/staff/terminated', methods=['GET'])
@login_required
def api_staff_terminated_list():
    """離職員工清單"""
    with get_db() as conn:
        # Check if column exists
        try:
            rows = conn.execute("""
                SELECT id, name, employee_code, department, role,
                       hire_date, termination_date, termination_reason
                FROM punch_staff
                WHERE active = FALSE
                ORDER BY termination_date DESC NULLS LAST, name
            """).fetchall()
        except Exception:
            rows = conn.execute(
                "SELECT id, name, employee_code, department, role, hire_date FROM punch_staff WHERE active=FALSE"
            ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for f in ('hire_date','termination_date'):
            if d.get(f): d[f] = str(d[f])
        result.append(d)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════
# Feature: Salary Formula Builder support (公式說明 API)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/salary/formula/preview', methods=['POST'])
@require_module('salary')
def api_formula_preview():
    """即時預覽公式計算結果"""
    b             = request.get_json(force=True)
    formula       = b.get('formula', '').strip()
    base_salary   = float(b.get('base_salary', 30000))
    insured_salary= float(b.get('insured_salary', 30000))
    service_years = float(b.get('service_years', 1))
    extra = {
        'actual_days':   float(b.get('actual_days', 22)),
        'work_days':     float(b.get('work_days', 22)),
        'leave_days':    float(b.get('leave_days', 0)),
        'unpaid_days':   float(b.get('unpaid_days', 0)),
        'personal_days': float(b.get('personal_days', 0)),
        'sick_days':     float(b.get('sick_days', 0)),
        'daily_wage':    base_salary / 30 if base_salary > 0 else 0,
    }

    if not formula:
        return jsonify({'result': 0, 'error': None})
    try:
        result = _eval_formula(formula, base_salary, insured_salary, service_years, extra)
        return jsonify({'result': round(result, 2), 'error': None})
    except Exception as e:
        return jsonify({'result': None, 'error': str(e)})

# ═══════════════════════════════════════════════════════════════════
# Finance Module (財務模組)
# ═══════════════════════════════════════════════════════════════════

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

def init_finance_db():
    migrations = [
        """CREATE TABLE IF NOT EXISTS finance_categories (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            type        TEXT NOT NULL DEFAULT 'expense',
            color       TEXT DEFAULT '#4a7bda',
            sort_order  INT DEFAULT 0,
            active      BOOLEAN DEFAULT TRUE,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_records (
            id              SERIAL PRIMARY KEY,
            record_date     DATE NOT NULL,
            category_id     INT REFERENCES finance_categories(id) ON DELETE SET NULL,
            type            TEXT NOT NULL DEFAULT 'expense',
            title           TEXT NOT NULL,
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            tax_amount      NUMERIC(14,2) DEFAULT 0,
            vendor          TEXT DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            document_id     INT,
            created_by      TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_documents (
            id              SERIAL PRIMARY KEY,
            filename        TEXT NOT NULL,
            doc_type        TEXT DEFAULT '',
            ocr_raw         JSONB DEFAULT '{}',
            upload_date     DATE DEFAULT CURRENT_DATE,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_recurring (
            id              SERIAL PRIMARY KEY,
            title           TEXT NOT NULL,
            type            TEXT NOT NULL DEFAULT 'expense',
            category_id     INT REFERENCES finance_categories(id) ON DELETE SET NULL,
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            tax_amount      NUMERIC(14,2) DEFAULT 0,
            vendor          TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            frequency       TEXT NOT NULL DEFAULT 'monthly',
            day_of_month    INT DEFAULT 1,
            start_date      DATE NOT NULL,
            end_date        DATE,
            last_generated  TEXT DEFAULT '',
            active          BOOLEAN DEFAULT TRUE,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS bank_statements (
            id                  SERIAL PRIMARY KEY,
            account_name        TEXT DEFAULT '',
            txn_date            DATE NOT NULL,
            amount              NUMERIC(14,2) NOT NULL,
            txn_type            TEXT DEFAULT 'debit',
            description         TEXT DEFAULT '',
            reconciled          BOOLEAN DEFAULT FALSE,
            matched_record_id   INT REFERENCES finance_records(id) ON DELETE SET NULL,
            import_batch        TEXT DEFAULT '',
            created_at          TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_payables (
            id              SERIAL PRIMARY KEY,
            payable_type    TEXT NOT NULL DEFAULT 'payable',
            title           TEXT NOT NULL,
            party_name      TEXT DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            amount          NUMERIC(14,2) NOT NULL DEFAULT 0,
            due_date        DATE,
            status          TEXT NOT NULL DEFAULT 'open',
            paid_date       DATE,
            linked_record_id INT REFERENCES finance_records(id) ON DELETE SET NULL,
            note            TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS finance_budgets (
            id              SERIAL PRIMARY KEY,
            year            INT NOT NULL,
            month           INT NOT NULL,
            category_id     INT REFERENCES finance_categories(id) ON DELETE CASCADE,
            budget_amount   NUMERIC(14,2) NOT NULL DEFAULT 0,
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(year, month, category_id)
        )""",
        "ALTER TABLE salary_records ADD COLUMN IF NOT EXISTS finance_synced BOOLEAN DEFAULT FALSE",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[finance_init] {str(e)[:80]}")

    # Seed default categories
    defaults_income = [
        ('餐飲內用收入', 'income', '#2e9e6b', 1),
        ('外帶收入',     'income', '#0ea5e9', 2),
        ('外送收入',     'income', '#8b5cf6', 3),
        ('其他收入',     'income', '#c8a96e', 4),
    ]
    defaults_expense = [
        ('食材成本',   'expense', '#d64242', 10),
        ('薪資支出',   'expense', '#e07b2a', 11),
        ('租金',       'expense', '#8892a4', 12),
        ('水電費',     'expense', '#4a7bda', 13),
        ('設備維修',   'expense', '#e05c8a', 14),
        ('消耗品',     'expense', '#6366f1', 15),
        ('廣告行銷',   'expense', '#f59e0b', 16),
        ('其他支出',   'expense', '#64748b', 17),
    ]
    try:
        with get_db() as conn:
            cnt = conn.execute("SELECT COUNT(*) as c FROM finance_categories").fetchone()['c']
            if cnt == 0:
                for name, ftype, color, sort in (defaults_income + defaults_expense):
                    conn.execute(
                        "INSERT INTO finance_categories (name,type,color,sort_order) VALUES (%s,%s,%s,%s)",
                        (name, ftype, color, sort)
                    )
    except Exception as e:
        print(f"[finance_seed] {e}")

init_finance_db()

def _finance_cat_row(r):
    if not r: return None
    d = dict(r)
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

def _finance_rec_row(r):
    if not r: return None
    d = dict(r)
    if d.get('record_date'): d['record_date'] = str(d['record_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    for f in ('amount','tax_amount'):
        if d.get(f) is not None: d[f] = float(d[f])
    return d

# ── Finance Categories ─────────────────────────────────────────

@app.route('/api/finance/categories', methods=['GET'])
@require_module('finance')
def api_finance_categories_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM finance_categories ORDER BY sort_order, id").fetchall()
    return jsonify([_finance_cat_row(r) for r in rows])

@app.route('/api/finance/categories', methods=['POST'])
@require_module('finance')
def api_finance_category_create():
    b = request.get_json(force=True)
    if not b.get('name','').strip(): return jsonify({'error': '名稱為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_categories (name,type,color,sort_order,active,statement_section)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'].strip(), b.get('type','expense'), b.get('color','#4a7bda'),
              int(b.get('sort_order',0)), bool(b.get('active',True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type')=='income' else 'operating_expense')
             )).fetchone()
    return jsonify(_finance_cat_row(row)), 201

@app.route('/api/finance/categories/<int:cid>', methods=['PUT'])
@require_module('finance')
def api_finance_category_update(cid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_categories SET name=%s,type=%s,color=%s,sort_order=%s,active=%s,statement_section=%s
            WHERE id=%s RETURNING *
        """, (b.get('name','').strip(), b.get('type','expense'), b.get('color','#4a7bda'),
              int(b.get('sort_order',0)), bool(b.get('active',True)),
              b.get('statement_section') or ('operating_revenue' if b.get('type')=='income' else 'operating_expense'),
              cid)).fetchone()
    return jsonify(_finance_cat_row(row)) if row else ('', 404)

@app.route('/api/finance/categories/<int:cid>', methods=['DELETE'])
@require_module('finance')
def api_finance_category_delete(cid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_categories WHERE id=%s", (cid,))
    return jsonify({'deleted': cid})

# ── Finance Records ────────────────────────────────────────────

@app.route('/api/finance/records', methods=['GET'])
@require_module('finance')
def api_finance_records_list():
    month  = request.args.get('month', '')
    ftype  = request.args.get('type', '')
    cat_id = request.args.get('category_id', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("to_char(fr.record_date,'YYYY-MM')=%s"); params.append(month)
    if ftype:
        conds.append("fr.type=%s"); params.append(ftype)
    if cat_id:
        conds.append("fr.category_id=%s"); params.append(int(cat_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.*, fc.name as category_name, fc.color as category_color,
                   fd.filename as doc_filename, fd.ocr_raw as ocr_raw
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            LEFT JOIN finance_documents fd ON fd.id=fr.document_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date DESC, fr.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _finance_rec_row(r)
        d['category_name']  = r['category_name']
        d['category_color'] = r['category_color']
        d['doc_filename']   = r['doc_filename']
        d['ocr_raw']        = r['ocr_raw'] if r['ocr_raw'] else None
        result.append(d)
    return jsonify(result)


@app.route('/api/finance/documents', methods=['GET'])
@require_module('finance')
def api_finance_documents_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fd.*,
                   COUNT(fr.id) as linked_count,
                   MAX(fr.title) as linked_title,
                   MAX(fr.id) as linked_record_id
            FROM finance_documents fd
            LEFT JOIN finance_records fr ON fr.document_id = fd.id
            GROUP BY fd.id
            ORDER BY fd.created_at DESC
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('upload_date'): d['upload_date'] = str(d['upload_date'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        d['linked_count'] = int(d['linked_count'] or 0)
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/records', methods=['POST'])
@require_module('finance')
def api_finance_record_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('record_date'):      return jsonify({'error': '日期為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_records
              (record_date, category_id, type, title, amount, tax_amount, vendor, invoice_no, note, document_id, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type','expense'),
              b['title'].strip(), float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('invoice_no','').strip(),
              b.get('note','').strip(), b.get('document_id') or None,
              session.get('admin_display_name',''))).fetchone()
    return jsonify(_finance_rec_row(row)), 201

@app.route('/api/finance/records/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_finance_record_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_records SET
              record_date=%s, category_id=%s, type=%s, title=%s, amount=%s,
              tax_amount=%s, vendor=%s, invoice_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b['record_date'], b.get('category_id') or None, b.get('type','expense'),
              b.get('title','').strip(), float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('invoice_no','').strip(),
              b.get('note','').strip(), rid)).fetchone()
    return jsonify(_finance_rec_row(row)) if row else ('', 404)

@app.route('/api/finance/records/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_finance_record_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_records WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

# ── Finance P&L Summary ────────────────────────────────────────

@app.route('/api/finance/summary/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_summary(year, month):
    period = f"{year}-{month.zfill(2)}"
    with get_db() as conn:
        totals = conn.execute("""
            SELECT type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE to_char(record_date,'YYYY-MM')=%s
            GROUP BY type
        """, (period,)).fetchall()
        income  = next((float(r['total']) for r in totals if r['type']=='income'), 0.0)
        expense = next((float(r['total']) for r in totals if r['type']=='expense'), 0.0)

        by_cat = conn.execute("""
            SELECT fc.name, fc.color, fr.type, COALESCE(SUM(fr.amount),0) as total
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE to_char(fr.record_date,'YYYY-MM')=%s
            GROUP BY fc.name, fc.color, fr.type
            ORDER BY total DESC
        """, (period,)).fetchall()

        # Last 6 months trend
        trend = conn.execute("""
            SELECT to_char(record_date,'YYYY-MM') as mon,
                   type, COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE record_date >= (DATE_TRUNC('month', %s::date) - INTERVAL '5 months')
              AND record_date <  (DATE_TRUNC('month', %s::date) + INTERVAL '1 month')
            GROUP BY to_char(record_date,'YYYY-MM'), type
            ORDER BY mon
        """, (f"{period}-01", f"{period}-01")).fetchall()

    return jsonify({
        'income':  income,
        'expense': expense,
        'net':     income - expense,
        'by_category': [
            {'name': r['name'] or '未分類', 'color': r['color'] or '#8892a4',
             'type': r['type'], 'total': float(r['total'])}
            for r in by_cat
        ],
        'trend': [
            {'month': r['mon'], 'type': r['type'], 'total': float(r['total'])}
            for r in trend
        ],
    })

# ── Finance OCR ────────────────────────────────────────────────

@app.route('/api/finance/ocr', methods=['POST'])
@require_module('finance')
def api_finance_ocr():
    import anthropic as _ant
    import base64, re as _re

    if not ANTHROPIC_API_KEY:
        return jsonify({'error': '尚未設定 ANTHROPIC_API_KEY 環境變數'}), 500

    file = request.files.get('file')
    if not file:
        return jsonify({'error': '請上傳圖片或 PDF 檔案'}), 400

    raw = file.read()
    media_type = file.content_type or 'image/jpeg'
    # Only image types supported by Claude vision
    if media_type not in ('image/jpeg','image/png','image/gif','image/webp'):
        media_type = 'image/jpeg'

    img_b64 = base64.standard_b64encode(raw).decode()

    client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1024,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': img_b64}},
                    {'type': 'text', 'text': (
                        '請辨識此文件，以JSON格式回傳以下欄位（找不到的欄位填null）：\n'
                        '{"date":"YYYY-MM-DD","vendor":"廠商名稱","invoice_no":"發票或單據號碼",'
                        '"total_amount":含稅總金額數字,"tax_amount":稅額數字,"pre_tax_amount":未稅金額數字,'
                        '"doc_type":"invoice或receipt或expense之一",'
                        '"title":"建議記帳標題（簡短）",'
                        '"items":[{"name":"品項","qty":數量,"unit_price":單價,"amount":小計}],'
                        '"currency":"TWD"}\n只回傳JSON，不要其他文字或markdown。'
                    )}
                ]
            }]
        )
        text = msg.content[0].text.strip()
        text = _re.sub(r'^```json\s*', '', text, flags=_re.MULTILINE)
        text = _re.sub(r'\s*```$', '', text, flags=_re.MULTILINE)
        result = _json.loads(text)
    except _json.JSONDecodeError:
        result = {'raw_text': text, 'error': 'OCR 回傳格式無法解析'}
    except Exception as e:
        return jsonify({'error': f'OCR 失敗：{str(e)}'}), 500

    try:
        with get_db() as conn:
            doc = conn.execute("""
                INSERT INTO finance_documents (filename, doc_type, ocr_raw, upload_date)
                VALUES (%s,%s,%s,CURRENT_DATE) RETURNING id
            """, (file.filename, result.get('doc_type',''), _json.dumps(result))).fetchone()
        result['document_id'] = doc['id']
    except Exception as e:
        print(f"[finance_ocr doc save] {e}")

    return jsonify(result)

# ── Finance Export ─────────────────────────────────────────────

@app.route('/api/finance/export', methods=['GET'])
@require_module('finance')
def api_finance_export():
    """匯出財務記錄 Excel"""
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("to_char(fr.record_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT fr.record_date, fr.type, fr.title, fr.amount, fr.tax_amount,
                   fr.vendor, fr.invoice_no, fr.note, fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE {' AND '.join(conds)}
            ORDER BY fr.record_date, fr.id
        """, params).fetchall()

    wb, ws = _xl_workbook('財務記錄')
    headers = ['日期','類型','類別','標題','金額','稅額','廠商','單據號碼','備註']
    widths  = [12, 7, 14, 24, 12, 10, 16, 14, 24]
    _xl_write_header(ws, headers, widths)

    total_income = 0.0; total_expense = 0.0
    data = []
    for r in rows:
        amt = float(r['amount'] or 0); tax = float(r['tax_amount'] or 0)
        t = '收入' if r['type'] == 'income' else '支出'
        if r['type'] == 'income': total_income += amt
        else: total_expense += amt
        data.append([str(r['record_date']), t, r['category_name'] or '',
                     r['title'], amt, tax if tax else '',
                     r['vendor'] or '', r['invoice_no'] or '', r['note'] or ''])
    _xl_write_rows(ws, data, len(headers), number_cols={5, 6})

    # 合計列
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value='合計').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value='收入').font = Font(bold=True, color='2E7D32')
    ws.cell(row=summary_row, column=5, value=total_income).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=5).font = Font(bold=True, color='2E7D32')
    ws.cell(row=summary_row+1, column=2, value='支出').font = Font(bold=True, color='C62828')
    ws.cell(row=summary_row+1, column=5, value=total_expense).number_format = '#,##0.00'
    ws.cell(row=summary_row+1, column=5).font = Font(bold=True, color='C62828')
    net = total_income - total_expense
    ws.cell(row=summary_row+2, column=2, value='淨額').font = Font(bold=True)
    ws.cell(row=summary_row+2, column=5, value=net).number_format = '#,##0.00'
    ws.cell(row=summary_row+2, column=5).font = Font(bold=True, color='0D47A1' if net >= 0 else 'C62828')

    return _xl_response(wb, f'finance_{month or "all"}.xlsx')

# ── Finance Settings & Financial Statements ────────────────────

def init_finance_settings_db():
    migrations = [
        "ALTER TABLE finance_categories ADD COLUMN IF NOT EXISTS statement_section TEXT",
        """CREATE TABLE IF NOT EXISTS finance_settings (
            id            SERIAL PRIMARY KEY,
            setting_key   TEXT UNIQUE NOT NULL,
            setting_value TEXT DEFAULT ''
        )""",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[finance_settings_init] {str(e)[:80]}")

    # Set default statement_section based on type for existing rows with NULL
    section_defaults = {
        '餐飲內用收入': 'operating_revenue',
        '外帶收入':     'operating_revenue',
        '外送收入':     'operating_revenue',
        '其他收入':     'other_revenue',
        '食材成本':     'cogs',
        '薪資支出':     'operating_expense',
        '租金':         'operating_expense',
        '水電費':       'operating_expense',
        '設備維修':     'operating_expense',
        '消耗品':       'operating_expense',
        '廣告行銷':     'operating_expense',
        '其他支出':     'other_expense',
    }
    try:
        with get_db() as conn:
            # Fill named defaults
            for name, sec in section_defaults.items():
                conn.execute(
                    "UPDATE finance_categories SET statement_section=%s WHERE name=%s AND statement_section IS NULL",
                    (sec, name)
                )
            # Remaining NULLs: income → operating_revenue, expense → operating_expense
            conn.execute("""
                UPDATE finance_categories
                SET statement_section = CASE WHEN type='income' THEN 'operating_revenue' ELSE 'operating_expense' END
                WHERE statement_section IS NULL
            """)
    except Exception as e:
        print(f"[finance_settings_seed] {e}")

    # Seed settings defaults
    for k, v in [('company_name', ''), ('opening_cash', '0'), ('opening_equity', '0'),
                  ('company_tax_id', ''), ('company_address', '')]:
        try:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO finance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT (setting_key) DO NOTHING",
                    (k, v)
                )
        except Exception as e:
            print(f"[finance_settings_default] {e}")

init_finance_settings_db()


def init_insurance_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS insurance_settings (
                    setting_key   TEXT PRIMARY KEY,
                    setting_value TEXT DEFAULT ''
                )
            """)
        for k, v in [('labor_insurance_no', ''), ('health_insurance_no', ''),
                     ('employer_name', ''), ('employer_id', '')]:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO insurance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (k, v))
    except Exception as e:
        print(f"[insurance_init] {e}")

init_insurance_db()


# ═══════════════════════════════════════════════════════════════════════════════
# 教育訓練追蹤 (Training & Certificate Tracking)
# ═══════════════════════════════════════════════════════════════════════════════

def init_training_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS training_records (
                    id              SERIAL PRIMARY KEY,
                    staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    course_name     TEXT NOT NULL,
                    category        TEXT NOT NULL DEFAULT 'general',
                    completed_date  DATE,
                    expiry_date     DATE,
                    certificate_no  TEXT DEFAULT '',
                    note            TEXT DEFAULT '',
                    created_at      TIMESTAMPTZ DEFAULT NOW(),
                    updated_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f"[training_init] {e}")

init_training_db()

TRAINING_CATEGORIES = {
    'food_safety':  '食品安全',
    'fire_safety':  '消防安全',
    'first_aid':    '急救訓練',
    'hygiene':      '衛生管理',
    'service':      '服務禮儀',
    'equipment':    '設備操作',
    'general':      '一般訓練',
    'other':        '其他',
}

@app.route('/api/training/records', methods=['GET'])
@login_required
def api_training_list():
    staff_id  = request.args.get('staff_id')
    category  = request.args.get('category', '')
    expiring  = request.args.get('expiring')   # days, e.g. 60
    expired   = request.args.get('expired')    # '1' = show only expired

    sql = """
        SELECT tr.*, ps.name AS staff_name, ps.department
        FROM training_records tr
        JOIN punch_staff ps ON tr.staff_id = ps.id
        WHERE 1=1
    """
    params = []
    if staff_id:
        sql += " AND tr.staff_id = %s"; params.append(int(staff_id))
    if category:
        sql += " AND tr.category = %s"; params.append(category)
    if expiring:
        days = int(expiring)
        sql += " AND tr.expiry_date IS NOT NULL AND tr.expiry_date <= CURRENT_DATE + INTERVAL '%s days' AND tr.expiry_date >= CURRENT_DATE"
        params.append(days)
    if expired == '1':
        sql += " AND tr.expiry_date IS NOT NULL AND tr.expiry_date < CURRENT_DATE"
    sql += " ORDER BY tr.expiry_date ASC NULLS LAST, tr.completed_date DESC"

    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in ('completed_date', 'expiry_date', 'created_at', 'updated_at'):
            if d.get(k): d[k] = str(d[k])
        today = date.today()
        if d.get('expiry_date'):
            ed = _dt.strptime(d['expiry_date'], '%Y-%m-%d').date()
            days_left = (ed - today).days
            d['days_left'] = days_left
            d['status'] = 'expired' if days_left < 0 else 'expiring_soon' if days_left <= 60 else 'valid'
        else:
            d['days_left'] = None
            d['status'] = 'no_expiry'
        result.append(d)
    return jsonify(result)

@app.route('/api/training/records', methods=['POST'])
@login_required
def api_training_create():
    b = request.get_json(force=True) or {}
    staff_id       = b.get('staff_id')
    course_name    = (b.get('course_name') or '').strip()
    category       = b.get('category', 'general')
    completed_date = b.get('completed_date') or None
    expiry_date    = b.get('expiry_date') or None
    certificate_no = (b.get('certificate_no') or '').strip()
    note           = (b.get('note') or '').strip()
    if not staff_id or not course_name:
        return jsonify({'error': '缺少必填欄位'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO training_records
              (staff_id, course_name, category, completed_date, expiry_date, certificate_no, note)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (staff_id, course_name, category, completed_date, expiry_date, certificate_no, note)).fetchone()
    return jsonify({'ok': True, 'id': row['id']})

@app.route('/api/training/records/<int:rid>', methods=['PUT'])
@login_required
def api_training_update(rid):
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        conn.execute("""
            UPDATE training_records SET
              course_name=%s, category=%s, completed_date=%s, expiry_date=%s,
              certificate_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s
        """, (
            b.get('course_name'), b.get('category', 'general'),
            b.get('completed_date') or None, b.get('expiry_date') or None,
            b.get('certificate_no', ''), b.get('note', ''), rid
        ))
    return jsonify({'ok': True})

@app.route('/api/training/records/<int:rid>', methods=['DELETE'])
@login_required
def api_training_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM training_records WHERE id=%s", (rid,))
    return jsonify({'ok': True})

@app.route('/api/training/summary', methods=['GET'])
@login_required
def api_training_summary():
    """每位員工的訓練狀況摘要"""
    with get_db() as conn:
        staff_all = conn.execute(
            "SELECT id, name, department FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        records = conn.execute("""
            SELECT staff_id, category, expiry_date,
                   CASE
                     WHEN expiry_date IS NULL THEN 'no_expiry'
                     WHEN expiry_date < CURRENT_DATE THEN 'expired'
                     WHEN expiry_date <= CURRENT_DATE + INTERVAL '60 days' THEN 'expiring_soon'
                     ELSE 'valid'
                   END AS status
            FROM training_records
        """).fetchall()
    from collections import defaultdict
    by_staff = defaultdict(list)
    for r in records:
        by_staff[r['staff_id']].append(dict(r))

    result = []
    for s in staff_all:
        recs = by_staff[s['id']]
        result.append({
            'id': s['id'], 'name': s['name'], 'department': s['department'],
            'total': len(recs),
            'valid': sum(1 for r in recs if r['status'] in ('valid', 'no_expiry')),
            'expiring_soon': sum(1 for r in recs if r['status'] == 'expiring_soon'),
            'expired': sum(1 for r in recs if r['status'] == 'expired'),
        })
    return jsonify(result)

# ── 薪資計算預覽 (Salary Preview without saving) ───────────────────────────────

@app.route('/api/salary/records/preview', methods=['POST'])
@require_module('salary')
def api_salary_preview():
    """預覽薪資計算結果（不儲存）"""
    b     = request.get_json(force=True) or {}
    month = b.get('month', '').strip()
    if not month:
        return jsonify({'error': '請指定月份'}), 400
    result = []
    try:
        with get_db() as conn:
            staff_list = conn.execute(
                "SELECT * FROM punch_staff WHERE active=TRUE ORDER BY name"
            ).fetchall()
            for staff in staff_list:
                data = _auto_generate_salary(conn, dict(staff), month)
                punch_days = conn.execute("""
                    SELECT COUNT(DISTINCT (punched_at AT TIME ZONE 'Asia/Taipei')::date) AS n
                    FROM punch_records WHERE staff_id=%s
                      AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
                """, (staff['id'], month)).fetchone()['n']
                approved_ot = conn.execute("""
                    SELECT COUNT(*) AS n, COALESCE(SUM(ot_hours),0) AS hrs
                    FROM overtime_requests WHERE staff_id=%s
                      AND status='approved'
                      AND to_char(request_date,'YYYY-MM')=%s
                """, (staff['id'], month)).fetchone()
                result.append({
                    'staff_id':        data['staff_id'],
                    'staff_name':      staff['name'],
                    'department':      staff['department'],
                    'salary_type':     staff['salary_type'],
                    'punch_days':      punch_days,
                    'work_days':       float(data['work_days']),
                    'actual_days':     float(data['actual_days']),
                    'leave_days':      float(data['leave_days']),
                    'unpaid_days':     float(data['unpaid_days']),
                    'ot_count':        int(approved_ot['n']),
                    'ot_hours':        float(approved_ot['hrs']),
                    'ot_pay':          float(data['ot_pay']),
                    'base_salary':     float(data['base_salary']),
                    'allowance_total': float(data['allowance_total']),
                    'deduction_total': float(data['deduction_total']),
                    'net_pay':         float(data['net_pay']),
                })
    except Exception as e:
        import traceback
        print(f"[salary_preview] ERROR: {e}\n{traceback.format_exc()}")
        return jsonify({'ok': False, 'error': f'計算失敗：{e}'}), 500
    return jsonify({'ok': True, 'month': month, 'records': result})


@app.route('/api/finance/settings', methods=['GET'])
@require_module('finance')
def api_finance_settings_get():
    with get_db() as conn:
        rows = conn.execute("SELECT setting_key, setting_value FROM finance_settings").fetchall()
    return jsonify({r['setting_key']: r['setting_value'] for r in rows})


@app.route('/api/finance/settings', methods=['POST'])
@require_module('finance')
def api_finance_settings_save():
    data = request.get_json(force=True)
    with get_db() as conn:
        for k, v in data.items():
            conn.execute(
                "INSERT INTO finance_settings (setting_key, setting_value) VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(v))
            )
    return jsonify({'ok': True})


def _get_finance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM finance_settings").fetchall()
            return {r['setting_key']: r['setting_value'] for r in rows}
    except:
        return {}


def _roc_year(year):
    return int(year) - 1911


def _month_last_day(year, month):
    import calendar
    return calendar.monthrange(int(year), int(month))[1]


def _compute_statements(year, month):
    """Compute all three financial statements for the given year/month."""
    from collections import defaultdict
    period = f"{year}-{str(month).zfill(2)}"
    settings = _get_finance_settings()
    opening_cash   = float(settings.get('opening_cash',   0) or 0)
    opening_equity = float(settings.get('opening_equity', 0) or 0)
    company_name   = settings.get('company_name', '') or '公司名稱'

    with get_db() as conn:
        records = conn.execute("""
            SELECT fr.type, fr.amount,
                   fc.name            AS cat_name,
                   fc.statement_section AS section
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id = fr.category_id
            WHERE to_char(fr.record_date,'YYYY-MM') = %s
        """, (period,)).fetchall()

        # Cumulative net before this month (for balance sheet period-start)
        prev = conn.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN type='income'  THEN amount ELSE 0 END), 0) AS cum_income,
                COALESCE(SUM(CASE WHEN type='expense' THEN amount ELSE 0 END), 0) AS cum_expense
            FROM finance_records
            WHERE record_date < DATE_TRUNC('month', %s::date)
        """, (f"{period}-01",)).fetchone()

    cum_net_before = float(prev['cum_income']) - float(prev['cum_expense'])

    by_section = defaultdict(float)
    by_cat     = defaultdict(float)
    for r in records:
        sec = r['section'] or ('operating_revenue' if r['type'] == 'income' else 'operating_expense')
        by_section[sec]               += float(r['amount'])
        by_cat[(sec, r['cat_name'] or '未分類')] += float(r['amount'])

    operating_revenue = by_section['operating_revenue']
    other_revenue     = by_section['other_revenue']
    cogs              = by_section['cogs']
    operating_expense = by_section['operating_expense']
    other_expense     = by_section['other_expense']

    gross_profit      = operating_revenue - cogs
    operating_income  = gross_profit - operating_expense
    net_income        = operating_income + other_revenue - other_expense
    total_income      = operating_revenue + other_revenue
    total_expense     = cogs + operating_expense + other_expense

    cum_net_total     = cum_net_before + net_income
    cash_balance      = opening_cash + opening_equity + cum_net_total
    total_equity      = opening_equity + cum_net_total

    def cat_lines(section):
        return [{'name': k[1], 'amount': round(v, 2)}
                for k, v in sorted(by_cat.items()) if k[0] == section]

    return {
        'company_name': company_name,
        'year': int(year), 'month': int(month),
        'roc_year': _roc_year(year),
        'last_day': _month_last_day(year, month),
        'income_statement': {
            'operating_revenue':       round(operating_revenue, 2),
            'operating_revenue_lines': cat_lines('operating_revenue'),
            'other_revenue':           round(other_revenue, 2),
            'other_revenue_lines':     cat_lines('other_revenue'),
            'cogs':                    round(cogs, 2),
            'cogs_lines':              cat_lines('cogs'),
            'gross_profit':            round(gross_profit, 2),
            'operating_expense':       round(operating_expense, 2),
            'operating_expense_lines': cat_lines('operating_expense'),
            'operating_income':        round(operating_income, 2),
            'other_expense':           round(other_expense, 2),
            'other_expense_lines':     cat_lines('other_expense'),
            'net_income':              round(net_income, 2),
        },
        'balance_sheet': {
            'cash':                    round(cash_balance, 2),
            'total_assets':            round(cash_balance, 2),
            'total_liabilities':       0,
            'opening_equity':          round(opening_equity, 2),
            'retained_earnings':       round(cum_net_total, 2),
            'total_equity':            round(total_equity, 2),
            'total_liabilities_equity': round(total_equity, 2),
        },
        'cash_flow': {
            'operating_inflow':        round(total_income, 2),
            'operating_inflow_lines':  cat_lines('operating_revenue') + cat_lines('other_revenue'),
            'operating_outflow':       round(total_expense, 2),
            'operating_outflow_lines': cat_lines('cogs') + cat_lines('operating_expense') + cat_lines('other_expense'),
            'operating_net':           round(total_income - total_expense, 2),
            'investing_net':           0,
            'financing_net':           0,
            'net_change':              round(total_income - total_expense, 2),
            'opening_cash':            round(opening_cash + opening_equity + cum_net_before, 2),
            'closing_cash':            round(cash_balance, 2),
        },
    }


@app.route('/api/finance/statements/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_statements(year, month):
    try:
        return jsonify(_compute_statements(year, month))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/finance/export/statements/<year>/<month>', methods=['GET'])
@require_module('finance')
def api_finance_export_statements(year, month):
    import openpyxl, io as _io
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    d  = _compute_statements(year, month)
    co = d['company_name']
    ry = d['roc_year']
    m  = d['month']
    ld = d['last_day']
    IS = d['income_statement']
    BS = d['balance_sheet']
    CF = d['cash_flow']

    wb = openpyxl.Workbook()

    NAVY   = '1C3557'
    AMT    = '#,##0'
    FONT   = '標楷體'
    thin   = Side(style='thin')
    medium = Side(style='medium')

    def _border(top=False, bottom=False, dbl=False):
        return Border(
            top    = thin   if top else None,
            bottom = medium if dbl  else (thin if bottom else None),
        )

    def setup_ws(ws, title, date_str):
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 18
        for row_vals, styles in [
            ([co, '', ''],        {'font': Font(FONT, bold=True, size=14), 'align': 'center', 'merge': True}),
            ([title, '', ''],     {'font': Font(FONT, bold=True, size=13), 'align': 'center', 'merge': True}),
            ([date_str, '', ''],  {'font': Font(FONT, size=11),            'align': 'center', 'merge': True}),
        ]:
            ws.append(row_vals)
            r = ws.max_row
            ws.cell(r, 1).font      = styles['font']
            ws.cell(r, 1).alignment = Alignment(horizontal=styles['align'])
            if styles.get('merge'):
                ws.merge_cells(f'A{r}:C{r}')
        ws.append([])  # blank

        # column headers
        ws.append(['項　　目', '', '金　額（元）'])
        r = ws.max_row
        ws.cell(r, 1).font      = Font(FONT, bold=True, size=11, color='FFFFFF')
        ws.cell(r, 3).font      = Font(FONT, bold=True, size=11, color='FFFFFF')
        ws.cell(r, 1).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 3).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 2).fill      = PatternFill('solid', fgColor=NAVY)
        ws.cell(r, 1).alignment = Alignment(horizontal='center')
        ws.cell(r, 3).alignment = Alignment(horizontal='right')

    def row(ws, label, amount=None, indent=0, bold=False, subtotal=False, total=False, dbl=False):
        prefix = '　' * indent
        ws.append([prefix + label, '', amount])
        r = ws.max_row
        b = bold or subtotal or total
        ws.cell(r, 1).font      = Font(FONT, bold=b, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal='left')
        if amount is not None:
            ws.cell(r, 3).font           = Font(FONT, bold=b, size=11)
            ws.cell(r, 3).number_format  = AMT
            ws.cell(r, 3).alignment      = Alignment(horizontal='right')
        if subtotal or total:
            ws.cell(r, 3).border = _border(top=(subtotal or total), bottom=(subtotal or total), dbl=dbl)
        elif amount is None:
            ws.cell(r, 1).font = Font(FONT, bold=True, size=11)

    # ─── 損益表 ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '損益表'
    setup_ws(ws1, '損益表', f'中華民國{ry}年{m}月份')

    row(ws1, '一、營業收入', bold=True)
    for l in IS['operating_revenue_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業收入合計', IS['operating_revenue'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '二、營業成本', bold=True)
    for l in IS['cogs_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業成本合計', IS['cogs'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '毛　利', IS['gross_profit'], indent=1, total=True)
    ws1.append([])

    row(ws1, '三、營業費用', bold=True)
    for l in IS['operating_expense_lines']:
        row(ws1, l['name'], l['amount'], indent=2)
    row(ws1, '營業費用合計', IS['operating_expense'], indent=1, subtotal=True)
    ws1.append([])

    row(ws1, '營業利益（損失）', IS['operating_income'], indent=1, total=True)
    ws1.append([])

    if IS['other_revenue'] or IS['other_revenue_lines']:
        row(ws1, '四、營業外收入', bold=True)
        for l in IS['other_revenue_lines']:
            row(ws1, l['name'], l['amount'], indent=2)
        row(ws1, '營業外收入合計', IS['other_revenue'], indent=1, subtotal=True)
        ws1.append([])

    if IS['other_expense'] or IS['other_expense_lines']:
        row(ws1, '五、營業外費用', bold=True)
        for l in IS['other_expense_lines']:
            row(ws1, l['name'], l['amount'], indent=2)
        row(ws1, '營業外費用合計', IS['other_expense'], indent=1, subtotal=True)
        ws1.append([])

    row(ws1, '本期淨利（損）', IS['net_income'], bold=True, total=True, dbl=True)

    # ─── 資產負債表 ───────────────────────────────────────────────
    ws2 = wb.create_sheet('資產負債表')
    setup_ws(ws2, '資產負債表', f'中華民國{ry}年{m}月{ld}日')

    row(ws2, '【資　產】', bold=True)
    row(ws2, '流動資產', indent=1, bold=True)
    row(ws2, '現金及約當現金', BS['cash'], indent=2)
    row(ws2, '資產合計', BS['total_assets'], indent=1, total=True)
    ws2.append([])

    row(ws2, '【負　債】', bold=True)
    row(ws2, '流動負債', indent=1, bold=True)
    row(ws2, '應付帳款', 0, indent=2)
    row(ws2, '負債合計', BS['total_liabilities'], indent=1, total=True)
    ws2.append([])

    row(ws2, '【股東權益】', bold=True)
    row(ws2, '資本額', BS['opening_equity'], indent=2)
    row(ws2, '保留盈餘', BS['retained_earnings'], indent=2)
    row(ws2, '股東權益合計', BS['total_equity'], indent=1, total=True)
    ws2.append([])

    row(ws2, '負債及股東權益合計', BS['total_liabilities_equity'], bold=True, total=True, dbl=True)

    # ─── 現金流量表 ───────────────────────────────────────────────
    ws3 = wb.create_sheet('現金流量表')
    setup_ws(ws3, '現金流量表（直接法）', f'中華民國{ry}年{m}月份')

    row(ws3, '一、營業活動之現金流量', bold=True)
    row(ws3, '（一）收現收入', indent=1, bold=True)
    for l in CF['operating_inflow_lines']:
        row(ws3, l['name'], l['amount'], indent=3)
    row(ws3, '收現合計', CF['operating_inflow'], indent=2, subtotal=True)
    ws3.append([])
    row(ws3, '（二）付現費用', indent=1, bold=True)
    for l in CF['operating_outflow_lines']:
        row(ws3, l['name'], -l['amount'], indent=3)
    row(ws3, '付現合計', -CF['operating_outflow'], indent=2, subtotal=True)
    ws3.append([])
    row(ws3, '營業活動淨現金流量', CF['operating_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '二、投資活動之現金流量', bold=True)
    row(ws3, '投資活動淨現金流量', CF['investing_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '三、籌資活動之現金流量', bold=True)
    row(ws3, '籌資活動淨現金流量', CF['financing_net'], indent=1, total=True)
    ws3.append([])

    row(ws3, '四、本期現金增減', CF['net_change'], bold=True, total=True)
    row(ws3, '五、期初現金及約當現金', CF['opening_cash'], bold=True)
    row(ws3, '六、期末現金及約當現金', CF['closing_cash'], bold=True, total=True, dbl=True)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"statements_{year}{str(month).zfill(2)}.xlsx"
    return (buf.read(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{fname}"',
    })


# ═══════════════════════════════════════════════════════════════════
# Feature 1: 定期自動分錄 (Recurring Entries)
# ═══════════════════════════════════════════════════════════════════

def _recurring_row(r):
    if not r: return None
    d = dict(r)
    for f in ('amount', 'tax_amount'):
        if d.get(f) is not None: d[f] = float(d[f])
    if d.get('start_date'): d['start_date'] = str(d['start_date'])
    if d.get('end_date'):   d['end_date']   = str(d['end_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    return d

@app.route('/api/finance/recurring', methods=['GET'])
@require_module('finance')
def api_recurring_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.*, fc.name as category_name
            FROM finance_recurring fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            ORDER BY fr.active DESC, fr.id
        """).fetchall()
    result = []
    for r in rows:
        d = _recurring_row(r)
        d['category_name'] = r['category_name']
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/recurring', methods=['POST'])
@require_module('finance')
def api_recurring_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    if not b.get('start_date'):       return jsonify({'error': '開始日期為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_recurring
              (title, type, category_id, amount, tax_amount, vendor, note,
               frequency, day_of_month, start_date, end_date, active)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE) RETURNING *
        """, (b['title'].strip(), b.get('type','expense'), b.get('category_id') or None,
              float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('note','').strip(),
              b.get('frequency','monthly'), int(b.get('day_of_month',1) or 1),
              b['start_date'], b.get('end_date') or None)).fetchone()
    return jsonify(_recurring_row(row)), 201

@app.route('/api/finance/recurring/<int:rid>', methods=['PUT'])
@require_module('finance')
def api_recurring_update(rid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_recurring SET
              title=%s, type=%s, category_id=%s, amount=%s, tax_amount=%s,
              vendor=%s, note=%s, frequency=%s, day_of_month=%s,
              start_date=%s, end_date=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b.get('title','').strip(), b.get('type','expense'), b.get('category_id') or None,
              float(b.get('amount',0)), float(b.get('tax_amount',0)),
              b.get('vendor','').strip(), b.get('note','').strip(),
              b.get('frequency','monthly'), int(b.get('day_of_month',1) or 1),
              b.get('start_date'), b.get('end_date') or None,
              bool(b.get('active', True)), rid)).fetchone()
    return jsonify(_recurring_row(row)) if row else ('', 404)

@app.route('/api/finance/recurring/<int:rid>', methods=['DELETE'])
@require_module('finance')
def api_recurring_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_recurring WHERE id=%s", (rid,))
    return jsonify({'deleted': rid})

@app.route('/api/finance/recurring/generate', methods=['POST'])
@require_module('finance')
def api_recurring_generate():
    """為指定月份產生定期分錄（冪等：已產生則跳過）"""
    from datetime import date as _d, timedelta as _td
    import calendar as _cal
    b = request.get_json(force=True)
    month = b.get('month', '')  # YYYY-MM
    if not month:
        from datetime import datetime, timezone, timedelta
        month = datetime.now(timezone(timedelta(hours=8))).strftime('%Y-%m')
    y, m = int(month[:4]), int(month[5:])
    days_in_month = _cal.monthrange(y, m)[1]

    created, skipped = 0, 0
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM finance_recurring
            WHERE active=TRUE
              AND start_date <= %s
              AND (end_date IS NULL OR end_date >= %s)
        """, (f"{month}-28", f"{month}-01")).fetchall()

        for r in rows:
            # Check already generated this month
            if r['last_generated'] == month:
                skipped += 1
                continue
            # Check frequency
            freq = r['frequency']
            start_m = r['start_date'].month
            if freq == 'quarterly' and (m - start_m) % 3 != 0:
                skipped += 1
                continue
            if freq == 'yearly' and m != start_m:
                skipped += 1
                continue
            # Determine record date
            day = min(int(r['day_of_month'] or 1), days_in_month)
            rec_date = _d(y, m, day)
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, vendor, note, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'auto-recurring')
            """, (rec_date, r['category_id'], r['type'], r['title'],
                  r['amount'], r['tax_amount'] or 0, r['vendor'] or '', r['note'] or ''))
            conn.execute("UPDATE finance_recurring SET last_generated=%s WHERE id=%s",
                         (month, r['id']))
            created += 1

    return jsonify({'created': created, 'skipped': skipped, 'month': month})


# ═══════════════════════════════════════════════════════════════════
# Feature 2: 銀行對帳 (Bank Reconciliation)
# ═══════════════════════════════════════════════════════════════════

def _bank_row(r):
    if not r: return None
    d = dict(r)
    if d.get('txn_date'):   d['txn_date']   = str(d['txn_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('amount') is not None: d['amount'] = float(d['amount'])
    return d

@app.route('/api/finance/bank/import', methods=['POST'])
@require_module('finance')
def api_bank_import():
    """匯入銀行對帳單 CSV"""
    import csv, io as _io
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳 CSV 檔案'}), 400
    raw = file.read().decode('utf-8-sig', errors='replace')
    account_name = request.form.get('account_name', '').strip() or '銀行帳戶'
    import_batch = _dt.now(TW_TZ).strftime('%Y%m%d%H%M%S')

    reader = csv.reader(_io.StringIO(raw))
    rows_data = [r for r in reader if any(c.strip() for c in r)]
    if not rows_data: return jsonify({'error': 'CSV 無資料'}), 400

    # Auto-detect header row (skip rows where first column is not a date-like string)
    def _is_date(s):
        s = s.strip().replace('/', '-').replace('.', '-')
        for fmt in ('%Y-%m-%d','%Y-%m-%d','%m-%d-%Y','%d-%m-%Y'):
            try: _dt.strptime(s, fmt); return True
            except: pass
        # ROC year: 民國 e.g. 113/01/15
        import re
        if re.match(r'^\d{2,3}[/-]\d{1,2}[/-]\d{1,2}$', s):
            parts = re.split(r'[/-]', s)
            if int(parts[0]) < 200: return True
        return False

    def _parse_date(s):
        import re
        s = s.strip()
        parts = re.split(r'[/\-\.]', s)
        if len(parts) == 3:
            if int(parts[0]) < 200:
                # ROC year
                y2 = int(parts[0]) + 1911
                return f"{y2}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if int(parts[0]) > 31:
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if int(parts[2]) > 31:
                return f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
        return None

    def _parse_amount(s):
        import re
        s = re.sub(r'[,$\s]', '', str(s).strip())
        try: return float(s)
        except: return None

    inserted = 0
    with get_db() as conn:
        for row in rows_data:
            if len(row) < 2: continue
            date_str = _parse_date(row[0])
            if not date_str: continue
            desc = row[1].strip() if len(row) > 1 else ''
            # Format: date, desc, debit, credit  OR  date, desc, amount
            if len(row) >= 4:
                debit  = _parse_amount(row[2])
                credit = _parse_amount(row[3])
                if debit and debit > 0:
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,'debit',%s,%s)
                    """, (account_name, date_str, debit, desc, import_batch))
                    inserted += 1
                if credit and credit > 0:
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,'credit',%s,%s)
                    """, (account_name, date_str, credit, desc, import_batch))
                    inserted += 1
            elif len(row) >= 3:
                amt = _parse_amount(row[2])
                if amt is not None and amt != 0:
                    txn_type = 'credit' if amt > 0 else 'debit'
                    conn.execute("""INSERT INTO bank_statements
                        (account_name,txn_date,amount,txn_type,description,import_batch)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (account_name, date_str, abs(amt), txn_type, desc, import_batch))
                    inserted += 1
    return jsonify({'inserted': inserted, 'batch': import_batch})

@app.route('/api/finance/bank/statements', methods=['GET'])
@require_module('finance')
def api_bank_statements():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("TO_CHAR(bs.txn_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT bs.*, fr.title as matched_title, fr.amount as matched_amount,
                   fr.record_date as matched_date
            FROM bank_statements bs
            LEFT JOIN finance_records fr ON fr.id=bs.matched_record_id
            WHERE {' AND '.join(conds)}
            ORDER BY bs.txn_date DESC, bs.id DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _bank_row(r)
        d['matched_title']  = r['matched_title']
        d['matched_amount'] = float(r['matched_amount']) if r['matched_amount'] else None
        d['matched_date']   = str(r['matched_date']) if r['matched_date'] else None
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/bank/statements/<int:sid>', methods=['DELETE'])
@require_module('finance')
def api_bank_statement_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM bank_statements WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

@app.route('/api/finance/bank/match', methods=['POST'])
@require_module('finance')
def api_bank_match():
    b = request.get_json(force=True)
    sid = b.get('statement_id')
    rid = b.get('record_id')  # None to unmatch
    with get_db() as conn:
        if rid:
            conn.execute("UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s WHERE id=%s",
                         (rid, sid))
        else:
            conn.execute("UPDATE bank_statements SET reconciled=FALSE, matched_record_id=NULL WHERE id=%s",
                         (sid,))
    return jsonify({'ok': True})

@app.route('/api/finance/bank/auto-match', methods=['POST'])
@require_module('finance')
def api_bank_auto_match():
    """自動比對：相同金額且日期在 3 天內"""
    b = request.get_json(force=True)
    month = b.get('month', '')
    matched = 0
    with get_db() as conn:
        stmts = conn.execute("""
            SELECT * FROM bank_statements
            WHERE reconciled=FALSE
            """ + ("AND TO_CHAR(txn_date,'YYYY-MM')=%s" if month else ""),
            ([month] if month else [])).fetchall()
        for s in stmts:
            # Find finance record: same amount, date within ±3 days, same type direction
            ftype = 'income' if s['txn_type'] == 'credit' else 'expense'
            rec = conn.execute("""
                SELECT id FROM finance_records
                WHERE type=%s AND amount=%s
                  AND ABS(record_date - %s::date) <= 3
                  AND id NOT IN (
                      SELECT matched_record_id FROM bank_statements
                      WHERE matched_record_id IS NOT NULL
                  )
                ORDER BY ABS(record_date - %s::date), id
                LIMIT 1
            """, (ftype, s['amount'], s['txn_date'], s['txn_date'])).fetchone()
            if rec:
                conn.execute("""UPDATE bank_statements SET reconciled=TRUE, matched_record_id=%s
                                WHERE id=%s""", (rec['id'], s['id']))
                matched += 1
    return jsonify({'matched': matched})

@app.route('/api/finance/bank/summary', methods=['GET'])
@require_module('finance')
def api_bank_summary():
    month = request.args.get('month', '')
    cond = "AND TO_CHAR(txn_date,'YYYY-MM')=%s" if month else ""
    params = [month] if month else []
    with get_db() as conn:
        r = conn.execute(f"""
            SELECT
              COUNT(*) as total,
              SUM(CASE WHEN reconciled THEN 1 ELSE 0 END) as matched,
              SUM(CASE WHEN txn_type='credit' THEN amount ELSE 0 END) as total_credit,
              SUM(CASE WHEN txn_type='debit'  THEN amount ELSE 0 END) as total_debit,
              SUM(CASE WHEN reconciled AND txn_type='credit' THEN amount ELSE 0 END) as matched_credit,
              SUM(CASE WHEN reconciled AND txn_type='debit'  THEN amount ELSE 0 END) as matched_debit
            FROM bank_statements WHERE TRUE {cond}
        """, params).fetchone()
    d = dict(r)
    for k in d:
        if d[k] is not None: d[k] = float(d[k]) if isinstance(d[k], type(r['total_credit'])) else int(d[k])
    return jsonify(d)


# ═══════════════════════════════════════════════════════════════════
# Feature 3: 稅務申報準備 (Tax Filing Prep — Taiwan VAT 401)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/finance/tax/<int:year>/<int:period>', methods=['GET'])
@require_module('finance')
def api_finance_tax(year, period):
    """
    period: 1=Jan-Feb, 2=Mar-Apr, 3=May-Jun, 4=Jul-Aug, 5=Sep-Oct, 6=Nov-Dec
    """
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT fr.type, fr.amount, fr.tax_amount, fr.title,
                   fr.vendor, fr.invoice_no, fr.record_date,
                   fc.name as category_name
            FROM finance_records fr
            LEFT JOIN finance_categories fc ON fc.id=fr.category_id
            WHERE TO_CHAR(fr.record_date,'YYYY-MM') = ANY(%s)
            ORDER BY fr.record_date, fr.type
        """, (months,)).fetchall()

    sales_rows    = [r for r in rows if r['type'] == 'income']
    purchase_rows = [r for r in rows if r['type'] == 'expense']

    sales_amount    = sum(float(r['amount'])     for r in sales_rows)
    sales_tax       = sum(float(r['tax_amount'] or 0) for r in sales_rows)
    purchase_amount = sum(float(r['amount'])     for r in purchase_rows)
    purchase_tax    = sum(float(r['tax_amount'] or 0) for r in purchase_rows)
    tax_payable     = round(sales_tax - purchase_tax, 2)

    def _fmt_row(r):
        return {
            'date':     str(r['record_date']),
            'title':    r['title'],
            'vendor':   r['vendor'] or '',
            'invoice_no': r['invoice_no'] or '',
            'amount':   float(r['amount']),
            'tax_amount': float(r['tax_amount'] or 0),
            'category': r['category_name'] or '未分類',
        }

    return jsonify({
        'year': year, 'period': period,
        'roc_year': year - 1911,
        'months': months,
        'sales': {
            'rows':   [_fmt_row(r) for r in sales_rows],
            'amount': round(sales_amount, 2),
            'tax':    round(sales_tax, 2),
        },
        'purchases': {
            'rows':   [_fmt_row(r) for r in purchase_rows],
            'amount': round(purchase_amount, 2),
            'tax':    round(purchase_tax, 2),
        },
        'tax_payable': tax_payable,
        'is_refund':   tax_payable < 0,
    })


# ═══════════════════════════════════════════════════════════════════
# Feature 4: 應收/應付帳款 (AR/AP Tracking)
# ═══════════════════════════════════════════════════════════════════

def _payable_row(r):
    if not r: return None
    d = dict(r)
    if d.get('due_date'):   d['due_date']   = str(d['due_date'])
    if d.get('paid_date'):  d['paid_date']  = str(d['paid_date'])
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
    if d.get('amount') is not None: d['amount'] = float(d['amount'])
    return d

@app.route('/api/finance/payables', methods=['GET'])
@require_module('finance')
def api_payables_list():
    from datetime import date as _d
    ptype  = request.args.get('type', '')      # receivable / payable
    status = request.args.get('status', '')    # open / paid / overdue
    conds, params = ['TRUE'], []
    if ptype:  conds.append("payable_type=%s"); params.append(ptype)
    if status == 'overdue':
        conds.append("status='open' AND due_date < CURRENT_DATE")
    elif status:
        conds.append("status=%s"); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT *, CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE {' AND '.join(conds)}
            ORDER BY
              CASE WHEN status='open' AND due_date < CURRENT_DATE THEN 0
                   WHEN status='open' THEN 1
                   ELSE 2 END,
              due_date
        """, params).fetchall()
    result = []
    for r in rows:
        d = _payable_row(r)
        d['days_overdue'] = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        # Compute effective status
        if d['status'] == 'open' and d.get('due_date') and str(_d.today()) > d['due_date']:
            d['effective_status'] = 'overdue'
        else:
            d['effective_status'] = d['status']
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/payables', methods=['POST'])
@require_module('finance')
def api_payable_create():
    b = request.get_json(force=True)
    if not b.get('title','').strip(): return jsonify({'error': '標題為必填'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO finance_payables
              (payable_type, title, party_name, invoice_no, amount, due_date, status, note)
            VALUES (%s,%s,%s,%s,%s,%s,'open',%s) RETURNING *
        """, (b.get('payable_type','payable'), b['title'].strip(),
              b.get('party_name','').strip(), b.get('invoice_no','').strip(),
              float(b.get('amount',0)), b.get('due_date') or None,
              b.get('note','').strip())).fetchone()
    return jsonify(_payable_row(row)), 201

@app.route('/api/finance/payables/<int:pid>', methods=['PUT'])
@require_module('finance')
def api_payable_update(pid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE finance_payables SET
              payable_type=%s, title=%s, party_name=%s, invoice_no=%s,
              amount=%s, due_date=%s, status=%s,
              paid_date=%s, note=%s, updated_at=NOW()
            WHERE id=%s RETURNING *
        """, (b.get('payable_type','payable'), b.get('title','').strip(),
              b.get('party_name','').strip(), b.get('invoice_no','').strip(),
              float(b.get('amount',0)), b.get('due_date') or None,
              b.get('status','open'), b.get('paid_date') or None,
              b.get('note','').strip(), pid)).fetchone()
    return jsonify(_payable_row(row)) if row else ('', 404)

@app.route('/api/finance/payables/<int:pid>', methods=['DELETE'])
@require_module('finance')
def api_payable_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM finance_payables WHERE id=%s", (pid,))
    return jsonify({'deleted': pid})

@app.route('/api/finance/payables/aging', methods=['GET'])
@require_module('finance')
def api_payables_aging():
    ptype = request.args.get('type', 'payable')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT *,
              CURRENT_DATE - due_date AS days_overdue
            FROM finance_payables
            WHERE payable_type=%s AND status='open'
        """, (ptype,)).fetchall()
    buckets = {'current': 0, 'd1_30': 0, 'd31_60': 0, 'd61_90': 0, 'd90plus': 0}
    bucket_rows = {'current': [], 'd1_30': [], 'd31_60': [], 'd61_90': [], 'd90plus': []}
    for r in rows:
        do = int(r['days_overdue']) if r['days_overdue'] is not None else 0
        d = _payable_row(r)
        d['days_overdue'] = do
        if do <= 0:    k = 'current'
        elif do <= 30: k = 'd1_30'
        elif do <= 60: k = 'd31_60'
        elif do <= 90: k = 'd61_90'
        else:          k = 'd90plus'
        buckets[k]      += float(r['amount'])
        bucket_rows[k].append(d)
    return jsonify({'buckets': buckets, 'rows': bucket_rows, 'type': ptype})


# ═══════════════════════════════════════════════════════════════════
# Feature 5: 預算管理 (Budget vs Actual)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/finance/budgets', methods=['GET'])
@require_module('finance')
def api_budgets_list():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fb.*, fc.name as category_name, fc.type as category_type, fc.color
            FROM finance_budgets fb
            JOIN finance_categories fc ON fc.id=fb.category_id
            WHERE fb.year=%s AND fb.month=%s
            ORDER BY fc.type, fc.sort_order
        """, (int(year), int(month))).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['budget_amount'] = float(d['budget_amount'])
        if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
        if d.get('updated_at'): d['updated_at'] = d['updated_at'].isoformat()
        result.append(d)
    return jsonify(result)

@app.route('/api/finance/budgets', methods=['POST'])
@require_module('finance')
def api_budgets_save():
    """Upsert list of budgets: [{category_id, budget_amount}]"""
    b = request.get_json(force=True)
    year  = int(b.get('year',  0))
    month = int(b.get('month', 0))
    items = b.get('items', [])
    if not year or not month: return jsonify({'error': '年月為必填'}), 400
    with get_db() as conn:
        for it in items:
            cid = it.get('category_id')
            amt = float(it.get('budget_amount', 0))
            if cid is None: continue
            if amt == 0:
                conn.execute("DELETE FROM finance_budgets WHERE year=%s AND month=%s AND category_id=%s",
                             (year, month, cid))
            else:
                conn.execute("""
                    INSERT INTO finance_budgets (year, month, category_id, budget_amount, updated_at)
                    VALUES (%s,%s,%s,%s,NOW())
                    ON CONFLICT (year, month, category_id)
                    DO UPDATE SET budget_amount=EXCLUDED.budget_amount, updated_at=NOW()
                """, (year, month, cid, amt))
    return jsonify({'ok': True})

@app.route('/api/finance/budgets/vs-actual', methods=['GET'])
@require_module('finance')
def api_budgets_vs_actual():
    year  = request.args.get('year',  '')
    month = request.args.get('month', '')
    if not year or not month:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone(timedelta(hours=8)))
        year, month = str(now.year), str(now.month)
    period = f"{year}-{str(month).zfill(2)}"
    with get_db() as conn:
        cats = conn.execute("""
            SELECT id, name, type, color FROM finance_categories WHERE active=TRUE ORDER BY type, sort_order
        """).fetchall()
        budgets = conn.execute("""
            SELECT category_id, budget_amount FROM finance_budgets WHERE year=%s AND month=%s
        """, (int(year), int(month))).fetchall()
        actuals = conn.execute("""
            SELECT category_id, SUM(amount) as total
            FROM finance_records
            WHERE TO_CHAR(record_date,'YYYY-MM')=%s
            GROUP BY category_id
        """, (period,)).fetchall()
    budget_map = {r['category_id']: float(r['budget_amount']) for r in budgets}
    actual_map = {r['category_id']: float(r['total']) for r in actuals}
    result = []
    for c in cats:
        cid = c['id']
        bgt = budget_map.get(cid, 0)
        act = actual_map.get(cid, 0)
        pct = round(act / bgt * 100, 1) if bgt > 0 else None
        result.append({
            'category_id':   cid,
            'category_name': c['name'],
            'category_type': c['type'],
            'color':         c['color'],
            'budget':        bgt,
            'actual':        act,
            'remaining':     round(bgt - act, 2),
            'pct':           pct,
            'over_budget':   bgt > 0 and act > bgt,
        })
    return jsonify({'year': year, 'month': month, 'items': result})


# ═══════════════════════════════════════════════════════════════════
# Feature 6: 薪資費用連動 (Payroll → Finance)
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/finance/payroll/status', methods=['GET'])
@require_module('finance')
def api_payroll_status():
    """列出各月薪資是否已同步至財務"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month,
                   COUNT(*) as total,
                   SUM(CASE WHEN finance_synced THEN 1 ELSE 0 END) as synced,
                   SUM(net_pay) as total_net_pay
            FROM salary_records
            WHERE status IN ('confirmed','draft')
            GROUP BY month ORDER BY month DESC LIMIT 24
        """).fetchall()
    return jsonify([{
        'month':         r['month'],
        'total':         int(r['total']),
        'synced':        int(r['synced']),
        'total_net_pay': float(r['total_net_pay'] or 0),
        'all_synced':    int(r['synced']) == int(r['total']),
    } for r in rows])

@app.route('/api/finance/payroll/sync', methods=['POST'])
@require_module('finance')
def api_payroll_sync():
    """將指定月份已確認薪資寫入財務記錄"""
    b     = request.get_json(force=True)
    month = b.get('month', '')
    if not month: return jsonify({'error': '請提供月份'}), 400
    # Find or create 薪資支出 category
    with get_db() as conn:
        cat = conn.execute("""
            SELECT id FROM finance_categories WHERE name='薪資支出' AND type='expense' LIMIT 1
        """).fetchone()
        if not cat:
            cat = conn.execute("""
                INSERT INTO finance_categories (name,type,color,sort_order)
                VALUES ('薪資支出','expense','#e07b2a',11) RETURNING *
            """).fetchone()
        cat_id = cat['id']

        records = conn.execute("""
            SELECT sr.*, ps.name as staff_name
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s AND sr.finance_synced=FALSE
        """, (month,)).fetchall()

        if not records:
            return jsonify({'created': 0, 'message': '無需同步的薪資記錄'})

        record_date = f"{month}-{str(28).zfill(2)}"  # Use 28th as payroll date
        created = 0
        for sr in records:
            # Main salary entry
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,%s,'payroll-sync')
            """, (record_date, cat_id,
                  f"{sr['staff_name']} {month} 薪資",
                  float(sr['net_pay']),
                  f"薪資記錄 #{sr['id']}"))
            conn.execute("UPDATE salary_records SET finance_synced=TRUE WHERE id=%s", (sr['id'],))
            created += 1

    return jsonify({'created': created, 'month': month})


# ── Tax → Finance sync ──────────────────────────────────────────

@app.route('/api/finance/tax/<int:year>/<int:period>/sync', methods=['POST'])
@require_module('finance')
def api_finance_tax_sync(year, period):
    """將應繳/退稅金額建立為財務分錄，流入損益表"""
    if period < 1 or period > 6:
        return jsonify({'error': '期別需為 1-6'}), 400
    m_start = (period - 1) * 2 + 1
    m_end   = m_start + 1
    months  = [f"{year}-{str(m).zfill(2)}" for m in range(m_start, m_end + 1)]
    roc_year = year - 1911

    with get_db() as conn:
        rows = conn.execute("""
            SELECT type, SUM(tax_amount) as tax_total
            FROM finance_records
            WHERE TO_CHAR(record_date,'YYYY-MM') = ANY(%s)
              AND tax_amount IS NOT NULL AND tax_amount <> 0
            GROUP BY type
        """, (months,)).fetchall()

    sales_tax    = sum(float(r['tax_total']) for r in rows if r['type'] == 'income')
    purchase_tax = sum(float(r['tax_total']) for r in rows if r['type'] == 'expense')
    tax_payable  = round(sales_tax - purchase_tax, 2)

    if tax_payable == 0:
        return jsonify({'created': 0, 'message': '稅額為零，無需建立分錄'})

    # Record date = last day of period's last month
    import calendar as _cal
    record_date = f"{year}-{str(m_end).zfill(2)}-{_cal.monthrange(year, m_end)[1]}"
    note = f"銷項稅 ${round(sales_tax,0):,.0f} − 進項稅 ${round(purchase_tax,0):,.0f} = {'應繳' if tax_payable>0 else '退稅'} ${abs(round(tax_payable,0)):,.0f}"
    period_label = f"民國{roc_year}年第{period}期（{months[0]}～{months[-1]}）"

    created = 0
    with get_db() as conn:
        if tax_payable > 0:
            # 應繳稅款 → expense under 稅費
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='稅費' AND type='expense' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('稅費','expense','#8892a4', 99,'operating_expense') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'expense',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'],
                  f"應繳營業稅 {period_label}", tax_payable, note))
            created += 1
        else:
            # 退稅 → income under 其他收入
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE name='其他收入' AND type='income' LIMIT 1"
            ).fetchone()
            if not cat:
                cat = conn.execute("""
                    INSERT INTO finance_categories (name, type, color, sort_order, statement_section)
                    VALUES ('其他收入','income','#c8a96e', 99,'other_revenue') RETURNING *
                """).fetchone()
            conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, tax_amount, note, created_by)
                VALUES (%s,%s,'income',%s,%s,0,%s,'tax-sync')
            """, (record_date, cat['id'],
                  f"營業稅退稅 {period_label}", abs(tax_payable), note))
            created += 1

    return jsonify({'created': created, 'tax_payable': tax_payable, 'record_date': record_date})


# ═══════════════════════════════════════════════════════════════════
# LINE Broadcast Helper
# ═══════════════════════════════════════════════════════════════════

def _broadcast_announcement_line(title, content):
    """廣播公告給所有已綁定 LINE 的在職員工"""
    try:
        with get_db() as conn:
            cfg = conn.execute("SELECT * FROM line_punch_config WHERE id=1").fetchone()
            if not cfg or not cfg.get('enabled') or not cfg.get('channel_access_token'):
                return
            staff_rows = conn.execute(
                "SELECT line_user_id FROM punch_staff WHERE active=TRUE AND line_user_id IS NOT NULL"
            ).fetchall()
        if not staff_rows:
            return
        api = LineBotApi(cfg['channel_access_token'])
        snippet = content[:60] + ('…' if len(content) > 60 else '')
        msg = f"[公告] {title}\n{snippet}\n\n請至員工系統查看完整公告。"
        for s in staff_rows:
            try:
                api.push_message(s['line_user_id'], TextSendMessage(text=msg))
            except Exception as e:
                print(f"[LINE broadcast] {s['line_user_id']}: {e}")
    except Exception as e:
        print(f"[LINE broadcast] error: {e}")


# ═══════════════════════════════════════════════════════════════════
# Expense Claims 費用報帳申請
# ═══════════════════════════════════════════════════════════════════

def _init_expense_db():
    sqls = [
        """CREATE TABLE IF NOT EXISTS expense_claims (
            id                SERIAL PRIMARY KEY,
            staff_id          INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            title             TEXT NOT NULL,
            amount            NUMERIC(12,2) NOT NULL DEFAULT 0,
            expense_date      DATE NOT NULL,
            category          TEXT DEFAULT '',
            note              TEXT DEFAULT '',
            status            TEXT NOT NULL DEFAULT 'pending',
            document_id       INT REFERENCES finance_documents(id) ON DELETE SET NULL,
            review_note       TEXT DEFAULT '',
            reviewed_by       TEXT DEFAULT '',
            reviewed_at       TIMESTAMPTZ,
            finance_record_id INT REFERENCES finance_records(id) ON DELETE SET NULL,
            created_at        TIMESTAMPTZ DEFAULT NOW()
        )""",
    ]
    for sql in sqls:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[expense_init] {e}")

_init_expense_db()


def _expense_row(r):
    if not r: return None
    d = dict(r)
    if d.get('expense_date'): d['expense_date'] = str(d['expense_date'])
    if d.get('reviewed_at'): d['reviewed_at'] = d['reviewed_at'].isoformat()
    if d.get('created_at'):  d['created_at']  = d['created_at'].isoformat()
    if d.get('amount') is not None: d['amount'] = float(d['amount'])
    return d


# ── Employee endpoints ──────────────────────────────────────────

@app.route('/api/expense/my-claims', methods=['GET'])
def api_expense_my_list():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': '請先登入'}), 401
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM expense_claims WHERE staff_id=%s ORDER BY created_at DESC LIMIT 50
        """, (sid,)).fetchall()
    return jsonify([_expense_row(r) for r in rows])


@app.route('/api/expense/my-claims', methods=['POST'])
def api_expense_submit():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': '請先登入'}), 401
    b = request.get_json(force=True)
    if not b.get('title','').strip():  return jsonify({'error': '請填寫標題'}), 400
    if not b.get('expense_date'):      return jsonify({'error': '請填寫費用日期'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO expense_claims
              (staff_id, title, amount, expense_date, category, note, document_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (sid, b['title'].strip(), float(b.get('amount', 0)),
              b['expense_date'], b.get('category','').strip(),
              b.get('note','').strip(), b.get('document_id') or None)).fetchone()
    return jsonify(_expense_row(row)), 201


@app.route('/api/expense/ocr', methods=['POST'])
def api_expense_ocr():
    """員工自助 OCR — 複用 finance OCR 邏輯"""
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': '請先登入'}), 401
    import anthropic as _ant, base64, re as _re2
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': '尚未設定 ANTHROPIC_API_KEY'}), 500
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳圖片'}), 400
    raw = file.read()
    media_type = file.content_type or 'image/jpeg'
    if media_type not in ('image/jpeg','image/png','image/gif','image/webp'):
        media_type = 'image/jpeg'
    img_b64 = base64.standard_b64encode(raw).decode()
    client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model='claude-sonnet-4-6', max_tokens=512,
            messages=[{'role':'user','content':[
                {'type':'image','source':{'type':'base64','media_type':media_type,'data':img_b64}},
                {'type':'text','text':'請辨識此收據或發票，以JSON格式回傳：{"date":"YYYY-MM-DD","vendor":"廠商","title":"建議標題","total_amount":數字,"doc_type":"receipt或invoice"}\n只回傳JSON。'}
            ]}]
        )
        text = msg.content[0].text.strip()
        text = _re2.sub(r'^```json\s*','',text,flags=_re2.MULTILINE)
        text = _re2.sub(r'\s*```$','',text,flags=_re2.MULTILINE)
        result = _json.loads(text)
    except Exception as e:
        return jsonify({'error': f'OCR 失敗：{e}'}), 500
    try:
        with get_db() as conn:
            doc = conn.execute("""
                INSERT INTO finance_documents (filename, doc_type, ocr_raw)
                VALUES (%s,%s,%s) RETURNING id
            """, (file.filename, result.get('doc_type',''), _json.dumps(result))).fetchone()
        result['document_id'] = doc['id']
    except Exception as e:
        print(f"[expense_ocr doc] {e}")
    return jsonify(result)


# ── Leave: medical certificate upload ───────────────────────────

@app.route('/api/leave/upload-cert', methods=['POST'])
def api_leave_upload_cert():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': '請先登入'}), 401
    file = request.files.get('file')
    if not file: return jsonify({'error': '請上傳圖片'}), 400
    raw = file.read()
    if len(raw) > 10 * 1024 * 1024:
        return jsonify({'error': '檔案不可超過 10MB'}), 400
    import base64 as _b64c
    image_data = 'data:' + (file.content_type or 'image/jpeg') + ';base64,' + _b64c.b64encode(raw).decode()
    try:
        with get_db() as conn:
            doc = conn.execute("""
                INSERT INTO finance_documents (filename, doc_type, image_data, upload_date)
                VALUES (%s, 'medical_cert', %s, CURRENT_DATE) RETURNING id
            """, (file.filename, image_data)).fetchone()
        return jsonify({'document_id': doc['id'], 'filename': file.filename})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/documents/<int:doc_id>/image', methods=['GET'])
def api_document_image(doc_id):
    """Return a simple HTML page embedding the stored image as a data URL."""
    if not (session.get('logged_in') or session.get('punch_staff_id')):
        return jsonify({'error': 'unauthorized'}), 401
    with get_db() as conn:
        doc = conn.execute("SELECT image_data, filename FROM finance_documents WHERE id=%s", (doc_id,)).fetchone()
    if not doc or not doc['image_data']:
        return jsonify({'error': '找不到圖片'}), 404
    from flask import Response
    fname = (doc['filename'] or '').replace('"', '')
    html = (
        '<!doctype html><html><head><meta charset="utf-8">'
        f'<title>{fname}</title>'
        '<style>body{margin:0;background:#111;display:flex;justify-content:center;align-items:flex-start}'
        'img{max-width:100%;height:auto}</style></head>'
        f'<body><img src="{doc["image_data"]}" alt="{fname}"></body></html>'
    )
    return Response(html, mimetype='text/html')


# ── Admin endpoints ─────────────────────────────────────────────

@app.route('/api/expense/claims', methods=['GET'])
@login_required
def api_expense_admin_list():
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if status: conds.append("ec.status=%s"); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ec.*, ps.name as staff_name, ps.employee_code
            FROM expense_claims ec
            JOIN punch_staff ps ON ps.id=ec.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ec.created_at DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _expense_row(r)
        d['staff_name']    = r['staff_name']
        d['employee_code'] = r['employee_code']
        result.append(d)
    return jsonify(result)


@app.route('/api/expense/claims/<int:cid>', methods=['PUT'])
@login_required
def api_expense_review(cid):
    b      = request.get_json(force=True)
    action = b.get('action')  # approve / reject
    if action not in ('approve','reject'):
        return jsonify({'error': 'invalid action'}), 400
    reviewed_by  = session.get('admin_display_name','管理員')
    review_note  = b.get('review_note','').strip()
    new_status   = 'approved' if action == 'approve' else 'rejected'
    finance_rid  = None

    with get_db() as conn:
        claim = conn.execute("SELECT * FROM expense_claims WHERE id=%s", (cid,)).fetchone()
        if not claim: return ('', 404)

        if action == 'approve' and b.get('create_finance_record', True):
            cat = conn.execute(
                "SELECT id FROM finance_categories WHERE type='expense' AND active=TRUE ORDER BY sort_order LIMIT 1"
            ).fetchone()
            frec = conn.execute("""
                INSERT INTO finance_records
                  (record_date, category_id, type, title, amount, note, document_id, created_by)
                VALUES (%s,%s,'expense',%s,%s,%s,%s,'expense-claim') RETURNING id
            """, (claim['expense_date'], cat['id'] if cat else None,
                  claim['title'], claim['amount'],
                  f"報帳申請 #{cid}：{claim['note'] or ''}",
                  claim['document_id'])).fetchone()
            finance_rid = frec['id']

        row = conn.execute("""
            UPDATE expense_claims SET
              status=%s, reviewed_by=%s, review_note=%s,
              reviewed_at=NOW(), finance_record_id=%s
            WHERE id=%s RETURNING *
        """, (new_status, reviewed_by, review_note, finance_rid, cid)).fetchone()

    if row:
        extra = f"標題：{claim['title']}　金額：${float(claim['amount']):,.0f}"
        if review_note: extra += f"\n意見：{review_note}"
        _notify_review_result(claim['staff_id'], '費用報帳', action, extra)

    return jsonify(_expense_row(row)) if row else ('', 404)


# ═══════════════════════════════════════════════════════════════════════════
# 績效考核模組
# ═══════════════════════════════════════════════════════════════════════════

def _init_performance_db():
    sqls = [
        """CREATE TABLE IF NOT EXISTS performance_templates (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL,
            description TEXT DEFAULT '',
            period      TEXT DEFAULT 'quarterly',
            items       JSONB DEFAULT '[]',
            active      BOOLEAN DEFAULT TRUE,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS performance_reviews (
            id              SERIAL PRIMARY KEY,
            staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
            template_id     INT REFERENCES performance_templates(id) ON DELETE SET NULL,
            period_label    TEXT NOT NULL,
            scores          JSONB DEFAULT '{}',
            total_score     NUMERIC(6,2) DEFAULT 0,
            max_score       NUMERIC(6,2) DEFAULT 100,
            grade           TEXT DEFAULT '',
            comments        TEXT DEFAULT '',
            reviewer        TEXT DEFAULT '',
            salary_adjusted BOOLEAN DEFAULT FALSE,
            salary_delta    NUMERIC(12,2) DEFAULT 0,
            reviewed_at     TIMESTAMPTZ DEFAULT NOW(),
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS performance_config (
            key        TEXT PRIMARY KEY,
            value      JSONB NOT NULL,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )""",
    ]
    for sql in sqls:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[perf_init] {e}")

_init_performance_db()

_DEFAULT_GRADE_CONFIG = [
    {'grade': 'A', 'label': '優秀', 'min_pct': 90},
    {'grade': 'B', 'label': '良好', 'min_pct': 75},
    {'grade': 'C', 'label': '待加強', 'min_pct': 60},
    {'grade': 'D', 'label': '需改善', 'min_pct':  0},
]

def _get_grade_config():
    """從 DB 讀取評級設定，若未設定則回傳預設值（按門檻由高到低排序）。"""
    try:
        with get_db() as conn:
            row = conn.execute(
                "SELECT value FROM performance_config WHERE key='grade_config'"
            ).fetchone()
        if row:
            cfg = row['value']
            if isinstance(cfg, str):
                cfg = _json.loads(cfg)
            if isinstance(cfg, list) and cfg:
                return sorted(cfg, key=lambda x: -float(x.get('min_pct', 0)))
    except Exception:
        pass
    return _DEFAULT_GRADE_CONFIG

def _grade_labels():
    return {c['grade']: c['label'] for c in _get_grade_config()}


def _perf_template_row(r):
    if not r: return None
    d = dict(r)
    if d.get('created_at'): d['created_at'] = d['created_at'].isoformat()
    if isinstance(d.get('items'), str):
        try: d['items'] = _json.loads(d['items'])
        except: d['items'] = []
    return d

def _perf_review_row(r):
    if not r: return None
    d = dict(r)
    for f in ('reviewed_at', 'created_at'):
        if d.get(f): d[f] = d[f].isoformat()
    if isinstance(d.get('scores'), str):
        try: d['scores'] = _json.loads(d['scores'])
        except: d['scores'] = {}
    if d.get('total_score') is not None: d['total_score'] = float(d['total_score'])
    if d.get('max_score')   is not None: d['max_score']   = float(d['max_score'])
    if d.get('salary_delta')is not None: d['salary_delta']= float(d['salary_delta'])
    return d

def _score_to_grade(pct):
    for cfg in _get_grade_config():
        if pct >= cfg['min_pct']:
            return cfg['grade']
    return _get_grade_config()[-1]['grade']

# ── 考核範本 CRUD ───────────────────────────────────────────────

@app.route('/api/performance/templates', methods=['GET'])
@login_required
def api_perf_templates_list():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM performance_templates ORDER BY created_at DESC"
        ).fetchall()
    return jsonify([_perf_template_row(r) for r in rows])

@app.route('/api/performance/templates', methods=['POST'])
@login_required
def api_perf_template_create():
    b = request.get_json(force=True)
    name = (b.get('name') or '').strip()
    if not name: return jsonify({'error': '請填寫範本名稱'}), 400
    items = b.get('items', [])
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO performance_templates (name, description, period, items)
            VALUES (%s,%s,%s,%s) RETURNING *
        """, (name, b.get('description',''), b.get('period','quarterly'),
              _json.dumps(items))).fetchone()
    return jsonify(_perf_template_row(row)), 201

@app.route('/api/performance/templates/<int:tid>', methods=['PUT'])
@login_required
def api_perf_template_update(tid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE performance_templates
            SET name=%s, description=%s, period=%s, items=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b.get('name','').strip(), b.get('description',''),
              b.get('period','quarterly'), _json.dumps(b.get('items',[])),
              bool(b.get('active', True)), tid)).fetchone()
    return jsonify(_perf_template_row(row)) if row else ('', 404)

@app.route('/api/performance/templates/<int:tid>', methods=['DELETE'])
@login_required
def api_perf_template_delete(tid):
    with get_db() as conn:
        conn.execute("DELETE FROM performance_templates WHERE id=%s", (tid,))
    return jsonify({'deleted': tid})

# ── 考核記錄 CRUD ───────────────────────────────────────────────

@app.route('/api/performance/reviews', methods=['GET'])
@login_required
def api_perf_reviews_list():
    staff_id = request.args.get('staff_id')
    period   = request.args.get('period')
    conds, params = ['TRUE'], []
    if staff_id: conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    if period:   conds.append("pr.period_label ILIKE %s"); params.append(f'%{period}%')
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*,
                   ps.name  AS staff_name,  ps.role   AS staff_role,
                   pt.name  AS tpl_name
            FROM performance_reviews pr
            JOIN punch_staff         ps ON ps.id = pr.staff_id
            LEFT JOIN performance_templates pt ON pt.id = pr.template_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.reviewed_at DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = _perf_review_row(r)
        d['staff_name']   = r['staff_name']
        d['staff_role']   = r['staff_role']
        d['template_name'] = r['tpl_name'] or ''
        result.append(d)
    return jsonify(result)

@app.route('/api/performance/reviews', methods=['POST'])
@login_required
def api_perf_review_create():
    b           = request.get_json(force=True)
    staff_id    = b.get('staff_id')
    template_id = b.get('template_id')
    period_label= (b.get('period_label') or '').strip()
    scores      = b.get('scores', {})
    comments    = (b.get('comments') or '').strip()
    reviewer    = (b.get('reviewer') or '').strip() or session.get('admin_display_name', '管理員')

    if not staff_id or not period_label:
        return jsonify({'error': '請選擇員工及考核期間'}), 400

    # Calculate total & grade from template items
    total = 0.0; max_s = 100.0
    if template_id:
        with get_db() as conn:
            tpl = conn.execute(
                "SELECT items FROM performance_templates WHERE id=%s", (template_id,)
            ).fetchone()
        if tpl:
            items = tpl.get('items') or []
            if isinstance(items, str):
                try: items = _json.loads(items)
                except: items = []
            if items:
                max_s = sum(float(it.get('max_score', 10)) for it in items)
                total = sum(
                    float(scores.get(str(it.get('id', it.get('name',''))), 0))
                    for it in items
                )
    else:
        total = float(b.get('total_score', 0))
        max_s = float(b.get('max_score', 100))

    pct   = (total / max_s * 100) if max_s > 0 else 0
    grade = _score_to_grade(pct)

    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO performance_reviews
              (staff_id, template_id, period_label, scores, total_score,
               max_score, grade, comments, reviewer, reviewed_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()) RETURNING *
        """, (staff_id, template_id or None, period_label,
              _json.dumps(scores), round(total, 2), round(max_s, 2),
              grade, comments, reviewer)).fetchone()
        staff = conn.execute(
            "SELECT name FROM punch_staff WHERE id=%s", (staff_id,)
        ).fetchone()

    # LINE 通知
    grade_labels = _grade_labels()
    msg = (f"[績效考核] {period_label} 考核結果\n"
           f"總分：{total:.1f} / {max_s:.0f}（{pct:.0f}%）\n"
           f"評級：{grade} {grade_labels.get(grade,'')}\n"
           f"考核人：{reviewer}\n"
           + (f"備注：{comments[:60]}\n" if comments else '')
           + "請至員工系統查看詳情。")
    _notify_staff_line(staff_id, msg)

    d = _perf_review_row(row)
    d['staff_name'] = staff['name'] if staff else ''
    return jsonify(d), 201

@app.route('/api/performance/reviews/<int:rid>', methods=['PUT'])
@login_required
def api_perf_review_update(rid):
    b        = request.get_json(force=True)
    scores   = b.get('scores', {})
    comments = (b.get('comments') or '').strip()
    with get_db() as conn:
        rev = conn.execute(
            "SELECT * FROM performance_reviews WHERE id=%s", (rid,)
        ).fetchone()
        if not rev: return ('', 404)
        # Recalculate score
        total = float(b.get('total_score', rev['total_score']))
        max_s = float(b.get('max_score',   rev['max_score']))
        pct   = (total / max_s * 100) if max_s > 0 else 0
        grade = _score_to_grade(pct)
        row = conn.execute("""
            UPDATE performance_reviews
            SET scores=%s, total_score=%s, max_score=%s, grade=%s,
                comments=%s, reviewed_at=NOW()
            WHERE id=%s RETURNING *
        """, (_json.dumps(scores), round(total,2), round(max_s,2),
              grade, comments, rid)).fetchone()
    return jsonify(_perf_review_row(row)) if row else ('', 404)

@app.route('/api/performance/reviews/<int:rid>/adjust-salary', methods=['POST'])
@login_required
def api_perf_adjust_salary(rid):
    """依考核結果調薪 — 直接更新員工底薪並記錄"""
    b     = request.get_json(force=True)
    delta = float(b.get('salary_delta', b.get('delta', 0)))
    note  = (b.get('note') or '').strip()
    if delta == 0: return jsonify({'error': '調薪金額不可為 0'}), 400
    with get_db() as conn:
        rev = conn.execute(
            "SELECT * FROM performance_reviews WHERE id=%s", (rid,)
        ).fetchone()
        if not rev: return ('', 404)
        staff = conn.execute(
            "SELECT id, name, base_salary FROM punch_staff WHERE id=%s", (rev['staff_id'],)
        ).fetchone()
        if not staff: return ('', 404)
        new_salary = float(staff['base_salary'] or 0) + delta
        conn.execute(
            "UPDATE punch_staff SET base_salary=%s WHERE id=%s",
            (new_salary, staff['id'])
        )
        conn.execute("""
            UPDATE performance_reviews
            SET salary_adjusted=TRUE, salary_delta=%s
            WHERE id=%s
        """, (delta, rid))

    direction = '調升' if delta > 0 else '調降'
    msg = (f"[薪資調整] 績效考核連動\n"
           f"考核期：{rev['period_label']}　評級：{rev['grade']}\n"
           f"{direction} NT$ {abs(delta):,.0f}\n"
           f"新底薪：NT$ {new_salary:,.0f}\n"
           + (f"說明：{note}" if note else ''))
    _notify_staff_line(staff['id'], msg)

    return jsonify({'ok': True, 'new_salary': new_salary, 'delta': delta})

# ── 員工查自己的考核 ────────────────────────────────────────────

@app.route('/api/performance/my-reviews', methods=['GET'])
def api_perf_my_reviews():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    with get_db() as conn:
        rows = conn.execute("""
            SELECT pr.*, pt.name AS tpl_name
            FROM performance_reviews pr
            LEFT JOIN performance_templates pt ON pt.id=pr.template_id
            WHERE pr.staff_id=%s
            ORDER BY pr.reviewed_at DESC LIMIT 10
        """, (sid,)).fetchall()
    result = []
    for r in rows:
        d = _perf_review_row(r)
        d['template_name'] = r['tpl_name'] or ''
        result.append(d)
    return jsonify(result)


# ── 評級設定 CRUD ───────────────────────────────────────────────

@app.route('/api/performance/config', methods=['GET'])
@login_required
def api_perf_config_get():
    return jsonify({'grades': _get_grade_config()})

@app.route('/api/performance/config', methods=['PUT'])
@login_required
def api_perf_config_update():
    b      = request.get_json(force=True)
    grades = b.get('grades', [])
    if not grades:
        return jsonify({'error': '請至少設定一個評級'}), 400
    for g in grades:
        if not str(g.get('grade', '')).strip() or not str(g.get('label', '')).strip():
            return jsonify({'error': '評級代碼與標籤不可為空'}), 400
        pct = g.get('min_pct')
        if pct is None or not (0 <= float(pct) <= 100):
            return jsonify({'error': '門檻百分比需介於 0~100'}), 400
    # 確保至少有一個門檻為 0，避免無法分級
    if not any(float(g.get('min_pct', -1)) == 0 for g in grades):
        return jsonify({'error': '必須有一個評級的門檻設為 0%（作為最低等級）'}), 400
    grades_sorted = sorted(
        [{'grade': str(g['grade']).strip(), 'label': str(g['label']).strip(),
          'min_pct': float(g['min_pct'])} for g in grades],
        key=lambda x: -x['min_pct']
    )
    with get_db() as conn:
        conn.execute("""
            INSERT INTO performance_config (key, value, updated_at)
            VALUES ('grade_config', %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value, updated_at=NOW()
        """, (_json.dumps(grades_sorted),))
    return jsonify({'ok': True, 'grades': grades_sorted})


# ═══════════════════════════════════════════════════════════════════════════
# LINE Bot 雙向查詢擴充
# ═══════════════════════════════════════════════════════════════════════════

def _line_query_leave_balance(staff, user_id):
    """查詢員工本年度假期餘額"""
    from datetime import date as _dlb
    year = _dlb.today().year
    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT lb.total_days, lb.used_days, lt.name AS type_name
                FROM leave_balances lb
                JOIN leave_types lt ON lt.id=lb.leave_type_id
                WHERE lb.staff_id=%s AND lb.year=%s
                ORDER BY lt.sort_order
            """, (staff['id'], year)).fetchall()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗：{e}')
        return
    if not rows:
        _send_line_punch(user_id, f'📋 {staff["name"]} {year} 年\n尚無假期餘額記錄，請聯絡管理員。')
        return
    lines = [f'📋 {staff["name"]} {year} 年假期餘額']
    for r in rows:
        total = float(r['total_days'] or 0)
        used  = float(r['used_days']  or 0)
        remain= total - used
        bar   = '▓' * int(remain) + '░' * max(0, int(total - remain))
        lines.append(f'\n【{r["type_name"]}】\n  剩餘 {remain:.1f} 天 / 共 {total:.0f} 天\n  {bar}')
    _send_line_punch(user_id, '\n'.join(lines))


def _line_query_salary(staff, user_id):
    """查詢員工最近一筆薪資記錄"""
    try:
        with get_db() as conn:
            row = conn.execute("""
                SELECT month, net_pay, base_salary, allowance_total, deduction_total, status
                FROM salary_records
                WHERE staff_id=%s
                ORDER BY month DESC LIMIT 1
            """, (staff['id'],)).fetchone()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗：{e}')
        return
    if not row:
        _send_line_punch(user_id, f'📊 {staff["name"]}\n尚無薪資記錄。')
        return
    status_map = {'draft':'草稿', 'confirmed':'已確認', 'paid':'已發放'}
    _send_line_punch(user_id,
        f'📊 {staff["name"]} {row["month"]} 薪資\n\n'
        f'底薪：NT$ {float(row["base_salary"] or 0):,.0f}\n'
        f'津貼：NT$ {float(row["allowance_total"] or 0):,.0f}\n'
        f'扣除：NT$ {float(row["deduction_total"] or 0):,.0f}\n'
        f'━━━━━━━━━━━━\n'
        f'實領：NT$ {float(row["net_pay"] or 0):,.0f}\n'
        f'狀態：{status_map.get(row["status"], row["status"])}\n\n'
        f'詳細資訊請至員工系統薪資單查看。')


def _line_submit_leave(staff, user_id, text):
    """
    解析並建立請假申請。
    格式：請假 [假別] [開始日期] [結束日期(選填)] [原因(選填)]
    範例：請假 特休 2026-04-01
         請假 事假 2026-04-01 2026-04-02 家庭事務
    """
    import re as _re_lv
    from datetime import date as _dlv, timedelta as _tdlv
    WDAY_LV = ['一', '二', '三', '四', '五', '六', '日']
    parts = text.strip().split()
    # parts[0] = '請假'

    # Step 1: only "請假" → Quick Reply with leave types + remaining balance
    if len(parts) == 1:
        year = _dlv.today().year
        with get_db() as conn:
            types = conn.execute(
                "SELECT id, name FROM leave_types WHERE active=TRUE ORDER BY sort_order"
            ).fetchall()
            balances = {
                r['leave_type_id']: (float(r['total_days'] or 0) - float(r['used_days'] or 0))
                for r in conn.execute("""
                    SELECT leave_type_id, total_days, used_days FROM leave_balances
                    WHERE staff_id=%s AND year=%s
                """, (staff['id'], year)).fetchall()
            }
        if not types:
            _send_line_punch(user_id, '目前無可用假別，請聯絡管理員。')
            return
        lines = ['🌿 請假申請\n\n可用假別（剩餘天數）：']
        items = []
        for r in types:
            rem = balances.get(r['id'])
            rem_str = f' {rem:.1f}天' if rem is not None else ''
            lines.append(f'• {r["name"]}{rem_str}')
            items.append({'label': f'{r["name"]}{rem_str}', 'text': f'請假 {r["name"]}'})
        lines.append('\n請點下方按鈕選擇假別：')
        _send_line_with_quick_reply(user_id, '\n'.join(lines), items[:13])
        return

    # Step 2: "請假 假別" (no date) → Quick Reply with date options
    if len(parts) == 2:
        leave_type_name = parts[1]
        today = _dlv.today()
        date_items = []
        for i in range(7):
            d = today + _tdlv(days=i)
            if d.weekday() == 6:  # skip Sunday
                continue
            label = ('今天 ' if i == 0 else '明天 ' if i == 1 else '') + f'{d.strftime("%m/%d")}({WDAY_LV[d.weekday()]})'
            date_items.append({'label': label, 'text': f'請假 {leave_type_name} {d.isoformat()}'})
            if len(date_items) == 6:
                break
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n\n請選擇日期，或手動輸入多天：\n'
            f'請假 {leave_type_name} 開始日 結束日',
            date_items)
        return

    # Step 2.5: "請假 假別 DATE" (one date, no period) → Quick Reply: 全天/上午/下午/自訂時段
    if len(parts) == 3 and _re_lv.match(r'^\d{4}-\d{2}-\d{2}$', parts[2]):
        leave_type_name = parts[1]
        date_str = parts[2]
        items_period = [
            {'label': '全天',     'text': f'請假 {leave_type_name} {date_str} 全天'},
            {'label': '上午半天', 'text': f'請假 {leave_type_name} {date_str} 上午'},
            {'label': '下午半天', 'text': f'請假 {leave_type_name} {date_str} 下午'},
            {'label': '⏰ 自訂時段', 'text': f'請假 {leave_type_name} {date_str} 09:00'},
        ]
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n日期：{date_str}\n\n請選擇時段：',
            items_period)
        return

    # Step 2.6: "請假 假別 DATE HH:MM" → Quick Reply 選結束時間
    if (len(parts) == 4
            and _re_lv.match(r'^\d{4}-\d{2}-\d{2}$', parts[2])
            and _re_lv.match(r'^\d{2}:\d{2}$', parts[3])):
        leave_type_name = parts[1]
        date_str = parts[2]
        start_str = parts[3]
        sh, sm = map(int, start_str.split(':'))
        end_options = []
        for delta_h in [1, 1.5, 2, 2.5, 3, 4, 6, 8]:
            total_m = sh * 60 + sm + int(delta_h * 60)
            eh, em = (total_m // 60) % 24, total_m % 60
            end_options.append((f'{eh:02d}:{em:02d}', delta_h))
        items_end = [
            {'label': f'至 {t}（{d}h）', 'text': f'請假 {leave_type_name} {date_str} {start_str} {t}'}
            for t, d in end_options
        ]
        _send_line_with_quick_reply(user_id,
            f'🌿 請假 · {leave_type_name}\n日期：{date_str}　開始：{start_str}\n\n請選擇結束時間：',
            items_end)
        return

    leave_type_name = parts[1]
    date_str1 = parts[2]

    # 判斷是否為「幾點到幾點」的自訂時段格式：請假 假別 DATE HH:MM HH:MM
    leave_start_time = None
    leave_end_time   = None
    start_half = False; end_half = False
    period_token = None
    date_str2 = date_str1

    if len(parts) >= 5 and _re_lv.match(r'^\d{2}:\d{2}$', parts[3]) and _re_lv.match(r'^\d{2}:\d{2}$', parts[4]):
        # 自訂時段：DATE START_TIME END_TIME
        leave_start_time = parts[3]
        leave_end_time   = parts[4]
    elif len(parts) > 3:
        tok = parts[3].strip()
        if tok in ('全天', '上午', '下午'):
            period_token = tok
        elif _re_lv.match(r'^\d{4}-\d{2}-\d{2}$', tok):
            date_str2 = tok

    if period_token == '上午':
        start_half = True; end_half = True
    elif period_token == '下午':
        start_half = False; end_half = True

    reason = '（LINE 請假）'

    # Validate dates
    try:
        _dlv.fromisoformat(date_str1)
        _dlv.fromisoformat(date_str2)
    except ValueError:
        _send_line_punch(user_id, f'日期格式錯誤，請使用 YYYY-MM-DD，例：{_dlv.today().isoformat()}')
        return

    # Find leave type (fuzzy: exact or contains)
    with get_db() as conn:
        lt = conn.execute(
            "SELECT * FROM leave_types WHERE active=TRUE AND name=%s", (leave_type_name,)
        ).fetchone()
        if not lt:
            lt = conn.execute(
                "SELECT * FROM leave_types WHERE active=TRUE AND name ILIKE %s LIMIT 1",
                (f'%{leave_type_name}%',)
            ).fetchone()
        if not lt:
            avail = conn.execute(
                "SELECT name FROM leave_types WHERE active=TRUE ORDER BY sort_order"
            ).fetchall()
            names = '、'.join(r['name'] for r in avail)
            _send_line_punch(user_id, f'找不到假別「{leave_type_name}」\n\n可用假別：{names}')
            return

        # Check leave balance
        year = date_str1[:4]
        bal = conn.execute("""
            SELECT total_days, used_days FROM leave_balances
            WHERE staff_id=%s AND leave_type_id=%s AND year=%s
        """, (staff['id'], lt['id'], int(year))).fetchone()

        # Calculate requested days
        if leave_start_time and leave_end_time:
            # 自訂時段：依小時換算天數（員工每日工時）
            daily_hours = float(staff.get('daily_hours') or 8)
            sh, sm = map(int, leave_start_time.split(':'))
            eh, em = map(int, leave_end_time.split(':'))
            leave_minutes = (eh * 60 + em) - (sh * 60 + sm)
            if leave_minutes <= 0:
                leave_minutes += 24 * 60
            days = round(leave_minutes / 60 / daily_hours, 2)
        else:
            s = _dlv.fromisoformat(date_str1); e = _dlv.fromisoformat(date_str2)
            days = sum(1 for i in range((e - s).days + 1)
                       if (s + _tdlv(days=i)).weekday() != 6)
            if start_half or end_half:
                days = max(0.5, days - 0.5)

        remain = None
        if bal:
            remain = float(bal['total_days'] or 0) - float(bal['used_days'] or 0)
            if remain < days:
                _send_line_punch(user_id,
                    f'⚠️ {lt["name"]} 餘額不足\n剩餘 {remain:.1f} 天，申請 {days} 天\n\n'
                    f'請至員工系統調整後再申請。')
                return

        # Create leave request
        row = conn.execute("""
            INSERT INTO leave_requests
              (staff_id, leave_type_id, start_date, end_date, total_days,
               start_half, end_half, reason, status, leave_start_time, leave_end_time, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s,NOW()) RETURNING id
        """, (staff['id'], lt['id'], date_str1, date_str2, days,
              start_half, end_half, reason, leave_start_time, leave_end_time)).fetchone()

    if leave_start_time and leave_end_time:
        period_label = f'（{leave_start_time} ～ {leave_end_time}）'
    elif start_half and end_half and date_str1 == date_str2:
        period_label = '（上午半天）'
    elif end_half and not start_half and date_str1 == date_str2:
        period_label = '（下午半天）'
    else:
        period_label = ''

    bal_str = f'（剩餘 {remain:.1f} 天）' if remain is not None else ''
    _send_line_punch(user_id,
        f'✅ 請假申請已送出\n\n'
        f'假別：{lt["name"]} {bal_str}\n'
        f'日期：{date_str1}' + (f' ～ {date_str2}' if date_str2 != date_str1 else '') +
        f'{period_label}\n'
        f'天數：{days} 天\n\n'
        f'申請號：#{row["id"]}，等待管理員審核。')


def _line_query_performance(staff, user_id):
    """查詢員工最近一次績效考核"""
    try:
        with get_db() as conn:
            row = conn.execute("""
                SELECT pr.period_label, pr.grade, pr.total_score, pr.max_score,
                       pr.comments, pr.salary_adjusted, pr.salary_delta,
                       pr.reviewed_at, pt.name AS tpl_name
                FROM performance_reviews pr
                LEFT JOIN performance_templates pt ON pt.id=pr.template_id
                WHERE pr.staff_id=%s
                ORDER BY pr.reviewed_at DESC LIMIT 1
            """, (staff['id'],)).fetchone()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗：{e}')
        return
    if not row:
        _send_line_punch(user_id, f'{staff["name"]}\n尚無績效考核記錄。')
        return
    grade_label = _grade_labels()
    pct = float(row['total_score']) / float(row['max_score']) * 100 if row['max_score'] else 0
    adj = f"\n薪資調整：NT$ {float(row['salary_delta']):+,.0f}" if row['salary_adjusted'] else ''
    reviewed = str(row['reviewed_at'])[:10] if row['reviewed_at'] else ''
    _send_line_punch(user_id,
        f'{staff["name"]} 最近考核\n\n'
        f'期間：{row["period_label"]}\n'
        f'範本：{row["tpl_name"] or "—"}\n'
        f'得分：{float(row["total_score"]):.1f} / {float(row["max_score"]):.0f}（{pct:.0f}%）\n'
        f'評級：{row["grade"]} {grade_label.get(row["grade"],"")}'
        f'{adj}\n'
        + (f'備注：{row["comments"][:60]}\n' if row['comments'] else '')
        + f'考核日：{reviewed}')


def _line_query_monthly_records(staff, user_id, text):
    """查詢員工月出勤記錄與打卡明細。
    格式：出勤紀錄 [YYYY-MM]（省略月份則查本月）
    """
    import re as _rem
    from datetime import date as _dm, timezone as _tzm, timedelta as _tdm, datetime as _dtm
    TW = _tzm(_tdm(hours=8))

    # 解析月份
    parts = text.strip().split()
    month = None
    if len(parts) >= 2:
        m = _rem.match(r'^(\d{4})-(\d{1,2})$', parts[1])
        if m:
            month = f"{m.group(1)}-{m.group(2).zfill(2)}"
    if not month:
        month = _dtm.now(TW).strftime('%Y-%m')

    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT punch_type, punched_at, is_manual
                FROM punch_records
                WHERE staff_id=%s
                  AND to_char(punched_at AT TIME ZONE 'Asia/Taipei', 'YYYY-MM') = %s
                ORDER BY punched_at ASC
            """, (staff['id'], month)).fetchall()
    except Exception as e:
        _send_line_punch(user_id, f'查詢失敗：{e}')
        return

    if not rows:
        _send_line_punch(user_id, f'📋 {staff["name"]} {month}\n該月尚無打卡記錄。')
        return

    WDAY = ['一', '二', '三', '四', '五', '六', '日']

    # 依日期分組
    days = {}
    for r in rows:
        pa = r['punched_at']
        if pa.tzinfo is None:
            from datetime import timezone as _utzm
            pa = pa.replace(tzinfo=_utzm.utc)
        pa_tw = pa.astimezone(TW)
        ds = pa_tw.strftime('%Y-%m-%d')
        if ds not in days:
            days[ds] = []
        days[ds].append({'type': r['punch_type'], 'time': pa_tw.strftime('%H:%M'), 'manual': bool(r['is_manual'])})

    total_mins = 0
    anomaly_days = 0
    lines = []

    for ds in sorted(days.keys()):
        recs = days[ds]
        d = _dm.fromisoformat(ds)
        wday = WDAY[d.weekday()]

        clock_in  = next((r['time'] for r in recs if r['type'] == 'in'),  None)
        clock_out = next((r['time'] for r in recs if r['type'] == 'out'), None)
        has_manual = any(r['manual'] for r in recs)

        if clock_in and clock_out:
            ci = _dtm.strptime(f'{ds} {clock_in}',  '%Y-%m-%d %H:%M')
            co = _dtm.strptime(f'{ds} {clock_out}', '%Y-%m-%d %H:%M')
            mins = max(0, int((co - ci).total_seconds() / 60))
            total_mins += mins
            h, m = divmod(mins, 60)
            dur = f'{h}h{m:02d}' if m else f'{h}h'
        elif clock_in:
            dur = '⚠️缺下班'
            anomaly_days += 1
        else:
            dur = '⚠️缺上班'
            anomaly_days += 1

        manual_mark = '【補】' if has_manual else ''
        ci_str = clock_in  or '--:--'
        co_str = clock_out or '--:--'
        lines.append(f'{ds[5:]}({wday}) {ci_str}↑{co_str}↓ {dur}{manual_mark}')

    th, tm = divmod(total_mins, 60)
    total_str = f'{th}h{tm:02d}' if tm else f'{th}h'
    anomaly_str = f'｜異常 {anomaly_days} 天' if anomaly_days else ''
    header = (f'📋 {staff["name"]} {month} 出勤\n'
              f'出勤 {len(days)} 天｜工時 {total_str}{anomaly_str}\n'
              + '─' * 20)

    # 訊息過長時分批送出（LINE 單則上限約 5000 字）
    full = header + '\n' + '\n'.join(lines)
    if len(full) <= 4500:
        _send_line_punch(user_id, full)
    else:
        _send_line_punch(user_id, header)
        chunk, chunk_len = [], 0
        for line in lines:
            if chunk_len + len(line) + 1 > 1800:
                _send_line_punch(user_id, '\n'.join(chunk))
                chunk, chunk_len = [], 0
            chunk.append(line)
            chunk_len += len(line) + 1
        if chunk:
            _send_line_punch(user_id, '\n'.join(chunk))


def _line_overtime_start(staff, user_id):
    """加班 button → Quick Reply with date options."""
    from datetime import date as _dot_s, timedelta as _tdot_s
    WDAY_OT = ['一', '二', '三', '四', '五', '六', '日']
    today = _dot_s.today()
    items = []
    for i in range(-1, 5):
        d = today + _tdot_s(days=i)
        label = ('昨天 ' if i == -1 else '今天 ' if i == 0 else '明天 ' if i == 1 else '') + \
                f'{d.strftime("%m/%d")}({WDAY_OT[d.weekday()]})'
        items.append({'label': label, 'text': f'申請加班 {d.isoformat()}'})
    _send_line_with_quick_reply(user_id, '⏰ 加班申請\n\n請選擇加班日期：', items)


def _line_submit_overtime(staff, user_id, text):
    """
    LINE 加班申請流程（幾點到幾點）：
      申請加班 DATE           → Quick Reply 選開始時間
      申請加班 DATE HH:MM     → Quick Reply 選結束時間
      申請加班 DATE HH:MM HH:MM → 送出申請
    """
    import re as _re_ot
    from datetime import date as _dot, datetime as _dtt
    parts = text.strip().split(None, 3)

    if len(parts) < 2:
        _line_overtime_start(staff, user_id)
        return

    date_str = parts[1]
    try:
        _dot.fromisoformat(date_str)
    except ValueError:
        _send_line_punch(user_id, f'日期格式錯誤，請使用 YYYY-MM-DD，例：{_dot.today().isoformat()}')
        return

    # Step 2: date only → select start time
    if len(parts) == 2:
        start_options = ['08:00','09:00','17:00','18:00','19:00','20:00','21:00','22:00']
        items = [{'label': t, 'text': f'申請加班 {date_str} {t}'} for t in start_options]
        _send_line_with_quick_reply(user_id,
            f'⏰ 加班申請 · {date_str}\n\n請選擇開始時間：', items)
        return

    start_str = parts[2]
    if not _re_ot.match(r'^\d{2}:\d{2}$', start_str):
        _send_line_punch(user_id, '時間格式錯誤，請使用 HH:MM，例：18:00')
        return

    # Step 3: date + start time → select end time
    if len(parts) == 3:
        sh, sm = map(int, start_str.split(':'))
        end_options = []
        for delta_h in [1, 1.5, 2, 2.5, 3, 4, 5, 6]:
            total_m = sh * 60 + sm + int(delta_h * 60)
            eh, em = (total_m // 60) % 24, total_m % 60
            end_options.append(f'{eh:02d}:{em:02d}')
        items = [{'label': f'至 {t}（+{d}h）', 'text': f'申請加班 {date_str} {start_str} {t}'}
                 for t, d in zip(end_options, [1, 1.5, 2, 2.5, 3, 4, 5, 6])]
        _send_line_with_quick_reply(user_id,
            f'⏰ 加班申請 · {date_str} {start_str} 開始\n\n請選擇結束時間：', items)
        return

    # Step 4: date + start + end → submit
    end_str = parts[3].strip().split()[0]  # take only first token (HH:MM)
    if not _re_ot.match(r'^\d{2}:\d{2}$', end_str):
        _send_line_punch(user_id, '時間格式錯誤，請使用 HH:MM，例：20:00')
        return

    try:
        sh, sm = map(int, start_str.split(':'))
        eh, em = map(int, end_str.split(':'))
        hours = ((eh * 60 + em) - (sh * 60 + sm)) / 60
        if hours <= 0:
            hours += 24  # crosses midnight
        if hours <= 0 or hours > 24:
            raise ValueError
    except ValueError:
        _send_line_punch(user_id, '時間計算錯誤，請重新選擇。')
        return

    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO overtime_requests
              (staff_id, request_date, start_time, end_time, ot_hours, reason, status)
            VALUES (%s, %s, %s, %s, %s, %s, 'pending')
            RETURNING id
        """, (staff['id'], date_str, start_str, end_str, round(hours, 2), '（LINE 加班申請）')).fetchone()

    _send_line_punch(user_id,
        f'✅ 加班申請已送出\n\n'
        f'日期：{date_str}\n'
        f'時段：{start_str} ～ {end_str}（{hours:.1f} 小時）\n'
        f'申請編號：#{row["id"]}\n\n'
        '請等候管理員審核，審核結果將通知您。')


def _line_show_help(staff, user_id):
    _send_line_punch(user_id,
        f'哈囉 {staff["name"]}！以下是可用的指令：\n\n'
        '─── 打卡 ───\n'
        '📍 傳送位置 → 自動打卡\n'
        '💬 上班 / 下班 / 休息 / 回來\n'
        '📋 狀態 → 今日打卡記錄\n\n'
        '─── 查詢 ───\n'
        '🌿 查餘假 → 本年假期餘額\n'
        '💰 查薪資 → 最近薪資單\n'
        '📊 出勤紀錄 → 本月出勤明細\n'
        '   出勤紀錄 2026-03 → 指定月份\n'
        '考核 → 最近績效考核\n\n'
        '─── 申請 ───\n'
        '📝 請假 [假別] [日期] → 送出請假\n'
        '   範例：請假 特休 2026-04-01\n'
        '⏰ 申請加班 [日期] [時數] → 加班申請\n'
        '   範例：申請加班 2026-04-05 3\n'
        '🗂️ 假別 → 查看可用假別清單\n\n'
        '─── 其他 ───\n'
        '🔓 解除綁定')


def _line_show_leave_types(staff, user_id):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT name, max_days FROM leave_types WHERE active=TRUE ORDER BY sort_order"
        ).fetchall()
    if not rows:
        _send_line_punch(user_id, '目前無可用假別。'); return
    lines = ['🗂️ 可用假別清單\n']
    for r in rows:
        limit = f'（年限 {r["max_days"]} 天）' if r['max_days'] else ''
        lines.append(f'• {r["name"]} {limit}')
    lines.append('\n申請方式：請假 [假別] [日期]')
    _send_line_punch(user_id, '\n'.join(lines))



# ═══════════════════════════════════════════════════════════════════════════════
# 📱 MOBILE API  (JWT-based, /api/mobile/*)
# ═══════════════════════════════════════════════════════════════════════════════
import jwt as _pyjwt

_MOBILE_JWT_SECRET = os.environ.get('MOBILE_JWT_SECRET', app.secret_key)
_JWT_EXPIRE_HOURS  = 24 * 7   # 7 days

def _make_jwt(payload: dict) -> str:
    payload['exp'] = _dt.now(_tz.utc) + _td(hours=_JWT_EXPIRE_HOURS)
    return _pyjwt.encode(payload, _MOBILE_JWT_SECRET, algorithm='HS256')

def _decode_jwt(token: str):
    return _pyjwt.decode(token, _MOBILE_JWT_SECRET, algorithms=['HS256'])

def mobile_jwt_required(f):
    """Decorator: reads Bearer token, sets g.mobile_user = {id, role, ...}"""
    from flask import g
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'error': '未授權'}), 401
        token = auth[7:]
        try:
            payload = _decode_jwt(token)
        except _pyjwt.ExpiredSignatureError:
            return jsonify({'error': 'token 已過期，請重新登入'}), 401
        except Exception:
            return jsonify({'error': 'token 無效'}), 401
        g.mobile_user = payload
        return f(*args, **kwargs)
    return decorated

def mobile_admin_required(f):
    """Decorator: must be admin role"""
    from flask import g
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'error': '未授權'}), 401
        token = auth[7:]
        try:
            payload = _decode_jwt(token)
        except Exception:
            return jsonify({'error': 'token 無效'}), 401
        if payload.get('role') != 'admin':
            return jsonify({'error': '需要管理員權限'}), 403
        g.mobile_user = payload
        return f(*args, **kwargs)
    return decorated

# ── Login ──────────────────────────────────────────────────────────────────────

@app.route('/api/mobile/login', methods=['POST'])
def mobile_login():
    b = request.get_json(force=True) or {}
    username = b.get('username', '').strip()
    password = b.get('password', '').strip()
    if not username or not password:
        return jsonify({'error': '請輸入帳號及密碼'}), 400

    # Try admin accounts first
    with get_db() as conn:
        admin = conn.execute(
            "SELECT * FROM admin_accounts WHERE username=%s AND active=TRUE", (username,)
        ).fetchone()
    if admin and admin['password_hash'] == _hash_pw(password):
        perms = admin['permissions']
        if isinstance(perms, str):
            try: perms = _json.loads(perms)
            except: perms = []
        token = _make_jwt({
            'sub': str(admin['id']), 'role': 'admin',
            'username': admin['username'],
            'display_name': admin['display_name'] or admin['username'],
            'is_super': bool(admin['is_super']),
            'permissions': perms,
        })
        with get_db() as conn:
            conn.execute("UPDATE admin_accounts SET last_login_at=NOW() WHERE id=%s", (admin['id'],))
        return jsonify({
            'token': token,
            'role': 'admin',
            'user': {
                'id': admin['id'],
                'username': admin['username'],
                'display_name': admin['display_name'] or admin['username'],
                'is_super': bool(admin['is_super']),
                'permissions': perms,
            }
        })

    # Try employee accounts
    with get_db() as conn:
        staff = conn.execute(
            "SELECT * FROM punch_staff WHERE username=%s AND active=TRUE", (username,)
        ).fetchone()
    if staff and staff['password_hash'] == _hash_pw(password):
        token = _make_jwt({
            'sub': str(staff['id']), 'role': 'employee',
            'staff_id': staff['id'],
            'name': staff['name'],
            'username': staff['username'],
        })
        return jsonify({
            'token': token,
            'role': 'employee',
            'user': {
                'id': staff['id'],
                'name': staff['name'],
                'username': staff['username'],
                'role': staff['role'],
                'department': staff['department'],
                'position_title': staff['position_title'],
                'employee_code': staff['employee_code'],
            }
        })

    return jsonify({'error': '帳號或密碼錯誤'}), 401

# ── Employee: Me & Profile ─────────────────────────────────────────────────────

@app.route('/api/mobile/me', methods=['GET'])
@mobile_jwt_required
def mobile_me():
    from flask import g
    u = g.mobile_user
    if u['role'] == 'employee':
        with get_db() as conn:
            staff = conn.execute(
                """SELECT id, name, username, role, department, position_title,
                          employee_code, hire_date, birth_date, base_salary,
                          insured_salary, daily_hours, salary_type, active
                   FROM punch_staff WHERE id=%s""", (int(u['sub']),)
            ).fetchone()
        if not staff:
            return jsonify({'error': '帳號不存在'}), 404
        d = dict(staff)
        for k in ('hire_date', 'birth_date'):
            if d.get(k): d[k] = str(d[k])
        return jsonify(d)
    else:
        with get_db() as conn:
            admin = conn.execute(
                "SELECT id, username, display_name, is_super, permissions FROM admin_accounts WHERE id=%s",
                (int(u['sub']),)
            ).fetchone()
        if not admin:
            return jsonify({'error': '帳號不存在'}), 404
        d = dict(admin)
        if isinstance(d['permissions'], str):
            try: d['permissions'] = _json.loads(d['permissions'])
            except: d['permissions'] = []
        return jsonify(d)

# ── Employee: Punch ────────────────────────────────────────────────────────────

@app.route('/api/mobile/punch', methods=['POST'])
@mobile_jwt_required
def mobile_punch():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可打卡'}), 403
    staff_id = int(u['sub'])
    b = request.get_json(force=True) or {}
    punch_type = b.get('punch_type', 'in')
    lat  = b.get('latitude')
    lng  = b.get('longitude')
    note = b.get('note', '')

    # GPS validation (same logic as web)
    with get_db() as conn:
        cfg = conn.execute("SELECT gps_required FROM punch_config WHERE id=1").fetchone()
        gps_required = cfg['gps_required'] if cfg else False
        locs = conn.execute("SELECT * FROM punch_locations WHERE active=TRUE").fetchall()

    gps_distance = None
    location_name = ''
    if lat is not None and lng is not None and locs:
        def haversine(la1, lo1, la2, lo2):
            R = 6371000
            p = math.pi / 180
            a = (math.sin((la2-la1)*p/2)**2 +
                 math.cos(la1*p)*math.cos(la2*p)*math.sin((lo2-lo1)*p/2)**2)
            return int(2*R*math.asin(math.sqrt(a)))
        best = min(locs, key=lambda l: haversine(float(l['lat']), float(l['lng']), float(lat), float(lng)))
        gps_distance = haversine(float(best['lat']), float(best['lng']), float(lat), float(lng))
        location_name = best['location_name']
        if gps_required and gps_distance > best['radius_m']:
            return jsonify({'error': f'距離打卡地點 {gps_distance}m，超出範圍 {best["radius_m"]}m'}), 400
    elif gps_required:
        return jsonify({'error': '此門市需要 GPS 定位才能打卡'}), 400

    with get_db() as conn:
        conn.execute(
            """INSERT INTO punch_records
               (staff_id, punch_type, note, latitude, longitude, gps_distance, location_name)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (staff_id, punch_type, note, lat, lng, gps_distance, location_name)
        )
    return jsonify({'ok': True, 'location_name': location_name, 'gps_distance': gps_distance})

@app.route('/api/mobile/punch/status', methods=['GET'])
@mobile_jwt_required
def mobile_punch_status():
    """Return today's punch records for the employee."""
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可查詢'}), 403
    staff_id = int(u['sub'])
    today = date.today().isoformat()
    with get_db() as conn:
        rows = conn.execute(
            """SELECT punch_type, punched_at, note, gps_distance, location_name
               FROM punch_records WHERE staff_id=%s
               AND punched_at::date = %s::date ORDER BY punched_at""",
            (staff_id, today)
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        d['punched_at'] = d['punched_at'].isoformat() if d.get('punched_at') else None
        data.append(d)
    return jsonify(data)

# ── Employee: Attendance ───────────────────────────────────────────────────────

@app.route('/api/mobile/attendance', methods=['GET'])
@mobile_jwt_required
def mobile_attendance():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可查詢'}), 403
    staff_id = int(u['sub'])
    month = request.args.get('month', _dt.now(TW_TZ).strftime('%Y-%m'))
    try:
        y, m = map(int, month.split('-'))
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400
    with get_db() as conn:
        rows = conn.execute(
            """SELECT punch_type, punched_at, note, gps_distance, location_name, is_manual
               FROM punch_records WHERE staff_id=%s
               AND date_trunc('month', punched_at) = %s::date
               ORDER BY punched_at""",
            (staff_id, f'{y}-{m:02d}-01')
        ).fetchall()

    # Group by day
    from collections import defaultdict
    days = defaultdict(list)
    for r in rows:
        day = r['punched_at'].date().isoformat()
        days[day].append({
            'type': r['punch_type'],
            'time': r['punched_at'].strftime('%H:%M'),
            'note': r['note'],
            'gps_distance': r['gps_distance'],
            'location_name': r['location_name'],
            'is_manual': r['is_manual'],
        })

    result = []
    for day in sorted(days.keys()):
        records = days[day]
        ins  = [r for r in records if r['type'] == 'in']
        outs = [r for r in records if r['type'] == 'out']
        clock_in  = ins[0]['time']  if ins  else None
        clock_out = outs[-1]['time'] if outs else None
        hours = None
        if clock_in and clock_out:
            ci = _dt.strptime(clock_in,  '%H:%M')
            co = _dt.strptime(clock_out, '%H:%M')
            diff = (co - ci).seconds / 3600
            hours = round(diff, 2)
        result.append({'date': day, 'clock_in': clock_in, 'clock_out': clock_out,
                       'hours': hours, 'records': records})
    return jsonify(result)

# ── Employee: Leave ────────────────────────────────────────────────────────────

@app.route('/api/mobile/leave/types', methods=['GET'])
@mobile_jwt_required
def mobile_leave_types():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, max_days FROM leave_types WHERE active=TRUE ORDER BY sort_order"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mobile/leave', methods=['GET'])
@mobile_jwt_required
def mobile_leave_list():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可查詢'}), 403
    staff_id = int(u['sub'])
    with get_db() as conn:
        rows = conn.execute(
            """SELECT lr.id, lt.name AS leave_type, lr.start_date, lr.end_date,
                      lr.days, lr.reason, lr.status, lr.created_at
               FROM leave_requests lr
               JOIN leave_types lt ON lr.leave_type_id = lt.id
               WHERE lr.staff_id=%s ORDER BY lr.created_at DESC LIMIT 50""",
            (staff_id,)
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        for k in ('start_date','end_date','created_at'):
            if d.get(k): d[k] = str(d[k])
        data.append(d)
    return jsonify(data)

@app.route('/api/mobile/leave', methods=['POST'])
@mobile_jwt_required
def mobile_leave_apply():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可申請'}), 403
    staff_id = int(u['sub'])
    b = request.get_json(force=True) or {}
    leave_type_id = b.get('leave_type_id')
    start_date    = b.get('start_date')
    end_date      = b.get('end_date', start_date)
    reason        = b.get('reason', '')
    if not leave_type_id or not start_date:
        return jsonify({'error': '缺少必填欄位'}), 400
    try:
        sd = _dt.strptime(start_date, '%Y-%m-%d').date()
        ed = _dt.strptime(end_date,   '%Y-%m-%d').date()
        days = (ed - sd).days + 1
    except Exception:
        return jsonify({'error': '日期格式錯誤'}), 400
    with get_db() as conn:
        conn.execute(
            """INSERT INTO leave_requests (staff_id, leave_type_id, start_date, end_date, days, reason, status)
               VALUES (%s, %s, %s, %s, %s, %s, 'pending')""",
            (staff_id, leave_type_id, start_date, end_date, days, reason)
        )
    return jsonify({'ok': True})

# ── Employee: Schedule ─────────────────────────────────────────────────────────

@app.route('/api/mobile/schedule', methods=['GET'])
@mobile_jwt_required
def mobile_schedule():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可查詢'}), 403
    staff_id = int(u['sub'])
    month = request.args.get('month', _dt.now(TW_TZ).strftime('%Y-%m'))
    try:
        y, m = map(int, month.split('-'))
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400
    with get_db() as conn:
        rows = conn.execute(
            """SELECT sa.shift_date, st.name AS shift_name, st.start_time, st.end_time, st.color
               FROM shift_assignments sa
               JOIN shift_types st ON sa.shift_type_id = st.id
               WHERE sa.staff_id=%s AND date_trunc('month', sa.shift_date) = %s::date
               ORDER BY sa.shift_date""",
            (staff_id, f'{y}-{m:02d}-01')
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        d['shift_date'] = str(d['shift_date'])
        if d.get('start_time'): d['start_time'] = str(d['start_time'])
        if d.get('end_time'):   d['end_time']   = str(d['end_time'])
        data.append(d)
    return jsonify(data)

# ── Employee: Salary ───────────────────────────────────────────────────────────

@app.route('/api/mobile/salary', methods=['GET'])
@mobile_jwt_required
def mobile_salary():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可查詢'}), 403
    staff_id = int(u['sub'])
    with get_db() as conn:
        rows = conn.execute(
            """SELECT id, period_year, period_month, base_salary, bonus, deductions,
                      net_salary, status, paid_at, created_at
               FROM salary_records WHERE staff_id=%s ORDER BY period_year DESC, period_month DESC LIMIT 12""",
            (staff_id,)
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        for k in ('paid_at','created_at'):
            if d.get(k): d[k] = str(d[k])
        for k in ('base_salary','bonus','deductions','net_salary'):
            if d.get(k) is not None: d[k] = float(d[k])
        data.append(d)
    return jsonify(data)

# ── Employee: Overtime ─────────────────────────────────────────────────────────

@app.route('/api/mobile/overtime', methods=['POST'])
@mobile_jwt_required
def mobile_overtime():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可申請'}), 403
    staff_id = int(u['sub'])
    b = request.get_json(force=True) or {}
    ot_date = b.get('ot_date')
    hours   = b.get('hours')
    reason  = b.get('reason', '')
    if not ot_date or not hours:
        return jsonify({'error': '缺少必填欄位'}), 400
    try:
        _dt.strptime(ot_date, '%Y-%m-%d')
        hours = float(hours)
    except Exception:
        return jsonify({'error': '格式錯誤'}), 400
    with get_db() as conn:
        conn.execute(
            """INSERT INTO overtime_requests
               (staff_id, request_date, ot_hours, reason, status)
               VALUES (%s, %s, %s, %s, 'pending')""",
            (staff_id, ot_date, hours, reason)
        )
    return jsonify({'ok': True})

@app.route('/api/mobile/overtime', methods=['GET'])
@mobile_jwt_required
def mobile_overtime_list():
    from flask import g
    u = g.mobile_user
    if u['role'] != 'employee':
        return jsonify({'error': '僅員工可查詢'}), 403
    staff_id = int(u['sub'])
    with get_db() as conn:
        rows = conn.execute(
            """SELECT id, request_date AS ot_date, ot_hours, reason, status, created_at
               FROM overtime_requests WHERE staff_id=%s ORDER BY request_date DESC LIMIT 30""",
            (staff_id,)
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        for k in ('ot_date','created_at'):
            if d.get(k): d[k] = str(d[k])
        if d.get('ot_hours'): d['ot_hours'] = float(d['ot_hours'])
        data.append(d)
    return jsonify(data)

# ── Admin: Dashboard ───────────────────────────────────────────────────────────

@app.route('/api/mobile/admin/dashboard', methods=['GET'])
@mobile_admin_required
def mobile_admin_dashboard():
    today = date.today().isoformat()
    with get_db() as conn:
        total_staff = conn.execute("SELECT COUNT(*) AS n FROM punch_staff WHERE active=TRUE").fetchone()['n']
        punched_today = conn.execute(
            "SELECT COUNT(DISTINCT staff_id) AS n FROM punch_records WHERE punched_at::date=%s::date", (today,)
        ).fetchone()['n']
        pending_leaves = conn.execute(
            "SELECT COUNT(*) AS n FROM leave_requests WHERE status='pending'"
        ).fetchone()['n']
        pending_ot = conn.execute(
            "SELECT COUNT(*) AS n FROM overtime_requests WHERE status='pending'"
        ).fetchone()['n']
        # Last 7 days attendance rate
        rows_7d = conn.execute(
            """SELECT punched_at::date AS day, COUNT(DISTINCT staff_id) AS cnt
               FROM punch_records
               WHERE punched_at::date >= (CURRENT_DATE - INTERVAL '6 days')
               GROUP BY day ORDER BY day""",
        ).fetchall()
    attendance_trend = [{'date': str(r['day']), 'count': r['cnt']} for r in rows_7d]
    return jsonify({
        'total_staff': total_staff,
        'punched_today': punched_today,
        'pending_leaves': pending_leaves,
        'pending_ot': pending_ot,
        'attendance_trend': attendance_trend,
    })

# ── Admin: Today's Attendance ──────────────────────────────────────────────────

@app.route('/api/mobile/admin/attendance/today', methods=['GET'])
@mobile_admin_required
def mobile_admin_attendance_today():
    today = date.today().isoformat()
    with get_db() as conn:
        staff_all = conn.execute(
            "SELECT id, name, department, position_title FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        records = conn.execute(
            """SELECT staff_id, punch_type, punched_at
               FROM punch_records WHERE punched_at::date=%s::date ORDER BY punched_at""",
            (today,)
        ).fetchall()
    from collections import defaultdict
    by_staff = defaultdict(list)
    for r in records:
        by_staff[r['staff_id']].append(r)

    result = []
    for s in staff_all:
        recs = by_staff[s['id']]
        ins  = [r for r in recs if r['punch_type'] == 'in']
        outs = [r for r in recs if r['punch_type'] == 'out']
        clock_in  = ins[0]['punched_at'].strftime('%H:%M')  if ins  else None
        clock_out = outs[-1]['punched_at'].strftime('%H:%M') if outs else None
        result.append({
            'id': s['id'], 'name': s['name'],
            'department': s['department'], 'position': s['position_title'],
            'clock_in': clock_in, 'clock_out': clock_out,
            'status': 'present' if clock_in else 'absent',
        })
    return jsonify(result)

# ── Admin: Leave Requests ──────────────────────────────────────────────────────

@app.route('/api/mobile/admin/leaves', methods=['GET'])
@mobile_admin_required
def mobile_admin_leaves():
    status = request.args.get('status', 'pending')
    with get_db() as conn:
        rows = conn.execute(
            """SELECT lr.id, ps.name AS staff_name, lt.name AS leave_type,
                      lr.start_date, lr.end_date, lr.days, lr.reason, lr.status, lr.created_at
               FROM leave_requests lr
               JOIN punch_staff ps ON lr.staff_id = ps.id
               JOIN leave_types lt ON lr.leave_type_id = lt.id
               WHERE (%s = '' OR lr.status = %s)
               ORDER BY lr.created_at DESC LIMIT 50""",
            (status, status)
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        for k in ('start_date','end_date','created_at'):
            if d.get(k): d[k] = str(d[k])
        data.append(d)
    return jsonify(data)

@app.route('/api/mobile/admin/leaves/<int:lid>', methods=['PUT'])
@mobile_admin_required
def mobile_admin_leave_action(lid):
    from flask import g
    b = request.get_json(force=True) or {}
    action = b.get('action')  # 'approve' | 'reject'
    if action not in ('approve', 'reject'):
        return jsonify({'error': '無效操作'}), 400
    status = 'approved' if action == 'approve' else 'rejected'
    reviewer = g.mobile_user.get('display_name', g.mobile_user.get('username', ''))
    with get_db() as conn:
        conn.execute(
            "UPDATE leave_requests SET status=%s, reviewed_by=%s, reviewed_at=NOW() WHERE id=%s",
            (status, reviewer, lid)
        )
    return jsonify({'ok': True})

# ── Admin: Overtime Requests ───────────────────────────────────────────────────

@app.route('/api/mobile/admin/overtime', methods=['GET'])
@mobile_admin_required
def mobile_admin_overtime():
    status = request.args.get('status', 'pending')
    with get_db() as conn:
        rows = conn.execute(
            """SELECT ot.id, ps.name AS staff_name, ot.request_date AS ot_date, ot.ot_hours,
                      ot.reason, ot.status, ot.created_at
               FROM overtime_requests ot
               JOIN punch_staff ps ON ot.staff_id = ps.id
               WHERE (%s = '' OR ot.status = %s)
               ORDER BY ot.created_at DESC LIMIT 50""",
            (status, status)
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        for k in ('ot_date','created_at'):
            if d.get(k): d[k] = str(d[k])
        if d.get('ot_hours'): d['ot_hours'] = float(d['ot_hours'])
        data.append(d)
    return jsonify(data)

@app.route('/api/mobile/admin/overtime/<int:oid>', methods=['PUT'])
@mobile_admin_required
def mobile_admin_overtime_action(oid):
    from flask import g
    b = request.get_json(force=True) or {}
    action = b.get('action')
    if action not in ('approve', 'reject'):
        return jsonify({'error': '無效操作'}), 400
    status = 'approved' if action == 'approve' else 'rejected'
    reviewer = g.mobile_user.get('display_name', g.mobile_user.get('username', ''))
    with get_db() as conn:
        conn.execute(
            "UPDATE overtime_requests SET status=%s, reviewed_by=%s, reviewed_at=NOW() WHERE id=%s",
            (status, reviewer, oid)
        )
    return jsonify({'ok': True})

# ── Admin: Staff List ──────────────────────────────────────────────────────────

@app.route('/api/mobile/admin/staff', methods=['GET'])
@mobile_admin_required
def mobile_admin_staff():
    with get_db() as conn:
        rows = conn.execute(
            """SELECT id, name, username, department, position_title, employee_code,
                      role, active, hire_date
               FROM punch_staff ORDER BY active DESC, name"""
        ).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        if d.get('hire_date'): d['hire_date'] = str(d['hire_date'])
        data.append(d)
    return jsonify(data)

# ── Admin: Anomaly Summary ─────────────────────────────────────────────────────

@app.route('/api/mobile/admin/anomalies', methods=['GET'])
@mobile_admin_required
def mobile_admin_anomalies():
    month = request.args.get('month', _dt.now(TW_TZ).strftime('%Y-%m'))
    try:
        y, m = map(int, month.split('-'))
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400
    import calendar
    total_days = calendar.monthrange(y, m)[1]
    with get_db() as conn:
        staff_all = conn.execute(
            "SELECT id, name, department FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        records = conn.execute(
            """SELECT staff_id, punch_type, punched_at::date AS day
               FROM punch_records
               WHERE date_trunc('month', punched_at) = %s::date""",
            (f'{y}-{m:02d}-01',)
        ).fetchall()
    from collections import defaultdict
    by_staff = defaultdict(set)
    for r in records:
        by_staff[r['staff_id']].add(str(r['day']))

    result = []
    for s in staff_all:
        work_days = len(by_staff[s['id']])
        result.append({
            'id': s['id'], 'name': s['name'], 'department': s['department'],
            'work_days': work_days, 'missing_days': max(0, 22 - work_days),
        })
    return jsonify(result)



# ═══════════════════════════════════════════════════════════════════════════════
# WebAuthn (Face ID / 指紋) — 網頁生物辨識登入
# ═══════════════════════════════════════════════════════════════════════════════
import base64 as _b64
import struct as _struct

# RP_ID 必須與瀏覽器的網域完全一致
_WEBAUTHN_RP_ID   = os.environ.get('WEBAUTHN_RP_ID', 'punch-system.onrender.com')
_WEBAUTHN_RP_NAME = '打卡系統'
_WEBAUTHN_ORIGIN  = os.environ.get('WEBAUTHN_ORIGIN', 'https://punch-system.onrender.com')

def _b64url_encode(data: bytes) -> str:
    return _b64.urlsafe_b64encode(data).rstrip(b'=').decode('ascii')

def _b64url_decode(s: str) -> bytes:
    s = s.replace(' ', '+').replace('-', '+').replace('_', '/')
    padding = 4 - len(s) % 4
    if padding != 4:
        s += '=' * padding
    return _b64.b64decode(s)

def _init_webauthn_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS webauthn_credentials (
                    id            SERIAL PRIMARY KEY,
                    user_key      TEXT NOT NULL,
                    credential_id TEXT NOT NULL UNIQUE,
                    public_key    BYTEA NOT NULL,
                    sign_count    BIGINT DEFAULT 0,
                    device_name   TEXT DEFAULT '',
                    created_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f'[webauthn_init] {e}')

_init_webauthn_db()

# ── Registration Begin ─────────────────────────────────────────────────────────

@app.route('/api/webauthn/register/begin', methods=['POST'])
def webauthn_register_begin():
    """登入後呼叫：產生 WebAuthn 註冊挑戰"""
    # 判斷來源：session admin 或 session staff
    user_key = None
    user_name = None
    user_display = None

    if session.get('logged_in'):
        user_key     = f"admin_{session['admin_id']}"
        user_name    = session.get('admin_username', '')
        user_display = session.get('admin_display_name', user_name)
    elif session.get('punch_staff_id'):
        sid = session['punch_staff_id']
        user_key     = f"staff_{sid}"
        user_name    = session.get('punch_staff_name', str(sid))
        user_display = user_name
    else:
        return jsonify({'error': '請先登入'}), 401

    # 動態偵測實際 origin/rp_id，優先使用環境變數，否則從請求推導
    from urllib.parse import urlparse as _urlparse
    _req_origin = request.headers.get('Origin') or request.url_root.rstrip('/')
    rp_id  = os.environ.get('WEBAUTHN_RP_ID')  or (_urlparse(_req_origin).hostname or _WEBAUTHN_RP_ID)
    origin = os.environ.get('WEBAUTHN_ORIGIN') or _req_origin

    challenge = secrets.token_bytes(32)
    session['webauthn_reg_challenge'] = _b64url_encode(challenge)
    session['webauthn_reg_user_key']  = user_key
    session['webauthn_reg_rp_id']     = rp_id
    session['webauthn_reg_origin']    = origin

    user_id_bytes = user_key.encode('utf-8')

    options = {
        'rp': {'id': rp_id, 'name': _WEBAUTHN_RP_NAME},
        'user': {
            'id': _b64url_encode(user_id_bytes),
            'name': user_name,
            'displayName': user_display,
        },
        'challenge': _b64url_encode(challenge),
        'pubKeyCredParams': [
            {'type': 'public-key', 'alg': -7},    # ES256
            {'type': 'public-key', 'alg': -257},   # RS256
        ],
        'timeout': 60000,
        'authenticatorSelection': {
            'authenticatorAttachment': 'platform',   # 僅使用裝置內建（Face ID / 指紋）
            'userVerification': 'required',
            'residentKey': 'preferred',
        },
        'attestation': 'none',
    }
    return jsonify(options)

# ── Registration Complete ──────────────────────────────────────────────────────

@app.route('/api/webauthn/register/complete', methods=['POST'])
def webauthn_register_complete():
    import json as _json2, hashlib as _hs2
    challenge_b64 = session.get('webauthn_reg_challenge')
    user_key      = session.get('webauthn_reg_user_key')
    if not challenge_b64 or not user_key:
        return jsonify({'error': '找不到挑戰，請重新開始'}), 400

    # 使用 begin 時儲存的 rp_id / origin
    webauthn_rp_id = session.get('webauthn_reg_rp_id', _WEBAUTHN_RP_ID)
    webauthn_origin = session.get('webauthn_reg_origin', _WEBAUTHN_ORIGIN)

    b = request.get_json(force=True) or {}
    try:
        credential_id = b['id']
        resp          = b['response']
        client_data   = _b64url_decode(resp['clientDataJSON'])
        attestation   = _b64url_decode(resp['attestationObject'])
        client_json   = _json2.loads(client_data)

        # Verify clientData
        assert client_json['type'] == 'webauthn.create', 'wrong type'
        recv_challenge = client_json['challenge']
        # normalize both sides
        assert recv_challenge.rstrip('=') == challenge_b64.rstrip('='), 'challenge mismatch'
        assert client_json['origin'] == webauthn_origin, f"origin mismatch: {client_json['origin']}"

        # Parse CBOR attestation object to get public key
        try:
            import cbor2
            att_obj = cbor2.loads(attestation)
        except ImportError:
            # Minimal CBOR parser for attestation (none format)
            att_obj = _minimal_cbor_decode(attestation)

        auth_data = att_obj[b'authData'] if b'authData' in att_obj else att_obj.get('authData', b'')

        # authData layout: rpIdHash(32) + flags(1) + signCount(4) + [AAGUID(16) + credLen(2) + credId + coseKey]
        rp_id_hash = auth_data[:32]
        expected_hash = _hs2.sha256(webauthn_rp_id.encode()).digest()
        assert rp_id_hash == expected_hash, 'rpIdHash mismatch'

        flags = auth_data[32]
        assert flags & 0x01, 'User Presence not set'
        assert flags & 0x04, 'User Verification not set'

        # Extract credential data
        cred_data = auth_data[37:]  # skip rpIdHash + flags + signCount
        aaguid    = cred_data[:16]
        cred_id_len = _struct.unpack('>H', cred_data[16:18])[0]
        cred_id_bytes = cred_data[18:18 + cred_id_len]
        cose_key_bytes = cred_data[18 + cred_id_len:]

        # Verify credential_id matches
        assert _b64url_encode(cred_id_bytes).rstrip('=') == credential_id.rstrip('='), 'credentialId mismatch'

        device_name = b.get('device_name', '我的裝置')
        with get_db() as conn:
            conn.execute("""
                INSERT INTO webauthn_credentials
                  (user_key, credential_id, public_key, sign_count, device_name)
                VALUES (%s, %s, %s, 0, %s)
                ON CONFLICT (credential_id) DO UPDATE
                  SET sign_count=0, device_name=%s
            """, (user_key, credential_id, cose_key_bytes, device_name, device_name))

        session.pop('webauthn_reg_challenge', None)
        session.pop('webauthn_reg_user_key', None)
        session.pop('webauthn_reg_rp_id', None)
        session.pop('webauthn_reg_origin', None)
        return jsonify({'ok': True})

    except Exception as ex:
        return jsonify({'error': f'綁定失敗：{ex}'}), 400

# ── Authentication Begin ───────────────────────────────────────────────────────

@app.route('/api/webauthn/auth/begin', methods=['POST'])
def webauthn_auth_begin():
    b        = request.get_json(force=True) or {}
    username = (b.get('username') or '').strip()

    allow_credentials = []

    if username:
        # Find user_key from username (try admin first, then staff)
        with get_db() as conn:
            admin = conn.execute(
                "SELECT id FROM admin_accounts WHERE username=%s AND active=TRUE", (username,)
            ).fetchone()
            if admin:
                user_key = f"admin_{admin['id']}"
            else:
                staff = conn.execute(
                    "SELECT id FROM punch_staff WHERE username=%s AND active=TRUE", (username,)
                ).fetchone()
                user_key = f"staff_{staff['id']}" if staff else None

        if user_key:
            with get_db() as conn:
                creds = conn.execute(
                    "SELECT credential_id FROM webauthn_credentials WHERE user_key=%s", (user_key,)
                ).fetchall()
            allow_credentials = [{'type': 'public-key', 'id': r['credential_id']} for r in creds]

    if not allow_credentials and not username:
        # Discoverable credential (resident key) — no allowCredentials needed
        pass

    # 動態偵測實際 origin/rp_id
    from urllib.parse import urlparse as _urlparse2
    _req_origin2 = request.headers.get('Origin') or request.url_root.rstrip('/')
    rp_id  = os.environ.get('WEBAUTHN_RP_ID')  or (_urlparse2(_req_origin2).hostname or _WEBAUTHN_RP_ID)
    origin = os.environ.get('WEBAUTHN_ORIGIN') or _req_origin2

    challenge = secrets.token_bytes(32)
    session['webauthn_auth_challenge'] = _b64url_encode(challenge)
    session['webauthn_auth_rp_id']     = rp_id
    session['webauthn_auth_origin']    = origin

    options = {
        'challenge': _b64url_encode(challenge),
        'timeout': 60000,
        'rpId': rp_id,
        'allowCredentials': allow_credentials,
        'userVerification': 'required',
    }
    return jsonify(options)

# ── Authentication Complete ────────────────────────────────────────────────────

@app.route('/api/webauthn/auth/complete', methods=['POST'])
def webauthn_auth_complete():
    import json as _json3, hashlib as _hs3
    challenge_b64 = session.get('webauthn_auth_challenge')
    if not challenge_b64:
        return jsonify({'error': '找不到挑戰，請重新開始'}), 400

    # 使用 begin 時儲存的 rp_id / origin
    webauthn_rp_id  = session.get('webauthn_auth_rp_id', _WEBAUTHN_RP_ID)
    webauthn_origin = session.get('webauthn_auth_origin', _WEBAUTHN_ORIGIN)

    b = request.get_json(force=True) or {}
    try:
        credential_id = b['id']
        resp          = b['response']
        client_data   = _b64url_decode(resp['clientDataJSON'])
        auth_data     = _b64url_decode(resp['authenticatorData'])
        signature     = _b64url_decode(resp['signature'])
        client_json   = _json3.loads(client_data)

        assert client_json['type'] == 'webauthn.get', 'wrong type'
        recv_challenge = client_json['challenge']
        assert recv_challenge.rstrip('=') == challenge_b64.rstrip('='), 'challenge mismatch'
        assert client_json['origin'] == webauthn_origin, f"origin mismatch: {client_json['origin']}"

        # Verify rpIdHash
        rp_id_hash = auth_data[:32]
        assert rp_id_hash == _hs3.sha256(webauthn_rp_id.encode()).digest(), 'rpIdHash mismatch'
        flags = auth_data[32]
        assert flags & 0x01, 'User Presence not set'
        assert flags & 0x04, 'User Verification not set'

        # Lookup credential
        with get_db() as conn:
            cred = conn.execute(
                "SELECT * FROM webauthn_credentials WHERE credential_id=%s", (credential_id,)
            ).fetchone()
        if not cred:
            return jsonify({'error': '找不到已綁定的裝置，請先綁定'}), 401

        # Verify signature using stored COSE public key
        client_data_hash = _hs3.sha256(client_data).digest()
        signed_data = auth_data + client_data_hash
        _verify_cose_signature(cred['public_key'], signed_data, signature)

        # Update sign count
        new_sign_count = _struct.unpack('>I', auth_data[33:37])[0]
        with get_db() as conn:
            conn.execute(
                "UPDATE webauthn_credentials SET sign_count=%s WHERE id=%s",
                (new_sign_count, cred['id'])
            )

        session.pop('webauthn_auth_challenge', None)
        session.pop('webauthn_auth_rp_id', None)
        session.pop('webauthn_auth_origin', None)

        # Create session based on user_key
        user_key = cred['user_key']
        if user_key.startswith('admin_'):
            admin_id = int(user_key[6:])
            with get_db() as conn:
                admin = conn.execute(
                    "SELECT * FROM admin_accounts WHERE id=%s AND active=TRUE", (admin_id,)
                ).fetchone()
            if not admin:
                return jsonify({'error': '帳號不存在或已停用'}), 401
            perms = admin['permissions']
            if isinstance(perms, str):
                try: perms = _json3.loads(perms)
                except: perms = []
            session['logged_in']          = True
            session['admin_id']           = admin['id']
            session['admin_username']     = admin['username']
            session['admin_display_name'] = admin['display_name'] or admin['username']
            session['admin_permissions']  = perms
            session['admin_is_super']     = bool(admin['is_super'])
            return jsonify({'ok': True, 'redirect': '/admin', 'role': 'admin'})

        elif user_key.startswith('staff_'):
            staff_id = int(user_key[6:])
            with get_db() as conn:
                staff = conn.execute(
                    "SELECT id, name, role FROM punch_staff WHERE id=%s AND active=TRUE", (staff_id,)
                ).fetchone()
            if not staff:
                return jsonify({'error': '帳號不存在或已停用'}), 401
            session['punch_staff_id']   = staff['id']
            session['punch_staff_name'] = staff['name']
            return jsonify({'ok': True, 'role': 'staff', 'user': dict(staff)})

        return jsonify({'error': '未知帳號類型'}), 400

    except Exception as ex:
        return jsonify({'error': f'驗證失敗：{ex}'}), 400

# ── 已綁定裝置列表 & 刪除 ────────────────────────────────────────────────────────

@app.route('/api/webauthn/credentials', methods=['GET'])
def webauthn_list_credentials():
    if session.get('logged_in'):
        user_key = f"admin_{session['admin_id']}"
    elif session.get('punch_staff_id'):
        user_key = f"staff_{session['punch_staff_id']}"
    else:
        return jsonify({'error': '請先登入'}), 401
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, device_name, created_at FROM webauthn_credentials WHERE user_key=%s ORDER BY created_at DESC",
            (user_key,)
        ).fetchall()
    return jsonify([{'id': r['id'], 'device_name': r['device_name'],
                     'created_at': str(r['created_at'])} for r in rows])

@app.route('/api/webauthn/credentials/<int:cid>', methods=['DELETE'])
def webauthn_delete_credential(cid):
    if session.get('logged_in'):
        user_key = f"admin_{session['admin_id']}"
    elif session.get('punch_staff_id'):
        user_key = f"staff_{session['punch_staff_id']}"
    else:
        return jsonify({'error': '請先登入'}), 401
    with get_db() as conn:
        conn.execute(
            "DELETE FROM webauthn_credentials WHERE id=%s AND user_key=%s", (cid, user_key)
        )
    return jsonify({'ok': True})

# ── Crypto helpers ─────────────────────────────────────────────────────────────

def _verify_cose_signature(cose_key_bytes: bytes, message: bytes, signature: bytes):
    """驗證 COSE 格式公鑰的簽名（支援 ES256 / RS256）"""
    try:
        import cbor2
        cose = cbor2.loads(cose_key_bytes)
    except ImportError:
        cose = _minimal_cbor_decode(cose_key_bytes)

    from cryptography.hazmat.primitives.asymmetric import ec, padding
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives.asymmetric.utils import decode_dss_signature

    kty = cose.get(1) or cose.get(b'\x01')
    alg = cose.get(3) or cose.get(b'\x03')

    if alg == -7 or kty == 2:  # ES256 / EC2
        x = cose.get(-2) or cose.get(b'\x21') or b''
        y = cose.get(-3) or cose.get(b'\x22') or b''
        pub_numbers = ec.EllipticCurvePublicNumbers(
            x=int.from_bytes(x, 'big'),
            y=int.from_bytes(y, 'big'),
            curve=ec.SECP256R1()
        )
        pub_key = pub_numbers.public_key(default_backend())
        pub_key.verify(signature, message, ec.ECDSA(hashes.SHA256()))

    elif alg == -257 or kty == 3:  # RS256 / RSA
        from cryptography.hazmat.primitives.asymmetric import rsa
        n = cose.get(-1) or cose.get(b'\x20') or b''
        e_bytes = cose.get(-2) or cose.get(b'\x21') or b''
        pub_numbers = rsa.RSAPublicNumbers(
            e=int.from_bytes(e_bytes, 'big'),
            n=int.from_bytes(n, 'big')
        )
        pub_key = pub_numbers.public_key(default_backend())
        pub_key.verify(signature, message, padding.PKCS1v15(), hashes.SHA256())
    else:
        raise ValueError(f'Unsupported alg: {alg}')

def _minimal_cbor_decode(data: bytes) -> dict:
    """極簡 CBOR map decoder（僅處理 attestation none 格式所需）"""
    import io
    buf = io.BytesIO(data)
    return _cbor_read(buf)

def _cbor_read(buf):
    import io
    b0 = ord(buf.read(1))
    major = b0 >> 5
    info  = b0 & 0x1f
    if info <= 23:
        val = info
    elif info == 24:
        val = ord(buf.read(1))
    elif info == 25:
        val = _struct.unpack('>H', buf.read(2))[0]
    elif info == 26:
        val = _struct.unpack('>I', buf.read(4))[0]
    elif info == 27:
        val = _struct.unpack('>Q', buf.read(8))[0]
    else:
        val = 0
    if major == 0:   return val
    if major == 1:   return -1 - val
    if major == 2:   return buf.read(val)      # bytes
    if major == 3:   return buf.read(val).decode('utf-8', errors='replace')  # str
    if major == 4:   return [_cbor_read(buf) for _ in range(val)]  # array
    if major == 5:   return {_cbor_read(buf): _cbor_read(buf) for _ in range(val)}  # map
    if major == 6:   _cbor_read(buf); return None  # tag
    if major == 7:
        if info == 20: return False
        if info == 21: return True
        if info == 22: return None
    return None

